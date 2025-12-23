import logging
import var_stx
import time
from selenium.webdriver.common.by import By
import module_other_stx
import login_stx
import minitor_stx
import vehicle_driver_stx
from retry import retry
import minitor_stx
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import promotion_stx
import re
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from selenium.webdriver.common.keys import Keys
from datetime import datetime, timedelta
wait = WebDriverWait(var_stx.driver, 10)

#19/12
def get_info_web():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 1
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table table-hover table-bordered']/tbody/tr[1]/th[" + n + "]"
        path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[2]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web:" .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)




def croll_path(path):
    n = 0
    while (n < 7):
        n = n + 1
        try:
            var_stx.driver.find_element(By.XPATH, path)
            break
        except:
            scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
            ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
            time.sleep(1.5)

    try:
        var_stx.driver.find_element(By.XPATH, path)
    except:
        n = 0
        while (n < 7):
            n = n + 1
            try:
                var_stx.driver.find_element(By.XPATH, path)
                break
            except:
                scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
                ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(-500, 0).release().perform()
                time.sleep(1.5)



class list_customer:

    def list_customer(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])


        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_customer).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.list_customer).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page)))
            time.sleep(2)
        except:
            pass

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                  var_stx.title_page, "7.1 Danh sách khách hàng", "_DanhSachKhachHang.png")


    def list_customer_g7_test(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx_g7test(self, var_stx.data['login']['tk_admin_test'],  var_stx.data['login']['mk_admin_test'])
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.list_customer).click()
        time.sleep(2.5)


    def list_customer_x(self):
        var_stx.driver.implicitly_wait(0.2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.name_sdt_email).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        except:
            pass
        var_stx.driver.find_element(By.XPATH, var_stx.groupid0).click()
        time.sleep(0.3)
        var_stx.driver.find_element(By.XPATH, var_stx.custrank0).click()
        time.sleep(0.3)
        var_stx.driver.find_element(By.XPATH, var_stx.name_sdt_email).click()
        time.sleep(0.3)


    def list_customer_from_to_day(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")

        list_customer.list_customer_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(var_stx.data['customer']['from_day'])
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(var_stx.data['customer']['to_day'])
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        module_other_stx.write_result_text_try_if_cut1(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                              var_stx.list_data2_7, "22/11/2024", "_DanhSachKhachHang_TuNgayDenNgay.png", -10)


    def list_customer_search(self, code, eventname, result, path_data, path_check, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")

        list_customer.list_customer_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        data = var_stx.driver.find_element(By.XPATH, path_data).text
        var_stx.driver.find_element(By.XPATH, var_stx.name_sdt_email).send_keys(data)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)



        module_other_stx.write_result_text_try_if_in(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                              path_check, data, name_image)


    def list_customer_group(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")

        list_customer.list_customer_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.groupid).click()
        time.sleep(1)
        name_group = var_stx.driver.find_element(By.XPATH, var_stx.groupid2).text
        print("name group"+ name_group)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.groupid2).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(10)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_12).click()
        time.sleep(2.5)

        module_other_stx.write_result_text_try_if_in(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                              var_stx.check_list_customer_group, name_group, "_DanhSachKhachHang_Nhom.png")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.exit).click()
            time.sleep(2)
        except:
            pass


    def list_customer_rank(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")

        list_customer.list_customer_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.custrank).click()
        time.sleep(1.5)
        name_rank = var_stx.driver.find_element(By.XPATH, var_stx.custrank4).text
        print("rank: "+ name_rank)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.custrank4).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(8)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                              var_stx.list_data2_4, name_rank, "_DanhSachKhachHang_Hang.png")


    def list_customer_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer_group(self, "", "", "")

        list_customer.list_customer_x(self)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.groupid2).click()
        time.sleep(1)

        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(10)
        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(13)
        module_other_stx.write_result_dowload_file(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                        "_DanhSachKhachHang.xls", "_DanhSachKhachHang_XuatExcel.png")

        # minitor_stx.get_info_web()
        # try:
        #     minitor_stx.get_info_excel(5, "Sheet 1")
        # except:
        #     var_stx.driver.refresh()
        #     time.sleep(7)
        #     var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        #     time.sleep(7)
        #     minitor_stx.get_info_web()
        #     minitor_stx.get_info_excel(5, "Sheet 1")
        # minitor_stx.check_info_web_excel(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng")
        #
        # var_stx.driver.back()
        # time.sleep(5)


    def list_customer_addnew(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer_rank(self, "", "", "")

        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))

        var_stx.driver.find_element(By.XPATH, var_stx.add_new3).click()
        time.sleep(5)

        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.phone_number)))

        var_stx.driver.find_element(By.XPATH, var_stx.phone_number).send_keys("0835210"+number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.full_name).send_keys("Auto_customer_"+number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.email).send_keys("Auto" + number + "@gmail.com")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.list_customer_addnew_group).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.save).click()
        time.sleep(3)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 28, 2, "Auto_customer_"+number)
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 14, "Auto_customer_"+number)

        print("message: " + "Khách hàng Auto_customer_"+number+" đã được thêm mới thành công")

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                              var_stx.text_success, "Khách hàng Auto_customer_"+number+" đã được thêm mới thành công", "_DanhSachKhachHang_ThemMoi.png")


        try:
            var_stx.driver.find_element(By.XPATH, var_stx.back1).click()
            time.sleep(7)
        except:
            pass

        #Khách hàng Auto_customer163 đã được thêm mới thành công


    def list_customer_lock(self, code, eventname, result, desire, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")


        data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 28, 2))
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[2]//*[text()='"+data+"']")
        except:
            list_customer.list_customer_x(self)
            var_stx.driver.find_element(By.XPATH, var_stx.name_sdt_email).send_keys(data)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(15)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_10_button).click()
        time.sleep(7)
        var_stx.driver.find_element(By.XPATH, var_stx.igree2).click()
        time.sleep(2.5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                        var_stx.message, desire, name_image)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.close1).click()
            time.sleep(12)
        except:
            pass


    def list_customer_get_code(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer_addnew(self, "", "", "")

        data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 28, 2))
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[2]//*[text()='"+data+"']")
        except:
            list_customer.list_customer_x(self)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_11_button).click()
        time.sleep(3)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                        var_stx.popup_message, "Bạn chưa đăng ký mã kích hoạt", "_DanhSachKhachHang_LayMaKichHoat.png")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.exit).click()
            time.sleep(7)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.popup_message)
            var_stx.driver.refresh()
            time.sleep(7)
        except:
            pass



    def list_customer_assign_group(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")

        data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 28, 2))
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[2]//*[text()='"+data+"']")
        except:
            list_customer.list_customer_x(self)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)

        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_12_button).click()
        time.sleep(3)
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, var_stx.list_customer_assign_group_x).click()
        except:
            pass
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.list_customer_addnew_group_select)))
        time.sleep(1)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_customer_addnew_group1_x).click()
            time.sleep(2)
        except:
            pass
        var_stx.driver.find_element(By.XPATH, var_stx.list_customer_addnew_group_select).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.list_customer_addnew_group1).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.igree2).click()
        time.sleep(3)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                        var_stx.message, "Gán nhóm khách hàng thành công!", "_DanhSachKhachHang_GanNhom.png")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.close1).click()
            time.sleep(7)
        except:
            pass


    def list_customer_detail_device(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            # list_customer.list_customer_addnew(self, "", "", "")
            list_customer.list_customer(self, "", "", "")

        data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 28, 2))
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[2]//*[text()='"+data+"']")
        except:
            list_customer.list_customer_x(self)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)

        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_13_button).click()
        time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                        var_stx.check_list_customer_detail_device, "Thông tin thiết bị",
                                                  "_DanhSachKhachHang_ThongTinthietBi.png")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_customer_detail_device_exit).click()
            time.sleep(8)
        except:
            pass


    def list_customer_detail_icon(self, code, eventname, result, path_icon, desire, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")


        data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 28, 2))
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[2]//*[text()='"+data+"']")
        except:

            list_customer.list_customer_x(self)
            var_stx.driver.find_element(By.XPATH, var_stx.name_sdt_email).send_keys(data)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
            try:
                var_stx.driver.find_element(By.XPATH, path_icon)
            except:
                list_customer.list_customer_x(self)
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                time.sleep(5)

        try:
            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, path_icon)))
        except:
            pass


        var_stx.driver.find_element(By.XPATH, path_icon).click()
        time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                  var_stx.title_page3, desire, name_image)



        check_page = var_stx.driver.find_element(By.XPATH, var_stx.title_page3).text
        if check_page == desire:
            var_stx.driver.back()
            time.sleep(7)



    def list_customer_detail_delete(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_customer)
        except:
            list_customer.list_customer(self, "", "", "")


        data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 28, 2))
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[2]//*[text()='"+data+"']")
        except:

            list_customer.list_customer_x(self)
            var_stx.driver.find_element(By.XPATH, var_stx.name_sdt_email).send_keys(data)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)

        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_19_button).click()
        time.sleep(3)
        var_stx.driver.find_element(By.XPATH, var_stx.delete).click()
        time.sleep(1.5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.1 Danh sách khách hàng",
                                                  var_stx.check_list_customer_detail_delete, "Xóa khách hàng thành công!", "_DanhSachKhachHang_LichSuKhoa.png")







class contract_card:



    def partner(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.partner).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.1 Đối tác",
                                                  var_stx.title_page1, "7.7.1 Đối tác", "_Doitac.png")


    def partner_g7_test(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx_g7test(self, var_stx.data['login']['tk_admin_test'],  var_stx.data['login']['mk_admin_test'])
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.partner).click()
        time.sleep(3.5)


    def partner_x(self):
        var_stx.driver.implicitly_wait(0.2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.partner_name).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.code_partner).click()
            time.sleep(0.3)
        except:
            pass


    def partner_search(self, code, eventname, result, path_input, path_data, path_check, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_partner)
        except:
            contract_card.partner(self, "", "", "")

        contract_card.partner_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        data = var_stx.driver.find_element(By.XPATH, path_data).text

        var_stx.driver.find_element(By.XPATH, path_input).send_keys(data)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.1 Đối tác",
                                              path_check, data, name_image)


    def partner_combobox(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_partner)
        except:
            contract_card.partner(self, "", "", "")

        contract_card.partner_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId_select).click()
        time.sleep(1.5)
        name_company = var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId_2).text
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId_2).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.1 Đối tác",
                                              var_stx.list_data2_5, name_company, "_Doitac_CongTy.png")


    def partner_addnew(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_partner)
        except:
            contract_card.partner(self, "", "", "")

        contract_card.partner_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.add_new1).click()
        time.sleep(2.5)

        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))

        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_code).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.3)
        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_code).send_keys("Auto" + number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_name).send_keys("Auto_name_partner_" + number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_phone).send_keys("0835219" + number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_get_email)

        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_email).send_keys(f"Auto_{number}@gmail.com")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.save3).click()
        time.sleep(0.5)
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 14, "Auto_name_partner_" + number)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 29, 2, "Auto_name_partner_"+number)
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        time.sleep(0.5)

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.1 Đối tác",
                                                  var_stx.toast_message, "Thêm mới thông tin đối tác thành công!", "_Doitac_ThemMoi.png")


    def partner_update(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_partner)
        except:
            contract_card.partner_addnew(self, "", "", "")


        contract_card.partner_x(self)
        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))
        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 29, 2))

        var_stx.driver.find_element(By.XPATH, var_stx.partner_name).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_3_a).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_phone).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.partner_addnew_phone).send_keys(" 0995219" + number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.save3).click()
        time.sleep(0.5)
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        time.sleep(0.5)

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.1 Đối tác",
                                                  var_stx.toast_message, "Cập nhật thông tin đối tác thành công!", "_Doitac_CapNhat.png")
        time.sleep(1.5)


    def partner_delete(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_partner)
            var_stx.driver.find_element(By.XPATH, var_stx.STAXI)
        except:
            contract_card.partner_addnew(self, "", "", "")


        contract_card.partner_x(self)
        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))
        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 29, 2))

        var_stx.driver.find_element(By.XPATH, var_stx.partner_name).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_delete).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.delete).click()
        time.sleep(0.5)
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        time.sleep(0.5)

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.1 Đối tác",
                                                  var_stx.toast_message, "Xóa thông tin đối tác thành công!", "_Doitac_Xoa.png")
        time.sleep(1.5)










    def contract(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.2 Hợp đồng",
                                                  var_stx.title_page1, "7.7.2 Hợp đồng", "_HopDong.png")


    def contract_g7_test(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx_g7test(self, var_stx.data['login']['tk_admin_test'],  var_stx.data['login']['mk_admin_test'])
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract).click()
        time.sleep(3.5)


    def contract_x(self):
        var_stx.driver.implicitly_wait(0.2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.code_contract).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.code_partner).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.partner_name).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId).click()
            time.sleep(0.3)
        except:
            pass

    def contract_search(self, code, eventname, result, path_data, path_input, path_check, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_contract)
        except:
            contract_card.contract(self, "", "", "")

        contract_card.contract_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        data = var_stx.driver.find_element(By.XPATH, path_data).text

        var_stx.driver.find_element(By.XPATH, path_input).send_keys(data)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.2 Hợp đồng",
                                              path_check, data, name_image)


    def contract_combobox(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_contract)
        except:
            contract_card.contract(self, "", "", "")

        contract_card.contract_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId_select).click()
        time.sleep(1.5)
        name_company = var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId_2).text
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId_2).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(6)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.2 Hợp đồng",
                                              var_stx.list_data2_3, name_company, "_HopDong_CongTy.png")


    def contract_addnew(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_contract)
        except:
            contract_card.contract(self, "", "", "")

        contract_card.partner_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.add_new1).click()
        time.sleep(3)

        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_contract_icon).click()
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_contract).click()
            time.sleep(1.5)
        except:
            pass

        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_code).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_code).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_code).send_keys("Auto" + number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_sdt).send_keys("083521" + number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_hd).send_keys("100.000.000")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_kyquy).send_keys("10.000.000")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_status).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_image).send_keys("https://g7test.staxi.vn/Media/Uploads/%E2%80%94Pngtree%E2%80%94car_flag_icon_333837.png")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_company).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_note).send_keys("Test ghi chú tạo hợp đồng")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_tk_company).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_get_email)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_auto_export)
        var_stx.driver.find_element(By.XPATH, var_stx.search_mst)

        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_mst).send_keys("350080"+number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_name_company).send_keys("Công ty test số "+number)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_adress).send_keys("Lô 14 Nguyễn Cảnh Dị, Đại Kim, Hoàng Mai, Hà Nội")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_email).send_keys(f"Auto_{number}@gmail.com")
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_addnew_email1).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.save_1).click()
        time.sleep(1.5)
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 14, "Auto" + number)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 30, 2, "Auto" + number)

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.2 Hợp đồng",
                                                  var_stx.contract_addnew, "Thêm mới hợp đồng thành công", "_HopDong_ThemMoi.png")


    def contract_update(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_contract)
        except:
            contract_card.contract_addnew(self, "", "", "")


        contract_card.contract_x(self)
        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))
        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 30, 2))

        var_stx.driver.find_element(By.XPATH, var_stx.code_contract).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_2_a1).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys("035921" + number)
        print("095219" + number)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.save_1).click()
        time.sleep(0.5)
        try:
            var_stx.driver.implicitly_wait(0.3)
            button = var_stx.driver.find_element(By.XPATH, var_stx.save_1)
            var_stx.driver.execute_script("arguments[0].click();", button)
            time.sleep(0.5)
        except:
            pass
        try:
            var_stx.driver.implicitly_wait(0.3)
            button = var_stx.driver.find_element(By.XPATH, var_stx.save_1_a)
            var_stx.driver.execute_script("arguments[0].click();", button)
            time.sleep(0.5)
        except:
            pass


        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.contract_update)))
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.2 Hợp đồng",
                                                  var_stx.contract_update, "Cập nhật hợp đồng thành công", "_HopDong_CapNhat.png")
        time.sleep(1.5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.close_contact).click()
            time.sleep(3)
        except:
            pass



    def contract_see_file(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_contract)
        except:
            contract_card.contract_addnew(self, "", "", "")


        contract_card.contract_x(self)
        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 30, 2))

        var_stx.driver.find_element(By.XPATH, var_stx.code_contract).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(6)
        var_stx.driver.find_element(By.XPATH, var_stx.list_data2_4_button).click()
        time.sleep(2.5)
        module_other_stx.write_result_text_try_if_src(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.2 Hợp đồng",
                                              var_stx.contract_see_file, "https://g7test.staxi.vn/Media/Uploads/%E2%80%94Pngtree%E2%80%94car_flag_icon_333837.png",
                                                      "_HopDong_XemFile.png")

        try:
            button = var_stx.driver.find_element(By.XPATH, var_stx.contract_see_file_close)
            var_stx.driver.execute_script("arguments[0].click();", button)
            time.sleep(5)
        except:
            pass


    def contract_delete(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_contract)
        except:
            contract_card.contract_addnew(self, "", "", "")
            # contract_card.contract(self, "", "", "")



        contract_card.contract_x(self)
        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 30, 2))

        var_stx.driver.find_element(By.XPATH, var_stx.code_contract).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(6)
        check_name = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_2).text
        if name == check_name:
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2_delete).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.delete).click()
            time.sleep(0.5)
            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.contract_delete)))
            time.sleep(0.5)
            module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.2 Hợp đồng",
                                                      var_stx.contract_delete, "Xóa hợp đồng thành công", "_HopDong_Xoa.png")
            time.sleep(1.5)










    def card_management(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.title_page1)
        except:
            var_stx.driver.refresh()
            time.sleep(5)


        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.title_page1, "7.7.3 Quản lý thẻ", "_QuanLyThe.png")


    def card_management_g7_test(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx_g7test(self, var_stx.data['login']['tk_admin_test'],  var_stx.data['login']['mk_admin_test'])
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management).click()
        time.sleep(5)


    def card_management_x(self):
        var_stx.driver.implicitly_wait(0.2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.code_hd).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId).click()
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.CardControl).click()
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.StatusSearch).click()
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.PaidMethod).click()
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.StatusBFA).click()
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.Status_1time).click()
            time.sleep(0.3)
        except:
            pass



    def card_management_x1(self):
        var_stx.driver.implicitly_wait(0.2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.code_hd).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass

        var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId).click()
        time.sleep(0.3)
        var_stx.driver.find_element(By.XPATH, var_stx.CardControl).click()
        time.sleep(0.3)
        var_stx.driver.find_element(By.XPATH, var_stx.StatusSearch).click()
        time.sleep(0.3)
        var_stx.driver.find_element(By.XPATH, var_stx.PaidMethod).click()
        time.sleep(0.3)
        var_stx.driver.find_element(By.XPATH, var_stx.StatusBFA).click()
        time.sleep(0.3)

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys("0987")
            time.sleep(0.3)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        except:
            pass


    def card_management_search(self, code, eventname, result, path_input, path_data, path_check, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_page2)
            contract_card.card_management_x1(self)
        except:
            pass

        if name_image == "_QuanLyThe_SoDienThoai.png":
            contract_card.card_management_x(self)

        else:
            contract_card.card_management_x(self)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(6)

        try:
            data = var_stx.driver.find_element(By.XPATH, path_data)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys("09")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)

        data = var_stx.driver.find_element(By.XPATH, path_data).text
        var_stx.driver.find_element(By.XPATH, path_input).send_keys(data)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                              path_check, data, name_image)


    def card_management_company(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")

        # try:
        #     var_stx.driver.find_element(By.XPATH, var_stx.check_page2)
        #     contract_card.card_management_x1(self)
        # except:
        #     pass

        contract_card.card_management_x(self)

        var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId).click()
        time.sleep(1)
        name = var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId1).text
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchCompanyId1).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)


        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                              var_stx.col_id_company2, name, "_QuanLyThe_CongTy.png")


    def card_management_status(self, code, eventname, result, path_card_management_status, path_check, desire, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")


        # try:
        #     var_stx.driver.find_element(By.XPATH, var_stx.check_page2)
        #     contract_card.card_management_x1(self)
        # except:
        #     pass

        contract_card.card_management_x(self)

        var_stx.driver.find_element(By.XPATH, path_card_management_status).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        croll_path(path_check)

        # scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        # time.sleep(1.5)
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        # time.sleep(1.5)

        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            check_text = var_stx.driver.find_element(By.XPATH, path_check).get_attribute("title")
            logging.info(check_text)
            logging.info(desire)
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)
            if check_text == desire:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)


    def card_management_card(self, code, eventname, result, card_management_card, path_check, desire, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")

        # try:
        #     var_stx.driver.find_element(By.XPATH, var_stx.check_page2)
        #     contract_card.card_management_x1(self)
        # except:
        #     pass

        contract_card.card_management_x(self)

        var_stx.driver.find_element(By.XPATH, card_management_card).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        croll_path(path_check)
        #
        # scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        time.sleep(1.5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                              path_check, desire, name_image)


    def card_management_qr(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.col_id_serial2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.name_customer).clear()
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys("auto")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(8)


        var_stx.driver.find_element(By.XPATH, var_stx.card_management_qr).click()
        try:
            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        except:
            pass
        module_other_stx.write_result_text_try_if_in(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                              var_stx.toast_message, "Thành công lưu lại:", "_QuanLyThe_QRCode.png")


    def card_management_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")


        try:
            data = var_stx.driver.find_element(By.XPATH, var_stx.col_id_mobile2).text
            var_stx.driver.find_element(By.XPATH, var_stx.MobileSearch).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            var_stx.driver.find_element(By.XPATH, var_stx.MobileSearch).send_keys(data)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
            time.sleep(5)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
            time.sleep(10)
        module_other_stx.write_result_dowload_file(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                        "_QuanLyThe.xlsx", "_QuanLyThe_XuatExcel.png")




    def card_management_card_1time(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")

        contract_card.card_management_x(self)

        var_stx.driver.find_element(By.XPATH, var_stx.PaidMethod_1time).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        croll_path(var_stx.col_id_CardId_1)
        # scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        # time.sleep(1.5)
        module_other_stx.write_result_displayed_try(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                              var_stx.col_id_CardId_1, "_QuanLyThe_CapThe1Lan.png")


    def card_management_addnew(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")


        contract_card.partner_x(self)
        try:
            # var_stx.driver.find_element(By.XPATH, var_stx.add_new1).click()
            element = var_stx.driver.find_element(By.XPATH, var_stx.card_add_new)
            var_stx.driver.execute_script("arguments[0].click();", element)
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            # var_stx.driver.find_element(By.XPATH, var_stx.card_add_new).click()
            element = var_stx.driver.find_element(By.XPATH, var_stx.card_add_new)
            var_stx.driver.execute_script("arguments[0].click();", element)

        time.sleep(3)
        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.info_card_serial).click()
        except:
            element = var_stx.driver.find_element(By.XPATH, var_stx.card_add_new)
            var_stx.driver.execute_script("arguments[0].click();", element)

            # var_stx.driver.find_element(By.XPATH, var_stx.card_add_new).click()
            time.sleep(3)
            var_stx.driver.find_element(By.XPATH, var_stx.info_card_serial).click()

        time.sleep(0.5)

        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_code_hd).click()
        time.sleep(0.5)

        var_stx.driver.find_element(By.XPATH, var_stx.info_card_name).clear()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_name).send_keys("Auto_" + number)
        time.sleep(0.5)

        var_stx.driver.find_element(By.XPATH, var_stx.info_card_day1).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_day2).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_serial).click()
        time.sleep(0.5)
        # var_stx.driver.find_element(By.XPATH, var_stx.info_card_pay).click()
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_pay1).click()
        time.sleep(0.5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.info_card_money1).send_keys("1.000.000")
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.info_card_money)

        var_stx.driver.find_element(By.XPATH, var_stx.LimitMoneyTrans).send_keys("1.500.000")

        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_link_app).click()
        time.sleep(0.5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.info_card_dat_ho).click()
        except:
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, "Mất checkbox Thanh toán đặt hộ")


        var_stx.driver.find_element(By.XPATH, var_stx.info_card_sdt).send_keys("0835210360")#0359123123
        # var_stx.driver.find_element(By.XPATH, var_stx.info_card_sdt).send_keys("0902200780")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_rank_card).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_confirm).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_rank_note).send_keys("Trường test Tạo mới thẻ")
        time.sleep(1.5)


        try:
            var_stx.driver.find_element(By.XPATH, var_stx.save_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.add_new_card_confirm).click()

        try:
            wait = WebDriverWait(var_stx.driver, 5)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))#Cập nhật thông tin thẻ thành công
            module_other_stx.write_result_text_try_if_or(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                             var_stx.toast_message, "Cập nhật thông tin thẻ thành công", "Thêm mới thẻ thành công!", "_QuanLyThe_ThemMoi.png")

        except:
            pass


        try:
            var_stx.driver.implicitly_wait(3)
            var_stx.driver.find_element(By.XPATH, var_stx.confirm2).click()
            wait = WebDriverWait(var_stx.driver, 5)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))#Cập nhật thông tin thẻ thành công
            module_other_stx.write_result_text_try_if_or(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                         var_stx.toast_message, "Cập nhật thông tin thẻ thành công", "Thêm mới thẻ thành công!", "_QuanLyThe_ThemMoi.png")
        except:
            pass

        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 14, "Auto_" + number)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 31, 2, "Auto_" + number)
        # try:
        #     wait = WebDriverWait(var_stx.driver, 10)
        #     element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.alert)))
        # except:
        #     pass
        # module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
        #                                           var_stx.card_management_addnew, "Thêm mới thẻ thành công!", "_QuanLyThe_ThemMoi.png")

        # try:
        #     wait = WebDriverWait(var_stx.driver, 5)
        #     element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))#Cập nhật thông tin thẻ thành công
        # except:
        #     pass
        # module_other_stx.write_result_text_try_if_or(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
        #                                           var_stx.toast_message, "Cập nhật thông tin thẻ thành công", "Thêm mới thẻ thành công!", "_QuanLyThe_ThemMoi.png")


        try:
            var_stx.driver.find_element(By.XPATH, var_stx.card_management_addnew_x).click()
            time.sleep(2)
        except:
            pass





    def card_management_update(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management_addnew(self, "", "", "")
            # contract_card.card_management(self, "", "", "")


        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))
        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 31, 2))
        var_stx.driver.refresh()
        time.sleep(5)
        contract_card.card_management_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)
        var_stx.driver.find_element(By.XPATH, var_stx.col_id_serial2_a).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_rank_note).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.info_card_rank_note).send_keys("Trường test cập nhật thẻ")
        time.sleep(1.5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.save_1).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.add_new_card_confirm).click()
        time.sleep(1)
        print("n0")
        try:
            var_stx.driver.implicitly_wait(0.3)
            var_stx.driver.find_element(By.XPATH, var_stx.confirm2).click()
            print("n1")
        except:
            pass

        try:
            print("n2")
            wait = WebDriverWait(var_stx.driver, 5)
            print("n3")
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))#Cập nhật thông tin thẻ thành công
            print("n4")
            print(element)
            print("n5")
        except:
            pass#card_management_update

        module_other_stx.write_result_text_try_if_or(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.toast_message, "Cập nhật thông tin thẻ thành công", "Cập nhật thẻ thành công!", "_QuanLyThe_CapNhat.png")
        time.sleep(1.5)


    def card_management_see_file(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management_addnew(self, "", "", "")

        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 31, 2))

        contract_card.card_management_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)
        var_stx.driver.find_element(By.XPATH, var_stx.col_id_Serial_2).click()
        time.sleep(2.5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.card_management_see_file, "MÃ QRCODE", "_QuanLyThe_XemQrCode.png")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.qr_x).click()
            time.sleep(2)
        except:
            pass


    def card_management_clock(self, code, eventname, result, path_check, desire1, desire2, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management_addnew(self, "", "", "")

        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 31, 2))

        contract_card.card_management_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)
        croll_path(var_stx.col_id_IsLocked2)

        var_stx.driver.find_element(By.XPATH, var_stx.col_id_IsLocked2).click()
        time.sleep(2.5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.reason_clock).send_keys(var_stx.data['customer']['reason_lock'])
        except:
            pass
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.thuchien).click()
        time.sleep(0.2)
        try:
            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, path_check)))
        except:
            pass
        module_other_stx.write_result_text_try_if_or(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  path_check, desire1, desire2, name_image)


    def card_management_recharge(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management_addnew(self, "", "", "")
            # contract_card.card_management(self, "", "", "")

        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 31, 2))

        contract_card.card_management_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)

        croll_path(var_stx.col_id_recharge0)

        var_stx.driver.find_element(By.XPATH, var_stx.col_id_recharge0).click()
        time.sleep(1)
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.recharge_card_money)))
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.recharge_card_money).send_keys(var_stx.data['customer']['recharge_card_money'])
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.recharge_card_note).send_keys(var_stx.data['customer']['recharge_card_note'])
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management_recharge_save).click()
        time.sleep(2)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.card_management_recharge, "Thêm tiền vào thẻ thành công!", "_QuanLyThe_NapTien.png")


    def card_management_recharge_km(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management_addnew(self, "", "", "")

        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 31, 2))

        contract_card.card_management_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)

        croll_path(var_stx.col_id_rechargekm0)


        var_stx.driver.find_element(By.XPATH, var_stx.col_id_rechargekm0).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.recharge_km_money).send_keys(var_stx.data['customer']['recharge_card_money'])
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.recharge_km_note).send_keys(var_stx.data['customer']['recharge_km_note'])
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management_recharge_km_save).click()
        time.sleep(1.5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.check_card_management_recharge_km, "Chỉnh sửa tiền khuyến mại thành công!", "_QuanLyThe_NapTienKm.png")


    def card_management_active(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management_addnew(self, "", "", "")

        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 31, 2))

        contract_card.card_management_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)

        #
        # scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        # time.sleep(1.5)
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        # time.sleep(1.5)
        croll_path(var_stx.col_id_status2_a)

        var_stx.driver.find_element(By.XPATH, var_stx.col_id_status2_a).click()
        time.sleep(4)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management_active1).click()
        time.sleep(0.2)
        try:
            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.toast_message, "Kích hoạt thẻ thành công", "_QuanLyThe_KichHoat.png")


    def card_management_cancel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            # contract_card.card_management(self, "", "", "")
            contract_card.card_management_addnew(self, "", "", "")

        name = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 31, 2))

        contract_card.card_management_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.name_customer).send_keys(name)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)
        # scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        # time.sleep(1.5)
        # ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
        # time.sleep(1.5)
        # check_name = var_stx.driver.find_element(By.XPATH, var_stx.col_id_namecustomer2).text
        # if check_name == name:
        croll_path(var_stx.col_id_CardId_3)
        var_stx.driver.find_element(By.XPATH, var_stx.col_id_CardId_3).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.delete).click()
        try:
            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        except:
            pass
        module_other_stx.write_result_text_try_if_or(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.toast_message, "Xóa thẻ thành công", "Xóa thẻ thành công!", "_QuanLyThe_Xoa.png")




    def issue_white_cards(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")


        contract_card.partner_x(self)
        try:
            element = var_stx.driver.find_element(By.XPATH, var_stx.issue_white_cards)
            var_stx.driver.execute_script("arguments[0].click();", element)
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            element = var_stx.driver.find_element(By.XPATH, var_stx.issue_white_cards)
            var_stx.driver.execute_script("arguments[0].click();", element)

        time.sleep(3)
        vehicle_driver_stx.increase()
        number = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 7, 2))

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.IncreaseValue).click()
            time.sleep(0.5)
        except:
            element = var_stx.driver.find_element(By.XPATH, var_stx.issue_white_cards)
            var_stx.driver.execute_script("arguments[0].click();", element)
            time.sleep(3)
            var_stx.driver.find_element(By.XPATH, var_stx.IncreaseValue).click()
            time.sleep(0.5)

        # var_stx.driver.find_element(By.XPATH, var_stx.serial_before)
        # var_stx.driver.find_element(By.XPATH, var_stx.serial_after)

        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_hd).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_hd1).click()
        time.sleep(1)

        var_stx.driver.find_element(By.XPATH, var_stx.CustomerName).send_keys("Auto_" + number)
        time.sleep(1)

        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_start_date)
        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_start_end)

        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_pay).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_authentic).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_note).send_keys("Trường test tạo thẻ trắng")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_confirm).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.white_cards_confirm1).click()
        time.sleep(1)
        # wait = WebDriverWait(var_stx.driver, 10)
        # element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.card_management_addnew1)))

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.card_management_addnew1, "Thêm mới thẻ thành công", "_QuanLyThe_PhatHanhTheTrang.png")
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 14, "Auto_" + number)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 37, 2, "Auto_" + number)

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_card_while)
            var_stx.driver.find_element(By.XPATH, var_stx.HideDataTableCardWhite).click()
            time.sleep(1.5)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.list_card_while_cancel).click()
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.list_card_while_cancel1).click()
            time.sleep(1.5)

    def one_time_card(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            contract_card.card_management(self, "", "", "")


        contract_card.partner_x(self)
        try:
            element = var_stx.driver.find_element(By.XPATH, var_stx.one_time_card)
            var_stx.driver.execute_script("arguments[0].click();", element)
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.check_one_time_card)
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            element = var_stx.driver.find_element(By.XPATH, var_stx.one_time_card)
            var_stx.driver.execute_script("arguments[0].click();", element)
        time.sleep(5)

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.check_one_time_card, "Cấp thẻ dùng 1 lần", "_QuanLyThe_CapTheDung1Lan.png")

    def one_time_card_coppy(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_one_time_card)
        except:
            contract_card.one_time_card(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_coppy).click()
        time.sleep(0.2)
        try:
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_hanoi)))
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.toast_message, "Sao chép thông tin thành công", "_CapTheDung1Lan_CoppyThongTinThe.png")

    def one_time_card_print(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_one_time_card)
        except:
            contract_card.one_time_card(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_print).click()
        time.sleep(5)
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[1])
        time.sleep(1)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.id_qr, "", "_CapTheDung1Lan_In.png")

        module_other_stx.close_tab()
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(1)

    def one_time_card_save(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_one_time_card)
        except:
            contract_card.one_time_card(self, "", "", "")

        serial = var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_serial).text

        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_type_qr).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_name).send_keys("Trần Quang Trường")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_phone).send_keys("0359667692")
        time.sleep(0.5)
        now = datetime.now()

        # Luôn lớn hơn hiện tại 1 ngày
        xung = now + timedelta(days=1)
        tomorow = xung.strftime("%d/%m/%Y %H:%M")

        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_start_date)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_end_date).send_keys(Keys.CONTROL + "a")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_end_date).send_keys(tomorow)
        time.sleep(1)

        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_max).send_keys("3.000.000")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_count).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_count).send_keys("2")
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_authentic)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_pin)
        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_note).send_keys("Trường test thêm mới cấp thẻ 1 lần")
        time.sleep(0.5)
        var_stx.driver.save_screenshot(var_stx.imagepath + code + "_anh1.png")

        var_stx.driver.find_element(By.XPATH, var_stx.one_time_card_save).click()
        time.sleep(0.2)
        var_stx.driver.save_screenshot(var_stx.imagepath + code + "_anh2.png")
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        except:
            pass

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.3 Quản lý thẻ",
                                                  var_stx.toast_message, "Cập nhật thẻ thành công!", "_QuanLyThe_CapThe1Lan.png")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + "_anh3.png")
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 14, serial)




    def closing_debts(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                                  var_stx.title_page1, "7.7.4 Chốt công nợ", "_ChotCongNo.png")


    def closing_debts_x(self):
        var_stx.driver.implicitly_wait(0.2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.code_hd).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("20/10/2024 00:00")
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.3)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("20/12/2024 23:59")
            time.sleep(0.3)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.icon_search1).click()
        time.sleep(3.5)


    def closing_debts_g7_test(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx_g7test(self, var_stx.data['login']['tk_admin_test'],  var_stx.data['login']['mk_admin_test'])
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts).click()
        time.sleep(5)


    def closing_debts_search(self, code, eventname, result, type_check, path_data, path_check, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_closing_debts)
        except:
            contract_card.closing_debts(self, "", "", "")

        contract_card.closing_debts_x(self)


        try:
            var_stx.driver.find_element(By.XPATH, path_data)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("31/03/2023 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("31/03/2026 00:00")
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            except:
                var_stx.driver.find_element(By.XPATH, var_stx.icon_search1).click()
            time.sleep(3.5)



        if type_check == "0":
            data = var_stx.driver.find_element(By.XPATH, path_data).text
            if name_image == "_ChotCongNo_MaHopDong.png":
                var_stx.driver.find_element(By.XPATH, var_stx.code_hd).send_keys(data)

            if name_image == "_ChotCongNo_SoDienThoai.png":
                var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(data)

            time.sleep(1)
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            except:
                var_stx.driver.find_element(By.XPATH, var_stx.icon_search1).click()
            time.sleep(3.5)
            module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                                  path_check, data, name_image)

        if type_check == "1":
            module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                                  path_check, "", name_image)


    def closing_debts_reset(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_closing_debts)
        except:
            contract_card.closing_debts(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_reset).click()
        time.sleep(0.5)
        module_other_stx.write_result_displayed_try(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                              var_stx.closing_debts_reset, "_ChotCongNo_DatLai.png")


    def closing_debts_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_closing_debts)
        except:
            contract_card.closing_debts(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_2)
        except:
            contract_card.closing_debts_search(self, "", "", "", "0", var_stx.ag1_3, "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel1).click()
        time.sleep(5)
        module_other_stx.write_result_dowload_file(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                                        "_ChotCongNo.xls", "_DanhSachKhachHang_XuatExcel.png")

        # get_info_web()
        # try:
        #     minitor_stx.get_info_excel1(5, "Sheet 1")
        # except:
        #     var_stx.driver.refresh()
        #     time.sleep(7)
        #     contract_card.closing_debts_search(self, "", "", "", "0", var_stx.ag1_3, "", "")
        #
        #     var_stx.driver.find_element(By.XPATH, var_stx.export_excel1).click()
        #     time.sleep(7)
        #     get_info_web()
        #     minitor_stx.get_info_excel1(5, "Sheet 1")
        # minitor_stx.check_info_web_excel(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ")

        # var_stx.driver.back()
        # time.sleep(5)


    def closing_debts_link(self, code, eventname, result, button, path_check, desire, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_closing_debts)
        except:
            contract_card.closing_debts(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_2)
        except:
            contract_card.closing_debts_search(self, "", "", "", "0", var_stx.ag1_2, "", "")

        var_stx.driver.find_element(By.XPATH, button).click()
        time.sleep(10)
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[1])
        time.sleep(2)
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                              path_check, desire, name_image)

        module_other_stx.close_tab()
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(1)


    def closing_debts_closing_debts(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_closing_debts)
        except:
            contract_card.closing_debts(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.ag1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("31/03/2023 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("31/03/2026 00:00")
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            except:
                var_stx.driver.find_element(By.XPATH, var_stx.icon_search1).click()

        # contract_card.closing_debts_search(self, "", "", "", "0", var_stx.col_id_mobile3, var_stx.col_id_mobile2, "")




        time.sleep(3)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_closing_debts2_input1).click()
        time.sleep(1)

        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_closing_debts2).click()
        time.sleep(2.5)

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_note).click()
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_closing_debts2_input1).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_closing_debts2).click()
            time.sleep(2.5)
            var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_note).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_note).send_keys(var_stx.data['customer']['closing_debts_note'])
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_mail).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_mail_input).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_mail_input).send_keys(var_stx.data['customer']['closing_debts_email'])
        time.sleep(0.5)

        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts_day).click()
        time.sleep(0.5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                                  var_stx.check_closing_debts_closing_debts, "", "_ChotCongNo_ChotCongNo.png")





    def closing_debts_closing_debts_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_closing_debts_closing_debts)
        except:
            contract_card.closing_debts_closing_debts(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel7).click()
        time.sleep(5)
        module_other_stx.write_result_dowload_file(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.4 Chốt công nợ",
                                                        "ChotCongNo_ChotCongNo.xls", "_ChotCongNo_ChotCongNo_XuatExcel.png")


        try:
            name = var_stx.driver.find_element(By.XPATH, var_stx.not_role).text
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 14, name)
            var_stx.driver.back()
            time.sleep(5)
        except:
            pass


        try:
            var_stx.driver.find_element(By.XPATH, var_stx.close).click()
            time.sleep(2)
        except:
            pass










    def report_customer_card_details(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_customer_card_details).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.10 Báo cáo chi tiết cuốc khách thẻ",
                                                  var_stx.title_page1, "7.7.10 Báo cáo chi tiết cuốc khách thẻ", "_BaoCaoCHiTietCuocKhachNo.png")


    def report_customer_card_details_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_customer_card_details)
        except:
            contract_card.report_customer_card_details(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.month_before).click()
        # time.sleep(1.5)
        # var_stx.driver.find_element(By.XPATH, var_stx.moth_).click()
        # time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.10 Báo cáo chi tiết cuốc khách thẻ",
                                              var_stx.presentation1_2, "", "_BaoCaoCHiTietCuocKhachNo_TimKiem.png")


    def report_customer_card_details_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_customer_card_details)
        except:
            contract_card.report_customer_card_details_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel_1).click()
        time.sleep(7)
        try:
            # minitor_stx.get_info_web()
            # minitor_stx.get_info_excel1(5, "Sheet 1")
            # minitor_stx.check_info_web_excel(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.10 Báo cáo chi tiết cuốc khách thẻ")
            module_other_stx.write_result_dowload_file(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.10 Báo cáo chi tiết cuốc khách thẻ",
                                                            "_BaoCaoCHiTietCuocKhachNo_Excel.xlsx", "__BaoCaoCHiTietCuocKhachNo_XuatExcel.png")

        except:
            nodata = var_stx.driver.find_element(By.XPATH, var_stx.not_role).text
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, nodata)
            var_stx.driver.back()
            time.sleep(4)


    def report_customer_card_details_excel_full(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_customer_card_details)
        except:
            contract_card.report_customer_card_details_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel_2).click()
        time.sleep(7)
        try:
            nodata = var_stx.driver.find_element(By.XPATH, var_stx.not_role).text
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, nodata)
            var_stx.driver.back()
            time.sleep(4)
        except:
            module_other_stx.write_result_dowload_file(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.10 Báo cáo chi tiết cuốc khách thẻ",
                                                            "_BaoCaoCHiTietCuocKhachNo_ExcelFull.xlsx", "__BaoCaoCHiTietCuocKhachNo_XuatExcelDayDuThongTin.png")








    def report_card_top_up_transactions(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_card_top_up_transactions).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.11 Báo cáo giao dịch nạp thẻ",
                                                  var_stx.title_page1, "7.7.11 Báo cáo giao dịch nạp thẻ", "_BaoCaoGiaoDichNapThe.png")


    def report_card_top_up_transactions_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_card_top_up_transactions)
        except:
            contract_card.report_card_top_up_transactions(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(2.5)


        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("01/07/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("01/08/2025 11:27")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.11 Báo cáo giao dịch nạp thẻ",
                                              var_stx.list_data2_2, "", "_BaoCaoGiaoDichNapThe_TimKiem.png")


    def report_card_top_up_transactions_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_card_top_up_transactions)
        except:
            contract_card.report_card_top_up_transactions_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.serial).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        try:
            nodata = var_stx.driver.find_element(By.XPATH, var_stx.not_role).text
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, nodata)
            var_stx.driver.back()
            time.sleep(4)
        except:
            minitor_stx.get_info_web()
            minitor_stx.get_info_excel1(5, "Sheet 1")
            minitor_stx.check_info_web_excel(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.10 Báo cáo chi tiết cuốc khách thẻ")









    def report_the_trip_with_the_driver_card(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_the_trip_with_the_driver_card).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.12 Báo cáo cuốc đi thẻ theo lái xe",
                                                  var_stx.title_page1, "7.7.12 Báo cáo cuốc đi thẻ theo lái xe", "_BaoCaoCuocDiTheTheoLaiXe.png")


    def report_the_trip_with_the_driver_card_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_the_trip_with_the_driver_card)
        except:
            contract_card.report_the_trip_with_the_driver_card(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("22/10/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("28/10/2025 10:17")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
        time.sleep(1.5)
        button = var_stx.driver.find_element(By.XPATH, var_stx.search)
        var_stx.driver.execute_script("arguments[0].click();", button)
        time.sleep(3)

        try:
            wait = WebDriverWait(var_stx.driver, 15)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.col_id_serial2)))
        except:
            pass

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.12 Báo cáo cuốc đi thẻ theo lái xe",
                                              var_stx.col_id_serial2, "", "_BaoCaoCuocDiTheTheoLaiXe_TimKiem.png")


    def report_the_trip_with_the_driver_card_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_the_trip_with_the_driver_card)
        except:
            contract_card.report_the_trip_with_the_driver_card_search(self, "", "", "")

        button = var_stx.driver.find_element(By.XPATH, var_stx.export_excel2)
        var_stx.driver.execute_script("arguments[0].click();", button)

        # var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        minitor_stx.get_info_web_new()
        try:
            minitor_stx.get_info_excel1(5, "Sheet 1")
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
            button = var_stx.driver.find_element(By.XPATH, var_stx.export_excel2)
            var_stx.driver.execute_script("arguments[0].click();", button)
            time.sleep(7)
            minitor_stx.get_info_web_new()
            minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.12 Báo cáo cuốc đi thẻ theo lái xe")



    def report_the_trip_with_the_driver_card_payment(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.col_id_serial2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_the_trip_with_the_driver_card)
        except:
            contract_card.report_the_trip_with_the_driver_card_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.reportrange_30day).click()
        time.sleep(2)
        button = var_stx.driver.find_element(By.XPATH, var_stx.search)
        var_stx.driver.execute_script("arguments[0].click();", button)
        time.sleep(7)


        n = 0
        while (n < 25):
            n = n + 1
            n = str(n)
            path_checkbox = "//*[@name='left']/div[" + n + "]//*[@ref='eInput']"
            try:
                if var_stx.driver.find_element(By.XPATH, path_checkbox).is_selected() == False:
                    var_stx.driver.find_element(By.XPATH, path_checkbox).click()
                    time.sleep(0.5)
                    break
            except:
                pass
            n = int(n)


        var_stx.driver.find_element(By.XPATH, var_stx.payment).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.confirm).click()
        time.sleep(1)

        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.toast_message)))
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.12 Báo cáo cuốc đi thẻ theo lái xe",
                                                 var_stx.toast_message, "Cập nhật thành công!", "_BaoCaoCuocKhachTheoLaiXe_ThanhToan.png")





    def customer_summary_report(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.customer_summary_report).click()
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.13 Báo cáo tổng hợp khách hàng",
                                                  var_stx.title_page1, "7.7.13 Báo cáo tổng hợp khách hàng", "_BaoCaoTongHopKhachHang.png")


    def customer_summary_report_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_customer_summary_report)
        except:
            contract_card.customer_summary_report(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys("0984")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.13 Báo cáo tổng hợp khách hàng",
                                              var_stx.table_1_4, "", "_BaoCaoTongHopKhachHang_TimKiem.png")


    def customer_summary_report_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_customer_summary_report)
        except:
            contract_card.customer_summary_report_search(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_4)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).send_keys("0984")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        promotion_stx.get_info_web1()
        try:
            minitor_stx.get_info_excel(5, "Sheet")
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
            time.sleep(7)
            promotion_stx.get_info_web1()
            minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.13 Báo cáo tổng hợp khách hàng")









    def card_transaction_report(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_transaction_report).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass

        module_other_stx.write_result_text_try_if(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.14 Báo cáo giao dịch thẻ",
                                                  var_stx.title_page1, "7.7.14 Báo cáo giao dịch thẻ", "_BaoCaoGiaoDichThe.png")


    def card_transaction_report_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_transaction_report)
        except:
            contract_card.card_transaction_report(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(2.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("20/10/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("28/10/2025 10:43")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.14 Báo cáo giao dịch thẻ",
                                              var_stx.list_data2_7, "", "_BaoCaoGiaoDichThe_TimKiem.png")


    def card_transaction_report_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_transaction_report)
        except:
            contract_card.card_transaction_report_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        try:
            minitor_stx.get_info_web()
            minitor_stx.get_info_excel1(5, "Sheet 1")
            minitor_stx.check_info_web_excel(code, eventname, result, "KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.14 Báo cáo giao dịch thẻ")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_KhachHang_BaoCaoGiaoDichThe.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_KhachHang_BaoCaoGiaoDichThe.png")
            var_stx.driver.back()
            time.sleep(4)










    def check_prepaid_cards(self, code, eventname, result, number_serial):
        var_stx.driver.implicitly_wait(5)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 50, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 51, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 52, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 53, 2, 0)
        # number_serial = "9999209842501710"
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_transaction_report).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("10/01/2024 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)

        var_stx.driver.implicitly_wait(0.05)
        n = 1
        page = 1
        while (n < 23):
            n += 1
            n = str(n)
            status = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[8]/span"
            method = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[9]/span"
            monney = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[10]"
            naptien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 50, 2))
            duoctangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 51, 2))
            sudungtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 52, 2))
            tangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 53, 2))

            try:
                name_status = var_stx.driver.find_element(By.XPATH, status).text
                name_method = var_stx.driver.find_element(By.XPATH, method).text
                number_money = var_stx.driver.find_element(By.XPATH, monney).text
                number_money = ''.join(re.findall(r'\d+', number_money))
                print("Hàng: {}, Trạng thái: {}, Phương thức: {}, Số tiền: {}".format(n, name_status, name_method, number_money))

                if (name_status == "Nạp tiền") and (name_method == "Trả trước"):
                    check_number_money = var_stx.driver.find_element(By.XPATH, monney).text
                    print("test: {}".format(check_number_money[0:1]))
                    if check_number_money[0:1] == "-":
                        print('n1')
                        count_number_money = naptien - int(number_money)
                        print("n2: {}".format(count_number_money))
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 50, 2, int(count_number_money))
                    else:
                        print("n3")
                        count_number_money = naptien + int(number_money)
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 50, 2, int(count_number_money))

                if (name_status == "Tiền được tặng") and (name_method == "Trả trước"):
                    count_number_money = int(number_money) + duoctangtien
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 51, 2, int(count_number_money))

                if (name_status == "Sử dụng tiền") and (name_method == "Trả trước"):
                    count_number_money = int(number_money) + sudungtien
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 52, 2, int(count_number_money))

                if (name_status == "Tặng tiền") and (name_method == "Trả trước"):
                    count_number_money = int(number_money) + tangtien
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 53, 2, int(count_number_money))

            except:
                page += 1
                page = str(page)
                path_page = "//*[@class='pagermvc']//*[text()='"+page+"']"
                try:
                    var_stx.driver.find_element(By.XPATH, path_page).click()
                    time.sleep(5)
                    n = 1
                    print("đã sang trang:" + page)
                except:
                    break

            n = int(n)
            page = int(page)

        # contract_card.card_management_search(self, "", "", "", var_stx.number_serial, number_serial, "", "")


        contract_card.card_management(self, "", "", "")
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        sodu = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_8).text
        phuongthuc = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_13).text
        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)

        if sodu[0:1] == "-":
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraTruoc_CheckThongTin.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Số dư thẻ bị âm\nSố dư: {}".format(sodu))
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraTruoc_CheckThongTin.png")
        else:
            try:
                naptien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 50, 2))
                duoctangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 51, 2))
                sudungtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 52, 2))
                tangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 53, 2))
                tratruoc = (naptien + duoctangtien) - (sudungtien + tangtien)
                sodu = ''.join(re.findall(r'\d+', sodu))
                sodu = int(sodu)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "*7.7.14:\nNạp tiền: {}\nTiền được tặng: {}\nSử dụng tiền: {}\nTặng tiền: {}\nSố dư thẻ trả trước: {}\n*7.7.3:"
                                                                                        "\nSố dư: {}\nPhương thức: {}".format(naptien, duoctangtien, sudungtien, tangtien, tratruoc, sodu, phuongthuc))

                if (tratruoc == sodu) and (phuongthuc == "Trả trước"):
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraTruoc_CheckThongTin.png")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraTruoc_CheckThongTin.png")
            except:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraTruoc_CheckThongTin.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraTruoc_CheckThongTin.png")


    def check_prepaid_cards_status(self, code, eventname, result, number_serial):
        var_stx.driver.implicitly_wait(5)
        # number_serial = "9999209842501710"
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])



        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_transaction_report).click()
        time.sleep(5)

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("10/01/2024 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ - 7.7.14 Báo cáo giao dịch thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)

        var_stx.driver.implicitly_wait(0.05)
        n = 1
        m = 2
        page = 1
        while (n < 23):
            n += 1
            n = str(m)
            m += 1
            m = str(m)
            i = int(m) - 1
            j = int(n) - 1
            j = str(j)
            status_on = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[8]/span"
            monney_on = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[10]"
            balance_on = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[11]"

            status_under = "//*[@class='table table-hover table-bordered']/tbody/tr[" + m + "]/td[8]/span"
            monney_under = "//*[@class='table table-hover table-bordered']/tbody/tr[" + m + "]/td[10]"
            balance_under = "//*[@class='table table-hover table-bordered']/tbody/tr[" + m + "]/td[11]"

            try:
                name_status_on = var_stx.driver.find_element(By.XPATH, status_on).text
                number_money_on = var_stx.driver.find_element(By.XPATH, monney_on).text
                number_money_on = ''.join(re.findall(r'\d+', number_money_on))
                number_balance_on = var_stx.driver.find_element(By.XPATH, balance_on).text
                number_balance_on = ''.join(re.findall(r'\d+', number_balance_on))

                name_status_under = var_stx.driver.find_element(By.XPATH, status_under).text
                number_money_under = var_stx.driver.find_element(By.XPATH, monney_under).text
                number_money_under = ''.join(re.findall(r'\d+', number_money_under))
                number_balance_under = var_stx.driver.find_element(By.XPATH, balance_under).text
                number_balance_under = ''.join(re.findall(r'\d+', number_balance_under))

                if (name_status_on == "Sử dụng tiền") or (name_status_on == "Tặng tiền"):
                    money = int(number_balance_under) - int(number_money_on)
                    if int(money) == int(number_balance_on):
                        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                    else:
                        cuon = var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[" + j + "]")

                        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Trang: {}, Hàng: {}, Trạng thái: {}, Số dư trước đó: {}, Số tiền: {}\nSố dư còn lại trên web: {}\nSố dư còn lại sau trừ:  {}"
                                                   .format(page, i, name_status_under, number_balance_under, number_money_on,  number_balance_on, money))
                        logging.info("-----------------------------------")
                        logging.info("Mã thẻ: {}".format(number_serial))
                        logging.info("Trang: {}, Hàng: {}, Trạng thái: {}, Số dư trước đó: {}, Số tiền: {}".format(page, i, name_status_under, number_balance_under, number_money_on))
                        logging.info("Số dư còn lại trên web: {}".format(number_balance_on))
                        logging.info("Số dư còn lại sau trừ:  {}".format(money))
                        logging.info("False")
                        logging.info("-----------------------------------")
                        var_stx.driver.execute_script("arguments[0].scrollIntoView();", cuon)
                        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                        var_stx.driver.save_screenshot(var_stx.imagepath + code + "_BaoCaoGiaoDichThe_Trang{}_Dong{}.png".format(page, i))


                if (name_status_on == "Nạp tiền") or (name_status_on == "Tiền được tặng"):
                    check_number_money_on = var_stx.driver.find_element(By.XPATH, monney_on).text
                    print("test1: {}".format(check_number_money_on[0:1]))

                    if check_number_money_on[0:1] == "-":
                        money = int(number_balance_under) - int(number_money_on)
                        print("đã vào trừ")
                        print(money)
                    else:
                        money = int(number_balance_under) + int(number_money_on)

                    if int(money) == int(number_balance_on):
                        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                    else:
                        cuon = var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[" + j + "]")
                        var_stx.driver.execute_script("arguments[0].scrollIntoView();", cuon)

                        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Trang: {}, Hàng: {}, Trạng thái: {}, Số dư trước đó: {}, Số tiền: {}\nSố dư còn lại trên web: {}\nSố dư còn lại sau cộng: {}"
                                                   .format(page, i, name_status_under, number_balance_under, number_money_on,  number_balance_on, money))
                        logging.info("-----------------------------------")
                        logging.info("Mã thẻ: {}".format(number_serial))
                        logging.info("Trang: {}, Hàng: {}, Trạng thái: {}, Số dư trước đó: {}, Số tiền: {}".format(page, i, name_status_under, number_balance_under, number_money_on))
                        logging.info("Số dư còn lại trên web: {}".format(number_balance_on))
                        logging.info("Số dư còn lại sau cộng: {}".format(money))
                        logging.info("False")
                        logging.info("-----------------------------------")
                        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                        var_stx.driver.save_screenshot(var_stx.imagepath + code + "_BaoCaoGiaoDichThe_Trang{}_Dong{}.png".format(page, i))


            except:
                page += 1
                page = str(page)
                path_page = "//*[@class='pagermvc']//*[text()='"+page+"']"
                try:
                    var_stx.driver.find_element(By.XPATH, path_page).click()
                    time.sleep(5)
                    n = 1
                    m = 2
                    print("đã sang trang:" + page)
                except:
                    break

            n = int(n)
            m = int(m)
            page = int(page)



        wordbook = openpyxl.load_workbook(var_stx.checklistpath)
        sheet = wordbook.get_sheet_by_name("Checklist")
        rownum = 9
        while (rownum < 1000):
            rownum += 1
            rownum = str(rownum)
            if sheet["A" + rownum].value == code and sheet["F" + rownum].value != None:
                print(sheet["A" + rownum].value)
                print(sheet["F" + rownum].value)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            rownum = int(rownum)









    @retry(tries=2, delay=2, backoff=1, jitter=5)
    def login_the_tra_truoc(self):
        var_stx.driver.implicitly_wait(0.5)
        if var_stx.linktest[0:14] == "https://g7test":
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.STAXI)
            except:
                login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
            except:
                login_stx.login.login_stx(self, var_stx.data['login']['tk_admin'], var_stx.data['login']['mk_admin'])


    @retry(tries=2, delay=2, backoff=1, jitter=5)
    def check_the_tra_truoc1(self, rownum, number_card):
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 50, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 51, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 52, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 53, 2, 0)
        contract_card.login_the_tra_truoc(self)

        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_transaction_report).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/01/2021 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_card)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        var_stx.driver.implicitly_wait(0.05)
        n = 1
        m = 2
        page = 1
        while (n < 23):
            n += 1
            n = str(m)
            m += 1
            m = str(m)
            i = int(m) - 1
            j = int(n) - 1
            j = str(j)
            status_on = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[8]/span"
            monney_on = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[10]"
            balance_on = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[11]"

            status_under = "//*[@class='table table-hover table-bordered']/tbody/tr[" + m + "]/td[8]/span"
            monney_under = "//*[@class='table table-hover table-bordered']/tbody/tr[" + m + "]/td[10]"
            balance_under = "//*[@class='table table-hover table-bordered']/tbody/tr[" + m + "]/td[11]"

            status = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[8]/span"
            method = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[9]/span"
            monney = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[10]"

            naptien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 50, 2))
            duoctangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 51, 2))
            sudungtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 52, 2))
            tangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 53, 2))

            try:

                name_status = var_stx.driver.find_element(By.XPATH, status).text
                name_method = var_stx.driver.find_element(By.XPATH, method).text
                number_money = var_stx.driver.find_element(By.XPATH, monney).text
                number_money = ''.join(re.findall(r'\d+', number_money))
                print("Hàng: {}, Trạng thái: {}, Phương thức: {}, Số tiền: {}".format(n, name_status, name_method, number_money))

                if (name_status == "Nạp tiền") and (name_method == "Trả trước"):
                    check_number_money = var_stx.driver.find_element(By.XPATH, monney).text
                    print("test: {}".format(check_number_money[0:1]))
                    if check_number_money[0:1] == "-":
                        print('n1')
                        count_number_money = naptien - int(number_money)
                        print("n2: {}".format(count_number_money))
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 50, 2, int(count_number_money))
                    else:
                        print("n3")
                        count_number_money = naptien + int(number_money)
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 50, 2, int(count_number_money))

                if (name_status == "Tiền được tặng") and (name_method == "Trả trước"):
                    count_number_money = int(number_money) + duoctangtien
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 51, 2, int(count_number_money))

                if (name_status == "Sử dụng tiền") and (name_method == "Trả trước"):
                    count_number_money = int(number_money) + sudungtien
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 52, 2, int(count_number_money))

                if (name_status == "Tặng tiền") and (name_method == "Trả trước"):
                    count_number_money = int(number_money) + tangtien
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 53, 2, int(count_number_money))





                name_status_on = var_stx.driver.find_element(By.XPATH, status_on).text
                number_money_on = var_stx.driver.find_element(By.XPATH, monney_on).text
                number_money_on = ''.join(re.findall(r'\d+', number_money_on))
                number_balance_on = var_stx.driver.find_element(By.XPATH, balance_on).text
                number_balance_on = number_balance_on.replace(".", "").replace("₫", "").strip()

                name_status_under = var_stx.driver.find_element(By.XPATH, status_under).text
                number_money_under = var_stx.driver.find_element(By.XPATH, monney_under).text
                number_money_under = ''.join(re.findall(r'\d+', number_money_under))
                number_balance_under = var_stx.driver.find_element(By.XPATH, balance_under).text
                number_balance_under = number_balance_under.replace(".", "").replace("₫", "").strip()

                if (name_status_on == "Sử dụng tiền") or (name_status_on == "Tặng tiền"):
                    money = int(number_balance_under) - int(number_money_on)
                    if int(money) == int(number_balance_on):
                        pass
                    else:
                        cuon = var_stx.driver.find_element(By.XPATH,"//*[@class='table table-hover table-bordered']/tbody/tr[" + j + "]")
                        var_stx.driver.execute_script("arguments[0].scrollIntoView();", cuon)
                        var_stx.driver.save_screenshot(var_stx.imagepath + "MaThe{}_Trang{}_Dong{}.png".format(number_card, page, i))
                        rownum = int(rownum)
                        print("số dòng: ".format(rownum))
                        var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 2, "Trang: {}, Hàng: {}, Trạng thái: {}, Số dư trước đó: {}, Số tiền: {}\nSố dư còn lại trên web: {}\nSố dư còn lại sau trừ:  {}\n\n"
                                           .format(page, i, name_status_under, number_balance_under, number_money_on, number_balance_on, money))

                if (name_status_on == "Nạp tiền") or (name_status_on == "Tiền được tặng"):
                    check_number_money_on = var_stx.driver.find_element(By.XPATH, monney_on).text
                    print("test1: {}".format(check_number_money_on[0:1]))

                    if check_number_money_on[0:1] == "-":
                        money = int(number_balance_under) - int(number_money_on)
                        print("đã vào trừ")
                        print(money)
                    else:
                        money = int(number_balance_under) + int(number_money_on)

                    if int(money) == int(number_balance_on):
                        pass
                    else:
                        cuon = var_stx.driver.find_element(By.XPATH, "//*[@class='table table-hover table-bordered']/tbody/tr[" + j + "]")
                        var_stx.driver.execute_script("arguments[0].scrollIntoView();", cuon)
                        var_stx.driver.save_screenshot(var_stx.imagepath + "MaThe{}_Trang{}_Dong{}.png".format(number_card, page, i))
                        rownum = int(rownum)
                        print("số dòng: ".format(rownum))
                        var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 2, "Trang: {}, Hàng: {}, Trạng thái: {}, Số dư trước đó: {}, Số tiền: {}\nSố dư còn lại trên web: {}\nSố dư còn lại sau cộng: {}\n\n"
                                           .format(page, i, name_status_under, number_balance_under, number_money_on, number_balance_on, money))



            except:
                page += 1
                page = str(page)
                path_page = "//*[@class='pagermvc']//*[text()='" + page + "']"
                try:
                    var_stx.driver.find_element(By.XPATH, path_page).click()
                    time.sleep(5)
                    n = 1
                    m = 2
                    print("đã sang trang:" + page)
                except:
                    break

            n = int(n)
            m = int(m)
            page = int(page)


    @retry(tries=2, delay=2, backoff=1, jitter=5)
    def check_the_tra_truoc2(self, rownum, number_card):
        rownum = int(rownum)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 36, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 36, 3, "")
        contract_card.login_the_tra_truoc(self)

        #7.7.3
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management).click()
        time.sleep(5)
        try:
            wait = WebDriverWait(var_stx.driver, 15)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.number_serial)))
        except:
            var_stx.driver.refresh()
            time.sleep(5)
        time.sleep(2)

        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_card)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            sodu = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_8).text
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 36, 2, sodu)
            if sodu[0:1] == "-":
                var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 4, "Số dư thẻ bị âm\nSố dư: {}".format(sodu))
                var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 5, "Fail")

        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_card)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
            try:
                sodu = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_8).text
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 36, 2, sodu)
            except:
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 36, 3, "Thẻ đã xóa")



        naptien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 50, 2))
        duoctangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 51, 2))
        sudungtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 52, 2))
        tangtien = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 53, 2))
        trangthaithe = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 36, 3)
        try:
            sodu1 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 36, 2)
            sodu1 = int(sodu1.replace(".", "").replace(" ₫", ""))
        except:
            sodu1 = 0

        tratruoc = (naptien + duoctangtien) - (sudungtien + tangtien)
        var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 3, "*Card_before_06.\n 7.7.14: (Nạp tiền + Tiền được tặng) - (Sử dụng tiền + Tặng tiền)\nNạp tiền: {}\nTiền được tặng: {}\nSử dụng tiền: {}\nTặng tiền: {}\n\nSố dư 7.7.14: {}\nSố dư 7.7.3: {}"
                           .format(naptien, duoctangtien, sudungtien, tangtien, tratruoc, sodu1))

        if trangthaithe == "Thẻ đã xóa":
            var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 4, "Thẻ đã xóa")

        if tratruoc == sodu1:
            pass
        else:
            var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 5, "Fail")


    def the_tra_truoc(self):

        # Clear data
        wordbook = openpyxl.load_workbook('data_card.xlsx')
        sheet = wordbook.get_sheet_by_name("Sheet1")
        i = 1
        while (i < 2000):
            i += 1
            i = str(i)
            sheet["B" + i] = ""
            sheet["C" + i] = ""
            sheet["D" + i] = ""
            sheet["E" + i] = ""
            i = int(i)
        wordbook.save('data_card.xlsx')

        # 7.7.14
        wordbook = openpyxl.load_workbook('data_card.xlsx')
        sheet = wordbook.get_sheet_by_name("Sheet1")
        rownum = 1
        while (rownum < 1000):
            rownum += 1
            rownum = str(rownum)
            if sheet["A" + rownum].value != None:
                number_card = sheet["A" + rownum].value
                print(number_card)

                contract_card.check_the_tra_truoc1(self, rownum, number_card)
                contract_card.check_the_tra_truoc2(self, rownum, number_card)

                row_eror = sheet["B" + rownum].value
                rownum = int(rownum)
                if row_eror != None:
                    var_stx.appendData('data_card.xlsx', "Sheet1", rownum, 5, "Fail")


            else:
                break
            rownum = int(rownum)



























    def card_after1(self, code, eventname, result, number_serial, from_day, to_day, id_company):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 169, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 170, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 170, 4, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 171, 2, 0)

        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)



        # 7.7.4
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)
        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.ag1_8)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(number_serial)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
            time.sleep(3)


        try:
            var_stx.driver.implicitly_wait(0.5)
            sotienno7_7_4 = var_stx.driver.find_element(By.XPATH, var_stx.ag1_8).text
            sotienno7_7_4 = ''.join(re.findall(r'\d+', sotienno7_7_4))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 170, 2, int(sotienno7_7_4))
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 170, 4, "Không tìm thấy thẻ")




        #7.7.10
        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.ag1_10_a).click()
            time.sleep(3)
            var_stx.driver.switch_to.window(var_stx.driver.window_handles[1])

            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.check_report_customer_card_details)))
            time.sleep(1)
        except:
            pass

        var_stx.driver.implicitly_wait(0.05)
        n = 0
        page = 1
        while (n < 23):
            n += 1
            n = str(n)
            path_cuocxe = "//*[@class='ag-center-cols-viewport']/div/div[" + n + "]/div[11]"
            path_serial = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[1]"
            cuocxe_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 171, 2))
            try:
                serial = var_stx.driver.find_element(By.XPATH, path_serial).text
                cuocxe = var_stx.driver.find_element(By.XPATH, path_cuocxe).text
                cuocxe = ''.join(re.findall(r'\d+', cuocxe))
                print("Hàng: {}, Số Serial: {}, Cuốc xe: {},".format(n, serial, cuocxe))

                if serial == number_serial:
                    print("Hàng: {}, Số Serial: {}, Cuốc xe: {}".format(n, serial, cuocxe))
                    count_cuocxe = int(cuocxe) + cuocxe_excel
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 171, 2, int(count_cuocxe))

            except:
                page += 1
                page = str(page)
                try:
                    page1 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[2]").text
                    page2 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[4]").text
                    if page1 == page2:
                        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
                        module_other_stx.close_tab()
                        break
                    else:
                        var_stx.driver.find_element(By.XPATH, "//*[@class='ag-icon ag-icon-next']").click()
                        time.sleep(5)
                    n = 0
                    print("đã sang trang:" + page)
                except:
                    break

            n = int(n)
            page = int(page)

        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(0.5)
        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        sotienno7_7_4_excel = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 170, 2)
        tongcuocxe_7_7_10_excel = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 171, 2)
        status_card = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 170, 4)
        try:
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "- Số tiền nợ 7.7.4         : {}\n- Tổng cước xe(7.7.10): {}".format(sotienno7_7_4_excel, tongcuocxe_7_7_10_excel))
            if int(sotienno7_7_4_excel) == int(tongcuocxe_7_7_10_excel):
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_ChotCongNo.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_ChotCongNo.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau1.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_ChotCongNo.png")

        if status_card == "Không tìm thấy thẻ":
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không tìm thấy thẻ")


    def card_after2(self, code, eventname, result, number_serial, id_company):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 172, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 172, 4, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 173, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 174, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 175, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 176, 2, 0)

        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)


        #7.7.4
        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.check_closing_debts)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.closing_debts).click()
            time.sleep(5)

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/01/2023 00:00")
        time.sleep(0.5)
        # var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        # time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.tonghop7_7_4)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/01/2023 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(number_serial)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
            time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.ag1_8)
            tonghop7_7_4 = var_stx.driver.find_element(By.XPATH, var_stx.tonghop7_7_4).text
            tonghop7_7_4 = ''.join(re.findall(r'\d+', tonghop7_7_4))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 172, 2, int(tonghop7_7_4))
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 172, 4, "Không tìm thấy thẻ")








        #7.7.5
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.List_of_debts_finalized).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/01/2023 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.StatusDebt_1).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.StatusDebtTT_3).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)
        var_stx.driver.implicitly_wait(0.05)
        n = 0
        page = 1
        while (n < 23):
            n += 1
            n = str(n)
            path_congno = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[2]/span/a"
            try:
                var_stx.driver.find_element(By.XPATH, path_congno).click()
                time.sleep(2)
                wait = WebDriverWait(var_stx.driver, 10)
                element = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='Tổng tiền chốt:']")))
                time.sleep(0.5)
                print("n0")
                m = 0
                while (m < 21):
                    m += 1
                    m = str(m)
                    print("n0.1")
                    serial = "//*[@id='detailInfoDebt']/table/tbody/tr[" + m + "]/td[5]"
                    sotien = "//*[@id='detailInfoDebt']/table/tbody/tr[" + m + "]/td[7]"
                    print("n0.2")
                    print(serial)
                    print(sotien)
                    sotien7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 173, 2))

                    try:
                        check_serial = var_stx.driver.find_element(By.XPATH, serial).text
                        print("n0.3")
                        print(check_serial)
                        if str(check_serial) == str(number_serial):
                            print("n1")
                            try:
                                sotien7_7_5 = var_stx.driver.find_element(By.XPATH, sotien).text
                                sotien7_7_5 = ''.join(re.findall(r'\d+', sotien7_7_5))
                                count_sotien7_7_5 = sotien7_7_5_excel + int(sotien7_7_5)

                                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 173, 2, count_sotien7_7_5)
                                var_stx.driver.find_element(By.XPATH, var_stx.close2).click()
                                time.sleep(2)
                                print("n2")
                                break
                            except:
                                pass
                    except:
                        pass
                    m = int(m)

            except:
                break

            n = int(n)
            page = int(page)












        # 7.7.7
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.Payment_of_debts).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/01/2023 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.StatusDebt_2).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)
        var_stx.driver.implicitly_wait(0.05)
        n = 0
        while (n < 23):
            n += 1
            n = str(n)
            path_checkbox = "//*[@class='ag-pinned-left-cols-container']/div[" + n + "]/div[1]//*[@type='checkbox']"
            sotienno_7_7_7_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 174, 2))
            try:
                var_stx.driver.find_element(By.XPATH, path_checkbox).click()
                time.sleep(1)
                var_stx.driver.find_element(By.XPATH, var_stx.PaymentForm).click()
                time.sleep(2)

                wait = WebDriverWait(var_stx.driver, 10)
                element = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='  THANH TOÁN CÔNG NỢ']")))
                time.sleep(0.5)
                print("n0")
                m = 0
                while (m < 21):
                    m += 1
                    m = str(m)
                    print("n0.1")
                    serial = "//*[@id='detailInfoDebt']/table/tbody[" + m + "]/tr/td[3]"
                    sotien = "//*[@id='detailInfoDebt']/table/tbody[" + m + "]/tr/td[5]"
                    print("n0.2")
                    print(serial)
                    print(sotien)

                    try:
                        check_serial = var_stx.driver.find_element(By.XPATH, serial).text
                        print("n0.3")
                        print(check_serial)
                        if str(check_serial) == str(number_serial):
                            print("n1")
                            try:
                                sotienno_7_7_7 = var_stx.driver.find_element(By.XPATH, sotien).text
                                sotienno_7_7_7 = ''.join(re.findall(r'\d+', sotienno_7_7_7))
                                count_sotien1 = int(sotienno_7_7_7) + int(sotienno_7_7_7_excel)
                                print("n1.5")
                                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 174, 2, int(count_sotien1))
                                var_stx.driver.find_element(By.XPATH, var_stx.close).click()
                                time.sleep(2)
                                print("n2")
                                break
                            except:
                                sotienno_7_7_7 = 0
                                count_sotienno_7_7_7 = int(sotienno_7_7_7) + int(sotienno_7_7_7_excel)
                                print("n3")
                                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 174, 2, int(count_sotienno_7_7_7))
                                var_stx.driver.find_element(By.XPATH, var_stx.close).click()
                                time.sleep(2)
                                print("n3.5")
                                break
                    except:
                        pass
                    m = int(m)

            except:
                break

            n = int(n)



        #7.7.3
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management).click()
        time.sleep(5)
        try:
            wait = WebDriverWait(var_stx.driver, 15)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.number_serial)))
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            print("đã load lại trang 1")
            try:
                wait = WebDriverWait(var_stx.driver, 15)
                element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.number_serial)))
            except:
                var_stx.driver.refresh()
                time.sleep(7)
                print("đã load lại trang 2")

        time.sleep(2)

        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(3.5)

        try:
            sodu7_7_3 = var_stx.driver.find_element(By.XPATH, var_stx.sodu7_7_3).text
            sodu7_7_3 = int(sodu7_7_3.replace("₫", "").replace(".", "").strip())
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 175, 2, sodu7_7_3)

            hanmuc7_7_3 = var_stx.driver.find_element(By.XPATH, var_stx.hanmuc7_7_3).text
            hanmuc7_7_3 = int(hanmuc7_7_3.replace("₫", "").replace(".", "").strip())
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 176, 2, hanmuc7_7_3)
        except:
            pass


        try:
            var_stx.driver.implicitly_wait(0.5)
            ma_hd7_7_3 = var_stx.driver.find_element(By.XPATH, var_stx.ma_hd7_7_3).text
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 169, 2, ma_hd7_7_3)
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 169, 2, "Không tìm thấy mã hđ")


        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)

        tonghop_7_7_4_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 172, 2))
        tonghop_7_7_4_excel_edit = abs(tonghop_7_7_4_excel)
        status_card = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 172, 4)


        sotien_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 173, 2))
        sotien_7_7_5_excel_edit = abs(sotien_7_7_5_excel)

        sotienno_7_7_7_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 174, 2))
        sotienno_7_7_7_excel_edit = abs(sotienno_7_7_7_excel)

        sodu_7_7_3_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 175, 2))
        sodu_7_7_3_excel_edit = abs(sodu_7_7_3_excel)

        hanmuc_7_7_3_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 176, 2))

        check2_3_4 = (tonghop_7_7_4_excel_edit + sotien_7_7_5_excel_edit + sotienno_7_7_7_excel_edit)

        try:
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "- Hạn mức 7.7.3    : {}\n- Số dư 7.7.3    (1): {}\n- CN chưa chốt (2): {}\n- CN đã chốt nhưng chưa TT (3): {}\n- CN TT 1 phần_Nợ còn         (4): {}\n=>(2)+(3)+(4) = {}"
                                       .format(hanmuc_7_7_3_excel, sodu_7_7_3_excel, tonghop_7_7_4_excel, sotien_7_7_5_excel, sotienno_7_7_7_excel, check2_3_4))
            if sodu_7_7_3_excel_edit == (tonghop_7_7_4_excel_edit + sotien_7_7_5_excel_edit + sotienno_7_7_7_excel_edit):
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau2.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau2.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau2.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau2.png")

        if (sodu_7_7_3_excel > 0) or (sodu_7_7_3_excel > hanmuc_7_7_3_excel):
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau2.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau2.png")
        if status_card == "Không tìm thấy thẻ":
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không tìm thấy thẻ")


    def card_after3(self, code, eventname, result, number_serial, id_company):
        var_stx.driver.implicitly_wait(5)
        id_company = str(id_company)
        number_serial = str(number_serial)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 177, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 177, 4, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 178, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 179, 2, 0)

        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)


        #7.7.3
        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.check_card_management)
        except:
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.card_management).click()
                time.sleep(5)
            except:
                var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.card_management).click()
                time.sleep(5)





        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(3.5)
            var_stx.driver.find_element(By.XPATH, var_stx.sodu7_7_3)

        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(3.5)



        try:
            var_stx.driver.implicitly_wait(0.5)
            sodu7_7_3 = var_stx.driver.find_element(By.XPATH, var_stx.sodu7_7_3).text
            sodu7_7_3 = int(sodu7_7_3.replace("₫", "").replace(".", "").strip())
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 177, 2, sodu7_7_3)
            print("n1")
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 177, 4, "Không tìm thấy thẻ")
            print("n2")




        #7.7.10
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.Report_customer_card_details).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.reportrange_left_year).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, "//*[@class='drp-calendar left']//*[@class='yearselect']//*[text()='2022']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH,"//*[@class='drp-calendar left']//*[@class='table-condensed']//*[text()='15']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.reportrange_right_year).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH,"//*[@class='drp-calendar right']//*[@class='yearselect']//*[text()='2027']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH,"//*[@class='drp-calendar left']//*[@class='table-condensed']//*[text()='15']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
        time.sleep(0.5)
        # var_stx.driver.find_element(By.XPATH, var_stx.DebtStatus_3).click()
        # time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(3.5)

        #Chưa Thanh toán
        var_stx.driver.implicitly_wait(0.05)
        n = 0
        page = 1
        while (n < 23):
            n += 1
            n = str(n)
            path_trangthaithanhtoan = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[6]"
            path_cuocxe = "//*[@class='ag-center-cols-viewport']/div/div[" + n + "]/div[11]"
            path_serial = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[1]"
            cuocxe_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 178, 2))
            try:
                serial = var_stx.driver.find_element(By.XPATH, path_serial).text
                trangthaithanhtoan = var_stx.driver.find_element(By.XPATH, path_trangthaithanhtoan).text
                cuocxe = var_stx.driver.find_element(By.XPATH, path_cuocxe).text
                cuocxe = ''.join(re.findall(r'\d+', cuocxe))

                if (serial == number_serial) and (trangthaithanhtoan == "Chưa thanh toán"):
                    print("Hàng: {}, Số Serial: {}, Trạng thái thanh toán: {}, Cuốc xe: {},".format(n, serial, trangthaithanhtoan, cuocxe))

                    count_cuocxe = int(cuocxe) + cuocxe_excel
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 178, 2, int(count_cuocxe))

            except:
                page += 1
                page = str(page)
                try:
                    page1 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[2]").text
                    page2 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[4]").text
                    if page1 == page2:
                        break
                    else:
                        var_stx.driver.find_element(By.XPATH, "//*[@class='ag-icon ag-icon-next']").click()
                        time.sleep(5)
                    n = 0
                    print("đã sang trang:" + page)
                except:
                    break

            n = int(n)
            page = int(page)








        # 7.7.7
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.Payment_of_debts).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/01/2023 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.StatusDebt_2).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)
        var_stx.driver.implicitly_wait(0.05)
        n = 0
        while (n < 23):
            n += 1
            n = str(n)
            path_checkbox = "//*[@class='ag-pinned-left-cols-container']/div[" + n + "]/div[1]//*[@type='checkbox']"
            sotienno_7_7_7_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 179, 2))
            try:
                var_stx.driver.find_element(By.XPATH, path_checkbox).click()
                time.sleep(1)
                var_stx.driver.find_element(By.XPATH, var_stx.PaymentForm).click()
                time.sleep(2)

                wait = WebDriverWait(var_stx.driver, 10)
                element = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='  THANH TOÁN CÔNG NỢ']")))
                time.sleep(0.5)
                print("n0")
                m = 0
                while (m < 21):
                    m += 1
                    m = str(m)
                    print("n0.1")
                    serial = "//*[@id='detailInfoDebt']/table/tbody[" + m + "]/tr/td[3]"
                    sotien = "//*[@id='detailInfoDebt']/table/tbody[" + m + "]/tr/td[5]"
                    print("n0.2")
                    print(serial)
                    print(sotien)

                    try:
                        check_serial = var_stx.driver.find_element(By.XPATH, serial).text
                        print("n0.3")
                        print(check_serial)
                        if str(check_serial) == str(number_serial):
                            print("n1")
                            try:
                                sotienno_7_7_7 = var_stx.driver.find_element(By.XPATH, sotien).text
                                sotienno_7_7_7 = ''.join(re.findall(r'\d+', sotienno_7_7_7))
                                count_sotien1 = int(sotienno_7_7_7) + int(sotienno_7_7_7_excel)
                                print("n1.5")
                                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 179, 2, int(count_sotien1))
                                var_stx.driver.find_element(By.XPATH, var_stx.close).click()
                                time.sleep(2)
                                print("n2")
                                break
                            except:
                                sotienno_7_7_7 = 0
                                count_sotienno_7_7_7 = int(sotienno_7_7_7) + int(sotienno_7_7_7_excel)
                                print("n3")
                                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 174, 2, int(count_sotienno_7_7_7))
                                var_stx.driver.find_element(By.XPATH, var_stx.close).click()
                                time.sleep(2)
                                print("n3.5")
                                break
                    except:
                        pass
                    m = int(m)

            except:
                break

            n = int(n)



        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)

        sodu_7_7_3_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 177, 2))
        sodu_7_7_3_excel_edit = abs(sodu_7_7_3_excel)
        print(sodu_7_7_3_excel_edit)
        status_card = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 177, 4)

        tongcuoc_7_7_10_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 178, 2))
        tongcuoc_7_7_10_excel_edit = abs(tongcuoc_7_7_10_excel)
        print(tongcuoc_7_7_10_excel_edit)

        sotienno_7_7_7_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 179, 2))
        sotienno_7_7_7_excel_edit = abs(sotienno_7_7_7_excel)
        print(sotienno_7_7_7_excel_edit)

        check2_3 = (tongcuoc_7_7_10_excel_edit + sotienno_7_7_7_excel_edit)

        try:
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "- Số dư 7.7.3                          (1): {}\n- Tổng cuốc chưa thanh toán (2): {}\n- CN TT 1 phần_Nợ còn(3)          : {}\n=> (2)+(3) = {}"
                                       .format(sodu_7_7_3_excel, tongcuoc_7_7_10_excel, sotienno_7_7_7_excel, check2_3))
            if sodu_7_7_3_excel_edit == (tongcuoc_7_7_10_excel_edit + sotienno_7_7_7_excel_edit):
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau3.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau3.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau3.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau3.png")

        if  status_card == "Không tìm thấy thẻ":
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không tìm thấy thẻ")


    def card_after4(self, code, eventname, result, number_serial, from_day, to_day, id_company):
        var_stx.driver.implicitly_wait(5)
        id_company = str(id_company)
        number_serial = str(number_serial)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 180, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 181, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 182, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 183, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 183, 4, "")

        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)




        #7.7.5
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.List_of_debts_finalized).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.StatusDebt_1).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)


        try:
            var_stx.driver.implicitly_wait(0.05)
            var_stx.driver.find_element(By.XPATH, var_stx.path_congno).click()
            time.sleep(2)
            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='Gửi email thông báo']")))
            time.sleep(0.5)
            sotien_hopdong_7_7_5 = var_stx.driver.find_element(By.XPATH, var_stx.sotien_hopdong_7_7_5).text
            sotien_hopdong_7_7_5 = ''.join(re.findall(r'\d+', sotien_hopdong_7_7_5))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 183, 2, sotien_hopdong_7_7_5)
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 183, 4, "Không tìm thấy thẻ")



        print("n0")
        m = 1
        while (m < 21):
            m += 1
            m = str(m)
            print("n0.1")
            serial = "//*[@id='detailInfoDebt']/table/tbody/tr[" + m + "]/td[5]"
            sotien = "//*[@id='detailInfoDebt']/table/tbody/tr[" + m + "]/td[7]"
            print("n0.2")
            print(serial)
            print(sotien)
            sotien_all_serial_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 182, 2))

            try:
                check_serial = var_stx.driver.find_element(By.XPATH, serial).text
                print("n0.3")
                print(check_serial)
                sotien_all_serial_7_7_5 = var_stx.driver.find_element(By.XPATH, sotien).text
                sotien_all_serial_7_7_5 = ''.join(re.findall(r'\d+', sotien_all_serial_7_7_5))
                count_sotien_all_serial = int(sotien_all_serial_7_7_5) + int(sotien_all_serial_7_7_5_excel)
                print("n1.5")
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 182, 2, int(count_sotien_all_serial))

                if str(check_serial) == str(number_serial):
                    print("n1")
                    try:
                        sotien_1serial_7_7_5 = var_stx.driver.find_element(By.XPATH, sotien).text
                        sotien_1serial_7_7_5 = ''.join(re.findall(r'\d+', sotien_1serial_7_7_5))
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 180, 2, sotien_1serial_7_7_5)
                    except:
                        pass

            except:
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.edit_close).click()
                    time.sleep(2)
                except:
                    pass
                print("close1")
                break

            m = int(m)













        #7.7.10
        scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
        ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(800, 0).release().perform()
        time.sleep(1)

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.see_detail_7_7_5).click()
            time.sleep(3)
            var_stx.driver.switch_to.window(var_stx.driver.window_handles[1])

            wait = WebDriverWait(var_stx.driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.check_report_customer_card_details)))
            time.sleep(1)
        except:
            pass


        var_stx.driver.implicitly_wait(0.05)
        n = 0
        page = 1
        while (n < 23):
            n += 1
            n = str(n)
            path_cuocxe = "//*[@class='ag-center-cols-viewport']/div/div[" + n + "]/div[11]"
            path_serial = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[1]"
            # path_cuocxe = "//*[@class='ag-center-cols-viewport']/div/div[" + n + "]/div[11]"
            # path_serial = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[1]"
            cuocxe_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 181, 2))
            try:
                serial = var_stx.driver.find_element(By.XPATH, path_serial).text
                cuocxe = var_stx.driver.find_element(By.XPATH, path_cuocxe).text
                cuocxe = ''.join(re.findall(r'\d+', cuocxe))
                print("Hàng: {}, Số Serial: {}, Cuốc xe: {}".format(n, serial, cuocxe))

                if serial == number_serial:
                    print("Hàng: {}, Số Serial: {}, Cuốc xe: {}".format(n, serial, cuocxe))
                    count_cuocxe = int(cuocxe) + cuocxe_excel
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 181, 2, int(count_cuocxe))

            except:
                page += 1
                page = str(page)
                try:
                    page1 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[2]").text
                    page2 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[4]").text
                    if page1 == page2:
                        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
                        module_other_stx.close_tab()
                        break
                    else:
                        var_stx.driver.find_element(By.XPATH, "//*[@class='ag-icon ag-icon-next']").click()
                        time.sleep(5)
                    n = 0
                    print("đã sang trang:" + page)
                except:
                    break

            n = int(n)
            page = int(page)

        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(0.5)
        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)

        sotien_1serial_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 180, 2))
        tongcuoc_1_serial_7_7_10_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 181, 2))
        sotien_all_serial_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 182, 2))
        sotien_hopdong_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 183, 2))
        status_card = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 183, 4)


        try:
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "- Số tiền serial 7.7.5      (1): {}\n- Tổng cước xe(7.7.10) (2): {}\n\n- Tổng serial        (3): {}\n- Tổng hợp đồng (4): {}"
                                       .format(sotien_1serial_7_7_5_excel, tongcuoc_1_serial_7_7_10_excel, sotien_all_serial_7_7_5_excel, sotien_hopdong_7_7_5_excel))
            if (sotien_1serial_7_7_5_excel == tongcuoc_1_serial_7_7_10_excel) and (sotien_all_serial_7_7_5_excel == sotien_hopdong_7_7_5_excel):
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau4.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau4.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau4.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau4.png")

        if status_card == "Không tìm thấy thẻ":
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không tìm thấy thẻ")


    def check_card_after(self, rownum, code_hd):
        #after 1
        rownum = int(rownum)
        sotienno7_7_4_excel = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 170, 2)
        tongcuocxe_7_7_10_excel = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 171, 2)
        var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 5, "*Customer86: Số tiền nợ 7.7.4 = Tổng cước xe(7.7.10)\n- Số tiền nợ 7.7.4           : {}\n- Tổng cước xe(7.7.10): {}\n\n".format(sotienno7_7_4_excel, tongcuocxe_7_7_10_excel))

        if int(sotienno7_7_4_excel) == int(tongcuocxe_7_7_10_excel):
            pass
        else:
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 7, "Fail_86")
            print("n1")




        #after 2
        tonghop_7_7_4_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 172, 2))
        tonghop_7_7_4_excel_edit = abs(tonghop_7_7_4_excel)
        status_card2 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 172, 4)
        sotien_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 173, 2))
        sotien_7_7_5_excel_edit = abs(sotien_7_7_5_excel)
        sotienno_7_7_7_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 174, 2))
        sotienno_7_7_7_excel_edit = abs(sotienno_7_7_7_excel)
        sodu_7_7_3_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 175, 2))
        sodu_7_7_3_excel_edit = abs(sodu_7_7_3_excel)
        hanmuc_7_7_3_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 176, 2))
        check2_3_4 = (tonghop_7_7_4_excel_edit + sotien_7_7_5_excel_edit + sotienno_7_7_7_excel_edit)
        var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 5, "*Customer87: (1)=(2)+(3)+(4)\n- Hạn mức 7.7.3     : {}\n- Số dư 7.7.3    (1): {}\n- CN chưa chốt (2): {}\n- CN đã chốt nhưng chưa TT (3): {}\n- CN TT 1 phần_Nợ còn           (4): {}\n=>(2)+(3)+(4) = {}\n\n"
                                       .format(hanmuc_7_7_3_excel, sodu_7_7_3_excel, tonghop_7_7_4_excel, sotien_7_7_5_excel, sotienno_7_7_7_excel, check2_3_4))

        if sodu_7_7_3_excel_edit == (tonghop_7_7_4_excel_edit + sotien_7_7_5_excel_edit + sotienno_7_7_7_excel_edit):
            pass
        else:
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 7, ", Fail_87")
            print("n2")

        if (sodu_7_7_3_excel > 0):
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 7, ", Fail_87")
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 8, "\nSố dư > 0")

        if (sodu_7_7_3_excel > hanmuc_7_7_3_excel):
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 8, "\nSố dư > Hạn mức")
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 7, ", Fail_87")

        code_hd_excel = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 169, 2)
        if str(code_hd_excel) != str(code_hd):
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 8, "Sai mã hợp đồng, hđ web: {}".format(code_hd_excel))






        #after 3
        sodu_7_7_3_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 177, 2))
        sodu_7_7_3_excel_edit = abs(sodu_7_7_3_excel)
        print(sodu_7_7_3_excel_edit)
        status_card3 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 177, 4)

        tongcuoc_7_7_10_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 178, 2))
        tongcuoc_7_7_10_excel_edit = abs(tongcuoc_7_7_10_excel)
        print(tongcuoc_7_7_10_excel_edit)

        sotienno_7_7_7_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 179, 2))
        sotienno_7_7_7_excel_edit = abs(sotienno_7_7_7_excel)
        print(sotienno_7_7_7_excel_edit)
        check2_3 = (tongcuoc_7_7_10_excel_edit + sotienno_7_7_7_excel_edit)

        var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 5, "*Customer88:  (1)=(2)+(3)\n- Số dư 7.7.3                                (1): {}\n- Tổng cuốc chưa thanh toán (2): {}\n- CN TT 1 phần_Nợ còn(3)            : {}\n=> (2)+(3) = {}\n\n"
                                       .format(sodu_7_7_3_excel, tongcuoc_7_7_10_excel, sotienno_7_7_7_excel, check2_3))

        if sodu_7_7_3_excel_edit == (tongcuoc_7_7_10_excel_edit + sotienno_7_7_7_excel_edit):
            pass
        else:
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 7, ", Fail_88")
            print("n3")
        if status_card3 == "Không tìm thấy thẻ":
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 8, "\nKhông tìm thấy thẻ")





        #after 4
        sotien_1serial_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 180, 2))
        tongcuoc_1_serial_7_7_10_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 181, 2))
        sotien_all_serial_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 182, 2))
        sotien_hopdong_7_7_5_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 183, 2))
        status_card4 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 183, 4)

        var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 5, "*Customer89: (1 = 2) ; (3 = 4)\n- Số tiền serial 7.7.5      (1): {}\n- Tổng cước xe(7.7.10) (2): {}\n\n- Tổng serial        (3): {}\n- Tổng hợp đồng (4): {}"
                                       .format(sotien_1serial_7_7_5_excel, tongcuoc_1_serial_7_7_10_excel, sotien_all_serial_7_7_5_excel, sotien_hopdong_7_7_5_excel))

        if (sotien_1serial_7_7_5_excel == tongcuoc_1_serial_7_7_10_excel) and (sotien_all_serial_7_7_5_excel == sotien_hopdong_7_7_5_excel):
            pass
        else:
            var_stx.appendData('data_card.xlsx', "Thẻ trả sau", rownum, 7, ", Fail_89")
            print("n3")


    def card_after5(self, code, eventname, result):
        #clear data
        wordbook = openpyxl.load_workbook('data_card.xlsx')
        sheet = wordbook.get_sheet_by_name("Thẻ trả sau")
        i = 1
        while (i < 1000):
            i += 1
            i = str(i)
            sheet["E" + i] = ""
            sheet["G" + i] = ""
            sheet["H" + i] = ""
            i = int(i)
        wordbook.save('data_card.xlsx')



        wordbook = openpyxl.load_workbook('data_card.xlsx')
        sheet = wordbook.get_sheet_by_name("Thẻ trả sau")
        rownum = 1
        while (rownum < 1000):
            rownum += 1
            rownum = str(rownum)
            from_day = sheet["C" + rownum].value
            to_day = sheet["D" + rownum].value
            number_hd = sheet["A" + rownum].value
            number_serial = sheet["B" + rownum].value
            id_company = sheet["F" + rownum].value
            print("Đang chạy tới hàng: {}, Mã Hđ: {}, Số serial: {}, Từ ngày: {}, Đến ngày: {}, Công ty: {}".format(rownum, number_hd, number_serial, from_day, to_day, id_company))
            if number_serial == None:
                break

            try:
                try:
                    contract_card.card_after1(self, code, eventname, result, number_serial, from_day, to_day, id_company)
                except:
                    contract_card.card_after1(self, code, eventname, result, number_serial, from_day, to_day, id_company)

                try:
                    contract_card.card_after2(self, code, eventname, result, number_serial, id_company)
                except:
                    contract_card.card_after2(self, code, eventname, result, number_serial, id_company)

                try:
                    contract_card.card_after3(self, code, eventname, result, number_serial, id_company)
                except:
                    contract_card.card_after3(self, code, eventname, result, number_serial, id_company)

                try:
                    contract_card.card_after4(self, code, eventname, result, number_serial, from_day, to_day, id_company)
                except:
                    contract_card.card_after4(self, code, eventname, result, number_serial, from_day, to_day, id_company)

                contract_card.check_card_after(self, rownum, number_hd)
            except:
                module_other_stx.swich_tab_0()

            rownum = int(rownum)












    def card_debt_closing(self, code, eventname, result, number_serial, from_day, to_day, id_company):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 190, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 191, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 192, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 193, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 194, 2, 0)

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 196, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 197, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 198, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 199, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 200, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 201, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 202, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 203, 2, "")

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 205, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 206, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 207, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 208, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 209, 2, 0)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 210, 2, "")
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 211, 2, "")





        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)



        # 10.6.3
        var_stx.driver.find_element(By.XPATH, var_stx.admin).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.admin_10_6).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.admin).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.admin_10_6).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.admin_10_6_3).click()
        time.sleep(2)
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.name_config).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.name_config).send_keys("TRU_TIEN_KHI_CHOT_CN")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(3)
        name_config = var_stx.driver.find_element(By.XPATH, var_stx.table_1_3).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 190, 2, name_config)



        #7.7.4
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 191, 2, number_serial)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.closing_debts).click()
        time.sleep(2)
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, "//*[@class='ag-center-cols-viewport']/div/div[1]//*[text()='"+number_serial+"']")
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial1).send_keys(number_serial)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
            time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, "//*[@class='ag-center-cols-viewport']/div/div[1]//*[text()='"+number_serial+"']")
            var_stx.driver.find_element(By.XPATH, var_stx.ag1_1_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.icon_debt_closing).click()
            time.sleep(2.5)
            code_phieu = var_stx.driver.find_element(By.XPATH, var_stx.CodeDebt).get_attribute('value')      #Mã phiếu chốt (*)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 193, 2, code_phieu)

            summary_money = var_stx.driver.find_element(By.XPATH, var_stx.MoneyClosing).text       #Tổng tiền chốt
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 194, 2, summary_money)

            var_stx.driver.find_element(By.XPATH, var_stx.paymentbtn).click()
            time.sleep(5)
            try:
                var_stx.driver.find_element(By.XPATH, "//*[@class='ag-center-cols-viewport']/div/div[1]//*[text()='" + number_serial + "']")
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 192, 2, "Chưa mất bản ghi")
            except:
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 192, 2, "Đã mất bản ghi")
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 192, 3, f"Không tìm thấy serial {number_serial}")



        #7.7.5
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.List_of_debts_finalized).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)

        code_hd = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 193, 2)
        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, "//*[@class='ag-center-cols-viewport']/div/div[1]//*[text()='"+code_hd+"']")
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
            time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, "//*[@class='ag-center-cols-viewport']/div/div[1]//*[text()='"+code_hd+"']")
            code_congno = var_stx.driver.find_element(By.XPATH, var_stx.ag1_2).text       #Mã công nợ
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 196, 2, code_congno)

            code_hopdong = var_stx.driver.find_element(By.XPATH, var_stx.ag1_3).text       #Mã hợp đồng
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 197, 2, code_hopdong)

            time_chot = var_stx.driver.find_element(By.XPATH, var_stx.ag1_4).text       #Thời gian chốt
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 198, 2, time_chot)

            so_tien_no_chot = var_stx.driver.find_element(By.XPATH, var_stx.ag1_5).text       #Số tiền nợ chốt
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 199, 2, so_tien_no_chot)

            status_phieu = var_stx.driver.find_element(By.XPATH, var_stx.ag1_6).text       #Trạng thái phiếu
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 200, 2, status_phieu)

            status_pay = var_stx.driver.find_element(By.XPATH, var_stx.ag1_7).text       #Trạng thái thanh toán
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 201, 2, status_pay)

            person_chot = var_stx.driver.find_element(By.XPATH, var_stx.ag1_9).text       #người chốt
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 202, 2, person_chot)

        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 203, 2, f"Không tìm thấy hợp đồng")




        #7.7.6
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.history_debt_closing).click()
        time.sleep(5)


        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(code_hd)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_1)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(code_hd)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(3)


        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_1)
            code_congno = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_2).text       #Mã công nợ
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 205, 2, code_congno)

            code_hopdong = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_3).text       #Mã hợp đồng
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 206, 2, code_hopdong)

            time_thaotac = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_4).text       #Thời gian thao tác
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 207, 2, time_thaotac)

            person_thaotac = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_5).text       #Người thao tác
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 208, 2, person_thaotac)

            summary_no = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_6).text       #Tổng nợ
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 209, 2, summary_no)

            type_action = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_11).text       #Loại thao tác
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 210, 2, type_action)

        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 211, 2, f"Không tìm thấy hợp đồng")


        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)

        name_config_10_6_3 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 190, 2)
        mat_ban_ghi_7_7_4 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 192, 2)
        ma_phieu_chot_7_7_4 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 193, 2)
        tong_tien_chot_7_7_4 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 194, 2)
        tong_tien_chot_7_7_4 = ''.join(re.findall(r'\d+', tong_tien_chot_7_7_4))
        tong_tien_chot_7_7_4 = int(tong_tien_chot_7_7_4)


        ma_cong_no_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 196, 2)
        ma_hop_dong_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 197, 2)
        thoi_gian_chot_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 198, 2)
        thoi_gian_chot_7_7_5_check = datetime.strptime(str(thoi_gian_chot_7_7_5), "%H:%M:%S - %d/%m/%Y")
        # thoi_gian_chot_7_7_5_check = datetime.strptime(str(thoi_gian_chot_7_7_5), "%Y-%m-%d %H:%M:%S")

        so_tien_no_chot_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 199, 2)
        so_tien_no_chot_7_7_5 = ''.join(re.findall(r'\d+', so_tien_no_chot_7_7_5))
        so_tien_no_chot_7_7_5 = int(so_tien_no_chot_7_7_5)
        trang_thai_phieu_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 200, 2)
        trang_thai_phieu_7_7_5_check = trang_thai_phieu_7_7_5.lower().replace("đã", "").strip()

        trang_thai_thanh_toan_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 201, 2)
        nguoi_chot_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 202, 2)
        mat_ban_ghi_7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 203, 2)




        ma_cong_no_7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 205, 2)
        ma_hop_dong_7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 206, 2)
        thoi_gian_thao_tac_7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 207, 2)
        thoi_gian_thao_tac_7_7_6_check = datetime.strptime(str(thoi_gian_thao_tac_7_7_6), "%H:%M:%S %d/%m/%Y")

        # thoi_gian_thao_tac_7_7_6_check = datetime.strptime(str(thoi_gian_thao_tac_7_7_6), "%Y-%m-%d %H:%M:%S")

        nguoi_thao_tac_7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 208, 2)
        tong_no_7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 209, 2)
        tong_no_7_7_6 = ''.join(re.findall(r'\d+', tong_no_7_7_6))
        tong_no_7_7_6 = int(tong_no_7_7_6)
        loai_thao_tac_7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 210, 2)
        loai_thao_tac_7_7_6_check = loai_thao_tac_7_7_6.lower().strip()
        mat_ban_ghi_7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 211, 2)




        try:
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, f"1.Cấu hình trả về: {name_config_10_6_3}\n\n2.Trạng thái bản ghi: {mat_ban_ghi_7_7_4}\n\n3.Check dữ liệu:\n\n"
                                                                                    f"Mã phiếu chôt(Mã công nợ): {ma_phieu_chot_7_7_4}(7.7.4), {ma_cong_no_7_7_5}(7.7.5), {ma_cong_no_7_7_6}(7.7.6)\n\n"
                                                                                    f"Mã hợp đồng: {ma_hop_dong_7_7_5}(7.7.5), {ma_hop_dong_7_7_6}(7.7.6)\n\n"
                                                                                    f"Thời gian chốt(Thời gian thao tác): \n{thoi_gian_chot_7_7_5}(7.7.5), \n{thoi_gian_thao_tac_7_7_6}(7.7.6)\n\n"
                                                                                    f"Tổng tiền chốt(Số tiền nợ chốt, Tổng nợ): \n{tong_tien_chot_7_7_4}(7.7.4), \n{so_tien_no_chot_7_7_5}(7.7.5), \n{tong_no_7_7_6}(7.7.6)\n\n"
                                                                                    f"Trạng thái phiếu(Loại thao tác): {trang_thai_phieu_7_7_5}(7.7.5), {loai_thao_tac_7_7_6}(7.7.6)\n\n"
                                                                                    f"Trạng thái thanh toán: {trang_thai_thanh_toan_7_7_5}(7.7.5)\n\n"
                                                                                    f"Người chốt(Người thao tác): \n{nguoi_chot_7_7_5}(7.7.5), \n{nguoi_thao_tac_7_7_6}(7.7.6)\n")

            if (mat_ban_ghi_7_7_4 == "Đã mất bản ghi") \
                    and (ma_phieu_chot_7_7_4 == ma_cong_no_7_7_5 == ma_cong_no_7_7_6) \
                    and (ma_hop_dong_7_7_5 == ma_hop_dong_7_7_6) \
                    and (thoi_gian_chot_7_7_5_check == thoi_gian_thao_tac_7_7_6_check) \
                    and (tong_tien_chot_7_7_4 == so_tien_no_chot_7_7_5 == tong_no_7_7_6) \
                    and (trang_thai_phieu_7_7_5_check == loai_thao_tac_7_7_6_check) \
                    and (trang_thai_thanh_toan_7_7_5 == "Chưa thanh toán") \
                    and (nguoi_chot_7_7_5 == nguoi_thao_tac_7_7_6):
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_ChotCong_No.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_ChotCong_No.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_ChotCong_No.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_ChotCong_No.png")

        if mat_ban_ghi_7_7_5 and mat_ban_ghi_7_7_6 == "Không tìm thấy hợp đồng":
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không tìm thấy hợp đồng")


    def debt_closing_7_7_7(self, number_serial, from_day, to_day, id_company, type):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)
        code_cong_no = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 193, 2)

        if type == "Thanh toán 1 phần":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 213, 2, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 214, 2, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 215, 2, 0)

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 213, 3, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 214, 3, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 215, 3, 0)



        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)



        # 7.7.7
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.Payment_of_debts).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(code_cong_no)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)


        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, "//*[@class='ag-root-wrapper ag-layout-normal ag-ltr']//*[text()='"+code_cong_no+"']")
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(code_cong_no)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
            time.sleep(3)



        var_stx.driver.find_element(By.XPATH, var_stx.checkbox1_7_7_7).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.PaymentForm).click()
        time.sleep(1)
        wait = WebDriverWait(var_stx.driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.PaymentType)))
        time.sleep(1)
        element.click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, "//*[@id='PaymentType']//*[text()='"+type+"']").click()
        time.sleep(2)

        if type == "Thanh toán 1 phần":
            tong_no_chot_7_7_7 = var_stx.driver.find_element(By.XPATH, var_stx.Payment_MoneyClose1).get_attribute("value")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 213, 2, tong_no_chot_7_7_7)
            tong_no_chot_7_7_7 = ''.join(re.findall(r'\d+', tong_no_chot_7_7_7))
            tong_no_chot_7_7_7 = int(tong_no_chot_7_7_7)

            no_con_lai_7_7_7 = var_stx.driver.find_element(By.XPATH, var_stx.Payment_MoneyClose).get_attribute("value")
            no_con_lai_7_7_7 = ''.join(re.findall(r'\d+', no_con_lai_7_7_7))
            no_con_lai_7_7_7 = int(no_con_lai_7_7_7)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 216, 2, no_con_lai_7_7_7)

            so_tien_thanh_toan_7_7_7 = (no_con_lai_7_7_7/100) * 30
            so_tien_thanh_toan_7_7_7 = int(so_tien_thanh_toan_7_7_7)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 214, 2, so_tien_thanh_toan_7_7_7)

            var_stx.driver.find_element(By.XPATH, var_stx.Payment_MoneyPayment).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.Payment_MoneyPayment).send_keys(so_tien_thanh_toan_7_7_7)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.NoteTT).send_keys("Ghi chú thanh toán 1 phần")
            time.sleep(1.5)

            con_lai_7_7_7 = var_stx.driver.find_element(By.XPATH, var_stx.td_rema_sum).text
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 215, 2, con_lai_7_7_7)

        if type == "Thanh toán toàn bộ":
            tong_no_chot_7_7_7 = var_stx.driver.find_element(By.XPATH, var_stx.Payment_MoneyClose1).get_attribute("value")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 213, 3, tong_no_chot_7_7_7)

            no_con_lai_7_7_7 = var_stx.driver.find_element(By.XPATH, var_stx.Payment_MoneyClose).get_attribute("value")
            no_con_lai_7_7_7 = ''.join(re.findall(r'\d+', no_con_lai_7_7_7))
            no_con_lai_7_7_7 = int(no_con_lai_7_7_7)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 216, 3, no_con_lai_7_7_7)

            sotienphanbo_7_7_7 = var_stx.driver.find_element(By.XPATH, var_stx.td_allo_sum).text
            sotienphanbo_7_7_7 = ''.join(re.findall(r'\d+', sotienphanbo_7_7_7))
            sotienphanbo_7_7_7 = int(sotienphanbo_7_7_7)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 214, 3, sotienphanbo_7_7_7)


            con_lai_7_7_7 = var_stx.driver.find_element(By.XPATH, var_stx.td_rema_sum).text
            if con_lai_7_7_7 == "₫":
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 215, 3, 0)
            else:
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 215, 3, con_lai_7_7_7)

            var_stx.driver.find_element(By.XPATH, var_stx.NoteTT).send_keys("Ghi chú thanh toán toàn bộ")
            time.sleep(1.5)

        var_stx.driver.find_element(By.XPATH, var_stx.btnPayment_edit).click()
        time.sleep(3)


    def debt_closing_7_7_5(self, number_serial, from_day, to_day, id_company, type):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)
        if type == "Thanh toán 1 phần":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 217, 2, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 218, 2, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 219, 2, "")

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 217, 3, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 218, 3, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 219, 3, "")


        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)



        #7.7.5
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.List_of_debts_finalized).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
        time.sleep(3)

        code_hd = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 193, 2)
        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, "//*[@class='ag-center-cols-viewport']/div/div[1]//*[text()='"+code_hd+"']")
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial2).send_keys(number_serial)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search2).click()
            time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, "//*[@class='ag-center-cols-viewport']/div/div[1]//*[text()='"+code_hd+"']")

            if type == "Thanh toán 1 phần":
                status_phieu = var_stx.driver.find_element(By.XPATH, var_stx.ag1_6).text       #Trạng thái phiếu
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 217, 2, status_phieu)

                status_pay = var_stx.driver.find_element(By.XPATH, var_stx.ag1_7).text       #Trạng thái thanh toán
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 218, 2, status_pay)
            else:
                status_phieu = var_stx.driver.find_element(By.XPATH, var_stx.ag1_6).text  # Trạng thái phiếu
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 217, 3, status_phieu)

                status_pay = var_stx.driver.find_element(By.XPATH, var_stx.ag1_7).text  # Trạng thái thanh toán
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 218, 3, status_pay)

        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 219, 2, f"Không tìm thấy hợp đồng")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 219, 3, f"Không tìm thấy hợp đồng")


    def debt_closing_7_7_6(self, number_serial, from_day, to_day, id_company, type):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)
        code_hd = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 193, 2)

        if type == "Thanh toán 1 phần":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 221, 2, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 222, 2, "")

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 221, 3, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 222, 3, "")

        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)



        #7.7.6
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.history_debt_closing).click()
        time.sleep(5)


        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        # if type == "Thanh toán 1 phần":
        #     var_stx.driver.find_element(By.XPATH, var_stx.StatusDebtClose_1).click()
        # else:
        #     var_stx.driver.find_element(By.XPATH, var_stx.StatusDebtClose_2).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(code_hd)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(3)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_1)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            if type == "Thanh toán 1 phần":
                var_stx.driver.find_element(By.XPATH, var_stx.StatusDebtClose_1).click()
            else:
                var_stx.driver.find_element(By.XPATH, var_stx.StatusDebtClose_2).click()
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.SearchStr).send_keys(code_hd)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(3)


        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_1)
            if type == "Thanh toán 1 phần":
                type_action = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_11).text       #Loại thao tác
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 221, 2, type_action)
            else:
                type_action = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_11).text       #Loại thao tác
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 221, 3, type_action)
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 222, 2, f"Không tìm thấy hợp đồng")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 222, 3, f"Không tìm thấy hợp đồng")


    def debt_closing_7_7_10(self, number_serial, id_company, type):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)

        if type == "Thanh toán 1 phần":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 224, 2, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 225, 2, 0)

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 224, 3, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 225, 3, 0)

        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)




        #7.7.10
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.Report_customer_card_details).click()
        time.sleep(5)
        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.reportrange_left_year).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, "//*[@class='drp-calendar left']//*[@class='yearselect']//*[text()='2022']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH,"//*[@class='drp-calendar left']//*[@class='table-condensed']//*[text()='15']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.reportrange_right_year).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH,"//*[@class='drp-calendar right']//*[@class='yearselect']//*[text()='2027']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH,"//*[@class='drp-calendar left']//*[@class='table-condensed']//*[text()='15']").click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
        time.sleep(0.5)
        if type == "Thanh toán 1 phần":
            var_stx.driver.find_element(By.XPATH, var_stx.DebtStatus_2).click()
        else:
            var_stx.driver.find_element(By.XPATH, var_stx.DebtStatus_1).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(3.5)

        #Thanh toán 1 phần/Đã thanh toán
        var_stx.driver.implicitly_wait(0.05)
        n = 0
        page = 1
        while (n < 23):
            n += 1
            n = str(n)
            path_trangthaithanhtoan = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[7]"
            path_cuocxe = "//*[@class='ag-center-cols-viewport']/div/div[" + n + "]/div[12]"
            path_serial = "//*[@class='ag-center-cols-container']/div[" + n + "]/div[2]"
            cuocxe_excel = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 225, 2))
            try:
                serial = var_stx.driver.find_element(By.XPATH, path_serial).text
                trangthaithanhtoan = var_stx.driver.find_element(By.XPATH, path_trangthaithanhtoan).text
                cuocxe = var_stx.driver.find_element(By.XPATH, path_cuocxe).text
                cuocxe = ''.join(re.findall(r'\d+', cuocxe))
                print("Hàng: {}, Số Serial: {}, Trạng thái thanh toán: {}, Cuốc xe: {},".format(n, serial, trangthaithanhtoan, cuocxe))

                if type == "Thanh toán 1 phần":
                    if (serial == number_serial) and (trangthaithanhtoan == "Thanh toán 1 phần"):
                        count_cuocxe = int(cuocxe) + cuocxe_excel
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 224, 2, trangthaithanhtoan)
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 225, 2, int(count_cuocxe))
                else:
                    if (serial == number_serial) and (trangthaithanhtoan == "Đã thanh toán"):
                        count_cuocxe = int(cuocxe) + cuocxe_excel
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 224, 3, trangthaithanhtoan)
                        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 225, 3, int(count_cuocxe))

            except:
                page += 1
                page = str(page)
                try:
                    page1 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[2]").text
                    page2 = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-paging-description']/span[4]").text
                    if page1 == page2:
                        break
                    else:
                        var_stx.driver.find_element(By.XPATH, "//*[@class='ag-icon ag-icon-next']").click()
                        time.sleep(5)
                    n = 0
                    print("đã sang trang:" + page)
                except:
                    break

            n = int(n)
            page = int(page)


    def debt_closing_7_7_11(self, number_serial, from_day, to_day, id_company, type):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)

        if type == "Thanh toán 1 phần":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 227, 2, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 228, 2, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 229, 2, "")

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 227, 3, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 228, 3, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 229, 3, "")


        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)


        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_card_top_up_transactions).click()
        time.sleep(5)

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.serial).send_keys(number_serial)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.MobileSearch).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2_2)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.serial).send_keys(number_serial)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.MobileSearch).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)

        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2_2)

            if type == "Thanh toán 1 phần":
                status_phieu = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_8).text       #Trạng thái
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 227, 2, status_phieu)

                status_pay = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_10).text       #Số tiền
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 228, 2, status_pay)
            else:
                status_phieu = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_8).text  # Trạng thái
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 227, 3, status_phieu)

                status_pay = var_stx.driver.find_element(By.XPATH, var_stx.list_data2_10).text  # Số tiền
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 228, 3, status_pay)

        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 229, 2, f"Không tìm thấy hợp đồng")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 229, 3, f"Không tìm thấy hợp đồng")


    def debt_closing_7_7_14(self, number_serial, from_day, to_day, id_company, type):
        var_stx.driver.implicitly_wait(4)
        id_company = str(id_company)
        number_serial = str(number_serial)

        if type == "Thanh toán 1 phần":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 231, 2, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 232, 2, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 233, 2, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 234, 2, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 235, 2, "")

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 337, 2, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 338, 2, "")
        else:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 231, 3, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 232, 3, "")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 233, 3, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 234, 3, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 235, 3, "")

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 337, 3, 0)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 338, 3, "")


        if var_stx.linktest[0:14] == "https://g7test":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm1)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:17] == "https://g7staging":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.check_g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'], id_company)


        if var_stx.linktest[0:18] == "https://app.g7taxi":
            if id_company == "209":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hn)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)

            if id_company == "212":
                try:
                    var_stx.driver.find_element(By.XPATH, var_stx.g7_taxi_hcm)
                except:
                    login_stx.login.login_select_company(self, var_stx.data['login']['tk_admin1'], var_stx.data['login']['mk_admin1'], id_company)




        #7.7.14
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_transaction_report).click()
        time.sleep(5)

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.number_serial).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_1)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(from_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(to_day)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)


        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_1)

            if type == "Thanh toán 1 phần":
                status = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_8).text       #Trạng thái
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 231, 2, status)

                method1 = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_9).text       #Phương thức
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 232, 2, method1)

                money = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_10).text       #Số tiền
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 233, 2, money)

                money1 = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_11).text       #Số dư còn lại
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 234, 2, money1)
            else:
                status = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_8).text  # Trạng thái
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 231, 3, status)

                method1 = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_9).text  # Phương thức
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 232, 3, method1)

                money = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_10).text  # Số tiền
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 233, 3, money)

                money1 = var_stx.driver.find_element(By.XPATH, var_stx.table_table_2_11).text  # Số dư còn lại
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 234, 3, money1)

        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 235, 2, f"Không tìm thấy bản ghi")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 235, 3, f"Không tìm thấy bản ghi")







        #7.7.3
        var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.customer).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.contract_card).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.card_management).click()
        time.sleep(5)



        try:
            var_stx.driver.implicitly_wait(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(3.5)
            var_stx.driver.find_element(By.XPATH, var_stx.sodu7_7_3)

        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.number_serial).send_keys(number_serial)
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(3.5)

        try:
            var_stx.driver.implicitly_wait(0.5)
            if type == "Thanh toán 1 phần":
                sodu7_7_3 = var_stx.driver.find_element(By.XPATH, var_stx.sodu7_7_3).text
                sodu7_7_3 = int(sodu7_7_3.replace("₫", "").replace(".", "").strip())
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 237, 2, sodu7_7_3)
            else:
                sodu7_7_3 = var_stx.driver.find_element(By.XPATH, var_stx.sodu7_7_3).text
                sodu7_7_3 = int(sodu7_7_3.replace("₫", "").replace(".", "").strip())
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 237, 3, sodu7_7_3)
        except:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 238, 2, "Không tìm thấy thẻ")
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 238, 3, "Không tìm thấy thẻ")


    def check_debt_closing(self, code, eventname, result, type):
        logging.info("-------------------------")
        logging.info("KHÁCH HÀNG - 7.7 Hợp đồng thẻ & thẻ")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)

        if type == "Thanh toán 1 phần":
            #7.7.5
            trangthaithanhtoan7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 218, 2)
            trangthaithanhtoan7_7_5 = str(trangthaithanhtoan7_7_5)
            #7.7.6
            trangthaithanhtoan7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 221, 2)
            trangthaithanhtoan7_7_6 = str(trangthaithanhtoan7_7_6)
            #7.7.7
            sotienthanhtoan7_7_7 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 214, 2)
            sotienthanhtoan7_7_7 = int(sotienthanhtoan7_7_7)
            #7.7.10
            trangthaithanhtoan7_7_10 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 224, 2)
            trangthaithanhtoan7_7_10 = str(trangthaithanhtoan7_7_10)
            cuocxe_7_10 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 225, 2)
            cuocxe_7_10 = int(cuocxe_7_10)
            #7.7.11
            trangthai7_7_11 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 227, 2)
            trangthai7_7_11 = str(trangthai7_7_11)
            sotien7_7_11 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 228, 2)
            sotien7_7_11 = ''.join(re.findall(r'\d+', sotien7_7_11))
            sotien7_7_11 = int(sotien7_7_11)
            #7.7.14
            trangthai7_7_14 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 231, 2)
            trangthai7_7_14 = str(trangthai7_7_14)
            sotien7_7_14 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 233, 2)
            sotien7_7_14 = ''.join(re.findall(r'\d+', sotien7_7_14))
            sotien7_7_14 = int(sotien7_7_14)
            soduconlai7_7_14 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 234, 2)
            soduconlai7_7_14 = int(soduconlai7_7_14.replace(".", "").replace("₫", "").strip())
            #7.7.3
            sodu7_7_3 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 237, 2)

        else:
            #7.7.5
            trangthaithanhtoan7_7_5 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 218, 3)
            trangthaithanhtoan7_7_5 = str(trangthaithanhtoan7_7_5)
            #7.7.6
            trangthaithanhtoan7_7_6 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 221, 3)
            trangthaithanhtoan7_7_6 = str(trangthaithanhtoan7_7_6)
            #7.7.7
            sotienthanhtoan7_7_7 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 214, 3)
            # sotienthanhtoan7_7_7 = int(sotienthanhtoan7_7_7)
            #7.7.10
            trangthaithanhtoan7_7_10 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 224, 3)
            trangthaithanhtoan7_7_10 = str(trangthaithanhtoan7_7_10)
            cuocxe_7_10 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 225, 3)
            # cuocxe_7_10 = int(cuocxe_7_10)
            #7.7.11
            trangthai7_7_11 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 227, 3)
            trangthai7_7_11 = str(trangthai7_7_11)
            sotien7_7_11 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 228, 3)
            sotien7_7_11 = ''.join(re.findall(r'\d+', sotien7_7_11))
            sotien7_7_11 = int(sotien7_7_11)
            #7.7.14
            trangthai7_7_14 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 231, 3)
            trangthai7_7_14 = str(trangthai7_7_14)
            sotien7_7_14 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 233, 3)
            sotien7_7_14 = ''.join(re.findall(r'\d+', sotien7_7_14))
            sotien7_7_14 = int(sotien7_7_14)
            soduconlai7_7_14 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 234, 3)
            soduconlai7_7_14 = int(soduconlai7_7_14.replace(".", "").replace("₫", "").strip())
            #7.7.3
            sodu7_7_3 = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 237, 3)






        if type == "Thanh toán 1 phần":
            try:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, f"(1): {sotienthanhtoan7_7_7}\n"
                                                                                        f"(2): {trangthaithanhtoan7_7_5}\n(3): {trangthaithanhtoan7_7_6}\n"
                                                                                        f"(4): {trangthaithanhtoan7_7_10}\n(5): {trangthai7_7_11}\n"
                                                                                        f"(6): {sotien7_7_11}\n(7): {sotien7_7_14}\n"
                                                                                        f"(8): {soduconlai7_7_14}\n(9): {sodu7_7_3}\n")
                if (trangthaithanhtoan7_7_5 == "Thanh toán 1 phần") and (trangthaithanhtoan7_7_6 == "Thanh toán 1 phần") and \
                        (trangthaithanhtoan7_7_10 == "Thanh toán 1 phần") and (trangthai7_7_11 == "Nạp tiền") and \
                        (sotien7_7_11 == sotienthanhtoan7_7_7) and (sotien7_7_14 == sotienthanhtoan7_7_7) and \
                        (soduconlai7_7_14 == sodu7_7_3):
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_Checkthongtin.png")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_Checkthongtin.png")
            except:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_Checkthongtin.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_Checkthongtin.png")


        if type == "Thanh toán toàn bộ":
            try:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, f"(1): {sotienthanhtoan7_7_7}\n"
                                                                                        f"(2): {trangthaithanhtoan7_7_5}\n(3): {trangthaithanhtoan7_7_6}\n"
                                                                                        f"(4): {trangthaithanhtoan7_7_10}\n(5): {trangthai7_7_11}\n"
                                                                                        f"(6): {sotien7_7_11}\n(7): {sotien7_7_14}\n"
                                                                                        f"(8): {soduconlai7_7_14}\n(9): {sodu7_7_3}\n")
                if (trangthaithanhtoan7_7_5 == "Đã thanh toán") and (trangthaithanhtoan7_7_6 == "Thanh toán 1 phần") and \
                        (trangthaithanhtoan7_7_10 == "Đã thanh toán") and (trangthai7_7_11 == "Nạp tiền") and \
                        (sotien7_7_11 == sotienthanhtoan7_7_7) and (sotien7_7_14 == sotienthanhtoan7_7_7) and \
                        (soduconlai7_7_14 == sodu7_7_3):
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_Checkthongtin2.png")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_Checkthongtin2.png")
            except:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_Checkthongtin2.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_Checkthongtin2.png")


        if type == "Thanh toán toàn bộ_toàn bộ":
            try:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, f"(1): {sotienthanhtoan7_7_7}\n"
                                                                                        f"(2): {trangthaithanhtoan7_7_5}\n(3): {trangthaithanhtoan7_7_6}\n"
                                                                                        f"(4): {trangthaithanhtoan7_7_10}\n(5): {trangthai7_7_11}\n"
                                                                                        f"(6): {sotien7_7_11}\n(7): {sotien7_7_14}\n"
                                                                                        f"(8): {soduconlai7_7_14}\n(9): {sodu7_7_3}\n")
                if (trangthaithanhtoan7_7_5 == "Đã thanh toán") and (trangthaithanhtoan7_7_6 == "Thanh toán toàn bộ") and \
                        (trangthaithanhtoan7_7_10 == "Đã thanh toán") and (trangthai7_7_11 == "Nạp tiền") and \
                        (sotien7_7_11 == sotienthanhtoan7_7_7) and (sotien7_7_14 == sotienthanhtoan7_7_7) and \
                        (soduconlai7_7_14 == sodu7_7_3):
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_Checkthongtin3.png")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_Checkthongtin3.png")
            except:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_TheTraSau_Checkthongtin3.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_TheTraSau_Checkthongtin3.png")












    def get_number_serial(self, code):
        wordbook = openpyxl.load_workbook(var_stx.checklistpath)
        sheet = wordbook.get_sheet_by_name("Checklist")
        rownum = 1
        while (rownum < 1000):
            rownum += 1
            rownum = str(rownum)
            number_code = sheet["A" + rownum].value
            number_serial = sheet["N" + rownum].value
            if number_code == code:
                print(number_code)
                print(number_serial)
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 188, 2, number_serial)
                break
            rownum = int(rownum)


    def check_card_debt_closing_1phan(self, code, eventname, result):
        var_stx.driver.implicitly_wait(4)
        contract_card.get_number_serial(self, code)
        number_serial = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 188, 2)
        type_config = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 190, 2)


        if type_config == "false":
            contract_card.debt_closing_7_7_7(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_5(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_6(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_10(self, number_serial, "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_11(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_14(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.check_debt_closing(self, code, eventname, result, "Thanh toán 1 phần")

        if type_config == "true":
            contract_card.debt_closing_7_7_7(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_14(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_5(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_6(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_10(self, number_serial, "209", "Thanh toán 1 phần")
            contract_card.debt_closing_7_7_11(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán 1 phần")
            contract_card.check_debt_closing(self, code, eventname, result, "Thanh toán 1 phần")


    def check_card_debt_closing_toanbo_1phan(self, code, eventname, result):
        var_stx.driver.implicitly_wait(4)
        contract_card.get_number_serial(self, code)
        number_serial = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 188, 2)
        type_config = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 190, 2)

        if type_config == "false":
            contract_card.debt_closing_7_7_7(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_5(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_6(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_10(self, number_serial, "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_11(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_14(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.check_debt_closing(self, code, eventname, result, "Thanh toán toàn bộ")

        if type_config == "true":
            contract_card.debt_closing_7_7_7(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_14(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_5(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_6(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_10(self, number_serial, "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_11(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.check_debt_closing(self, code, eventname, result, "Thanh toán toàn bộ")


    def check_card_debt_closing_toanbo_toanbo(self, code, eventname, result):
        var_stx.driver.implicitly_wait(4)
        contract_card.get_number_serial(self, code)
        number_serial = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 188, 2)
        type_config = var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 190, 2)


        if type_config == "false":
            contract_card.card_debt_closing(self, "", "", "", number_serial, "01/01/2023 00:02","09/02/2027 00:00", "209")
            contract_card.debt_closing_7_7_7(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_5(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_6(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_10(self, number_serial, "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_11(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_14(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.check_debt_closing(self, code, eventname, result, "Thanh toán toàn bộ_toàn bộ")

        if type_config == "true":
            contract_card.card_debt_closing(self, "", "", "", number_serial, "01/01/2023 00:02","09/02/2027 00:00", "209")
            contract_card.debt_closing_7_7_7(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_14(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_5(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_6(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_10(self, number_serial, "209", "Thanh toán toàn bộ")
            contract_card.debt_closing_7_7_11(self, number_serial, "01/01/2023 00:02", "09/02/2027 00:00", "209", "Thanh toán toàn bộ")
            contract_card.check_debt_closing(self, code, eventname, result, "Thanh toán toàn bộ_toàn bộ")

