import logging
import var_stx
import time
import json
from selenium.webdriver.common.by import By
import module_other_stx
import login_stx
from seleniumwire.utils import decode as sw_decode
from selenium.webdriver.common.keys import Keys
import os
import shutil
import re
import openpyxl
from xls2xlsx import XLS2XLSX
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
wait = WebDriverWait(var_stx.driver, 10)
from selenium.webdriver.common.action_chains import ActionChains
from seleniumwire import webdriver




def write_from_moth(fromdate_mon_input):
    var_stx.driver.implicitly_wait(1)

    from_date = time.strftime("%d/%m/%Y, ", time.localtime())
    from_date = str(from_date)
    print(from_date)

    from_date_year = from_date[6:10]
    print(from_date_year)

    from_moth = from_date[3:5]
    from_moth = int(from_moth)

    from_date_day = from_date[0:2]
    print(from_date_day)
    from_date_day = int(from_date_day)
    if from_date_day == 31 or from_date_day == 30 or from_date_day == 29 or from_date_day == 28:
        from_date_day = "01"

        print("vao 0")
        from_full = str(from_date_day) + str(from_moth) + str(from_date_year)
        print(from_full)
        delete = var_stx.driver.find_element(By.XPATH, fromdate_mon_input)
        delete.send_keys(Keys.CONTROL, "a")
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, fromdate_mon_input).send_keys(from_full)
        time.sleep(2)
    else:
        from_moth1 = from_moth - 1
        if from_moth1 == 10:
            from_moth1 = "10"
        if from_moth1 == 9:
            from_moth1 = "09"
        if from_moth1 == 8:
            from_moth1 = "08"
        if from_moth1 == 7:
            from_moth1 = "07"
        if from_moth1 == 6:
            from_moth1 = "06"
        if from_moth1 == 5:
            from_moth1 = "05"
        if from_moth1 == 4:
            from_moth1 = "04"
        if from_moth1 == 3:
            from_moth1 = "03"
        if from_moth1 == 2:
            from_moth1 = "02"
        if from_moth1 == 1:
            from_moth1 = "01"
        if from_moth1 == 0:
            from_moth1 = "01"

        if from_date_day == 10:
            from_date_day = "10"
        if from_date_day == 9:
            from_date_day = "9"
        if from_date_day == 8:
            from_date_day = "8"
        if from_date_day == 7:
            from_date_day = "7"
        if from_date_day == 6:
            from_date_day = "6"
        if from_date_day == 5:
            from_date_day = "5"
        if from_date_day == 4:
            from_date_day = "4"
        if from_date_day == 3:
            from_date_day = "3"
        if from_date_day == 2:
            from_date_day = "2"
        if from_date_day == 1:
            from_date_day = "1"


        print(from_moth1)
        print("vào 1")
        from_full = str(from_date_day) + "/" + str(from_moth1) + "/" + str(from_date_year)
        print(from_full)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 3, 2, from_full)
        delete = var_stx.driver.find_element(By.XPATH, fromdate_mon_input)
        delete.send_keys(Keys.CONTROL, "a")
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, fromdate_mon_input).send_keys(from_full)
        time.sleep(2)



def write_from_date(fromdate_date_input):
    var_stx.driver.implicitly_wait(1)

    from_date = time.strftime("%d/%m/%Y", time.localtime())
    from_date = str(from_date)
    print(from_date)

    from_date_year = from_date[6:10]
    print(from_date_year)

    from_moth = from_date[3:5]
    from_moth = int(from_moth)
    print(from_moth)


    from_date_day = from_date[0:2]
    from_date_day = int(from_date_day)
    from_date_day = str(from_date_day)
    print(from_date_day)
    from_full = str(from_date_day) + "/" + str(from_moth) + "/" + str(from_date_year)
    print(from_full)

    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 3, 2, from_full)
    delete = var_stx.driver.find_element(By.XPATH, fromdate_date_input)
    delete.send_keys(Keys.CONTROL, "a")
    time.sleep(2)
    var_stx.driver.find_element(By.XPATH, fromdate_date_input).send_keys(from_full)
    time.sleep(2)



def clearData_luutamthoi_checkexcel(file, sheetName, column1, column2, column3, column4, column5, column6):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 119
    while (i < 150):
        i += 1
        i = str(i)
        sheet["A"+i] = column1
        sheet["B"+i] = column2
        sheet["C"+i] = column3
        sheet["D"+i] = column4
        sheet["E"+i] = column5
        sheet["F"+i] = column6
        i = int(i)
    wordbook.save(file)


def clearData_luutamthoi_checkexcel1(file, sheetName, row_to, row_end, column1, column2):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = row_to
    while (i < row_end):
        i += 1
        i = str(i)
        sheet["B"+i] = column1
        sheet["C"+i] = column2
        i = int(i)
    wordbook.save(file)


def get_info_web():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table table-hover table-bordered']/tbody/tr[1]/th[" + n + "]"
        path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[2]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web:" .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)



def get_info_web_new():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 1
    while (n < 50):
        n += 1
        n = str(n)
        row += 1

        path_column = f"(//div[@aria-colindex='{n}'])[1]"
        path_data = f"(//div[@aria-colindex='{n}'])[2]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web:" .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)






def get_info_excel(row, sheet):
    row2 = row + 1

    try:
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xls"))

        x2x = XLS2XLSX(var_stx.excelpath + "/baocao_stx.xls")
        x2x.to_xlsx(var_stx.excelpath + "/baocao_stx.xlsx")
    except:
        # var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xls"))

        x2x = XLS2XLSX(var_stx.excelpath + "/baocao_stx.xls")
        x2x.to_xlsx(var_stx.excelpath + "/baocao_stx.xlsx")

    # #Đọc check file excel
    bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                  'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL']

    print("r0")
    try:
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)
    except:
        var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)

    print("r1")
    row_tamthoi = 119
    for i in bangchucai:
        row_tamthoi += 1
        if str(sheet[str(i + str(row))].value) == "None":
            break

        cloumn = str(i + str(row))
        print("vị trí tên cột excel: ".format(cloumn))

        cloumn2 = str(i + str(row2))
        print("vị trí data cột excel: ".format(cloumn2))

        name_column = str(sheet[cloumn].value)
        print("Tên cột excel: ".format(name_column))

        data_column = str(sheet[cloumn2].value)
        print(data_column)
        print("Data cột excel: ".format(data_column))

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 3, name_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 4, data_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 5, cloumn)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 6, cloumn2)




def get_info_excel1(row, sheet):
    row2 = row + 1

    try:
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))

    except:
        # var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))


    # #Đọc check file excel
    bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                  'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL',
                  'AM', 'AN', 'AO']

    print("r0")
    try:
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)
    except:
        # var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)

    print("r1")
    row_tamthoi = 119
    for i in bangchucai:
        row_tamthoi += 1
        if str(sheet[str(i + str(row))].value) == "None":
            break

        cloumn = str(i + str(row))
        print("vị trí tên cột excel: ".format(cloumn))

        cloumn2 = str(i + str(row2))
        print("vị trí data cột excel: ".format(cloumn2))

        name_column = str(sheet[cloumn].value)
        print("Tên cột excel: ".format(name_column))

        data_column = str(sheet[cloumn2].value)
        print(data_column)
        print("Data cột excel: ".format(data_column))

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 3, name_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 4, data_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 5, cloumn)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 6, cloumn2)



def get_info_excel_skip(row, sheet, row_skip, row_end):
    row2 = row + row_skip

    try:
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))

    except:
        # var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))


    # #Đọc check file excel
    bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                  'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL',
                  'AM', 'AN', 'AO']

    print("r0")
    try:
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)
    except:
        # var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)

    print("r1")
    row_tamthoi = 119
    for i in bangchucai:
        row_tamthoi += 1
        # if str(sheet[str(i + str(row))].value) == "None":
        if row_tamthoi == row_end:
            break

        cloumn = str(i + str(row))
        print("vị trí tên cột excel: ".format(cloumn))

        cloumn2 = str(i + str(row2))
        print("vị trí data cột excel: ".format(cloumn2))

        name_column = str(sheet[cloumn].value)
        print("Tên cột excel: ".format(name_column))

        data_column = str(sheet[cloumn2].value)
        print(data_column)
        print("Data cột excel: ".format(data_column))

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 3, name_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 4, data_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 5, cloumn)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 6, cloumn2)



def check_info_web_excel(code, eventname, result, path_module):
    row = 119
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)


    while (row < 150):
        row += 1
        name_column_web = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 1))
        data_column_web = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 2))
        name_column_excel = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 3))
        data_column_excel = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 4))
        location_name_coloumn = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 5))
        location_data_coloumn = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 6))
        print(f" row1: {row}")

        if name_column_web == "None":
            break


        logging.info("-------------------------")
        logging.info("Tên cột web:   " + name_column_web)
        logging.info("Tên cột excel: " + name_column_excel)
        if ''.join(name_column_web.split()).lower() == ''.join(name_column_excel.split()).lower():
            logging.info("True")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            if name_column_web in ['Logout', 'Chi tiết', 'Xóa', 'Logout lái xe', 'Ảnh lái xe', 'Đổi mật khẩu', 'Xóa', 'Chi tiết thiết bị', 'Khen thưởng',
                                   'Gán xe', 'Tiện ích', 'Chi tiết cuốc', 'Cuốc khách GPS', 'Trạng thái thanh toán', 'Khóa thẻ', 'QRCode', 'Nạp tiền',
                                   'Nạp tiền KM', 'Đặt hộ', 'Hủy bỏ', 'Chi tiết cuốc', 'Lộ trình GPS', 'Chi tiết đề cử', 'Hợp đồng chuyến đi',
                                   'Bản đồ trước đề cử', 'Bản đồ sau đề cử', 'Bản đồ xe trước đề cử', 'Bản đồ xe sau đề cử', 'Bản đồ xe\nsau đề cử',
                                   'Bản đồ xe\ntrước đề cử', 'Tổng số cuốc đề cử (1)', 'Tỷ lệ không hoàn thành/ Tổng số cuốc đề cử (2) = (6 + 7 + 8 + 9 + 10)',
                                   'Tổng số cuốc nhận được (3)', 'Tỷ lệ lái xe hủy/ Tổng số cuốc nhận (4)', 'Xem hóa đơn', 'Xuất hóa đơn', 'Lộ trình',
                                   'Thực thu', 'Vai trò', 'Phân quyền', 'Cấu hình', 'Khóa', 'Xóa', 'Lấy mã kích hoạt', 'GPS', 'APP', 'Báo cáo']:
                pass
            else:
                logging.error("False")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Tên cột web: {}\nTên cột excel: {} \nDòng: {}"
                                          .format(name_column_web, name_column_excel, location_name_coloumn))
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_name_coloumn))





        logging.info("Dữ liệu web:   " + data_column_web)
        logging.info("Dữ liệu excel: " + data_column_excel)
        if name_column_web in ["STT", "Biển số", "Mã đàm", "Đang lái", "Phiên bản xe", "Thời điểm vào ca", "Số cuốc", "Thời điểm mất kết nối",
                               "Số lần mất kết nối", "Trạng thái kết nối", "Hộp đen", 'Tên xe', 'Trạng thái', 'Mô hình quản lý', 'Lý do khóa/mở khóa',
                               'Thời gian mở/khóa', 'Số sao', 'Số điểm', 'Tên đăng nhập', 'Mật khẩu', 'Mã kích hoạt', 'CCCD', 'Số điện thoại', 'Ngày tạo',
                               'Lần đăng nhập cuối', 'Mô hình quản lý', 'Phiên bản ứng dụng', 'Trạng thái khóa', 'Lý do khóa/mở khóa', 'Thời gian mở/khóa',
                               'Xe đã gán', 'Chi tiết ứng dụng', 'Mã ví', 'Ví lái xe', 'Tên lái xe', 'Ngày giao dịch', 'Mã giao dịch', 'Hình thức', 'Loại hình phí',
                               'Số serial thẻ', 'Số dư ví', 'Ghi chú', 'Mã khuyến mại', 'Giá trị', 'Số lần sử dụng', 'Công ty', 'Tên khuyến mại', 'Số tiền phải thu',
                               'Số tiền phải trả', 'Kiểu khuyến mại', 'Giá trị KM', 'Kiểu thời gian KM', 'Nguồn', 'Mã đàm', 'Biển số xe', 'SĐT', 'Thời gian đặt cuốc',
                               'Địa chỉ đón', 'Địa chỉ trả', 'Số tiền cuốc', 'Số tiền KM', 'Tổng tiền khuyến mại', 'Tổng số KM đã sử dụng', 'Tổng KM đã sử dụng',
                               'Tổng số khuyến mại', 'Tổng tiền chưa sử dụng khuyến mại', 'Tổng số mã khuyến mại chưa sử dụng', 'Ngày',
                               'Tổng tài khoản', 'KHACH HANG', 'LAI XE', 'NHAN VIEN', 'BAN GIÁM ĐỐC', 'MGT-HDQT-100k', 'MGT-DDPL-50k', 'Mã hợp đồng', 'Số Serial',
                               'Tên công ty', 'Tên khách hàng', 'Số dư', 'Hạn mức', 'Ngày mở', 'Hạn dùng', 'Loại thẻ', 'Mã cuốc khách', 'Mã cuốc', 'Tên đối tác',
                                'Chi tiết phụ phí', 'Phương thức', 'Thanh toán', 'Thanh toán thẻ', 'Ngày dùng', 'Điểm đón', 'Điểm trả', 'Loại',
                               'Ngày nạp tiền', 'Người nạp tiền', 'Mã lái xe', 'Số thẻ', 'Thời gian cuốc đặt', 'Ngày đăng ký', 'Số dư tài khoản chính', 'Số dư tài khoản khuyến mại'
                               'Tổng nạp', 'Tổng tiền sử dụng', 'Ngày nạp tiền gần nhất', 'Tổng số cuốc', 'Ngày phát sinh cuốc gần nhất', 'Serial tặng/nhận', 'Lái xe',
                               'Khách hàng', 'Cước xe', 'Tổng cuốc', 'Tổng không đề cử', 'Tổng khách huỷ', 'Tổng lái xe huỷ', 'Tổng thành công', 'Không có lái xe nhận',
                               'Cấu hình', 'Nguồn mobile', 'Nguồn điều hành', 'Nguồn vãng lai', 'Tổng khách hủy', 'Tổng lái xe hủy', 'Nguồn khách vẫy', 'Nguồn đối tác',
                               'Phương pháp điều', 'Số lần đề cử', 'Số lái xe đề cử', 'Địa chỉ đón khách', 'Hành Khách', 'Bộ cấu hình điều', 'Lần đề cử', 'Thời gian cuốc đến',
                               'Khoảng cách(m)', 'Thời gian xác nhận', 'Thời gian huỷ cuốc', 'Thời gian gặp khách', 'Địa chỉ trả khách', 'Trạng thái cuốc', 'Khoảng\ncách (m)',
                               'Tổng số cuốc đề cử', 'Tỷ lệ từ chối/\nTổng số cuốc đề cử', 'Tổng số cuốc nhận', 'Tỷ lệ hủy/\nTổng số cuốc nhận', 'Hoàn thành',
                               'Quá giờ không nhận', 'Không được chọn', 'Từ chối', 'Nhầm khách', 'Khách hàng hủy', 'Lái xe hủy', 'Tỷ lệ không hoàn thành/', 'Tỷ lệ lái xe hủy/',
                               'Hoàn thành (5)', 'Quá giờ không nhận (6)', 'Từ chối (7)', 'Nhầm khách (8)', 'Khách hàng hủy (9)', 'Lái xe hủy (10)', 'Số Km theo Gps',
                               'Số Km theo tablet', 'Doanh thu qua điều hành', 'Số cuốc qua điều hành', 'Doanh thu qua mobile', 'Số cuốc qua mobile', 'Doanh thu qua vãng lai',
                               'Số cuốc qua vãng lai', 'Doanh thu cuốc taxi3', 'Số cuốc taxi3', 'Tổng doanh thu', 'Số tiền thu của lái xe', 'Số tiền khuyến mại', 'Doanh thu cuốc từ định vị',
                               'Số cuốc theo định vị', 'Tổng KM', 'Nguồn cuốc', 'Loại cuốc', 'Thời gian kết thúc', 'Tổng cuốc',
                               'HTTT khi đặt cuốc', 'HTTT cuốc', 'Km có khách (GPS)', 'Biển số xe (GPS)', 'Số hiệu xe (GPS)', 'Loại xe (GPS)', 'Thời gian đón khách (GPS)', 'Thời gian trả khách (GPS)',
                               'Thời gian đón khách (Đồng hồ - GPS)', 'Thời gian trả khách (Đồng hồ - GPS)', 'Địa điểm đón khách (GPS)', 'Địa điểm trả khách (GPS)', 'Km rỗng (GPS)',
                               'Tiền đồng hồ (GPS)', 'Tiền thực thu (GPS)', 'Thời gian chờ (GPS)', 'Km có khách GPS (GPS)', 'Tài khoản', 'Nhóm khách hàng', 'S.cuốc app', 'S.cuốc v.lai',
                               'Hình thức thanh toán', 'Tên loại hàng hóa', 'Phụ phí(%)', 'Mô tả', 'Loại hàng hóa', 'Giá tiền nhỏ nhất', 'Giá tiền lớn nhất', 'Loại phụ phí',
                               'Giá tiền nhỏ nhất(kg)', 'Giá tiền lớn nhất(kg)', 'Loại ví', 'Mã thẻ KH', 'T.gian đặt', 'T.gian kết thúc', 'Mã KM', 'Mã GD', 'T.tin t.toán', 'KM có khách ',
                            'Phí phải nộp', 'TG thông báo LX', 'TG lái xe xác nhận', 'Nội dung lỗi', 'Tổng số tin nhắn', 'Tin đăng ký', 'Tin giới thiệu', 'Tin đường dài', 'Tin sân bay',
                               'Tin khác', 'Tin OTP mã pin', 'BookID', 'Mã cuốc đối tác', 'HTTT', 'Khách hủy', 'Số hiệu', 'Km thực hiện', 'SĐT khách hàng', 'Khuyến mãi']:

            print("name vao 1" + name_column_web)
            data_column_excel = data_column_excel.replace("_x000D_", "\n")
            data_column_excel = data_column_excel.strip()
            if data_column_web == data_column_excel:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.error("False")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
                                          .format(data_column_web, data_column_excel, location_data_coloumn))
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_data_coloumn))

        print("đã đến đây")
        print(''.join(name_column_web.split()).lower())
        if ''.join(name_column_web.split()).lower() in ["Sốđiệnthoạikháchhàng", 'cướcdichuyển(1)', 'cướcchờ(2)']:
            print("đã vào đây")
            if data_column_web == "None":
                data_column_web = "0"
            print(data_column_web)

            data_column_web = ''.join(data_column_web.split()).lower()
            data_column_web = ''.join(re.findall(r'\d+', data_column_web))

            data_column_excel = ''.join(data_column_excel.split()).lower()
            data_column_excel = ''.join(re.findall(r'\d+', data_column_excel))

            if data_column_web == data_column_excel:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.error("False")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
                                          .format(data_column_web, data_column_excel, location_data_coloumn))
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_data_coloumn))







        if name_column_web in ['Km GPS', 'NL tiêu thụ', 'Số tiền', 'Cước phí', 'Tiền nợ', 'Cước phí(cước thật của cuốc)', 'Cước xe (trừ của khách)', 'Số dư còn lại', 'Phụ phí', 'Loại xe',
                               'D.thu ĐH', 'S.cuốc ĐH', 'D.thu App', 'D.thu v.lai', 'D.thu đ.vị', 'Tổng d.thu', 'S.tiền KM', 'Khoảng cách', 'S.tiền thu LX', 'Khuyến mại', 'Cước xe (trừ của khách)',
                               'Doanh thu tính %\n(1) = (2) + (3)', 'Cước xe\n(2)', 'Khuyến mãi\n(3)', 'Số km', 'Cước phí (cước thật của cuốc)']:
            print("name vao 2" + name_column_web)
            try:
                data_column_web = ''.join(re.findall(r'\d+', data_column_web))[:3]
                data_column_excel = ''.join(re.findall(r'\d+', data_column_excel))[:3]
            except Exception as e:
                logging.error(f"Lỗi khi xử lý dữ liệu: {e}")

            if data_column_web == data_column_excel:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.error("False")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
                                          .format(data_column_web, data_column_excel, location_data_coloumn))
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_data_coloumn))


        if name_column_web in ['Đánh giá', 'Thời gian chờ', 'Cước di chuyển', 'Cước chờ', 'Doanh thu', 'Tiền típ', 'Số tiền thanh toán', 'Mã KM/Số tiền KM', 'Số Km']:
            print("name vao 2" + name_column_web)
            try:
                data_column_web = ''.join(re.findall(r'\d+', data_column_web))[:1]
                data_column_web = int(data_column_web)
                data_column_excel = ''.join(re.findall(r'\d+', data_column_excel))[:1]
                data_column_excel = int(data_column_excel)
            except Exception as e:
                logging.error(f"Lỗi khi xử lý dữ liệu: {e}")
            if data_column_web == data_column_excel:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.error("False")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
                                          .format(data_column_web, data_column_excel, location_data_coloumn))
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_data_coloumn))




def check_info_web_excel1(code, eventname, result, path_module):
    row = 119
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)


    while (row < 150):
        row += 1
        name_column_web = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 1))
        data_column_web = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 2))
        data_column_excel = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 4))
        location_data_coloumn = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 6))
        print(f" row1: {row}")

        if name_column_web == "None":
            break

        logging.info("-------------------------")
        logging.info("Dữ liệu web: " + data_column_web)
        logging.info("Dữ liệu excel: " + data_column_excel)
        if data_column_web == data_column_excel:
            logging.info("True")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            print("sửa data web: {}".format(data_column_web[0:2]))
            print("sửa data excel: {}".format(data_column_excel[0:2]))
            data_column_web = ''.join(re.findall(r'\d+', data_column_web))
            data_column_excel = ''.join(re.findall(r'\d+', data_column_excel))

            if data_column_web[0:2] == data_column_excel[0:2]:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.error("False")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Dữ liệu web:: {}\nDữ liệu excel: {} \nDòng: {}"
                                          .format(data_column_web, data_column_excel, location_data_coloumn))
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_data_coloumn))




def get_src_vehicle(type, src, from_cut):
    var_stx.driver.implicitly_wait(0.05)
    n = 0
    while (n < 153):
        n += 1
        n = str(n)
        print("n: " + n)
        path_vehicle_type = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]"
        path_vehicle_src = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[1]/img"
        path_vehicle_codedam = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[2]"
        path_vehicle_time = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[3]"
        path_vehicle_connect = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[4]"
        path_vehicle_v = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[5]"
        try:
            check_vehicle_class = var_stx.driver.find_element(By.XPATH, path_vehicle_type).get_attribute("class")
            check_vehicle_type = var_stx.driver.find_element(By.XPATH, path_vehicle_type).get_attribute("typedata")
            check_vehicle_src = var_stx.driver.find_element(By.XPATH, path_vehicle_src).get_attribute("src")
            print("---------")
            print("class: " + check_vehicle_class)
            print("type: " + check_vehicle_type)
            print("src: " + check_vehicle_src)
            print("src cut: " + check_vehicle_src[from_cut::])

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 4, 2, check_vehicle_type)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 5, 2, check_vehicle_src[from_cut::])
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 6, 2, path_vehicle_src)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 6, 3, path_vehicle_codedam)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 6, 4, path_vehicle_time)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 6, 5, path_vehicle_connect)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 6, 6, path_vehicle_v)


            if check_vehicle_type == type and check_vehicle_src[from_cut::] == src and check_vehicle_class != "notDisplay":
                print("break r")
                break

        except:
            pass
        n = int(n)



def get_info_vehicle_popup(field, row):
    var_stx.driver.implicitly_wait(0.05)
    n = 0
    while (n < 20):
        n += 1
        n = str(n)
        print("n: " + n)
        path_name = "//*[@id='tabsCurrent']/table/tbody/tr[" + n + "]/td[1]"
        path_detail = "//*[@id='tabsCurrent']/table/tbody/tr[" + n + "]/td[2]"

        try:
            name = var_stx.driver.find_element(By.XPATH, path_name).text
            detail = var_stx.driver.find_element(By.XPATH, path_detail).text
            print(name)
            print(detail)
            if name == field:
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 3, detail)
        except:
            pass
        n = int(n)















class vehicle_online:

    def vehicle_online(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        # login_stx.login.login_stx(self, var_stx.data['login']['tk_admin'], var_stx.data['login']['mk_admin'])
        var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.vehicle_online).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.vehicle_online).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                 var_stx.title_page, "1.2 Xe online", "_GiamSat_XeOnline.png")



    def vehicle_online1(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx2(self, var_stx.data['login']['tk_admin'], var_stx.data['login']['mk_admin'])

        var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.vehicle_online).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.vehicle_online).click()
        time.sleep(2.5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                 var_stx.title_page, "1.2 Xe online", "_GiamSat_XeOnline.png")




    def vehicle_online_test(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx_g7test(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])
        var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.vehicle_online).click()
        time.sleep(2.5)


    def search_vehicle_x(self):
        var_stx.driver.implicitly_wait(0.2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search_vehicle_input).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.2)
        except:
            pass

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.vehicle_input1).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.2)
        except:
            pass

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.vehicle_input).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.2)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search_driver_input).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.2)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search_into_recently_input).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.2)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search_lost_signal_input).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.2)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search_status_online).click()
            time.sleep(0.2)
        except:
            pass
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search_status_blackbox).click()
            time.sleep(0.2)
        except:
            pass



    def search_vehicle(self, code, eventname, result, type_search, path_data, search_input, path_check, desire, path_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_vehicle_online)
        except:
            vehicle_online.vehicle_online(self, "", "", "")


        vehicle_online.search_vehicle_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(10)

        if type_search == "0":
            data = var_stx.driver.find_element(By.XPATH, path_data).text
            var_stx.driver.find_element(By.XPATH, search_input).click()
            var_stx.driver.find_element(By.XPATH, search_input).send_keys(data)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  path_check, data, path_image)
            try:
                var_stx.driver.find_element(By.XPATH, path_check)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có dữ liệu")




        if type_search == "1":
            data = var_stx.driver.find_element(By.XPATH, path_data).text
            data = data.split('-')[0]
            print(result)
            var_stx.driver.find_element(By.XPATH, search_input).send_keys(data)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
            var_stx.driver.implicitly_wait(1)
            logging.info("-------------------------")
            logging.info("Giám sát - Xe online 1.2")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            try:
                check_text = var_stx.driver.find_element(By.XPATH, path_check).text
                check_text = check_text.split('-')[0]
                logging.info(check_text)
                logging.info(data)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

                if check_text == data:
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có dữ liệu")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)


        if type_search == "2":
            var_stx.driver.implicitly_wait(0.05)
            n = 1
            while (n < 50):
                n += 1
                n = str(n)
                path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[8]"
                print(n)
                try:
                    name_data = var_stx.driver.find_element(By.XPATH, path_data).text
                    if name_data != "":
                        name_data = name_data.split(' ')[1]
                        var_stx.driver.find_element(By.XPATH, search_input).send_keys(name_data)
                        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                        time.sleep(10)
                        var_stx.driver.implicitly_wait(1)
                        logging.info("-------------------------")
                        logging.info("Giám sát - Xe online 1.2")
                        logging.info("Mã - " + code)
                        logging.info("Tên sự kiện - " + eventname)
                        logging.info("Kết quả - " + result)
                        try:
                            check_text = var_stx.driver.find_element(By.XPATH, path_check).text
                            check_text = check_text.split(' ')[1]
                            logging.info(check_text)
                            logging.info(name_data)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

                            if check_text == name_data:
                                logging.info("True")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                            else:
                                logging.info("False")
                                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        except:
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6,"Không có dữ liệu")
                            var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        break
                    print(name_data)
                except:
                    pass
                n = int(n)


        if type_search == "6":
            var_stx.driver.implicitly_wait(0.05)
            n = 1
            while (n < 50):
                n += 1
                n = str(n)
                path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[7]"
                print(n)
                try:
                    name_data = var_stx.driver.find_element(By.XPATH, path_data).text
                    if name_data != "":
                        name_data = name_data.split(' ')[1]
                        var_stx.driver.find_element(By.XPATH, search_input).send_keys(name_data)
                        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                        time.sleep(10)
                        var_stx.driver.implicitly_wait(1)
                        logging.info("-------------------------")
                        logging.info("Giám sát - Xe online 1.2")
                        logging.info("Mã - " + code)
                        logging.info("Tên sự kiện - " + eventname)
                        logging.info("Kết quả - " + result)
                        try:
                            check_text = var_stx.driver.find_element(By.XPATH, path_check).text
                            check_text = check_text.split(' ')[1]
                            logging.info(check_text)
                            logging.info(name_data)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

                            if check_text == name_data:
                                logging.info("True")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                            else:
                                logging.info("False")
                                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        except:
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6,"Không có dữ liệu")
                            var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        break
                    print(name_data)
                except:
                    pass
                n = int(n)


        if type_search == "3":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
            module_other_stx.write_result_text_try_if_other(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  path_check, "", path_image)
            try:
                var_stx.driver.find_element(By.XPATH, path_check)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có dữ liệu")



        if type_search == "4":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  path_check, desire, path_image)
            try:
                var_stx.driver.find_element(By.XPATH, path_check)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có dữ liệu")


        if type_search == "5":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
            module_other_stx.write_result_not_displayed_try(code, eventname, result, "Giám sát - Xe online 1.2",
                                                   desire, path_image)
            try:
                var_stx.driver.find_element(By.XPATH, path_check)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có dữ liệu")


        if type_search == "7":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                   path_check, desire, path_image)
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_11)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có dữ liệu")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, "")


        if type_search == "8":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(10)
            var_stx.driver.implicitly_wait(1)
            logging.info("-------------------------")
            logging.info("Giám sát - Xe online 1.2")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            try:
                time1 = var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_8).text
                status = var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_10).text
                logging.info(time1)
                logging.info(status)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Thời điểm mất kết nối: {}\nTrạng thái kết nối: {}".format(time1, status))

                if time1 != "" and status == "Đóng":
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có dữ liệu")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)




    def search_vehicle1(self, code, eventname, result, type_search, path_data, search_input, path_check, desire, path_image):
        var_stx.driver.implicitly_wait(5)
        vehicle_online.vehicle_online1(self, "", "", "")


        vehicle_online.search_vehicle_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(2)

        if type_search == "0":
            data = var_stx.driver.find_element(By.XPATH, path_data).text
            var_stx.driver.find_element(By.XPATH, search_input).click()
            var_stx.driver.find_element(By.XPATH, search_input).send_keys(data)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(2)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  path_check, data, path_image)


        if type_search == "1":
            data = var_stx.driver.find_element(By.XPATH, path_data).text
            data = data.split('-')[0]
            print(result)
            var_stx.driver.find_element(By.XPATH, search_input).send_keys(data)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(2)
            var_stx.driver.implicitly_wait(1)
            logging.info("-------------------------")
            logging.info("Giám sát - Xe online 1.2")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            try:
                check_text = var_stx.driver.find_element(By.XPATH, path_check).text
                check_text = check_text.split('-')[0]
                logging.info(check_text)
                logging.info(data)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

                if check_text == data:
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
            except:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)

            #
            #
            #
            # module_other_stx.write_result_text_try_if_cut2(code, eventname, result, "Giám sát - Xe online 1.2",
            #                                       path_check, var_stx.data['minitor']['vehicle_online_search1'], path_image, 0, 10)


        if type_search == "2":
            # write_from_date(search_input)
            # write_from_date(var_stx.search_into_recently_input)
            # time_check = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 3, 2))
            # var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            # time.sleep(2)
            # from_date = time.strftime("%d", time.localtime())
            # from_date = int(from_date)
            # if from_date in [1, 2, 3, 4, 5, 6, 7, 8, 9]:
            #     module_other_stx.write_result_text_try_if_cut1(code, eventname, result, "Giám sát - Xe online 1.2",
            #                                           path_check, time_check, path_image, -9)
            # else:
            #     module_other_stx.write_result_text_try_if_cut1(code, eventname, result, "Giám sát - Xe online 1.2",
            #                                           path_check, time_check, path_image, -10)
            vehicle = var_stx.driver.find_element(By.XPATH, var_stx.list_data2).text
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 2, 2, vehicle)
            var_stx.driver.implicitly_wait(0.05)
            n = 1
            while (n < 50):
                n += 1
                n = str(n)
                path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[8]"
                print(n)
                try:
                    name_data = var_stx.driver.find_element(By.XPATH, path_data).text
                    if name_data != "":
                        name_data = name_data.split(' ')[1]
                        var_stx.driver.find_element(By.XPATH, search_input).send_keys(name_data)
                        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                        time.sleep(2)
                        var_stx.driver.implicitly_wait(1)
                        logging.info("-------------------------")
                        logging.info("Giám sát - Xe online 1.2")
                        logging.info("Mã - " + code)
                        logging.info("Tên sự kiện - " + eventname)
                        logging.info("Kết quả - " + result)
                        try:
                            check_text = var_stx.driver.find_element(By.XPATH, path_check).text
                            check_text = check_text.split(' ')[1]
                            logging.info(check_text)
                            logging.info(name_data)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

                            if check_text == name_data:
                                logging.info("True")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                            else:
                                logging.info("False")
                                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        except:
                            logging.info("False")
                            var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        break
                    print(name_data)
                except:
                    pass
                n = int(n)


        if type_search == "6":

            var_stx.driver.implicitly_wait(0.05)
            n = 1
            while (n < 50):
                n += 1
                n = str(n)
                path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[" + n + "]/td[7]"
                print(n)
                try:
                    name_data = var_stx.driver.find_element(By.XPATH, path_data).text
                    if name_data != "":
                        name_data = name_data.split(' ')[1]
                        var_stx.driver.find_element(By.XPATH, search_input).send_keys(name_data)
                        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                        time.sleep(2)
                        var_stx.driver.implicitly_wait(1)
                        logging.info("-------------------------")
                        logging.info("Giám sát - Xe online 1.2")
                        logging.info("Mã - " + code)
                        logging.info("Tên sự kiện - " + eventname)
                        logging.info("Kết quả - " + result)
                        try:
                            check_text = var_stx.driver.find_element(By.XPATH, path_check).text
                            check_text = check_text.split(' ')[1]
                            logging.info(check_text)
                            logging.info(name_data)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

                            if check_text == name_data:
                                logging.info("True")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                            else:
                                logging.info("False")
                                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        except:
                            logging.info("False")
                            var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
                        break
                    print(name_data)
                except:
                    pass
                n = int(n)


        if type_search == "3":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(2)
            module_other_stx.write_result_text_try_if_other(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  path_check, "", path_image)


        if type_search == "4":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(2)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  path_check, desire, path_image)


        if type_search == "5":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(2)
            module_other_stx.write_result_not_displayed_try(code, eventname, result, "Giám sát - Xe online 1.2",
                                                   desire, path_image)


        if type_search == "7":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(2)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                   path_check, desire, path_image)
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_11)
            except:
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Không có xe có khách")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, "")


        if type_search == "8":
            var_stx.driver.find_element(By.XPATH, search_input).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(2)
            var_stx.driver.implicitly_wait(1)
            logging.info("-------------------------")
            logging.info("Giám sát - Xe online 1.2")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            try:
                time1 = var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_8).text
                status = var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_10).text
                logging.info(time1)
                logging.info(status)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Thời điểm mất kết nối: {}\nTrạng thái kết nối: {}".format(time1, status))

                if time1 != "" and status == "Đóng":
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)
            except:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + path_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + path_image)







    def vehicle_online_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        module_other_stx.delete_excel()
        clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_vehicle_online)
        except:
            vehicle_online.vehicle_online(self, "", "", "")

        vehicle_online.search_vehicle_x(self)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(2)


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(20)
        get_info_web()
        try:
            get_info_excel(5, "Sheet 1")
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
            time.sleep(30)
            get_info_excel(5, "Sheet 1")

        check_info_web_excel(code, eventname, result, "Giám sát - Xe online 1.2")



    def vehicle_online_logout_all(self, code, eventname, result, type_logout, type_check, path_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_vehicle_online)
        except:
            vehicle_online.vehicle_online(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.logout_all).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, type_logout).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.next).click()
        time.sleep(1.5)
        if type_check == "0":
            var_stx.driver.find_element(By.XPATH, var_stx.resson_logout).send_keys(var_stx.data['minitor']['logout_resson'])
            time.sleep(1)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  var_stx.check_vehicle_online_logout_all0, "Bạn chắc nhắn muốn logout các lái xe đã chọn ?", path_image)

        if type_check == "1":
            var_stx.driver.find_element(By.XPATH, var_stx.handle).click()
            time.sleep(1.5)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Xe online 1.2",
                                                  var_stx.check_vehicle_online_logout_all1, "files is null.", path_image)


        var_stx.driver.find_element(By.XPATH, var_stx.icon_x1).click()
        time.sleep(2)



    def vehicle_online_logout_driver(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_vehicle_online)
        except:
            vehicle_online.vehicle_online(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.check_list_data12).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.resson_logout_driver).send_keys(var_stx.data['minitor']['logout_resson'])
        time.sleep(1)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "Giám sát - Xe online 1.2",
                                              var_stx.check_vehicle_online_logout_driver, "", "GiamSat_LogouTaiXe.png")

        var_stx.driver.find_element(By.XPATH, var_stx.icon_x1).click()
        time.sleep(1)





class route:

    def dowload_route(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        day = time.strftime("%d", time.localtime())
        day = int(day)
        day = day - 1
        if day == 0:
            day = "01"
        if day == 1:
            day = "01"
        if day == 2:
            day = "02"
        if day == 3:
            day = "03"
        if day == 4:
            day = "04"
        if day == 5:
            day = "05"
        if day == 6:
            day = "06"
        if day == 7:
            day = "07"
        if day == 8:
            day = "08"
        if day == 9:
            day = "09"
        if day == 10:
            day = "10"
        moth_year = time.strftime("/%m/%Y 20:00", time.localtime())
        from_date = str(day) + str(moth_year)
        print(from_date)

        vehicle_online.vehicle_online1(self, "", "", "")
        var_stx.driver.find_element(By.XPATH, var_stx.OnlineState).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.OnlineState_3).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(4)
        try:
            vehicle = var_stx.driver.find_element(By.XPATH, var_stx.list_data2).text
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 2, 2, vehicle)
        except:
            var_stx.driver.refresh()
            time.sleep(6)
            var_stx.driver.find_element(By.XPATH, var_stx.State).click()
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.state_2).click()
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(4)
            vehicle = var_stx.driver.find_element(By.XPATH, var_stx.list_data2).text
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 2, 2, vehicle)


        var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.route).click()
        time.sleep(5)
        if var_stx.driver.find_element(By.XPATH, var_stx.detail_route).is_selected() == False:
            var_stx.driver.find_element(By.XPATH, var_stx.detail_route).click()
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.detail_route).click()
            time.sleep(0.5)
        if var_stx.driver.find_element(By.XPATH, var_stx.route_blackbox).is_selected() == False:
            var_stx.driver.find_element(By.XPATH, var_stx.route_blackbox).click()
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.route_blackbox).click()
            time.sleep(0.5)

        vehicle = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 2, 2))
        delete = var_stx.driver.find_element(By.XPATH, var_stx.route_input)
        delete.send_keys(Keys.CONTROL, "a")

        var_stx.driver.find_element(By.XPATH, var_stx.route_input).send_keys(vehicle)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, "//*[@class='es-list']//*[@value='"+vehicle+"']").click()
        time.sleep(1)

        var_stx.driver.find_element(By.XPATH, var_stx.StrFromDate).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.StrFromDate).send_keys(from_date)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.icon_account).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.get_data).click()
        time.sleep(10)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.get_data_icon).click()
            time.sleep(10)

        logging.info("-------------------------")
        logging.info("Giám sát - Lộ trình 1.7")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            route_time = var_stx.driver.find_element(By.XPATH, var_stx.list_data2).text
            route_status = var_stx.driver.find_element(By.XPATH, var_stx.list_data3).text
            route_power = var_stx.driver.find_element(By.XPATH, var_stx.list_data4).text
            logging.info(route_time)
            logging.info(route_status)
            logging.info(route_power)
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Thời gian: {}\nTrạng thái: {}\nNguồn: {}"
                                       .format(route_time, route_status, route_power))

            if (route_time and route_status and route_power) != "":
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSat_LoTrinh_LayDuLieu.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSat_LoTrinh_LayDuLieu.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSat_LoTrinh_LayDuLieu.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSat_LoTrinh_LayDuLieu.png")


    def route_icon(self, code, eventname, result, path_icon, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2)
        except:
            route.dowload_route(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, path_icon).click()
        time.sleep(1)
        module_other_stx.write_result_displayed_try(code, eventname, result, "Giám sát - Lộ trình 1.7",
                                                        path_icon, name_image)


    def route_print(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2)
        except:
            route.dowload_route(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.icon_print).click()
        time.sleep(2)
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[1])
        time.sleep(3)
        logging.info("-------------------------")
        logging.info("Giám sát - Lộ trình 1.7")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            vehicle = var_stx.driver.find_element(By.XPATH, var_stx.route_print_vehicle).text
            date = var_stx.driver.find_element(By.XPATH, var_stx.route_print_date).text
            logging.info(vehicle)
            logging.info(date)
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Biển số xe: {}\nThời gian: {}"
                                       .format(vehicle, date))
            if (vehicle and date) != "":
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSat_LoTrinh_In.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSat_LoTrinh_In.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSat_LoTrinh_In.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSat_LoTrinh_In.png")
        module_other_stx.close_tab()
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(1)


    def route_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        module_other_stx.delete_excel()
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2)
        except:
            route.dowload_route(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel1).click()
        time.sleep(15)
        module_other_stx.write_result_dowload_file(code, eventname, result, "Giám sát - Lộ trình 1.7",
                                                        "_GiamSat_LoTrinh_XuatExcel.xls", "_GiamSat_LoTrinh_XuatExcel.png")

        try:
            check_route_excel = var_stx.driver.find_element(By.XPATH, var_stx.check_route_excel).text
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_route_excel)
            var_stx.driver.back()
            time.sleep(3)
        except:
            pass


    def route_help(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.route_speed)
        except:
            route.dowload_route(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.icon_help).click()
        time.sleep(2)
        logging.info("-------------------------")
        logging.info("Giám sát - Lộ trình 1.7")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            help1 = var_stx.driver.find_element(By.XPATH, var_stx.route_help1).text
            help2 = var_stx.driver.find_element(By.XPATH, var_stx.route_help2).text
            help3 = var_stx.driver.find_element(By.XPATH, var_stx.route_help3).text
            help4 = var_stx.driver.find_element(By.XPATH, var_stx.route_help4).text
            logging.info(help1)
            logging.info(help2)
            logging.info(help3)
            logging.info(help4)
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Màu Vàng: {}\nMàu đỏ: {}\nMàu xanh dương: {}\nMàu đen: {}"
                                       .format(help1, help2, help3, help4))
            if (help1 == "Cảnh báo") and (help2 == "Sẵn sàng") and (help3 == "Có khách") and (help4 == "Nghỉ KD"):
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSat_LoTrinh_TroGiup.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSat_LoTrinh_TroGiup.png")
        except:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSat_LoTrinh_TroGiup.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSat_LoTrinh_TroGiup.png")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.icon_x2).click()
            time.sleep(1)
        except:
            pass





class minitor_vehicle:


    def minitor_vehicle(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx_company(self, var_stx.data['login']['tk_company'], var_stx.data['login']['mk_company'])
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Giám sát xe 1.3",
                                                 var_stx.title_page, "1.3 Giám sát xe", "_GiamSatXe.png")


    def minitor_vehicle_icon_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle)
        except:
            minitor_vehicle.minitor_vehicle(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_status0).click()
        time.sleep(2)
        code_dam = var_stx.driver.find_element(By.XPATH, var_stx.list_data2).text
        print("vehicle:" + code_dam)
        code_dam = ''.join(re.findall(r'\d+', code_dam))
        print("vehicle:" + code_dam)
        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_input).send_keys(code_dam)
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.icon_search).click()
        time.sleep(2)
        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Giám sát xe 1.3",
                                                 "//*[@id='googleMap']//*[@class='gm-style']/div[1]/div[1]//*[text()='"+code_dam+"']",
                                                  code_dam, "_GiamSatXe_TimKiem.png")


    def minitor_vehicle_icon(self, code, eventname, result, path_icon, path_save, path_check, desire, name_image, type_check):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle)
        except:
            minitor_vehicle.minitor_vehicle(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, path_icon).click()
        time.sleep(2)
        if type_check == "0":
            var_stx.driver.find_element(By.XPATH, path_save).click()
            time.sleep(0.5)
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Giám sát xe 1.3",
                                                      path_check, desire, name_image)
            time.sleep(3)

        if type_check == "1":
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Giám sát xe 1.3",
                                                      path_check, desire, name_image)

            var_stx.driver.find_element(By.XPATH, path_save).click()
            time.sleep(4)

        if type_check == "2":
            module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - Giám sát xe 1.3",
                                                      path_check, desire, name_image)

            var_stx.driver.find_element(By.XPATH, var_stx.icon_x2).click()
            time.sleep(1)


    def minitor_vehicle_satus(self, code, eventname, result, path_status, type_check, path_check, desire, name_image, from_cut, type_status):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle)
        except:
            minitor_vehicle.minitor_vehicle(self, "", "", "")
            time.sleep(5)

        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, path_status).click()
        time.sleep(3)
        if type_check == "0":
            get_src_vehicle(type_status, desire, from_cut)
            path_src = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 6, 2))

            module_other_stx.write_result_text_try_if_src1(code, eventname, result, "Giám sát - Giám sát xe 1.3",
                                                  path_src, desire, name_image, from_cut)


        if type_check == "1":
            logging.info("-------------------------")
            logging.info("Giám sát - Giám sát xe 1.3")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            try:
                check_text = var_stx.driver.find_element(By.XPATH, path_check).get_attribute("src")
                logging.info(check_text)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)
                if check_text == "/Content/themes/img/BlueCar.png" or check_text == "/Content/themes/img/RedCar.png":
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
            except:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)


        if type_check == "2":
            summary_vehicle = var_stx.driver.find_element(By.XPATH, path_check).text
            summary_vehicle1 = summary_vehicle[0:1]
            summary_vehicle1 = int(summary_vehicle1)
            logging.info("-------------------------")
            logging.info("Giám sát - Giám sát xe 1.3")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            logging.info(summary_vehicle1)
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, summary_vehicle)
            if summary_vehicle1 == 0:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)


        if type_check == "3":
            summary_vehicle = var_stx.driver.find_element(By.XPATH, path_check).text
            summary_vehicle1 = summary_vehicle[0:1]
            summary_vehicle1 = int(summary_vehicle1)
            logging.info("-------------------------")
            logging.info("Giám sát - Giám sát xe 1.3")
            logging.info("Mã - " + code)
            logging.info("Tên sự kiện - " + eventname)
            logging.info("Kết quả - " + result)
            logging.info(summary_vehicle1)
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, summary_vehicle)
            if summary_vehicle1 > 0:
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)


    def minitor_vehicle_minitor_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle)
        except:
            minitor_vehicle.minitor_vehicle(self, "", "", "")
            time.sleep(5)

        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_status0).click()
        time.sleep(3)
        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_minitor_search_input).send_keys(var_stx.data['minitor']['minitor_search'])
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, "/html/body/ul[1]//*[text()='"+var_stx.data['minitor']['minitor_search']+"']").click()
        time.sleep(2)
        vehicle_popup = var_stx.driver.find_element(By.XPATH, var_stx.minitor_search_vehicle_popup).text
        vehicle_popup1 = ''.join(re.findall(r'\d+', vehicle_popup))
        vehicle_list = var_stx.data['minitor']['minitor_search']
        vehicle_list1 = ''.join(re.findall(r'\d+', vehicle_list))


        logging.info("Giám sát - Giám sát xe 1.3")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        logging.info(vehicle_popup)
        logging.info(vehicle_list)
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Biển số tìm kiếm: {}\nBiển số popup: {}".format(vehicle_list, vehicle_popup))
        if vehicle_popup1 == vehicle_list1:
            logging.info("True")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSatXe_GiamSat_TimKiem.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSatXe_GiamSat_TimKiem.png")


    def minitor_vehicle_route_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle)
        except:
            minitor_vehicle.minitor_vehicle(self, "", "", "")
            time.sleep(5)


        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_route).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(var_stx.data['minitor']['route_fromday'])
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_input).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(var_stx.data['minitor']['route_today'])
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_input).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_route_search_input).send_keys(var_stx.data['minitor']['minitor_search'])
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, "/html/body/ul[2]//*[text()='"+var_stx.data['minitor']['minitor_search']+"']").click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.icon_search2).click()
        time.sleep(2)
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[1])
        time.sleep(3)
        module_other_stx.write_result_text_try_if_title(code, eventname, result, "Giám sát - Giám sát xe 1.3",
                                                  "THEO DÕI LỘ TRÌNH - STAXI", "_GiamSatXe_LoTrinh_TimKiem.png")

        module_other_stx.close_tab()
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(1)








    def info_vehicle(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        wait = WebDriverWait(var_stx.driver, 10)
        #
        # try:
        #     var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle)
        # except:
        minitor_vehicle.minitor_vehicle(self, "", "", "")
        time.sleep(5)

        del var_stx.driver.requests
        time.sleep(1)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.minitor_vehicle_status1)))
            element.click()
        except:
            var_stx.driver.refresh()
            time.sleep(5)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.minitor_vehicle_status1)))
            element.click()
        time.sleep(3)
        get_src_vehicle("1", "/Content/themes/img/RedCar.png", -30)
        path_src = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 6, 2))
        path_code_dam = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 6, 3))
        path_time = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 6, 4))
        path_connect = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 6, 5))
        path_v = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 6, 6))

        var_stx.driver.execute_script("document.body.style.zoom='70%'")




        #Lấy thông tin Danh sách giám sát
        code_dam_list = var_stx.driver.find_element(By.XPATH, path_code_dam).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 9, 2, code_dam_list)

        path_time_list = var_stx.driver.find_element(By.XPATH, path_time).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 13, 2, path_time_list)

        path_v_list = var_stx.driver.find_element(By.XPATH, path_v).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 14, 2, path_v_list)

        path_connect_list = var_stx.driver.find_element(By.XPATH, path_connect).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 21, 2, path_connect_list)


        #Lấy thông tin popup
        var_stx.driver.find_element(By.XPATH, path_src).click()
        time.sleep(7)
        var_stx.driver.find_element(By.XPATH, var_stx.icon_location).click()
        time.sleep(7)

        path_code_dam_popup = var_stx.driver.find_element(By.XPATH, var_stx.path_code_dam_popup).text
        len_code_dam = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 9, 2))
        len_code_dam = len(len_code_dam)
        code_dam_popup = path_code_dam_popup[0:len_code_dam]
        vehicle_popup = path_code_dam_popup[len_code_dam + 1::]
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 9, 3, code_dam_popup)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 8, 3, vehicle_popup)

        location_popup = var_stx.driver.find_element(By.XPATH, var_stx.location_popup).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 10, 3, location_popup)

        latitude_popup = var_stx.driver.find_element(By.XPATH, var_stx.latitude_popup).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 11, 3, latitude_popup)

        longitude_popup = var_stx.driver.find_element(By.XPATH, var_stx.longitude_popup).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 12, 3, longitude_popup)

        get_info_vehicle_popup("Thời gian:", 13)
        time_full = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 13, 3))
        time1 = time_full[-8::]
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 13, 3, time1)

        get_info_vehicle_popup("Vận tốc GPS:", 14)

        # get_info_vehicle_popup("Loại xe:", 15)
        type_vehicle_popup = var_stx.driver.find_element(By.XPATH, var_stx.type_vehicle_popup).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 15, 3, type_vehicle_popup)


        get_info_vehicle_popup("Trạng thái:", 16)
        status = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 5, 2))
        if status == "/Content/themes/img/RedCar.png":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 16, 2, "Sẵn sàng")

        get_info_vehicle_popup("Cuốc khách:", 17)
        get_info_vehicle_popup("Doanh thu:", 18)

        # get_info_vehicle_popup("Lái xe:", 19)
        # driver_full = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 19, 3))
        # driver_full1 = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 19, 4))
        #
        # len_driver = len(driver_full1)
        # driver = driver_full[-len_driver::]
        # var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 19, 3, driver)
        driver_popup = var_stx.driver.find_element(By.XPATH, var_stx.driver_popup).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 19, 3, driver_popup)

        get_info_vehicle_popup("Điện thoại:", 20)
        get_info_vehicle_popup("Kết nối:", 21)
        get_info_vehicle_popup("Phiên bản ứng dụng:", 22)

        #Lấy thông tin api
        var_stx.driver.find_element(By.XPATH, path_src).click()
        time.sleep(2)


        # print("chuẩn bị lấy api")
        # for request in var_stx.driver.requests:
        #     print("chuẩn bị lấy api1")
        #     print(request.url)
        #     if request.url[0:44] == "https://app.staxi.vn/Trace/VehicleInfoDetail":
        #         print("đã lấy api1")
        #         data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
        #         data1 = data1.decode("utf8")
        #         res = json.loads(data1)
        #         print(res['VehicleTypeName'])
        #         var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 15, 4, res['VehicleTypeName'])
        #         var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 18, 4, res['TotalMoney'])
        #         var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 19, 4, res['DriverName'])
        #         var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 20, 4, res['DriverMobile'])
        #         var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 22, 4, res['Version'])
        #
        # for request in var_stx.driver.requests:
        #     print(request.url)
        #     if request.url == "https://app.staxi.vn/Base/GetAddress":
        #         print("đã lấy api2")
        #         data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
        #         data1 = data1.decode("utf8")
        #         print(data1)
        #         var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 10, 4, data1)
        # var_stx.driver.execute_script("document.body.style.zoom='100%'")
        logging.info("-------------------------")
        logging.info("Giám sát xe 1.3")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        logging.info("True")
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")















    def check_info_none(self, code, eventname, result, row):
        var_stx.driver.execute_script("document.body.style.zoom='100%'")
        info_data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 3))
        logging.info("-------------------------")
        logging.info("Giám sát xe 1.3")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        logging.info("Popup Hiện trạng:      - " + info_data)
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Popup Hiện trạng: {}".format(info_data))
        if info_data != "None":
            logging.info("True")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")


    def check_info_popup_list(self, code, eventname, result, row):
        info_popup = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 3))
        info_list = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 2))

        logging.info("-------------------------")
        logging.info("Giám sát xe 1.3")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        logging.info("Popup Hiện trạng:           - " + info_popup)
        logging.info("Danh sách giám sát:         - " + info_list)
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Popup Hiện trạng: {}\nDanh sách giám sát: {}"
                                   .format(info_popup, info_list))
        if info_popup == info_list:
            logging.info("True")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")


    def check_info_popup_api(self, code, eventname, result, row):
        info_popup = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 3))
        info_api = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 4))

        logging.info("-------------------------")
        logging.info("Giám sát xe 1.3")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        logging.info("Popup Hiện trạng:  - " + info_popup)
        logging.info("API:               - " + info_api)
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Popup Hiện trạng: {}\nApi: {}"
                                   .format(info_popup, info_api))
        if info_popup == info_api:
            logging.info("True")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")









    def minitor_vehicle_group(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx2(self, var_stx.data['login']['tk_admin'], var_stx.data['login']['mk_admin'])

        var_stx.driver.find_element(By.XPATH, var_stx.minitor).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group).click()
        time.sleep(10)
        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - 1.9 Giám sát xe (nhóm)",
                                                 var_stx.title_page, "1.9 Giám sát xe (nhóm)", "_GiamSatXeNhom.png")

        var_stx.driver.implicitly_wait(2)
        print("n0")
        try:
            var_stx.driver.implicitly_wait(0.3)
            element = var_stx.driver.find_element(By.XPATH, var_stx.icon_more)
            var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            var_stx.driver.execute_script("arguments[0].click();", element)
            # var_stx.driver.find_element(By.XPATH, var_stx.icon_more).click()
            print("n1")
        except:
            print("n2")
            pass

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.zomm_out).click()
            print("n3")
        except:
            print("n4")
            var_stx.driver.refresh()
            time.sleep(15)
            element = var_stx.driver.find_element(By.XPATH, var_stx.icon_more)
            var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            var_stx.driver.execute_script("arguments[0].click();", element)
            # var_stx.driver.find_element(By.XPATH, var_stx.icon_more).click()
            print("n5")
            var_stx.driver.find_element(By.XPATH, var_stx.zomm_out).click()
            print("n6")
        try:
            var_stx.driver.implicitly_wait(0.3)
            var_stx.driver.find_element(By.XPATH, var_stx.zomm_out).click()
            print("n7")
        except:
            element = var_stx.driver.find_element(By.XPATH, var_stx.icon_more)
            var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            var_stx.driver.execute_script("arguments[0].click();", element)
            # var_stx.driver.find_element(By.XPATH, var_stx.icon_more).click()
            # print("n8")

        time.sleep(1)
        print("n9")
        var_stx.driver.find_element(By.XPATH, var_stx.zomm_out).click()
        time.sleep(1)
        print("n10")
        var_stx.driver.find_element(By.XPATH, var_stx.zomm_out).click()
        time.sleep(2.5)
        print("n11")
        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search_vehicle_input1).click()
            print("n12")
        except:
            try:
                print("n12")
                element = var_stx.driver.find_element(By.XPATH, var_stx.icon_show)
                var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                var_stx.driver.execute_script("arguments[0].click();", element)
                print("n13")
            except:
                print("cách 1 fail")
            time.sleep(1.5)


    def minitor_vehicle_group_icon_config(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle_group)
        except:
            minitor_vehicle.minitor_vehicle_group(self, "", "", "")

        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, var_stx.icon_config).click()
        except:
            try:
                element = var_stx.driver.find_element(By.XPATH, var_stx.icon_show)
                var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                var_stx.driver.execute_script("arguments[0].click();", element)
            except:
                print("cách 1 fail")
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.icon_config).click()

        time.sleep(3)
        var_stx.driver.find_element(By.XPATH, var_stx.igree2).click()
        time.sleep(1)
        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - 1.9 Giám sát xe (nhóm)",
                                                 var_stx.update_success, "Cập nhật thành công!", "_GiamSatXeNhom_IConCauHinh.png")
        time.sleep(2)


    def minitor_vehicle_group_icon_help(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle_group)
        except:
            minitor_vehicle.minitor_vehicle_group(self, "", "", "")

        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, var_stx.icon_help).click()
        except:
            try:
                element = var_stx.driver.find_element(By.XPATH, var_stx.icon_show)
                var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                var_stx.driver.execute_script("arguments[0].click();", element)
            except:
                print("cách 1 fail")
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.icon_help).click()

        time.sleep(2)
        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - 1.9 Giám sát xe (nhóm)",
                                                 var_stx.help1, "Trợ giúp", "_GiamSatXeNhom_IConTroGiup.png")
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.icon_close1).click()
            time.sleep(1)
        except:
            pass


    def minitor_vehicle_group_status(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle_group)
        except:
            minitor_vehicle.minitor_vehicle_group(self, "", "", "")

        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group_status).click()
        except:
            try:
                element = var_stx.driver.find_element(By.XPATH, var_stx.icon_show)
                var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                var_stx.driver.execute_script("arguments[0].click();", element)
            except:
                print("cách 1 fail")
            time.sleep(2.3)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group_status).click()

        time.sleep(1.5)
        logging.info("-------------------------")
        logging.info("Giám sát - 1.9 Giám sát xe (nhóm)")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            status_0 = var_stx.driver.find_element(By.XPATH, var_stx.status_0).text
            status_1 = var_stx.driver.find_element(By.XPATH, var_stx.status_1).text
            status_2 = var_stx.driver.find_element(By.XPATH, var_stx.status_2).text
            status_3 = var_stx.driver.find_element(By.XPATH, var_stx.status_3).text
            status_5 = var_stx.driver.find_element(By.XPATH, var_stx.status_5).text
            status_8 = var_stx.driver.find_element(By.XPATH, var_stx.status_8).text
            status__1 = var_stx.driver.find_element(By.XPATH, var_stx.status__1).text

            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "{}, {}, {}, {}, {}, {}, {}"
                                       .format(status_0, status_1, status_2, status_3, status_5, status_8, status__1))
            if (status_0 == "Trực tuyến") and (status_1 == "Sẵn sàng") and (status_2 == "Có khách") and (status_3 == "Bận") and\
                    (status_5 == "Ngừng KD") and (status_8 == "Mất tín hiệu") and (status__1 == "Ngoại tuyến"):
                logging.info("True")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSatNhom_TrangThai.png")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSatNhom_TrangThai.png")
        except Exception as e:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + "_GiamSatNhom_TrangThai.png")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_GiamSatNhom_TrangThai.png")
            logging.error("Exception occurred: " + str(e))

        var_stx.driver.find_element(By.XPATH, var_stx.search_vehicle_input1).click()
        time.sleep(1)


    def minitor_vehicle_group_group(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle_group)
        except:
            minitor_vehicle.minitor_vehicle_group(self, "", "", "")

        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group_group).click()
        except:
            try:
                element = var_stx.driver.find_element(By.XPATH, var_stx.icon_show)
                var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                var_stx.driver.execute_script("arguments[0].click();", element)
            except:
                print("cách 1 fail")
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group_group).click()

        time.sleep(1.5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "Giám sát - 1.9 Giám sát xe (nhóm)",
                                                 var_stx.minitor_vehicle_group_group, "", "_GiamSatXeNhom.png")

        var_stx.driver.find_element(By.XPATH, var_stx.search_vehicle_input1).click()
        time.sleep(1)


    def minitor_vehicle_group_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle_group)
        except:
            minitor_vehicle.minitor_vehicle_group(self, "", "", "")


        try:
            var_stx.driver.implicitly_wait(1)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group_search_vehicle2)
        except:
            try:
                element = var_stx.driver.find_element(By.XPATH, var_stx.icon_show)
                var_stx.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                var_stx.driver.execute_script("arguments[0].click();", element)
            except:
                print("cách 1 fail")
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group_search_vehicle2)

        var_stx.driver.save_screenshot(var_stx.imagepath + code + "_.png")
        module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + "_.png")
        vehicle = var_stx.driver.find_element(By.XPATH, var_stx.minitor_vehicle_group_search_vehicle2a).get_attribute("data-plate")

        var_stx.driver.find_element(By.XPATH, var_stx.search_vehicle_input1).clear()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search_vehicle_input1).send_keys(vehicle)
        time.sleep(1.5)
        try:
            var_stx.driver.find_element(By.XPATH, "//*[@data-plate='"+vehicle+"']").click()
            time.sleep(1.5)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.check_minitor_vehicle_group_search)))
            time.sleep(1)
        except:
            pass

        module_other_stx.write_result_text_try_if(code, eventname, result, "Giám sát - 1.9 Giám sát xe (nhóm)",
                                                 var_stx.check_minitor_vehicle_group_search, vehicle, "_GiamSatXeNhom_TimXe.png")


    def check_info_group_none(self, code, eventname, result, path_check, name_image):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_minitor_vehicle_group_search)
        except:
            minitor_vehicle.minitor_vehicle_group_search(self, "", "", "")


        logging.info("-------------------------")
        logging.info("Giám sát - 1.9 Giám sát xe (nhóm)")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        if name_image == "_GiamSatXeNhom_HienTrang_LoaiXe.png":
            try:
                info_data = var_stx.driver.find_element(By.XPATH, path_check).text
                logging.info("Popup Hiện trạng:      - " + info_data)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Popup Hiện trạng: {}".format(info_data))
                if info_data != "":
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
            except Exception as e:
                var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
                logging.error("Exception occurred: " + str(e))

        else:
            try:
                info_data = var_stx.driver.find_element(By.XPATH, path_check).text
                logging.info("Popup Hiện trạng:      - " + info_data)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Popup Hiện trạng: {}".format(info_data))
                if info_data != "":
                    logging.info("True")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
                else:
                    logging.info("False")
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                    var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                    module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
            except Exception as e:
                logging.info("False")
                var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
                module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
                logging.error("Exception occurred: " + str(e))




