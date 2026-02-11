import logging
from selenium.common.exceptions import NoSuchElementException
import var_stx
import time
import openpyxl
import subprocess
from selenium.webdriver.common.keys import Keys
import os
from selenium.webdriver.common.by import By
from retry import retry
import os
import shutil
import module_stx
from playsound import playsound
from gtts import gTTS
import json
import requests
from requests.auth import HTTPBasicAuth
from selenium.common.exceptions import WebDriverException
# chiều 31/12/2025







CALL_BUTTON = ("533", "1931")   # tọa độ nút CALL đã đo
CHECK_INTERVAL = 1               # giây kiểm tra trạng thái cuộc gọi


def run_adb(cmd):
    """
    Chạy adb an toàn trên Windows.
    - Fix UnicodeDecodeError bằng utf-8 + ignore lỗi decode
    - Luôn trả về string, không bao giờ None
    """
    try:
        result = subprocess.run(
            ["adb"] + cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore"
        )
        return result.stdout or ""
    except Exception as e:
        print("ADB ERROR:", e)
        return ""


def get_call_state():
    """
    Trả trạng thái cuộc gọi hiện tại từ dumpsys telecom.
    """
    output = run_adb(["shell", "dumpsys", "telecom"])

    for line in output.splitlines():
        if "[Call" in line:
            if "state=ACTIVE" in line:
                return "ACTIVE"
            if "state=DISCONNECTED" in line:
                return "DISCONNECTED"
            if "state=RINGING" in line:
                return "RINGING"
            if "state=DIALING" in line:
                return "DIALING"

    return "NONE"


def auto_call(phone_number: str, timeout: int = 20) -> str:
    """
    Gọi điện → theo dõi → tự cúp.

    Returns:
        ANSWERED
        REJECTED
        TIMEOUT
    """

    if not phone_number:
        return "EMPTY"

    print(f"\n📞 Đang gọi: {phone_number}")

    # 1. Mở màn hình quay số
    run_adb([
        "shell", "am", "start",
        "-a", "android.intent.action.DIAL",
        "-d", f"tel:{phone_number}"
    ])

    time.sleep(1)

    # 2. Tap nút CALL
    run_adb([
        "shell", "input", "swipe",
        CALL_BUTTON[0], CALL_BUTTON[1],
        CALL_BUTTON[0], CALL_BUTTON[1], "100"
    ])

    # 3. Theo dõi trạng thái
    start_time = time.time()

    while time.time() - start_time < timeout:

        state = get_call_state()

        if state == "ACTIVE":
            print("✅ ANSWERED → tắt cuộc gọi")
            run_adb(["shell", "input", "keyevent", "6"])
            return "ANSWERED"

        if state == "DISCONNECTED":
            print("❌ REJECTED")
            return "REJECTED"

        time.sleep(CHECK_INTERVAL)

    # 4. Timeout
    print("⏱ TIMEOUT → tự tắt cuộc gọi")
    run_adb(["shell", "input", "keyevent", "6"])
    return "TIMEOUT"


def call_3_numbers(phone1, phone2, phone3, timeout=20):
    """
    Luồng gọi tuần tự 3 số.

    BREAK khi:
    - ANSWERED
    - REJECTED
    - số trống
    - gọi hết 3 số
    """

    phones = [phone1, phone2, phone3]

    for i, phone in enumerate(phones, start=1):

        # ===== số trống → break =====
        if not phone:
            print(f"📴 Phone {i} trống → dừng chuỗi gọi")
            return "EMPTY_NUMBER"

        print(f"\n➡️ Gọi phone {i}: {phone}")

        status = auto_call(phone, timeout)

        print(f"📊 Kết quả phone {i}: {status}")

        # ===== người nhận đã tương tác → break =====
        if status in ("ANSWERED", "REJECTED"):
            print("🛑 Người nhận đã xử lý cuộc gọi → dừng chuỗi")
            return status

        # TIMEOUT → thử số tiếp theo

    # ===== gọi hết 3 số =====
    print("❌ Đã gọi hết 3 số nhưng không ai phản hồi")
    return "NO_RESPONSE"


# result = module_other_stx.call_3_numbers("0359667694", "0379077024", "", timeout=30)
        # print("FINAL RESULT:", result)






def timerun():
    while True:
        time.sleep(3)
        timerun = time.strftime("%H:%M:%S", time.localtime())
        print("Thời gian hiện tại:", timerun)
        print("Thời gian chạy tool:", var_stx.timerun)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 47, 2, timerun)
        if var_stx.timerun == "":
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 47, 2, timerun)
        else:
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 47, 2, var_stx.timerun)


        if var_stx.timerun == time.strftime("%H:%M", time.localtime()):
            break
        if var_stx.timerun == "":
            break



def clear_log():
    logging.basicConfig(handlers=[logging.FileHandler(filename=var_stx.logpath,
                                                      encoding='utf-8', mode='w')],  # mode='a+', w
                        format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                        datefmt="%F %A %T",
                        level=logging.INFO)




def clearData(file, sheetName, ketqua, trangthai, tenanh):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 9
    while (i < 1000):
        i += 1
        i = str(i)
        sheet["F"+i] = ketqua
        sheet["G"+i] = trangthai
        sheet["M"+i] = tenanh
        i = int(i)
    wordbook.save(file)






def clearData_luutamthoi(file,sheetName, web, api, popup):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 6
    while (i < 100):
        i += 1
        i = str(i)
        sheet["B"+i] = web
        sheet["C"+i] = api
        sheet["D"+i] = popup
        i = int(i)
    wordbook.save(file)




def clearData_luutamthoi1(file,sheetName, column1, column2, column3, column4):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 119
    while (i < 150):
        i += 1
        i = str(i)
        sheet["A"+i] = column1
        sheet["B"+i] = column2
        sheet["C"+i] = column3
        sheet["D"+i] = column4
        i = int(i)
    wordbook.save(file)





def close_tab():
    var_stx.driver.implicitly_wait(5)
    # đóng tab không liên quan
    try:
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(1)
        curr = var_stx.driver.current_window_handle
        for handle in var_stx.driver.window_handles:
            var_stx.driver.switch_to.window(handle)
            if handle != curr:
                var_stx.driver.close()
    except:
        pass
    time.sleep(1)



def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)




def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)




def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value




def writeData(file,sheetName,caseid,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 0
    while (i < 5000):
        i += 1
        i = str(i)
        if sheet["A"+i].value == caseid:
            i = int(i)
            sheet.cell(row=i, column=columnno).value = data
            break
        i = int(i)
    wordbook.save(file)




def appendData(file, sheetName, caseid, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    found = False

    # Duyệt từng dòng, tìm caseid trong cột A
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if str(cell_value).strip() == str(caseid).strip():
            # Nếu đã có dữ liệu ở cột cần ghi -> nối thêm
            old_val = sheet.cell(row=row, column=columnno).value
            if old_val is None or str(old_val).strip() == '':
                new_val = data
            else:
                new_val = f"{old_val}, {data}"   # Nối thêm
            sheet.cell(row=row, column=columnno).value = new_val
            found = True
            break

    # Nếu chưa có caseid thì thêm dòng mới
    if not found:
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = caseid
        sheet.cell(row=next_row, column=columnno).value = data

    workbook.save(file)



def generate_audio(text, filename):
    tts = gTTS(text=text, lang='vi')
    tts.save(filename)
    print(f"Đã tạo file âm thanh: {filename}")



def allow_mic(driver2):
    try:
        allow_button = driver2.ele('xpath://button[contains(text(), "Allow while visiting the site")]', timeout=5)
        if allow_button:
            allow_button.click()
            print("Đã bấm Cho phép dùng Microphone.")
    except Exception as e:
        print(f"Không thấy popup xin quyền micro: {e}")



def play_mp3_hidden(filepath):
    playsound(filepath)



def tele_search(tag, phone, data):
    global driver2
    driver2.ele(var_stx.tele_search_input).send_keys(Keys.CONTROL, "a", Keys.DELETE)
    time.sleep(0.5)
    driver2.ele(var_stx.tele_search_input).input(tag)
    time.sleep(1)
    driver2.ele(var_stx.tele_search_name1, timeout=10).click()
    time.sleep(2)
    driver2.ele(var_stx.tele_profile_name).click()
    time.sleep(2.5)
    check_phone = driver2.ele(var_stx.tele_profile_phone).text
    check_tag = driver2.ele(var_stx.tele_profile_tag).text
    print(check_phone)
    print(phone)
    print(check_tag)
    print(tag)
    if (check_phone == phone) and (check_tag == tag):
        driver2.ele(var_stx.hopthoai_input).input(data)
        driver2.ele(var_stx.hopthoai_input).input(Keys.ENTER)
        time.sleep(1)



def tele_call(tag, phone, count, cases):
    global driver2
    check_phone = driver2.ele(var_stx.tele_profile_phone).text
    check_tag = driver2.ele(var_stx.tele_profile_tag).text
    print(check_phone)
    print(phone)
    print(check_tag)
    print(tag)
    if (check_phone == phone) and (check_tag == tag):
        driver2.ele(var_stx.tele_profile_call).click()
        time.sleep(1)
        print("n10")
        allow_mic(driver2)
        n = 0
        while n < 15:
            n += 1
            time.sleep(1)
            try:
                check_call = driver2.ele(var_stx.calling).text
                print("n11")
                if check_call not in ["waiting...", "exchanging encryption keys...", "Connected"]:
                    # if check_call not in ["Ringing...", "Connecting..."]:
                    generate_audio(f"Bạn đang có {count} lỗi cần xử lý, mã lỗi: {cases}", var_stx.uploadpath + "output.mp3")
                    print("n12")
                    play_mp3_hidden(var_stx.uploadpath + "output.mp3")
                    print("n13")
                    time.sleep(2)
                    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 293, 2, "Đã gọi")
                    break
            except Exception as e:
                print(f"Lỗi kiểm tra cuộc gọi: {e}")
                pass

    driver2.ele(var_stx.tele_profile_end_call).click()
    time.sleep(2)



def open_tele():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var_v3.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)
    driver2.get("https://web.telegram.org/a/")
    time.sleep(2)
    var_stx.driver.close()





def call_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var_stx.uploadpath+"Profile 30""")
    global driver2
    driver2 = ChromiumPage(addr_or_opts=do1)
    clearData_luutamthoi2(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "")
    module_stx.retest_serious()



    driver2.get("https://web.telegram.org/a/")
    time.sleep(2)
    try:
        driver2.ele("//*[text()='OK']").click()
    except:
        pass

    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 293, 2, "Đang check cuộc gọi")
    count = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 291, 2))
    data = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 292, 2))
    cases = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 290, 3))


    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 3
    while rownum < 6:
        rownum += 1
        row_str = str(rownum)
        phone_tag = sheet["M" + row_str].value
        print(f"Dòng {row_str}, Phone_tag: {phone_tag}")
        phone, tag = phone_tag.split('(')
        phone = phone.strip()
        tag = tag.replace(')', '').strip()
        print(phone)
        print(tag)
        try:
            tele_search(tag, phone, data)
            tele_call(tag, phone, count, cases)
            check_call = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 293, 2))
            if check_call == "Đã gọi":
                break

        except Exception as e:
            print(f"Lỗi kiểm tra cuộc gọi: {e}")
            pass




def check_info_telegram():
    global driver2
    driver2.get("https://web.telegram.org/a/")





@retry(tries=2, delay=2, backoff=1, jitter=5, )
def notification_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var_stx.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)


    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    module_stx.check_casenone()
    module_stx.check_casefail()
    module_stx.check_casepass()

    mucnghiemtrong = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 65, 2))
    tong_case_trong = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 66, 2))

    case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 77, 2))
    case_pass = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 87, 2))


    time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    thoigianbatdauchay = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 47, 2))

    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2 +
                                              "\n- LinkTest: " + var_stx.linktest +
                                              "\n- ModeTest: " + var_stx.modetest +
                                              "\n- Số case Pass: " + case_pass +
                                              "\n- Số case False: " + case_fail +
                                              "\n- Số case trống: " + tong_case_trong +
                                              "\n- Số case False nghiêm trọng: " + mucnghiemtrong)

    thoigianbatdauchay = str(var_stx.readData(var_stx.path_luutamthoi , 'Sheet1', 47, 2))

    if int(case_fail) >= 1 or int(tong_case_trong) >= 1:
        print("đã vào tele")
        driver2.get("https://web.telegram.org/a/")
        time.sleep(2)
        case_pass = str(case_pass)
        case_fail = str(case_fail)
        try:
            driver2.ele("//*[text()='OK']").click()
        except:
            pass

        # if var_stx.linktest[0:25] == "http://192.168.45.48:8000" or var_stx.linktest[0:25] == "http://192.168.45.82:8000":
        #     driver2.ele(var_stx.hopthoai1).click()
        # else:
        #     driver2.ele(var_stx.hopthoai2).click()

        driver2.ele(var_stx.hopthoai3).click()

        time.sleep(1)
        time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
        time_string1 = str(time_string1)
        time_string2 = time.strftime("%H:%M", time.localtime())
        time_string2 = str(time_string2)
        driver2.ele(var_stx.hopthoai_input).input("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2 +
                                                  "\n- LinkTest: " + var_stx.linktest +
                                                  "\n- ModeTest: " + var_stx.modetest +
                                                  "\n- Số case Pass: " + case_pass +
                                                  "\n- Số case False: " + case_fail +
                                                  "\n- Số case trống: " + tong_case_trong +
                                                  "\n- Số case False nghiêm trọng: " + mucnghiemtrong)
        driver2.ele(var_stx.hopthoai_input).input(Keys.ENTER)
        time.sleep(1)
        driver2.ele(var_stx.hopthoai_iconlink).click()
        time.sleep(1)
        driver2.ele(var_stx.hopthoai_iconlink_file).click()
        time.sleep(1)
        subprocess.Popen(var_stx.uploadpath+"checklist_staxi.exe")
        time.sleep(1)
        driver2.ele(var_stx.hopthoai_send).click()
        time.sleep(2)
        driver2.ele(var_stx.hopthoai_iconlink).click()
        time.sleep(1)
        driver2.ele(var_stx.hopthoai_iconlink_file).click()
        time.sleep(1)
        subprocess.Popen(var_stx.uploadpath+"staxi_log.exe")
        time.sleep(1)
        driver2.ele(var_stx.hopthoai_send).click()
        time.sleep(1)

        time.sleep(30)
        driver2.close()








def check_user_id():
    res = requests.post(
        "https://chatapi.viber.com/pa/get_account_info",
        headers={"X-Viber-Auth-Token": "54c42460ff31117d-cfaa6f5220cbd0b1-baced03cce3b6c9f"}
    )

    print(res.json())



def viber_send_text():
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    module_stx.check_casenone()
    module_stx.check_casefail()
    module_stx.check_casepass()

    mucnghiemtrong = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 65, 2))
    tong_case_trong = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 66, 2))

    case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 77, 2))
    case_pass = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 87, 2))


    time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    thoigianbatdauchay = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 47, 2))

    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2 +
                                              "\n- LinkTest: " + var_stx.linktest +
                                              "\n- ModeTest: " + var_stx.modetest +
                                              "\n- Số case Pass: " + case_pass +
                                              "\n- Số case False: " + case_fail +
                                              "\n- Số case trống: " + tong_case_trong +
                                              "\n- Số case False nghiêm trọng: " + mucnghiemtrong)

    AUTH_TOKEN = "54c42460ff31117d-cfaa6f5220cbd0b1-baced03cce3b6c9f"  # id Cảnh báo Autotest STAXI Customer
    FROM_USER_ID = "DCfDJTWxQ+Wa/imEkkOn6Q=="


    # AUTH_TOKEN = "54c527ff9ab507cb-2bba287f248eba8d-520d20523e7d2980"   #id nhóm test
    # FROM_USER_ID = "DCfDJTWxQ+Wa/imEkkOn6Q=="


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    print(AUTH_TOKEN)
    print(FROM_USER_ID)


    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return


    # 2. Gửi tin nhắn văn bản
    send_url = "https://chatapi.viber.com/pa/post"
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,
        "type": "text",
        "text": ("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2 +
                                              "\n- LinkTest: " + var_stx.linktest +
                                              "\n- ModeTest: " + var_stx.modetest +
                                              "\n- Số case Pass: " + case_pass +
                                              "\n- Số case False: " + case_fail +
                                              "\n- Số case trống: " + tong_case_trong +
                                              "\n- Số case False nghiêm trọng: " + mucnghiemtrong)}

    headers = {
        "Content-Type": "application/json"}

    response = requests.post(send_url, json=payload, headers=headers)

    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())







def upload_pixeldrain_auth(filepath):
    API_KEY = "c567bb13-f4c0-4aac-b9bd-c8add1e467fc"  # Thay bằng key thật

    with open(filepath, "rb") as f:
        res = requests.post(
            "https://pixeldrain.com/api/file",
            files={"file": f},
            auth=HTTPBasicAuth('', API_KEY)
        )
        res_json = json.loads(res.text)
        file_id = res_json["id"]
        link_download = (f"https://pixeldrain.com/api/file/{file_id}")
        print(link_download)
        return link_download





def send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, file_path):
    file_url = upload_pixeldrain_auth(file_path)
    # file_url = upload_to_catbox(file_path)

    if not file_url:
        print("⚠️ Không thể upload file. Hủy gửi.")
        return

    file_name = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return

    # 2. Gửi tin nhắn văn bản
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,           # Viber user ID người nhận
        "min_api_version": 1,
        "tracking_data": "",               # Có thể để chuỗi rỗng nếu không dùng tracking
        "type": "file",
        "media": file_url,
        "size": file_size,
        "file_name": file_name}

    headers = {
        "X-Viber-Auth-Token": AUTH_TOKEN,
        "Content-Type": "application/json"}

    response = requests.post("https://chatapi.viber.com/pa/post", json=payload, headers=headers)
    print("📨 Phản hồi từ Viber:", response.status_code, response.json())
    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())



def viber_send_file():
    # ==== Ví dụ sử dụng ====
    AUTH_TOKEN = "54c42460ff31117d-cfaa6f5220cbd0b1-baced03cce3b6c9f"  # id Cảnh báo Autotest STAXI
    FROM_USER_ID = "DCfDJTWxQ+Wa/imEkkOn6Q=="






    FILE_PATH_checklisst = var_stx.checklistpath  # Thay bằng đường dẫn file thật
    FILE_PATH_log = var_stx.logpath  # Thay bằng đường dẫn file thật

    print(var_stx.checklistpath)
    print(var_stx.logpath)


    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_checklisst)
    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_log)


def send_viber():
    viber_send_text()
    viber_send_file()







def call_viver1():
    AUTH_TOKEN = "54c527ff9ab507cb-2bba287f248eba8d-520d20523e7d2980"   #id nhóm test
    # FROM_USER_ID = "DCfDJTWxQ+Wa/imEkkOn6Q=="

    # AUTH_TOKEN = "DCfDJTWxQ+Wa/imEkkOn6Q=="
    receiver_id = "NcTubYkRh5GD2fyZFDTOjw=="

    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return

    # 2. Gửi tin nhắn văn bản
    payload = {
   "receiver":receiver_id,
   "min_api_version":1,
   "sender":{
      "name":"Bot auto test",
      "avatar":"https://media-direct.cdn.viber.com/download_photo?dlid=7nZRXMsHtZr_J8VUjfHf-45KJ0EAONowCQqAKm25lgVGphx7OYNUcJf8RgSfq-V7_wVZ13D9AHn8nOB1_vp8dr5YqEOA5Nw6QjQIAbZW--mf-fRIhWWW7hAs5X--4BhwL3wftA&fltp=jpg&imsz=0000"
   },
   "tracking_data":"tracking data",
   "type":"contact",
   "contact":{
      "name":"Trần Quang Trường",
      "phone_number":"+84359667694"
   }
}
    headers = {
        "X-Viber-Auth-Token": AUTH_TOKEN,
        "Content-Type": "application/json"}

    response = requests.post("https://chatapi.viber.com/pa/post", json=payload, headers=headers)
    print("📨 Phản hồi từ Viber:", response.status_code, response.json())
    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())



    #
    # response = requests.post("https://chatapi.viber.com/pa/send_message", json=payload, headers=headers)
    # return response.status_code, response.json()




def call_viver():

    # AUTH_TOKEN = "s3fnH/NlIBI2DwqBeVlhEQ=="
    AUTH_TOKEN = "54c527ff9ab507cb-2bba287f248eba8d-520d20523e7d2980"   #id nhóm test

    receiver_id = "DCfDJTWxQ+Wa/imEkkOn6Q=="

    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return

    payload = {
   "receiver":receiver_id,
   "min_api_version":1,
   "sender":{
      "name":"Bot auto test",
      "avatar":"https://media-direct.cdn.viber.com/download_photo?dlid=7nZRXMsHtZr_J8VUjfHf-45KJ0EAONowCQqAKm25lgVGphx7OYNUcJf8RgSfq-V7_wVZ13D9AHn8nOB1_vp8dr5YqEOA5Nw6QjQIAbZW--mf-fRIhWWW7hAs5X--4BhwL3wftA&fltp=jpg&imsz=0000"
   },
   "tracking_data":"tracking data",
   "type":"contact",
   "contact":{
      "name":"Trần Quang Trường",
      "phone_number":"+84359667694"
   }
}

    headers = {
        "X-Viber-Auth-Token": AUTH_TOKEN,
        "Content-Type": "application/json"}

    response = requests.post("https://chatapi.viber.com/pa/post", json=payload, headers=headers)
    print("📨 Phản hồi từ Viber:", response.status_code, response.json())
    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())









def clearData_luutamthoi2(file,sheetName, column1, column2, column3, column4, column5):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 89
    while (i < 100):
        i += 1
        i = str(i)
        sheet["B"+i] = column1
        sheet["C"+i] = column2
        sheet["D"+i] = column3
        sheet["E"+i] = column4
        sheet["F"+i] = column5
        i = int(i)
    wordbook.save(file)



def delete_image():
    path = os.path.join(var_stx.imagepath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def delete_excel():
    path = os.path.join(var_stx.excelpath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def get_datachecklist(ma):
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 3000):
        rownum += 1
        rownum = str(rownum)
        if sheet["A" + rownum].value == ma:
            tensukien = sheet["B" + rownum].value
            ketqua = sheet["E" + rownum].value
            print(ma)
            print(tensukien)
            print(ketqua)
            logging.info("đang chạy case: " + ma)
        rownum = int(rownum)




# @retry(tries=3, delay=2, backoff=1, jitter=5, )
def swich_tab_0():
    try:
        print("[INFO] ⚙️  Switching tab 0...")
        try:
            var_stx.driver.implicitly_wait(15)
        except:
            var_stx.restart_driver()

        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
    except WebDriverException as e:
        print(f"[ERROR] ❌ Lỗi khi chuyển tab hoặc khởi động driver: {e}")
        print("[INFO] 🔁 Restarting driver...")
        var_stx.restart_driver()
        time.sleep(3)
        var_stx.driver.get(var_stx.linktest)
        time.sleep(10)
        print("[INFO] ✅ Driver restarted và đã mở lại link test.")
        # Optional: Retry lại switch sau restart
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])

    try:
        var_stx.driver.implicitly_wait(1)
        var_stx.driver.switch_to.alert.accept()
        time.sleep(1)
    except:
        pass


    try:
        var_stx.driver.implicitly_wait(1)
        subprocess.Popen(var_stx.uploadpath+"cancel.exe")
    except:
        pass


    try:
        var_stx.driver.switch_to.window(var_stx.driver.window_handles[0])
        time.sleep(0.5)
    except:
        curr = var_stx.driver.current_window_handle
        for handle in var_stx.driver.window_handles:
            if handle != curr:
                var_stx.driver.switch_to.window(handle)
                var_stx.driver.close()
                time.sleep(1)
        var_stx.driver.switch_to.window(curr)
        time.sleep(1)


    # Khởi tạo lại trình duyệt
    var_stx.restart_driver()
    var_stx.driver.get(var_stx.linktest)
    time.sleep(10)








def write_caseif():
    n = 0
    while (n < 80):
        var_stx.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   if case == 'Accounting"+n+"':\n       caseid_stx.caseid_accounting"+n+"(self)\nexcept:\n    module_other_stx.swich_tab_0()")
        n = int(n)



def write_caseif1():
    n = 53
    while (n < 70):
        var_stx.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   caseid.caseid_report"+n+"(self)\nexcept:\n     pass")
        n = int(n)



def write_caseif2():
    n = 53
    while (n < 70):
        var_stx.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("caseid.caseid_report"+n+"(self)")
        n = int(n)



def write_result_text_try_if_color(code, eventname, result, path_module, path_color, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        element1 = var_stx.driver.find_element(By.XPATH, path_color)
        check_color = element1.value_of_css_property("color")
        logging.info(check_color)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_color)

        if check_color == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)



def write_result_text_try_if_boolean(code, eventname, result, path_module, path_selector, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_selector = var_stx.driver.find_element(By.XPATH, path_selector).is_selected()
        logging.info(check_selector)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_selector)

        if check_selector == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)




def write_result_dowload_file(code, eventname, result, path_module, file, name_image, times=20):
    var_stx.driver.implicitly_wait(1)
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    n = 0
    while (n < times):
        n = n + 1
        try:
            filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
            shutil.move(filename, os.path.join(var_stx.excelpath, f"{file}"))
            break
        except:
            pass
        time.sleep(1)


    try:
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, f"{file}"))
        logging.info("True")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        writeData(var_stx.checklistpath, "Checklist", code, 6, file)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_stx.write_result_dowload_file(code, eventname, result, "Giám sát - Lộ trình 1.7",
    #                                            "_GiamSat_LoTrinh_XuatExcel.xls", "_GiamSat_LoTrinh_XuatExcel.png")




def write_result_text_try_if_cut1(code, eventname, result, path_module, path_text, check_result, name_image, from_cut):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

        if check_text[from_cut::] == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_cut1(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_cut2(code, eventname, result, path_module, path_text, check_result, name_image, from_cut, to_cut):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

        if check_text[from_cut:to_cut] == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_cut2(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png", 0, 10)




def write_result_text_try_if(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

        if check_text == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_strip(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

        if check_text.strip().lower() == check_result.strip().lower():
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_in(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

        if check_result.lower() in check_text.lower():
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_or(code, eventname, result, path_module, path_text, check_result1, check_result2, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result1)
        logging.info(check_result2)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

        if (check_text == check_result1) or (check_text == check_result2):
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)



def write_result_text_try_if_value(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).get_attribute("value")
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)

        if check_text == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)




def write_result_text_try_if_src(code, eventname, result, path_module, path_text, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).get_attribute("src")
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)
        if check_text == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_src(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_src1(code, eventname, result, path_module, path_text, check_result, name_image, from_cut):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        print("n1")
        check_text = var_stx.driver.find_element(By.XPATH, path_text).get_attribute("src")
        print("n2")
        logging.info(check_text)
        print("n3")
        logging.info(check_result)
        print("n4")
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)
        print("n5")
        if check_text[from_cut::] == check_result:
            print("n6")
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            print("n7")
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        print("n8")
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_src(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")



def write_result_text_try_if_title(code, eventname, result, path_module, check_result, name_image):
    var_stx.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        title = str(var_stx.driver.title)
        logging.info(title)
        logging.info(check_result)
        writeData(var_stx.checklistpath, "Checklist", code, 6, title)
        if title == check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_title(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_other(code, eventname, result, path_module, path_text, check_result, name_image):
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_stx.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        writeData(var_stx.checklistpath, "Checklist", code, 6, check_text)
        if check_text != check_result:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_stx.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_displayed_try(code, eventname, result, path_module, path_text, name_image):
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_displayed = var_stx.driver.find_element(By.XPATH, path_text).is_displayed()
        logging.info("True")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
    except NoSuchElementException:
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)

    # logging.info("Tìm biển kiểm soát - " + data['quantri']['bienkiemsoat'])
    # module_other_v3.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_stx.check_hide_car, "_QuanTri_DsXe_AnXe.png")




def write_result_not_displayed_try(code, eventname, result, path_module, path_text, name_image):
    var_stx.driver.implicitly_wait(2)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_not_displayed = var_stx.driver.find_element(By.XPATH, path_text).is_displayed()
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        logging.info("True")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
    # module_other_v3.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_stx.check_hide_car, "_QuanTri_DsXe_AnXe.png")




def write_result_close_popup(code, eventname, result, path_module, path_text, name_image):
    var_stx.driver.implicitly_wait(2)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        var_stx.driver.find_element(By.XPATH, path_text).click()
        logging.info("False")
        var_stx.driver.save_screenshot(var_stx.imagepath + code + name_image)
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        logging.info("True")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")


    # module_other_v3.write_result_close_popup(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_stx.check_hide_car, "_QuanTri_DsXe_AnXe.png")




def write_result_excelreport(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+":  "+output+"")
        if str(sheet[column + number_column].value) == output:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # module_other_v3.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")




def write_result_excelreport1(code, sheet, column, name_sheet, number_column, number_row, output, number_row2, output2):
    data_excel = str(sheet[number_row2].value)
    output2 = str(output2)

    print("Dữ liệu web: " +output2)
    print("Dữ liệu excel: " +data_excel)
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+":  "+output+"")
        logging.info("Dữ liệu excel: "+ data_excel)
        logging.info("Dữ liệu web: "+ output2)
        if str(sheet[column + number_column].value) == output and data_excel == output2:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row2)
    # if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
    # chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "D5", "Ngày tháng", "D10", activity_synthesis_group_report_day_month)




def write_result_excelreport2(code, output_web, output_excel, name_output):
    logging.info(name_output + " web: " + output_web)
    logging.info(name_output + " excel: " + output_excel)
    if output_web == output_excel:
        logging.info("True")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
    else:
        logging.info("False")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai trường " + output_web)






def write_result_excel_checkweb(code, data_web, desire):
    logging.info("Dữ liệu web: " + data_web)
    logging.info("Dữ liệu mong muốn: " + desire)
    if data_web == desire:
        logging.info("True")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
    else:
        logging.info("False")
        writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_stx.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động mất sai trường" + desire)








def write_result_excelreport_clear_data(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+": "+output+"")
        if str(sheet[column + number_column].value) == output:
            logging.info("True")
            writeData(var_stx.checklistpath, "Checklist", code, 6, " ")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_stx.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")





