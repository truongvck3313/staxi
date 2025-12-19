import openpyxl

import module_other_stx
import var_stx
import re
import caseid_stx
from collections import defaultdict
#19/12







def ModuleTest():
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("so1", i)
        if i == "0":
            run_all(self='')
        else:
            moduletest = ''.join(re.findall(r'\d+', var_stx.moduletest))
            print(type(moduletest))
            print(moduletest)
            for i in moduletest:
                print("so2", i)
                if i == "1":
                    login(self='')
                if i == "2":
                    minitor(self='')
                if i == "3":
                    vehicle(self='')
                if i == "4":
                    wallet(self='')
                if i == "5":
                    promotion(self='')
                if i == "6":
                    customer(self='')
                if i == "7":
                    report(self='')
                if i == "8":
                    admin(self='')
                if i == "9":
                    accounting(self='')



def check_casenone():
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 60, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 61, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 62, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 63, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 64, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 65, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 61, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 62, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 63, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 64, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == None:
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 66, 2, count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 60, 2, count)


        else:
            muc1 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 61, 2))
            muc2 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 62, 2))
            muc3 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 63, 2))
            muc4 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 64, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 60, 2, sumarry_case_none)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 66, 2, sumarry_case_none)


    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 65, 2, count)



def check_casefail():
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 70, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 71, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 72, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 73, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 74, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 75, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 71, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 72, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 73, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 74, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Fail":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 76, 2, count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 70, 2, count)

        else:
            muc1 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 70, 2, sumarry_case_none)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 76, 2, sumarry_case_none)



    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 75, 2, count)


    if var_stx.modetest == "0":
        case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 70, 2))
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_stx.modetest == "1":
        case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 71, 2))
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_stx.modetest == "2":
        case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 72, 2))
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_stx.modetest == "3":
        case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 73, 2))
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_stx.modetest == "4":
        case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 74, 2))
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 77, 2, case_fail)



def check_casepass():
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 70, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 71, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 72, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 73, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 74, 2, "0")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 75, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 81, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 82, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 83, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 84, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Pass":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 86, 2, count)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 80, 2, count)

        else:
            muc1 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 80, 2, sumarry_case_none)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 86, 2, sumarry_case_none)

        if var_stx.modetest == "0":
            case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 80, 2))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_stx.modetest == "1":
            case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 81, 2))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_stx.modetest == "2":
            case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 82, 2))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_stx.modetest == "3":
            case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 83, 2))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_stx.modetest == "4":
            case_fail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 84, 2))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 87, 2, case_fail)



def retest_casenone(self):
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9

    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)

        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)

                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức0: ", count)

        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)

        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)

        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)

        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)

        for case in list_casefail:
            try:
                if case == 'Login01':
                    caseid_stx.caseid_login01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login02':
                    caseid_stx.caseid_login02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login03':
                    caseid_stx.caseid_login03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login04':
                    caseid_stx.caseid_login04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login05':
                    caseid_stx.caseid_login05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login06':
                    caseid_stx.caseid_login06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login07':
                    caseid_stx.caseid_login07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login08':
                    caseid_stx.caseid_login08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login09':
                    caseid_stx.caseid_login09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login10':
                    caseid_stx.caseid_login10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login11':
                    caseid_stx.caseid_login11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Login12':
                    caseid_stx.caseid_login12(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Minitor01':
                    caseid_stx.caseid_minitor01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor02':
                    caseid_stx.caseid_minitor02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor03':
                    caseid_stx.caseid_minitor03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor04':
                    caseid_stx.caseid_minitor04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor05':
                    caseid_stx.caseid_minitor05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor06':
                    caseid_stx.caseid_minitor06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor07':
                    caseid_stx.caseid_minitor07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor08':
                    caseid_stx.caseid_minitor08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor09':
                    caseid_stx.caseid_minitor09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor10':
                    caseid_stx.caseid_minitor10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor11':
                    caseid_stx.caseid_minitor11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor12':
                    caseid_stx.caseid_minitor12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor13':
                    caseid_stx.caseid_minitor13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor14':
                    caseid_stx.caseid_minitor14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor15':
                    caseid_stx.caseid_minitor15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor16':
                    caseid_stx.caseid_minitor16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor17':
                    caseid_stx.caseid_minitor17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor18':
                    caseid_stx.caseid_minitor18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor19':
                    caseid_stx.caseid_minitor19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor20':
                    caseid_stx.caseid_minitor20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor21':
                    caseid_stx.caseid_minitor21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor22':
                    caseid_stx.caseid_minitor22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor23':
                    caseid_stx.caseid_minitor23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor24':
                    caseid_stx.caseid_minitor24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor25':
                    caseid_stx.caseid_minitor25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor26':
                    caseid_stx.caseid_minitor26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor27':
                    caseid_stx.caseid_minitor27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor28':
                    caseid_stx.caseid_minitor28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor29':
                    caseid_stx.caseid_minitor29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor30':
                    caseid_stx.caseid_minitor30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor31':
                    caseid_stx.caseid_minitor31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor32':
                    caseid_stx.caseid_minitor32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor33':
                    caseid_stx.caseid_minitor33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor34':
                    caseid_stx.caseid_minitor34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor35':
                    caseid_stx.caseid_minitor35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor36':
                    caseid_stx.caseid_minitor36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor37':
                    caseid_stx.caseid_minitor37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor38':
                    caseid_stx.caseid_minitor38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor39':
                    caseid_stx.caseid_minitor39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor40':
                    caseid_stx.caseid_minitor40(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor41':
                    caseid_stx.caseid_minitor41(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor42':
                    caseid_stx.caseid_minitor42(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor43':
                    caseid_stx.caseid_minitor43(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor44':
                    caseid_stx.caseid_minitor44(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor45':
                    caseid_stx.caseid_minitor45(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor46':
                    caseid_stx.caseid_minitor46(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor47':
                    caseid_stx.caseid_minitor47(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor48':
                    caseid_stx.caseid_minitor48(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor49':
                    caseid_stx.caseid_minitor49(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor50':
                    caseid_stx.caseid_minitor50(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor51':
                    caseid_stx.caseid_minitor51(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor52':
                    caseid_stx.caseid_minitor52(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor53':
                    caseid_stx.caseid_minitor53(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor54':
                    caseid_stx.caseid_minitor54(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor55':
                    caseid_stx.caseid_minitor55(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor56':
                    caseid_stx.caseid_minitor56(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor57':
                    caseid_stx.caseid_minitor57(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor58':
                    caseid_stx.caseid_minitor58(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor59':
                    caseid_stx.caseid_minitor59(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor60':
                    caseid_stx.caseid_minitor60(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor61':
                    caseid_stx.caseid_minitor61(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor62':
                    caseid_stx.caseid_minitor62(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor63':
                    caseid_stx.caseid_minitor63(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor64':
                    caseid_stx.caseid_minitor64(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor65':
                    caseid_stx.caseid_minitor65(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor66':
                    caseid_stx.caseid_minitor66(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor67':
                    caseid_stx.caseid_minitor67(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor68':
                    caseid_stx.caseid_minitor68(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor69':
                    caseid_stx.caseid_minitor69(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor70':
                    caseid_stx.caseid_minitor70(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor71':
                    caseid_stx.caseid_minitor71(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor72':
                    caseid_stx.caseid_minitor72(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor73':
                    caseid_stx.caseid_minitor73(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor74':
                    caseid_stx.caseid_minitor74(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor75':
                    caseid_stx.caseid_minitor75(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor76':
                    caseid_stx.caseid_minitor76(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor77':
                    caseid_stx.caseid_minitor77(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor78':
                    caseid_stx.caseid_minitor78(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor79':
                    caseid_stx.caseid_minitor79(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Minitor80':
                    caseid_stx.caseid_minitor80(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Vehicle01':
                    caseid_stx.caseid_vehicle01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle02':
                    caseid_stx.caseid_vehicle02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle03':
                    caseid_stx.caseid_vehicle03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle04':
                    caseid_stx.caseid_vehicle04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle05':
                    caseid_stx.caseid_vehicle05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle06':
                    caseid_stx.caseid_vehicle06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle07':
                    caseid_stx.caseid_vehicle07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle08':
                    caseid_stx.caseid_vehicle08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle09':
                    caseid_stx.caseid_vehicle09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle10':
                    caseid_stx.caseid_vehicle10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle11':
                    caseid_stx.caseid_vehicle11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle12':
                    caseid_stx.caseid_vehicle12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle13':
                    caseid_stx.caseid_vehicle13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle14':
                    caseid_stx.caseid_vehicle14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle15':
                    caseid_stx.caseid_vehicle15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle16':
                    caseid_stx.caseid_vehicle16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle17':
                    caseid_stx.caseid_vehicle17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle18':
                    caseid_stx.caseid_vehicle18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle19':
                    caseid_stx.caseid_vehicle19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle20':
                    caseid_stx.caseid_vehicle20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle21':
                    caseid_stx.caseid_vehicle21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle22':
                    caseid_stx.caseid_vehicle22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle23':
                    caseid_stx.caseid_vehicle23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle24':
                    caseid_stx.caseid_vehicle24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle25':
                    caseid_stx.caseid_vehicle25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle26':
                    caseid_stx.caseid_vehicle26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle27':
                    caseid_stx.caseid_vehicle27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle28':
                    caseid_stx.caseid_vehicle28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle29':
                    caseid_stx.caseid_vehicle29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle30':
                    caseid_stx.caseid_vehicle30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle31':
                    caseid_stx.caseid_vehicle31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle32':
                    caseid_stx.caseid_vehicle32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle33':
                    caseid_stx.caseid_vehicle33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle34':
                    caseid_stx.caseid_vehicle34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle35':
                    caseid_stx.caseid_vehicle35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle36':
                    caseid_stx.caseid_vehicle36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle37':
                    caseid_stx.caseid_vehicle37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle38':
                    caseid_stx.caseid_vehicle38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle39':
                    caseid_stx.caseid_vehicle39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle40':
                    caseid_stx.caseid_vehicle40(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle41':
                    caseid_stx.caseid_vehicle41(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle42':
                    caseid_stx.caseid_vehicle42(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle43':
                    caseid_stx.caseid_vehicle43(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle44':
                    caseid_stx.caseid_vehicle44(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle45':
                    caseid_stx.caseid_vehicle45(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle46':
                    caseid_stx.caseid_vehicle46(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle47':
                    caseid_stx.caseid_vehicle47(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle48':
                    caseid_stx.caseid_vehicle48(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle49':
                    caseid_stx.caseid_vehicle49(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle50':
                    caseid_stx.caseid_vehicle50(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle51':
                    caseid_stx.caseid_vehicle51(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle52':
                    caseid_stx.caseid_vehicle52(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle53':
                    caseid_stx.caseid_vehicle53(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle54':
                    caseid_stx.caseid_vehicle54(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle55':
                    caseid_stx.caseid_vehicle55(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle56':
                    caseid_stx.caseid_vehicle56(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle57':
                    caseid_stx.caseid_vehicle57(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle58':
                    caseid_stx.caseid_vehicle58(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle59':
                    caseid_stx.caseid_vehicle59(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Vehicle60':
                    caseid_stx.caseid_vehicle60(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Wallet01':
                    caseid_stx.caseid_wallet01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet02':
                    caseid_stx.caseid_wallet02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet03':
                    caseid_stx.caseid_wallet03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet04':
                    caseid_stx.caseid_wallet04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet05':
                    caseid_stx.caseid_wallet05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet06':
                    caseid_stx.caseid_wallet06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet07':
                    caseid_stx.caseid_wallet07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet08':
                    caseid_stx.caseid_wallet08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet09':
                    caseid_stx.caseid_wallet09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet10':
                    caseid_stx.caseid_wallet10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet11':
                    caseid_stx.caseid_wallet11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet12':
                    caseid_stx.caseid_wallet12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet13':
                    caseid_stx.caseid_wallet13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet14':
                    caseid_stx.caseid_wallet14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet15':
                    caseid_stx.caseid_wallet15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet16':
                    caseid_stx.caseid_wallet16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet17':
                    caseid_stx.caseid_wallet17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet18':
                    caseid_stx.caseid_wallet18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet19':
                    caseid_stx.caseid_wallet19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet20':
                    caseid_stx.caseid_wallet20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet21':
                    caseid_stx.caseid_wallet21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet22':
                    caseid_stx.caseid_wallet22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet23':
                    caseid_stx.caseid_wallet23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet24':
                    caseid_stx.caseid_wallet24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet25':
                    caseid_stx.caseid_wallet25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet26':
                    caseid_stx.caseid_wallet26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet27':
                    caseid_stx.caseid_wallet27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet28':
                    caseid_stx.caseid_wallet28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet29':
                    caseid_stx.caseid_wallet29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet30':
                    caseid_stx.caseid_wallet30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet31':
                    caseid_stx.caseid_wallet31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet32':
                    caseid_stx.caseid_wallet32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet33':
                    caseid_stx.caseid_wallet33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet34':
                    caseid_stx.caseid_wallet34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet35':
                    caseid_stx.caseid_wallet35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet36':
                    caseid_stx.caseid_wallet36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet37':
                    caseid_stx.caseid_wallet37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet38':
                    caseid_stx.caseid_wallet38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet39':
                    caseid_stx.caseid_wallet39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Wallet40':
                    caseid_stx.caseid_wallet40(self)
            except:
                module_other_stx.swich_tab_0()



            try:
                if case == 'Promotion01':
                    caseid_stx.caseid_promotion01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion02':
                    caseid_stx.caseid_promotion02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion03':
                    caseid_stx.caseid_promotion03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion04':
                    caseid_stx.caseid_promotion04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion05':
                    caseid_stx.caseid_promotion05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion05_1':
                    caseid_stx.caseid_promotion05_1(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Promotion06':
                    caseid_stx.caseid_promotion06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion07':
                    caseid_stx.caseid_promotion07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion08':
                    caseid_stx.caseid_promotion08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion09':
                    caseid_stx.caseid_promotion09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion10':
                    caseid_stx.caseid_promotion10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion11':
                    caseid_stx.caseid_promotion11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion12':
                    caseid_stx.caseid_promotion12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion13':
                    caseid_stx.caseid_promotion13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion14':
                    caseid_stx.caseid_promotion14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion15':
                    caseid_stx.caseid_promotion15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion16':
                    caseid_stx.caseid_promotion16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion17':
                    caseid_stx.caseid_promotion17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion18':
                    caseid_stx.caseid_promotion18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion19':
                    caseid_stx.caseid_promotion19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion20':
                    caseid_stx.caseid_promotion20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion21':
                    caseid_stx.caseid_promotion21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion22':
                    caseid_stx.caseid_promotion22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion23':
                    caseid_stx.caseid_promotion23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion24':
                    caseid_stx.caseid_promotion24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion25':
                    caseid_stx.caseid_promotion25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion26':
                    caseid_stx.caseid_promotion26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion27':
                    caseid_stx.caseid_promotion27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion28':
                    caseid_stx.caseid_promotion28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion29':
                    caseid_stx.caseid_promotion29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion30':
                    caseid_stx.caseid_promotion30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion31':
                    caseid_stx.caseid_promotion31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion32':
                    caseid_stx.caseid_promotion32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion33':
                    caseid_stx.caseid_promotion33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion34':
                    caseid_stx.caseid_promotion34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion35':
                    caseid_stx.caseid_promotion35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion36':
                    caseid_stx.caseid_promotion36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion37':
                    caseid_stx.caseid_promotion37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion38':
                    caseid_stx.caseid_promotion38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion39':
                    caseid_stx.caseid_promotion39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion40':
                    caseid_stx.caseid_promotion40(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion41':
                    caseid_stx.caseid_promotion41(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion42':
                    caseid_stx.caseid_promotion42(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion43':
                    caseid_stx.caseid_promotion43(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion44':
                    caseid_stx.caseid_promotion44(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion45':
                    caseid_stx.caseid_promotion45(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion46':
                    caseid_stx.caseid_promotion46(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion47':
                    caseid_stx.caseid_promotion47(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion48':
                    caseid_stx.caseid_promotion48(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion49':
                    caseid_stx.caseid_promotion49(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion50':
                    caseid_stx.caseid_promotion50(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion51':
                    caseid_stx.caseid_promotion51(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion52':
                    caseid_stx.caseid_promotion52(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion53':
                    caseid_stx.caseid_promotion53(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion54':
                    caseid_stx.caseid_promotion54(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion55':
                    caseid_stx.caseid_promotion55(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion56':
                    caseid_stx.caseid_promotion56(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion57':
                    caseid_stx.caseid_promotion57(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion58':
                    caseid_stx.caseid_promotion58(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion59':
                    caseid_stx.caseid_promotion59(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion60':
                    caseid_stx.caseid_promotion60(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion61':
                    caseid_stx.caseid_promotion61(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion62':
                    caseid_stx.caseid_promotion62(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion63':
                    caseid_stx.caseid_promotion63(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion64':
                    caseid_stx.caseid_promotion64(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion65':
                    caseid_stx.caseid_promotion65(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Promotion66':
                    caseid_stx.caseid_promotion66(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Customer01':
                    caseid_stx.caseid_customer01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer02':
                    caseid_stx.caseid_customer02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer03':
                    caseid_stx.caseid_customer03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer04':
                    caseid_stx.caseid_customer04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer05':
                    caseid_stx.caseid_customer05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer06':
                    caseid_stx.caseid_customer06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer07':
                    caseid_stx.caseid_customer07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer08':
                    caseid_stx.caseid_customer08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer09':
                    caseid_stx.caseid_customer09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer10':
                    caseid_stx.caseid_customer10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer11':
                    caseid_stx.caseid_customer11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer12':
                    caseid_stx.caseid_customer12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer13':
                    caseid_stx.caseid_customer13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer14':
                    caseid_stx.caseid_customer14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer15':
                    caseid_stx.caseid_customer15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer16':
                    caseid_stx.caseid_customer16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer17':
                    caseid_stx.caseid_customer17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer18':
                    caseid_stx.caseid_customer18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer19':
                    caseid_stx.caseid_customer19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer20':
                    caseid_stx.caseid_customer20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer21':
                    caseid_stx.caseid_customer21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer22':
                    caseid_stx.caseid_customer22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer23':
                    caseid_stx.caseid_customer23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer24':
                    caseid_stx.caseid_customer24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer25':
                    caseid_stx.caseid_customer25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer26':
                    caseid_stx.caseid_customer26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer26_1':
                    caseid_stx.caseid_customer26_1(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer26_2':
                    caseid_stx.caseid_customer26_2(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer27':
                    caseid_stx.caseid_customer27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer28':
                    caseid_stx.caseid_customer28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer29':
                    caseid_stx.caseid_customer29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer30':
                    caseid_stx.caseid_customer30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer31':
                    caseid_stx.caseid_customer31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer32':
                    caseid_stx.caseid_customer32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer33':
                    caseid_stx.caseid_customer33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer34':
                    caseid_stx.caseid_customer34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer35':
                    caseid_stx.caseid_customer35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer36':
                    caseid_stx.caseid_customer36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer37':
                    caseid_stx.caseid_customer37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer38':
                    caseid_stx.caseid_customer38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer39':
                    caseid_stx.caseid_customer39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer40':
                    caseid_stx.caseid_customer40(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer41':
                    caseid_stx.caseid_customer41(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer42':
                    caseid_stx.caseid_customer42(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer43':
                    caseid_stx.caseid_customer43(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer44':
                    caseid_stx.caseid_customer44(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer45':
                    caseid_stx.caseid_customer45(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer46':
                    caseid_stx.caseid_customer46(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer47':
                    caseid_stx.caseid_customer47(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer48':
                    caseid_stx.caseid_customer48(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer49':
                    caseid_stx.caseid_customer49(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer50':
                    caseid_stx.caseid_customer50(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer51':
                    caseid_stx.caseid_customer51(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer52':
                    caseid_stx.caseid_customer52(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer53':
                    caseid_stx.caseid_customer53(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer54':
                    caseid_stx.caseid_customer54(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer55':
                    caseid_stx.caseid_customer55(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer56':
                    caseid_stx.caseid_customer56(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer57':
                    caseid_stx.caseid_customer57(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer57_1':
                    caseid_stx.caseid_customer57_1(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer57_2':
                    caseid_stx.caseid_customer57_2(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer57_3':
                    caseid_stx.caseid_customer57_3(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer57_4':
                    caseid_stx.caseid_customer57_4(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer57_5':
                    caseid_stx.caseid_customer57_5(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Customer58':
                    caseid_stx.caseid_customer58(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer59':
                    caseid_stx.caseid_customer59(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer60':
                    caseid_stx.caseid_customer60(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer61':
                    caseid_stx.caseid_customer61(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer62':
                    caseid_stx.caseid_customer62(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer63':
                    caseid_stx.caseid_customer63(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer64':
                    caseid_stx.caseid_customer64(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer65':
                    caseid_stx.caseid_customer65(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer66':
                    caseid_stx.caseid_customer66(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer67':
                    caseid_stx.caseid_customer67(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer68':
                    caseid_stx.caseid_customer68(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer69':
                    caseid_stx.caseid_customer69(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer70':
                    caseid_stx.caseid_customer70(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer71':
                    caseid_stx.caseid_customer71(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer72':
                    caseid_stx.caseid_customer72(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer73':
                    caseid_stx.caseid_customer73(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer74':
                    caseid_stx.caseid_customer74(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer75':
                    caseid_stx.caseid_customer75(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer76':
                    caseid_stx.caseid_customer76(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer77':
                    caseid_stx.caseid_customer77(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer77_1':
                    caseid_stx.caseid_customer77_1(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Customer78':
                    caseid_stx.caseid_customer78(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer79':
                    caseid_stx.caseid_customer79(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer80':
                    caseid_stx.caseid_customer80(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer81':
                    caseid_stx.caseid_customer81(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer82':
                    caseid_stx.caseid_customer82(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer83':
                    caseid_stx.caseid_customer83(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer84':
                    caseid_stx.caseid_customer84(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer85':
                    caseid_stx.caseid_customer85(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer86':
                    caseid_stx.caseid_customer86(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer87':
                    caseid_stx.caseid_customer87(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Customer88':
                    caseid_stx.caseid_customer88(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer89':
                    caseid_stx.caseid_customer89(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer90':
                    caseid_stx.caseid_customer90(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer91':
                    caseid_stx.caseid_customer91(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer92':
                    caseid_stx.caseid_customer92(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Customer93':
                    caseid_stx.caseid_customer93(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Report01':
                    caseid_stx.caseid_report01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report02':
                    caseid_stx.caseid_report02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report03':
                    caseid_stx.caseid_report03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report04':
                    caseid_stx.caseid_report04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report05':
                    caseid_stx.caseid_report05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report06':
                    caseid_stx.caseid_report06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report07':
                    caseid_stx.caseid_report07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report07_1':
                    caseid_stx.caseid_report07_1(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report07_2':
                    caseid_stx.caseid_report07_2(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Report08':
                    caseid_stx.caseid_report08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report09':
                    caseid_stx.caseid_report09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report10':
                    caseid_stx.caseid_report10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report11':
                    caseid_stx.caseid_report11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report12':
                    caseid_stx.caseid_report12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report13':
                    caseid_stx.caseid_report13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report14':
                    caseid_stx.caseid_report14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report15':
                    caseid_stx.caseid_report15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report16':
                    caseid_stx.caseid_report16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report17':
                    caseid_stx.caseid_report17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report18':
                    caseid_stx.caseid_report18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report19':
                    caseid_stx.caseid_report19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report20':
                    caseid_stx.caseid_report20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report21':
                    caseid_stx.caseid_report21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report22':
                    caseid_stx.caseid_report22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report23':
                    caseid_stx.caseid_report23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report24':
                    caseid_stx.caseid_report24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report25':
                    caseid_stx.caseid_report25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report26':
                    caseid_stx.caseid_report26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report27':
                    caseid_stx.caseid_report27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report28':
                    caseid_stx.caseid_report28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report29':
                    caseid_stx.caseid_report29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report30':
                    caseid_stx.caseid_report30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report31':
                    caseid_stx.caseid_report31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report32':
                    caseid_stx.caseid_report32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report33':
                    caseid_stx.caseid_report33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report34':
                    caseid_stx.caseid_report34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report35':
                    caseid_stx.caseid_report35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report36':
                    caseid_stx.caseid_report36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report37':
                    caseid_stx.caseid_report37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report38':
                    caseid_stx.caseid_report38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report39':
                    caseid_stx.caseid_report39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report40':
                    caseid_stx.caseid_report40(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report41':
                    caseid_stx.caseid_report41(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report42':
                    caseid_stx.caseid_report42(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report43':
                    caseid_stx.caseid_report43(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report44':
                    caseid_stx.caseid_report44(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report45':
                    caseid_stx.caseid_report45(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report46':
                    caseid_stx.caseid_report46(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report47':
                    caseid_stx.caseid_report47(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report48':
                    caseid_stx.caseid_report48(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report49':
                    caseid_stx.caseid_report49(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report50':
                    caseid_stx.caseid_report50(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report51':
                    caseid_stx.caseid_report51(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report52':
                    caseid_stx.caseid_report52(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report53':
                    caseid_stx.caseid_report53(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report54':
                    caseid_stx.caseid_report54(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report55':
                    caseid_stx.caseid_report55(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report56':
                    caseid_stx.caseid_report56(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report57':
                    caseid_stx.caseid_report57(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report58':
                    caseid_stx.caseid_report58(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report59':
                    caseid_stx.caseid_report59(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report60':
                    caseid_stx.caseid_report60(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report61':
                    caseid_stx.caseid_report61(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report62':
                    caseid_stx.caseid_report62(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report63':
                    caseid_stx.caseid_report63(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report64':
                    caseid_stx.caseid_report64(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report65':
                    caseid_stx.caseid_report65(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report66':
                    caseid_stx.caseid_report66(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report67':
                    caseid_stx.caseid_report67(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report68':
                    caseid_stx.caseid_report68(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report69':
                    caseid_stx.caseid_report69(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report70':
                    caseid_stx.caseid_report70(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report71':
                    caseid_stx.caseid_report71(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report72':
                    caseid_stx.caseid_report72(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report73':
                    caseid_stx.caseid_report73(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report74':
                    caseid_stx.caseid_report74(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report75':
                    caseid_stx.caseid_report75(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report76':
                    caseid_stx.caseid_report76(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report77':
                    caseid_stx.caseid_report77(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report77_1':
                    caseid_stx.caseid_report77_1(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report77_2':
                    caseid_stx.caseid_report77_2(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report78':
                    caseid_stx.caseid_report78(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report79':
                    caseid_stx.caseid_report79(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Report80':
                    caseid_stx.caseid_report80(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'Admin01':
                    caseid_stx.caseid_admin01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin02':
                    caseid_stx.caseid_admin02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin03':
                    caseid_stx.caseid_admin03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin04':
                    caseid_stx.caseid_admin04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin05':
                    caseid_stx.caseid_admin05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin06':
                    caseid_stx.caseid_admin06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin07':
                    caseid_stx.caseid_admin07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08':
                    caseid_stx.caseid_admin08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_1':
                    caseid_stx.caseid_admin08_1(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_2':
                    caseid_stx.caseid_admin08_2(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_3':
                    caseid_stx.caseid_admin08_3(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_4':
                    caseid_stx.caseid_admin08_4(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_5':
                    caseid_stx.caseid_admin08_5(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_6':
                    caseid_stx.caseid_admin08_6(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_7':
                    caseid_stx.caseid_admin08_7(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_8':
                    caseid_stx.caseid_admin08_8(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin08_9':
                    caseid_stx.caseid_admin08_9(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin09':
                    caseid_stx.caseid_admin09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin10':
                    caseid_stx.caseid_admin10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin11':
                    caseid_stx.caseid_admin11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin12':
                    caseid_stx.caseid_admin12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin13':
                    caseid_stx.caseid_admin13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin14':
                    caseid_stx.caseid_admin14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin15':
                    caseid_stx.caseid_admin15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin16':
                    caseid_stx.caseid_admin16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin17':
                    caseid_stx.caseid_admin17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin18':
                    caseid_stx.caseid_admin18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin19':
                    caseid_stx.caseid_admin19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin20':
                    caseid_stx.caseid_admin20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin21':
                    caseid_stx.caseid_admin21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin22':
                    caseid_stx.caseid_admin22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin23':
                    caseid_stx.caseid_admin23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin24':
                    caseid_stx.caseid_admin24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin25':
                    caseid_stx.caseid_admin25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin26':
                    caseid_stx.caseid_admin26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin27':
                    caseid_stx.caseid_admin27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin28':
                    caseid_stx.caseid_admin28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin29':
                    caseid_stx.caseid_admin29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin30':
                    caseid_stx.caseid_admin30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin31':
                    caseid_stx.caseid_admin31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin32':
                    caseid_stx.caseid_admin32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin33':
                    caseid_stx.caseid_admin33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin34':
                    caseid_stx.caseid_admin34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin35':
                    caseid_stx.caseid_admin35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin36':
                    caseid_stx.caseid_admin36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin37':
                    caseid_stx.caseid_admin37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin38':
                    caseid_stx.caseid_admin38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin39':
                    caseid_stx.caseid_admin39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin40':
                    caseid_stx.caseid_admin40(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin41':
                    caseid_stx.caseid_admin41(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin42':
                    caseid_stx.caseid_admin42(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin43':
                    caseid_stx.caseid_admin43(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin44':
                    caseid_stx.caseid_admin44(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin45':
                    caseid_stx.caseid_admin45(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin46':
                    caseid_stx.caseid_admin46(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin47':
                    caseid_stx.caseid_admin47(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin48':
                    caseid_stx.caseid_admin48(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin49':
                    caseid_stx.caseid_admin49(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin50':
                    caseid_stx.caseid_admin50(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin51':
                    caseid_stx.caseid_admin51(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin52':
                    caseid_stx.caseid_admin52(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin53':
                    caseid_stx.caseid_admin53(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin54':
                    caseid_stx.caseid_admin54(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin55':
                    caseid_stx.caseid_admin55(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin56':
                    caseid_stx.caseid_admin56(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin57':
                    caseid_stx.caseid_admin57(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin58':
                    caseid_stx.caseid_admin58(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin59':
                    caseid_stx.caseid_admin59(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin60':
                    caseid_stx.caseid_admin60(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin61':
                    caseid_stx.caseid_admin61(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin62':
                    caseid_stx.caseid_admin62(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin63':
                    caseid_stx.caseid_admin63(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin64':
                    caseid_stx.caseid_admin64(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin65':
                    caseid_stx.caseid_admin65(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin66':
                    caseid_stx.caseid_admin66(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin67':
                    caseid_stx.caseid_admin67(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin68':
                    caseid_stx.caseid_admin68(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin69':
                    caseid_stx.caseid_admin69(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin70':
                    caseid_stx.caseid_admin70(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin71':
                    caseid_stx.caseid_admin71(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin72':
                    caseid_stx.caseid_admin72(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin73':
                    caseid_stx.caseid_admin73(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin74':
                    caseid_stx.caseid_admin74(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin75':
                    caseid_stx.caseid_admin75(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin76':
                    caseid_stx.caseid_admin76(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin77':
                    caseid_stx.caseid_admin77(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin78':
                    caseid_stx.caseid_admin78(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin79':
                    caseid_stx.caseid_admin79(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin80':
                    caseid_stx.caseid_admin80(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin81':
                    caseid_stx.caseid_admin81(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin82':
                    caseid_stx.caseid_admin82(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin83':
                    caseid_stx.caseid_admin83(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin84':
                    caseid_stx.caseid_admin84(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin85':
                    caseid_stx.caseid_admin85(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin86':
                    caseid_stx.caseid_admin86(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin87':
                    caseid_stx.caseid_admin87(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin88':
                    caseid_stx.caseid_admin88(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin89':
                    caseid_stx.caseid_admin89(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin90':
                    caseid_stx.caseid_admin90(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin91':
                    caseid_stx.caseid_admin91(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin92':
                    caseid_stx.caseid_admin92(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin93':
                    caseid_stx.caseid_admin93(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin94':
                    caseid_stx.caseid_admin94(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin95':
                    caseid_stx.caseid_admin95(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin96':
                    caseid_stx.caseid_admin96(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin97':
                    caseid_stx.caseid_admin97(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin98':
                    caseid_stx.caseid_admin98(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin99':
                    caseid_stx.caseid_admin99(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin100':
                    caseid_stx.caseid_admin100(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin101':
                    caseid_stx.caseid_admin101(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin102':
                    caseid_stx.caseid_admin102(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin103':
                    caseid_stx.caseid_admin103(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin104':
                    caseid_stx.caseid_admin104(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin105':
                    caseid_stx.caseid_admin105(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin106':
                    caseid_stx.caseid_admin106(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin107':
                    caseid_stx.caseid_admin107(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin108':
                    caseid_stx.caseid_admin108(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin109':
                    caseid_stx.caseid_admin109(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin110':
                    caseid_stx.caseid_admin110(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin111':
                    caseid_stx.caseid_admin111(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin112':
                    caseid_stx.caseid_admin112(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin113':
                    caseid_stx.caseid_admin113(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin114':
                    caseid_stx.caseid_admin114(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin115':
                    caseid_stx.caseid_admin115(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin116':
                    caseid_stx.caseid_admin116(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin117':
                    caseid_stx.caseid_admin117(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin118':
                    caseid_stx.caseid_admin118(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin119':
                    caseid_stx.caseid_admin119(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin120':
                    caseid_stx.caseid_admin120(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin121':
                    caseid_stx.caseid_admin121(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin122':
                    caseid_stx.caseid_admin122(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin123':
                    caseid_stx.caseid_admin123(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin124':
                    caseid_stx.caseid_admin124(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin125':
                    caseid_stx.caseid_admin125(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin126':
                    caseid_stx.caseid_admin126(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin127':
                    caseid_stx.caseid_admin127(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin128':
                    caseid_stx.caseid_admin128(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Admin129':
                    caseid_stx.caseid_admin129(self)
            except:
                module_other_stx.swich_tab_0()

            try:
                if case == 'PartnerTrip01':
                    caseid_stx.caseid_PartnerTrip01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'PartnerTrip02':
                    caseid_stx.caseid_PartnerTrip02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'PartnerTrip03':
                    caseid_stx.caseid_PartnerTrip03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'PartnerTrip04':
                    caseid_stx.caseid_PartnerTrip04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'PartnerTrip05':
                    caseid_stx.caseid_PartnerTrip05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'PartnerTrip06':
                    caseid_stx.caseid_PartnerTrip06(self)
            except:
                module_other_stx.swich_tab_0()



            try:
                if case == 'Accounting01':
                    caseid_stx.caseid_accounting01(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting02':
                    caseid_stx.caseid_accounting02(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting03':
                    caseid_stx.caseid_accounting03(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting04':
                    caseid_stx.caseid_accounting04(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting05':
                    caseid_stx.caseid_accounting05(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting06':
                    caseid_stx.caseid_accounting06(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting07':
                    caseid_stx.caseid_accounting07(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting08':
                    caseid_stx.caseid_accounting08(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting09':
                    caseid_stx.caseid_accounting09(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting10':
                    caseid_stx.caseid_accounting10(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting11':
                    caseid_stx.caseid_accounting11(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting12':
                    caseid_stx.caseid_accounting12(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting13':
                    caseid_stx.caseid_accounting13(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting14':
                    caseid_stx.caseid_accounting14(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting15':
                    caseid_stx.caseid_accounting15(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting16':
                    caseid_stx.caseid_accounting16(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting17':
                    caseid_stx.caseid_accounting17(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting18':
                    caseid_stx.caseid_accounting18(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting19':
                    caseid_stx.caseid_accounting19(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting20':
                    caseid_stx.caseid_accounting20(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting21':
                    caseid_stx.caseid_accounting21(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting22':
                    caseid_stx.caseid_accounting22(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting23':
                    caseid_stx.caseid_accounting23(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting24':
                    caseid_stx.caseid_accounting24(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting25':
                    caseid_stx.caseid_accounting25(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting26':
                    caseid_stx.caseid_accounting26(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting27':
                    caseid_stx.caseid_accounting27(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting28':
                    caseid_stx.caseid_accounting28(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting29':
                    caseid_stx.caseid_accounting29(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting30':
                    caseid_stx.caseid_accounting30(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting31':
                    caseid_stx.caseid_accounting31(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting32':
                    caseid_stx.caseid_accounting32(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting33':
                    caseid_stx.caseid_accounting33(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting34':
                    caseid_stx.caseid_accounting34(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting35':
                    caseid_stx.caseid_accounting35(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting36':
                    caseid_stx.caseid_accounting36(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting37':
                    caseid_stx.caseid_accounting37(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting38':
                    caseid_stx.caseid_accounting38(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting39':
                    caseid_stx.caseid_accounting39(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting40':
                    caseid_stx.caseid_accounting40(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting41':
                    caseid_stx.caseid_accounting41(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting42':
                    caseid_stx.caseid_accounting42(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting43':
                    caseid_stx.caseid_accounting43(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting44':
                    caseid_stx.caseid_accounting44(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting45':
                    caseid_stx.caseid_accounting45(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting46':
                    caseid_stx.caseid_accounting46(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting47':
                    caseid_stx.caseid_accounting47(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting48':
                    caseid_stx.caseid_accounting48(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting49':
                    caseid_stx.caseid_accounting49(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting50':
                    caseid_stx.caseid_accounting50(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting51':
                    caseid_stx.caseid_accounting51(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting52':
                    caseid_stx.caseid_accounting52(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting53':
                    caseid_stx.caseid_accounting53(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting54':
                    caseid_stx.caseid_accounting54(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting55':
                    caseid_stx.caseid_accounting55(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting56':
                    caseid_stx.caseid_accounting56(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting57':
                    caseid_stx.caseid_accounting57(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting58':
                    caseid_stx.caseid_accounting58(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting59':
                    caseid_stx.caseid_accounting59(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting60':
                    caseid_stx.caseid_accounting60(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting61':
                    caseid_stx.caseid_accounting61(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting62':
                    caseid_stx.caseid_accounting62(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting63':
                    caseid_stx.caseid_accounting63(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting64':
                    caseid_stx.caseid_accounting64(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting65':
                    caseid_stx.caseid_accounting65(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting66':
                    caseid_stx.caseid_accounting66(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting67':
                    caseid_stx.caseid_accounting67(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting68':
                    caseid_stx.caseid_accounting68(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting69':
                    caseid_stx.caseid_accounting69(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting70':
                    caseid_stx.caseid_accounting70(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting71':
                    caseid_stx.caseid_accounting71(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting72':
                    caseid_stx.caseid_accounting72(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting73':
                    caseid_stx.caseid_accounting73(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting74':
                    caseid_stx.caseid_accounting74(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting75':
                    caseid_stx.caseid_accounting75(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting76':
                    caseid_stx.caseid_accounting76(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting77':
                    caseid_stx.caseid_accounting77(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting78':
                    caseid_stx.caseid_accounting78(self)
            except:
                module_other_stx.swich_tab_0()
            try:
                if case == 'Accounting79':
                    caseid_stx.caseid_accounting79(self)
            except:
                module_other_stx.swich_tab_0()



def retest_casefail(self):
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["G"+rownum].value == "Fail":
            print(sheet["A"+rownum].value)
            case_fail = sheet["A"+rownum].value
            list_casefail.append(case_fail)
        rownum = int(rownum)
    print(list_casefail)
    for case in list_casefail:
        try:
            if case == 'Login01':
                caseid_stx.caseid_login01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login02':
                caseid_stx.caseid_login02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login03':
                caseid_stx.caseid_login03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login04':
                caseid_stx.caseid_login04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login05':
                caseid_stx.caseid_login05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login06':
                caseid_stx.caseid_login06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login07':
                caseid_stx.caseid_login07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login08':
                caseid_stx.caseid_login08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login09':
                caseid_stx.caseid_login09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login10':
                caseid_stx.caseid_login10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login11':
                caseid_stx.caseid_login11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login12':
                caseid_stx.caseid_login12(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Minitor01':
                caseid_stx.caseid_minitor01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor02':
                caseid_stx.caseid_minitor02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor03':
                caseid_stx.caseid_minitor03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor04':
                caseid_stx.caseid_minitor04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor05':
                caseid_stx.caseid_minitor05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor06':
                caseid_stx.caseid_minitor06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor07':
                caseid_stx.caseid_minitor07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor08':
                caseid_stx.caseid_minitor08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor09':
                caseid_stx.caseid_minitor09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor10':
                caseid_stx.caseid_minitor10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor11':
                caseid_stx.caseid_minitor11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor12':
                caseid_stx.caseid_minitor12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor13':
                caseid_stx.caseid_minitor13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor14':
                caseid_stx.caseid_minitor14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor15':
                caseid_stx.caseid_minitor15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor16':
                caseid_stx.caseid_minitor16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor17':
                caseid_stx.caseid_minitor17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor18':
                caseid_stx.caseid_minitor18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor19':
                caseid_stx.caseid_minitor19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor20':
                caseid_stx.caseid_minitor20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor21':
                caseid_stx.caseid_minitor21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor22':
                caseid_stx.caseid_minitor22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor23':
                caseid_stx.caseid_minitor23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor24':
                caseid_stx.caseid_minitor24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor25':
                caseid_stx.caseid_minitor25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor26':
                caseid_stx.caseid_minitor26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor27':
                caseid_stx.caseid_minitor27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor28':
                caseid_stx.caseid_minitor28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor29':
                caseid_stx.caseid_minitor29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor30':
                caseid_stx.caseid_minitor30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor31':
                caseid_stx.caseid_minitor31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor32':
                caseid_stx.caseid_minitor32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor33':
                caseid_stx.caseid_minitor33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor34':
                caseid_stx.caseid_minitor34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor35':
                caseid_stx.caseid_minitor35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor36':
                caseid_stx.caseid_minitor36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor37':
                caseid_stx.caseid_minitor37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor38':
                caseid_stx.caseid_minitor38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor39':
                caseid_stx.caseid_minitor39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor40':
                caseid_stx.caseid_minitor40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor41':
                caseid_stx.caseid_minitor41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor42':
                caseid_stx.caseid_minitor42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor43':
                caseid_stx.caseid_minitor43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor44':
                caseid_stx.caseid_minitor44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor45':
                caseid_stx.caseid_minitor45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor46':
                caseid_stx.caseid_minitor46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor47':
                caseid_stx.caseid_minitor47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor48':
                caseid_stx.caseid_minitor48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor49':
                caseid_stx.caseid_minitor49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor50':
                caseid_stx.caseid_minitor50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor51':
                caseid_stx.caseid_minitor51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor52':
                caseid_stx.caseid_minitor52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor53':
                caseid_stx.caseid_minitor53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor54':
                caseid_stx.caseid_minitor54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor55':
                caseid_stx.caseid_minitor55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor56':
                caseid_stx.caseid_minitor56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor57':
                caseid_stx.caseid_minitor57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor58':
                caseid_stx.caseid_minitor58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor59':
                caseid_stx.caseid_minitor59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor60':
                caseid_stx.caseid_minitor60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor61':
                caseid_stx.caseid_minitor61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor62':
                caseid_stx.caseid_minitor62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor63':
                caseid_stx.caseid_minitor63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor64':
                caseid_stx.caseid_minitor64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor65':
                caseid_stx.caseid_minitor65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor66':
                caseid_stx.caseid_minitor66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor67':
                caseid_stx.caseid_minitor67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor68':
                caseid_stx.caseid_minitor68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor69':
                caseid_stx.caseid_minitor69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor70':
                caseid_stx.caseid_minitor70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor71':
                caseid_stx.caseid_minitor71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor72':
                caseid_stx.caseid_minitor72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor73':
                caseid_stx.caseid_minitor73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor74':
                caseid_stx.caseid_minitor74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor75':
                caseid_stx.caseid_minitor75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor76':
                caseid_stx.caseid_minitor76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor77':
                caseid_stx.caseid_minitor77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor78':
                caseid_stx.caseid_minitor78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor79':
                caseid_stx.caseid_minitor79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor80':
                caseid_stx.caseid_minitor80(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Vehicle01':
                caseid_stx.caseid_vehicle01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle02':
                caseid_stx.caseid_vehicle02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle03':
                caseid_stx.caseid_vehicle03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle04':
                caseid_stx.caseid_vehicle04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle05':
                caseid_stx.caseid_vehicle05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle06':
                caseid_stx.caseid_vehicle06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle07':
                caseid_stx.caseid_vehicle07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle08':
                caseid_stx.caseid_vehicle08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle09':
                caseid_stx.caseid_vehicle09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle10':
                caseid_stx.caseid_vehicle10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle11':
                caseid_stx.caseid_vehicle11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle12':
                caseid_stx.caseid_vehicle12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle13':
                caseid_stx.caseid_vehicle13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle14':
                caseid_stx.caseid_vehicle14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle15':
                caseid_stx.caseid_vehicle15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle16':
                caseid_stx.caseid_vehicle16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle17':
                caseid_stx.caseid_vehicle17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle18':
                caseid_stx.caseid_vehicle18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle19':
                caseid_stx.caseid_vehicle19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle20':
                caseid_stx.caseid_vehicle20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle21':
                caseid_stx.caseid_vehicle21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle22':
                caseid_stx.caseid_vehicle22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle23':
                caseid_stx.caseid_vehicle23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle24':
                caseid_stx.caseid_vehicle24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle25':
                caseid_stx.caseid_vehicle25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle26':
                caseid_stx.caseid_vehicle26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle27':
                caseid_stx.caseid_vehicle27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle28':
                caseid_stx.caseid_vehicle28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle29':
                caseid_stx.caseid_vehicle29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle30':
                caseid_stx.caseid_vehicle30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle31':
                caseid_stx.caseid_vehicle31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle32':
                caseid_stx.caseid_vehicle32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle33':
                caseid_stx.caseid_vehicle33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle34':
                caseid_stx.caseid_vehicle34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle35':
                caseid_stx.caseid_vehicle35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle36':
                caseid_stx.caseid_vehicle36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle37':
                caseid_stx.caseid_vehicle37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle38':
                caseid_stx.caseid_vehicle38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle39':
                caseid_stx.caseid_vehicle39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle40':
                caseid_stx.caseid_vehicle40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle41':
                caseid_stx.caseid_vehicle41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle42':
                caseid_stx.caseid_vehicle42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle43':
                caseid_stx.caseid_vehicle43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle44':
                caseid_stx.caseid_vehicle44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle45':
                caseid_stx.caseid_vehicle45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle46':
                caseid_stx.caseid_vehicle46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle47':
                caseid_stx.caseid_vehicle47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle48':
                caseid_stx.caseid_vehicle48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle49':
                caseid_stx.caseid_vehicle49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle50':
                caseid_stx.caseid_vehicle50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle51':
                caseid_stx.caseid_vehicle51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle52':
                caseid_stx.caseid_vehicle52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle53':
                caseid_stx.caseid_vehicle53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle54':
                caseid_stx.caseid_vehicle54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle55':
                caseid_stx.caseid_vehicle55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle56':
                caseid_stx.caseid_vehicle56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle57':
                caseid_stx.caseid_vehicle57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle58':
                caseid_stx.caseid_vehicle58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle59':
                caseid_stx.caseid_vehicle59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle60':
                caseid_stx.caseid_vehicle60(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Wallet01':
                caseid_stx.caseid_wallet01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet02':
                caseid_stx.caseid_wallet02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet03':
                caseid_stx.caseid_wallet03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet04':
                caseid_stx.caseid_wallet04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet05':
                caseid_stx.caseid_wallet05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet06':
                caseid_stx.caseid_wallet06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet07':
                caseid_stx.caseid_wallet07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet08':
                caseid_stx.caseid_wallet08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet09':
                caseid_stx.caseid_wallet09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet10':
                caseid_stx.caseid_wallet10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet11':
                caseid_stx.caseid_wallet11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet12':
                caseid_stx.caseid_wallet12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet13':
                caseid_stx.caseid_wallet13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet14':
                caseid_stx.caseid_wallet14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet15':
                caseid_stx.caseid_wallet15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet16':
                caseid_stx.caseid_wallet16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet17':
                caseid_stx.caseid_wallet17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet18':
                caseid_stx.caseid_wallet18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet19':
                caseid_stx.caseid_wallet19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet20':
                caseid_stx.caseid_wallet20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet21':
                caseid_stx.caseid_wallet21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet22':
                caseid_stx.caseid_wallet22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet23':
                caseid_stx.caseid_wallet23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet24':
                caseid_stx.caseid_wallet24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet25':
                caseid_stx.caseid_wallet25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet26':
                caseid_stx.caseid_wallet26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet27':
                caseid_stx.caseid_wallet27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet28':
                caseid_stx.caseid_wallet28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet29':
                caseid_stx.caseid_wallet29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet30':
                caseid_stx.caseid_wallet30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet31':
                caseid_stx.caseid_wallet31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet32':
                caseid_stx.caseid_wallet32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet33':
                caseid_stx.caseid_wallet33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet34':
                caseid_stx.caseid_wallet34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet35':
                caseid_stx.caseid_wallet35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet36':
                caseid_stx.caseid_wallet36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet37':
                caseid_stx.caseid_wallet37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet38':
                caseid_stx.caseid_wallet38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet39':
                caseid_stx.caseid_wallet39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet40':
                caseid_stx.caseid_wallet40(self)
        except:
            module_other_stx.swich_tab_0()



        try:
            if case == 'Promotion01':
                caseid_stx.caseid_promotion01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion02':
                caseid_stx.caseid_promotion02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion03':
                caseid_stx.caseid_promotion03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion04':
                caseid_stx.caseid_promotion04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion05':
                caseid_stx.caseid_promotion05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion06':
                caseid_stx.caseid_promotion06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion07':
                caseid_stx.caseid_promotion07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion08':
                caseid_stx.caseid_promotion08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion09':
                caseid_stx.caseid_promotion09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion10':
                caseid_stx.caseid_promotion10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion11':
                caseid_stx.caseid_promotion11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion12':
                caseid_stx.caseid_promotion12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion13':
                caseid_stx.caseid_promotion13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion14':
                caseid_stx.caseid_promotion14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion15':
                caseid_stx.caseid_promotion15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion16':
                caseid_stx.caseid_promotion16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion17':
                caseid_stx.caseid_promotion17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion18':
                caseid_stx.caseid_promotion18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion19':
                caseid_stx.caseid_promotion19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion20':
                caseid_stx.caseid_promotion20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion21':
                caseid_stx.caseid_promotion21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion22':
                caseid_stx.caseid_promotion22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion23':
                caseid_stx.caseid_promotion23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion24':
                caseid_stx.caseid_promotion24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion25':
                caseid_stx.caseid_promotion25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion26':
                caseid_stx.caseid_promotion26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion27':
                caseid_stx.caseid_promotion27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion28':
                caseid_stx.caseid_promotion28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion29':
                caseid_stx.caseid_promotion29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion30':
                caseid_stx.caseid_promotion30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion31':
                caseid_stx.caseid_promotion31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion32':
                caseid_stx.caseid_promotion32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion33':
                caseid_stx.caseid_promotion33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion34':
                caseid_stx.caseid_promotion34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion35':
                caseid_stx.caseid_promotion35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion36':
                caseid_stx.caseid_promotion36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion37':
                caseid_stx.caseid_promotion37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion38':
                caseid_stx.caseid_promotion38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion39':
                caseid_stx.caseid_promotion39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion40':
                caseid_stx.caseid_promotion40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion41':
                caseid_stx.caseid_promotion41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion42':
                caseid_stx.caseid_promotion42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion43':
                caseid_stx.caseid_promotion43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion44':
                caseid_stx.caseid_promotion44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion45':
                caseid_stx.caseid_promotion45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion46':
                caseid_stx.caseid_promotion46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion47':
                caseid_stx.caseid_promotion47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion48':
                caseid_stx.caseid_promotion48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion49':
                caseid_stx.caseid_promotion49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion50':
                caseid_stx.caseid_promotion50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion51':
                caseid_stx.caseid_promotion51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion52':
                caseid_stx.caseid_promotion52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion53':
                caseid_stx.caseid_promotion53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion54':
                caseid_stx.caseid_promotion54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion55':
                caseid_stx.caseid_promotion55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion56':
                caseid_stx.caseid_promotion56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion57':
                caseid_stx.caseid_promotion57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion58':
                caseid_stx.caseid_promotion58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion59':
                caseid_stx.caseid_promotion59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion60':
                caseid_stx.caseid_promotion60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion61':
                caseid_stx.caseid_promotion61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion62':
                caseid_stx.caseid_promotion62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion63':
                caseid_stx.caseid_promotion63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion64':
                caseid_stx.caseid_promotion64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion65':
                caseid_stx.caseid_promotion65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion66':
                caseid_stx.caseid_promotion66(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Customer01':
                caseid_stx.caseid_customer01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer02':
                caseid_stx.caseid_customer02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer03':
                caseid_stx.caseid_customer03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer04':
                caseid_stx.caseid_customer04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer05':
                caseid_stx.caseid_customer05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer06':
                caseid_stx.caseid_customer06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer07':
                caseid_stx.caseid_customer07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer08':
                caseid_stx.caseid_customer08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer09':
                caseid_stx.caseid_customer09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer10':
                caseid_stx.caseid_customer10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer11':
                caseid_stx.caseid_customer11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer12':
                caseid_stx.caseid_customer12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer13':
                caseid_stx.caseid_customer13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer14':
                caseid_stx.caseid_customer14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer15':
                caseid_stx.caseid_customer15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer16':
                caseid_stx.caseid_customer16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer17':
                caseid_stx.caseid_customer17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer18':
                caseid_stx.caseid_customer18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer19':
                caseid_stx.caseid_customer19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer20':
                caseid_stx.caseid_customer20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer21':
                caseid_stx.caseid_customer21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer22':
                caseid_stx.caseid_customer22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer23':
                caseid_stx.caseid_customer23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer24':
                caseid_stx.caseid_customer24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer25':
                caseid_stx.caseid_customer25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer26':
                caseid_stx.caseid_customer26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer26_1':
                caseid_stx.caseid_customer26_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer26_2':
                caseid_stx.caseid_customer26_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer27':
                caseid_stx.caseid_customer27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer28':
                caseid_stx.caseid_customer28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer29':
                caseid_stx.caseid_customer29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer30':
                caseid_stx.caseid_customer30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer31':
                caseid_stx.caseid_customer31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer32':
                caseid_stx.caseid_customer32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer33':
                caseid_stx.caseid_customer33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer34':
                caseid_stx.caseid_customer34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer35':
                caseid_stx.caseid_customer35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer36':
                caseid_stx.caseid_customer36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer37':
                caseid_stx.caseid_customer37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer38':
                caseid_stx.caseid_customer38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer39':
                caseid_stx.caseid_customer39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer40':
                caseid_stx.caseid_customer40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer41':
                caseid_stx.caseid_customer41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer42':
                caseid_stx.caseid_customer42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer43':
                caseid_stx.caseid_customer43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer44':
                caseid_stx.caseid_customer44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer45':
                caseid_stx.caseid_customer45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer46':
                caseid_stx.caseid_customer46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer47':
                caseid_stx.caseid_customer47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer48':
                caseid_stx.caseid_customer48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer49':
                caseid_stx.caseid_customer49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer50':
                caseid_stx.caseid_customer50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer51':
                caseid_stx.caseid_customer51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer52':
                caseid_stx.caseid_customer52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer53':
                caseid_stx.caseid_customer53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer54':
                caseid_stx.caseid_customer54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer55':
                caseid_stx.caseid_customer55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer56':
                caseid_stx.caseid_customer56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57':
                caseid_stx.caseid_customer57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_1':
                caseid_stx.caseid_customer57_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_2':
                caseid_stx.caseid_customer57_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_3':
                caseid_stx.caseid_customer57_3(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_4':
                caseid_stx.caseid_customer57_4(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_5':
                caseid_stx.caseid_customer57_5(self)
        except:
            module_other_stx.swich_tab_0()


        try:
            if case == 'Customer58':
                caseid_stx.caseid_customer58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer59':
                caseid_stx.caseid_customer59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer60':
                caseid_stx.caseid_customer60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer61':
                caseid_stx.caseid_customer61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer62':
                caseid_stx.caseid_customer62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer63':
                caseid_stx.caseid_customer63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer64':
                caseid_stx.caseid_customer64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer65':
                caseid_stx.caseid_customer65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer66':
                caseid_stx.caseid_customer66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer67':
                caseid_stx.caseid_customer67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer68':
                caseid_stx.caseid_customer68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer69':
                caseid_stx.caseid_customer69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer70':
                caseid_stx.caseid_customer70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer71':
                caseid_stx.caseid_customer71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer72':
                caseid_stx.caseid_customer72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer73':
                caseid_stx.caseid_customer73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer74':
                caseid_stx.caseid_customer74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer75':
                caseid_stx.caseid_customer75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer76':
                caseid_stx.caseid_customer76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer77':
                caseid_stx.caseid_customer77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer77_1':
                caseid_stx.caseid_customer77_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer78':
                caseid_stx.caseid_customer78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer79':
                caseid_stx.caseid_customer79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer80':
                caseid_stx.caseid_customer80(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer81':
                caseid_stx.caseid_customer81(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer82':
                caseid_stx.caseid_customer82(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer83':
                caseid_stx.caseid_customer83(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer84':
                caseid_stx.caseid_customer84(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer85':
                caseid_stx.caseid_customer85(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer86':
                caseid_stx.caseid_customer86(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer87':
                caseid_stx.caseid_customer87(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Customer88':
                caseid_stx.caseid_customer88(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer89':
                caseid_stx.caseid_customer89(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer90':
                caseid_stx.caseid_customer90(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer91':
                caseid_stx.caseid_customer91(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer92':
                caseid_stx.caseid_customer92(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer93':
                caseid_stx.caseid_customer93(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Report01':
                caseid_stx.caseid_report01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report02':
                caseid_stx.caseid_report02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report03':
                caseid_stx.caseid_report03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report04':
                caseid_stx.caseid_report04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report05':
                caseid_stx.caseid_report05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report06':
                caseid_stx.caseid_report06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report07':
                caseid_stx.caseid_report07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report07_1':
                caseid_stx.caseid_report07_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report07_2':
                caseid_stx.caseid_report07_2(self)
        except:
            module_other_stx.swich_tab_0()


        try:
            if case == 'Report08':
                caseid_stx.caseid_report08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report09':
                caseid_stx.caseid_report09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report10':
                caseid_stx.caseid_report10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report11':
                caseid_stx.caseid_report11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report12':
                caseid_stx.caseid_report12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report13':
                caseid_stx.caseid_report13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report14':
                caseid_stx.caseid_report14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report15':
                caseid_stx.caseid_report15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report16':
                caseid_stx.caseid_report16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report17':
                caseid_stx.caseid_report17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report18':
                caseid_stx.caseid_report18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report19':
                caseid_stx.caseid_report19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report20':
                caseid_stx.caseid_report20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report21':
                caseid_stx.caseid_report21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report22':
                caseid_stx.caseid_report22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report23':
                caseid_stx.caseid_report23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report24':
                caseid_stx.caseid_report24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report25':
                caseid_stx.caseid_report25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report26':
                caseid_stx.caseid_report26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report27':
                caseid_stx.caseid_report27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report28':
                caseid_stx.caseid_report28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report29':
                caseid_stx.caseid_report29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report30':
                caseid_stx.caseid_report30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report31':
                caseid_stx.caseid_report31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report32':
                caseid_stx.caseid_report32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report33':
                caseid_stx.caseid_report33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report34':
                caseid_stx.caseid_report34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report35':
                caseid_stx.caseid_report35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report36':
                caseid_stx.caseid_report36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report37':
                caseid_stx.caseid_report37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report38':
                caseid_stx.caseid_report38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report39':
                caseid_stx.caseid_report39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report40':
                caseid_stx.caseid_report40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report41':
                caseid_stx.caseid_report41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report42':
                caseid_stx.caseid_report42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report43':
                caseid_stx.caseid_report43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report44':
                caseid_stx.caseid_report44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report45':
                caseid_stx.caseid_report45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report46':
                caseid_stx.caseid_report46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report47':
                caseid_stx.caseid_report47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report48':
                caseid_stx.caseid_report48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report49':
                caseid_stx.caseid_report49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report50':
                caseid_stx.caseid_report50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report51':
                caseid_stx.caseid_report51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report52':
                caseid_stx.caseid_report52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report53':
                caseid_stx.caseid_report53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report54':
                caseid_stx.caseid_report54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report55':
                caseid_stx.caseid_report55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report56':
                caseid_stx.caseid_report56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report57':
                caseid_stx.caseid_report57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report58':
                caseid_stx.caseid_report58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report59':
                caseid_stx.caseid_report59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report60':
                caseid_stx.caseid_report60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report61':
                caseid_stx.caseid_report61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report62':
                caseid_stx.caseid_report62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report63':
                caseid_stx.caseid_report63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report64':
                caseid_stx.caseid_report64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report65':
                caseid_stx.caseid_report65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report66':
                caseid_stx.caseid_report66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report67':
                caseid_stx.caseid_report67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report68':
                caseid_stx.caseid_report68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report69':
                caseid_stx.caseid_report69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report70':
                caseid_stx.caseid_report70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report71':
                caseid_stx.caseid_report71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report72':
                caseid_stx.caseid_report72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report73':
                caseid_stx.caseid_report73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report74':
                caseid_stx.caseid_report74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report75':
                caseid_stx.caseid_report75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report76':
                caseid_stx.caseid_report76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report77':
                caseid_stx.caseid_report77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report77_1':
                caseid_stx.caseid_report77_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report77_2':
                caseid_stx.caseid_report77_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report78':
                caseid_stx.caseid_report78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report79':
                caseid_stx.caseid_report79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report80':
                caseid_stx.caseid_report80(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Admin01':
                caseid_stx.caseid_admin01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin02':
                caseid_stx.caseid_admin02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin03':
                caseid_stx.caseid_admin03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin04':
                caseid_stx.caseid_admin04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin05':
                caseid_stx.caseid_admin05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin06':
                caseid_stx.caseid_admin06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin07':
                caseid_stx.caseid_admin07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08':
                caseid_stx.caseid_admin08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_1':
                caseid_stx.caseid_admin08_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_2':
                caseid_stx.caseid_admin08_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_3':
                caseid_stx.caseid_admin08_3(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_4':
                caseid_stx.caseid_admin08_4(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_5':
                caseid_stx.caseid_admin08_5(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_6':
                caseid_stx.caseid_admin08_6(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_7':
                caseid_stx.caseid_admin08_7(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_8':
                caseid_stx.caseid_admin08_8(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_9':
                caseid_stx.caseid_admin08_9(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin09':
                caseid_stx.caseid_admin09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin10':
                caseid_stx.caseid_admin10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin11':
                caseid_stx.caseid_admin11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin12':
                caseid_stx.caseid_admin12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin13':
                caseid_stx.caseid_admin13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin14':
                caseid_stx.caseid_admin14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin15':
                caseid_stx.caseid_admin15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin16':
                caseid_stx.caseid_admin16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin17':
                caseid_stx.caseid_admin17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin18':
                caseid_stx.caseid_admin18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin19':
                caseid_stx.caseid_admin19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin20':
                caseid_stx.caseid_admin20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin21':
                caseid_stx.caseid_admin21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin22':
                caseid_stx.caseid_admin22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin23':
                caseid_stx.caseid_admin23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin24':
                caseid_stx.caseid_admin24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin25':
                caseid_stx.caseid_admin25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin26':
                caseid_stx.caseid_admin26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin27':
                caseid_stx.caseid_admin27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin28':
                caseid_stx.caseid_admin28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin29':
                caseid_stx.caseid_admin29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin30':
                caseid_stx.caseid_admin30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin31':
                caseid_stx.caseid_admin31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin32':
                caseid_stx.caseid_admin32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin33':
                caseid_stx.caseid_admin33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin34':
                caseid_stx.caseid_admin34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin35':
                caseid_stx.caseid_admin35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin36':
                caseid_stx.caseid_admin36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin37':
                caseid_stx.caseid_admin37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin38':
                caseid_stx.caseid_admin38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin39':
                caseid_stx.caseid_admin39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin40':
                caseid_stx.caseid_admin40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin41':
                caseid_stx.caseid_admin41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin42':
                caseid_stx.caseid_admin42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin43':
                caseid_stx.caseid_admin43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin44':
                caseid_stx.caseid_admin44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin45':
                caseid_stx.caseid_admin45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin46':
                caseid_stx.caseid_admin46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin47':
                caseid_stx.caseid_admin47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin48':
                caseid_stx.caseid_admin48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin49':
                caseid_stx.caseid_admin49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin50':
                caseid_stx.caseid_admin50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin51':
                caseid_stx.caseid_admin51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin52':
                caseid_stx.caseid_admin52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin53':
                caseid_stx.caseid_admin53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin54':
                caseid_stx.caseid_admin54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin55':
                caseid_stx.caseid_admin55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin56':
                caseid_stx.caseid_admin56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin57':
                caseid_stx.caseid_admin57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin58':
                caseid_stx.caseid_admin58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin59':
                caseid_stx.caseid_admin59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin60':
                caseid_stx.caseid_admin60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin61':
                caseid_stx.caseid_admin61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin62':
                caseid_stx.caseid_admin62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin63':
                caseid_stx.caseid_admin63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin64':
                caseid_stx.caseid_admin64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin65':
                caseid_stx.caseid_admin65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin66':
                caseid_stx.caseid_admin66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin67':
                caseid_stx.caseid_admin67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin68':
                caseid_stx.caseid_admin68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin69':
                caseid_stx.caseid_admin69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin70':
                caseid_stx.caseid_admin70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin71':
                caseid_stx.caseid_admin71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin72':
                caseid_stx.caseid_admin72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin73':
                caseid_stx.caseid_admin73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin74':
                caseid_stx.caseid_admin74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin75':
                caseid_stx.caseid_admin75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin76':
                caseid_stx.caseid_admin76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin77':
                caseid_stx.caseid_admin77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin78':
                caseid_stx.caseid_admin78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin79':
                caseid_stx.caseid_admin79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin80':
                caseid_stx.caseid_admin80(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin81':
                caseid_stx.caseid_admin81(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin82':
                caseid_stx.caseid_admin82(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin83':
                caseid_stx.caseid_admin83(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin84':
                caseid_stx.caseid_admin84(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin85':
                caseid_stx.caseid_admin85(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin86':
                caseid_stx.caseid_admin86(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin87':
                caseid_stx.caseid_admin87(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin88':
                caseid_stx.caseid_admin88(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin89':
                caseid_stx.caseid_admin89(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin90':
                caseid_stx.caseid_admin90(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin91':
                caseid_stx.caseid_admin91(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin92':
                caseid_stx.caseid_admin92(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin93':
                caseid_stx.caseid_admin93(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin94':
                caseid_stx.caseid_admin94(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin95':
                caseid_stx.caseid_admin95(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin96':
                caseid_stx.caseid_admin96(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin97':
                caseid_stx.caseid_admin97(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin98':
                caseid_stx.caseid_admin98(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin99':
                caseid_stx.caseid_admin99(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin100':
                caseid_stx.caseid_admin100(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin101':
                caseid_stx.caseid_admin101(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin102':
                caseid_stx.caseid_admin102(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin103':
                caseid_stx.caseid_admin103(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin104':
                caseid_stx.caseid_admin104(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin105':
                caseid_stx.caseid_admin105(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin106':
                caseid_stx.caseid_admin106(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin107':
                caseid_stx.caseid_admin107(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin108':
                caseid_stx.caseid_admin108(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin109':
                caseid_stx.caseid_admin109(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin110':
                caseid_stx.caseid_admin110(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin111':
                caseid_stx.caseid_admin111(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin112':
                caseid_stx.caseid_admin112(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin113':
                caseid_stx.caseid_admin113(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin114':
                caseid_stx.caseid_admin114(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin115':
                caseid_stx.caseid_admin115(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin116':
                caseid_stx.caseid_admin116(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin117':
                caseid_stx.caseid_admin117(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin118':
                caseid_stx.caseid_admin118(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin119':
                caseid_stx.caseid_admin119(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin120':
                caseid_stx.caseid_admin120(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin121':
                caseid_stx.caseid_admin121(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin122':
                caseid_stx.caseid_admin122(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin123':
                caseid_stx.caseid_admin123(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin124':
                caseid_stx.caseid_admin124(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin125':
                caseid_stx.caseid_admin125(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin126':
                caseid_stx.caseid_admin126(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin127':
                caseid_stx.caseid_admin127(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin128':
                caseid_stx.caseid_admin128(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin129':
                caseid_stx.caseid_admin129(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip01':
                caseid_stx.caseid_PartnerTrip01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip02':
                caseid_stx.caseid_PartnerTrip02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip03':
                caseid_stx.caseid_PartnerTrip03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip04':
                caseid_stx.caseid_PartnerTrip04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip05':
                caseid_stx.caseid_PartnerTrip05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip06':
                caseid_stx.caseid_PartnerTrip06(self)
        except:
            module_other_stx.swich_tab_0()



        try:
            if case == 'Accounting01':
                caseid_stx.caseid_accounting01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting02':
                caseid_stx.caseid_accounting02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting03':
                caseid_stx.caseid_accounting03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting04':
                caseid_stx.caseid_accounting04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting05':
                caseid_stx.caseid_accounting05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting06':
                caseid_stx.caseid_accounting06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting07':
                caseid_stx.caseid_accounting07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting08':
                caseid_stx.caseid_accounting08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting09':
                caseid_stx.caseid_accounting09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting10':
                caseid_stx.caseid_accounting10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting11':
                caseid_stx.caseid_accounting11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting12':
                caseid_stx.caseid_accounting12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting13':
                caseid_stx.caseid_accounting13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting14':
                caseid_stx.caseid_accounting14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting15':
                caseid_stx.caseid_accounting15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting16':
                caseid_stx.caseid_accounting16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting17':
                caseid_stx.caseid_accounting17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting18':
                caseid_stx.caseid_accounting18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting19':
                caseid_stx.caseid_accounting19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting20':
                caseid_stx.caseid_accounting20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting21':
                caseid_stx.caseid_accounting21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting22':
                caseid_stx.caseid_accounting22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting23':
                caseid_stx.caseid_accounting23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting24':
                caseid_stx.caseid_accounting24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting25':
                caseid_stx.caseid_accounting25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting26':
                caseid_stx.caseid_accounting26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting27':
                caseid_stx.caseid_accounting27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting28':
                caseid_stx.caseid_accounting28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting29':
                caseid_stx.caseid_accounting29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting30':
                caseid_stx.caseid_accounting30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting31':
                caseid_stx.caseid_accounting31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting32':
                caseid_stx.caseid_accounting32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting33':
                caseid_stx.caseid_accounting33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting34':
                caseid_stx.caseid_accounting34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting35':
                caseid_stx.caseid_accounting35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting36':
                caseid_stx.caseid_accounting36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting37':
                caseid_stx.caseid_accounting37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting38':
                caseid_stx.caseid_accounting38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting39':
                caseid_stx.caseid_accounting39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting40':
                caseid_stx.caseid_accounting40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting41':
                caseid_stx.caseid_accounting41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting42':
                caseid_stx.caseid_accounting42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting43':
                caseid_stx.caseid_accounting43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting44':
                caseid_stx.caseid_accounting44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting45':
                caseid_stx.caseid_accounting45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting46':
                caseid_stx.caseid_accounting46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting47':
                caseid_stx.caseid_accounting47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting48':
                caseid_stx.caseid_accounting48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting49':
                caseid_stx.caseid_accounting49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting50':
                caseid_stx.caseid_accounting50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting51':
                caseid_stx.caseid_accounting51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting52':
                caseid_stx.caseid_accounting52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting53':
                caseid_stx.caseid_accounting53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting54':
                caseid_stx.caseid_accounting54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting55':
                caseid_stx.caseid_accounting55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting56':
                caseid_stx.caseid_accounting56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting57':
                caseid_stx.caseid_accounting57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting58':
                caseid_stx.caseid_accounting58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting59':
                caseid_stx.caseid_accounting59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting60':
                caseid_stx.caseid_accounting60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting61':
                caseid_stx.caseid_accounting61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting62':
                caseid_stx.caseid_accounting62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting63':
                caseid_stx.caseid_accounting63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting64':
                caseid_stx.caseid_accounting64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting65':
                caseid_stx.caseid_accounting65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting66':
                caseid_stx.caseid_accounting66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting67':
                caseid_stx.caseid_accounting67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting68':
                caseid_stx.caseid_accounting68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting69':
                caseid_stx.caseid_accounting69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting70':
                caseid_stx.caseid_accounting70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting71':
                caseid_stx.caseid_accounting71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting72':
                caseid_stx.caseid_accounting72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting73':
                caseid_stx.caseid_accounting73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting74':
                caseid_stx.caseid_accounting74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting75':
                caseid_stx.caseid_accounting75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting76':
                caseid_stx.caseid_accounting76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting77':
                caseid_stx.caseid_accounting77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting78':
                caseid_stx.caseid_accounting78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting79':
                caseid_stx.caseid_accounting79(self)
        except:
            module_other_stx.swich_tab_0()







def retest_serious():
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 290, 2, "")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 290, 3, "")
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 291, 2, 0)
    var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 292, 2, "")

    list_casefail = []
    grouped_by_casefail = defaultdict(list)  # Nhóm theo case_fail

    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while rownum < 1200:
        rownum += 1
        row_str = str(rownum)

        if sheet["L" + row_str].value == "x" and sheet["G" + row_str].value == "Fail":
            case_fail = sheet["A" + row_str].value
            event_name = sheet["B" + row_str].value
            account = sheet["C" + row_str].value
            step = sheet["D" + row_str].value
            desired_result = sheet["E" + row_str].value
            actual_result = sheet["F" + row_str].value
            status = sheet["G" + row_str].value

            grouped_by_casefail[case_fail].append({
                'event_name': event_name,
                'account': account,
                'step': step,
                'desired_result': desired_result,
                'actual_result': actual_result,
                'status': status
            })

            list_casefail.append(case_fail)

    print(list_casefail)
    print("Số case fail nghiêm trọng: ", len(list_casefail))

    print("\n====== NHÓM THEO CASE FAIL ======")
    for case_fail, items in grouped_by_casefail.items():
        print(f"\nCase Fail: {case_fail} - Số dòng liên quan: {len(items)}")

        # Gom tên case_fail (chỉ 1 vì mỗi nhóm là duy nhất)
        casefail_str = case_fail

        # Gom tất cả các bước step lại
        all_steps = "\n".join([item['step'] for item in items if item['step']])

        # Gộp nội dung chi tiết lỗi
        message_bug = ""
        for item in items:
            message_bug += (f" -Mã: {case_fail}\n"
                            f" -Tên sự kiện: {item['event_name']}\n"
                            f" -Tài khoản: {item['account']}\n"
                            f" -Các bước thao tác: {item['step']}\n"
                            f" -Kết quả mong muốn: {item['desired_result']}\n"
                            f" -Kết quả thực tế: {item['actual_result']}\n"
                            f" -Trạng thái: {item['status']}\n"
                            f"---------------------------------------------------------\n")
            var_stx.writeData_append(var_stx.path_luutamthoi, "Sheet1", 290, 3, f"{case_fail} {item['event_name']}, \n")

        # Ghi vào Excel
        var_stx.writeData_append(var_stx.path_luutamthoi, "Sheet1", 290, 2, f"{casefail_str}, ")  # Tên case_fail

        count_casefail = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', 290, 2))
        parts = count_casefail.split(',')
        cleaned = [p.strip() for p in parts if p.strip()]
        count = len(cleaned)

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 291, 2, count)  # Số dòng liên quan
        var_stx.writeData_append(var_stx.path_luutamthoi, "Sheet1", 292, 2, message_bug)  # Ghi toàn bộ message
        print(message_bug)










def run_all(self):
    list_0 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        print(sheet["G" + rownum].value)
        print(sheet["H" + rownum].value)
        if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet["J" + rownum].value == "x"
            or sheet["K" + rownum].value == "x") and sheet["A" + rownum].value != None:
            print(sheet["A" + rownum].value)
            case_0 = sheet["A" + rownum].value
            list_0.append(case_0)
        rownum = int(rownum)
    print(list_0)
    for case in list_0:
        try:
            if case == 'Login01':
                caseid_stx.caseid_login01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login02':
                caseid_stx.caseid_login02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login03':
                caseid_stx.caseid_login03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login04':
                caseid_stx.caseid_login04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login05':
                caseid_stx.caseid_login05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login06':
                caseid_stx.caseid_login06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login07':
                caseid_stx.caseid_login07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login08':
                caseid_stx.caseid_login08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login09':
                caseid_stx.caseid_login09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login10':
                caseid_stx.caseid_login10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login11':
                caseid_stx.caseid_login11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Login12':
                caseid_stx.caseid_login12(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Minitor01':
                caseid_stx.caseid_minitor01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor02':
                caseid_stx.caseid_minitor02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor03':
                caseid_stx.caseid_minitor03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor04':
                caseid_stx.caseid_minitor04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor05':
                caseid_stx.caseid_minitor05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor06':
                caseid_stx.caseid_minitor06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor07':
                caseid_stx.caseid_minitor07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor08':
                caseid_stx.caseid_minitor08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor09':
                caseid_stx.caseid_minitor09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor10':
                caseid_stx.caseid_minitor10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor11':
                caseid_stx.caseid_minitor11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor12':
                caseid_stx.caseid_minitor12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor13':
                caseid_stx.caseid_minitor13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor14':
                caseid_stx.caseid_minitor14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor15':
                caseid_stx.caseid_minitor15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor16':
                caseid_stx.caseid_minitor16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor17':
                caseid_stx.caseid_minitor17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor18':
                caseid_stx.caseid_minitor18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor19':
                caseid_stx.caseid_minitor19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor20':
                caseid_stx.caseid_minitor20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor21':
                caseid_stx.caseid_minitor21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor22':
                caseid_stx.caseid_minitor22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor23':
                caseid_stx.caseid_minitor23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor24':
                caseid_stx.caseid_minitor24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor25':
                caseid_stx.caseid_minitor25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor26':
                caseid_stx.caseid_minitor26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor27':
                caseid_stx.caseid_minitor27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor28':
                caseid_stx.caseid_minitor28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor29':
                caseid_stx.caseid_minitor29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor30':
                caseid_stx.caseid_minitor30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor31':
                caseid_stx.caseid_minitor31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor32':
                caseid_stx.caseid_minitor32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor33':
                caseid_stx.caseid_minitor33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor34':
                caseid_stx.caseid_minitor34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor35':
                caseid_stx.caseid_minitor35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor36':
                caseid_stx.caseid_minitor36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor37':
                caseid_stx.caseid_minitor37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor38':
                caseid_stx.caseid_minitor38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor39':
                caseid_stx.caseid_minitor39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor40':
                caseid_stx.caseid_minitor40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor41':
                caseid_stx.caseid_minitor41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor42':
                caseid_stx.caseid_minitor42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor43':
                caseid_stx.caseid_minitor43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor44':
                caseid_stx.caseid_minitor44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor45':
                caseid_stx.caseid_minitor45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor46':
                caseid_stx.caseid_minitor46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor47':
                caseid_stx.caseid_minitor47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor48':
                caseid_stx.caseid_minitor48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor49':
                caseid_stx.caseid_minitor49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor50':
                caseid_stx.caseid_minitor50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor51':
                caseid_stx.caseid_minitor51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor52':
                caseid_stx.caseid_minitor52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor53':
                caseid_stx.caseid_minitor53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor54':
                caseid_stx.caseid_minitor54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor55':
                caseid_stx.caseid_minitor55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor56':
                caseid_stx.caseid_minitor56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor57':
                caseid_stx.caseid_minitor57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor58':
                caseid_stx.caseid_minitor58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor59':
                caseid_stx.caseid_minitor59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor60':
                caseid_stx.caseid_minitor60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor61':
                caseid_stx.caseid_minitor61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor62':
                caseid_stx.caseid_minitor62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor63':
                caseid_stx.caseid_minitor63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor64':
                caseid_stx.caseid_minitor64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor65':
                caseid_stx.caseid_minitor65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor66':
                caseid_stx.caseid_minitor66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor67':
                caseid_stx.caseid_minitor67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor68':
                caseid_stx.caseid_minitor68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor69':
                caseid_stx.caseid_minitor69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor70':
                caseid_stx.caseid_minitor70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor71':
                caseid_stx.caseid_minitor71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor72':
                caseid_stx.caseid_minitor72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor73':
                caseid_stx.caseid_minitor73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor74':
                caseid_stx.caseid_minitor74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor75':
                caseid_stx.caseid_minitor75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor76':
                caseid_stx.caseid_minitor76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor77':
                caseid_stx.caseid_minitor77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor78':
                caseid_stx.caseid_minitor78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor79':
                caseid_stx.caseid_minitor79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Minitor80':
                caseid_stx.caseid_minitor80(self)
        except:
            module_other_stx.swich_tab_0()


        try:
            if case == 'Vehicle01':
                caseid_stx.caseid_vehicle01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle02':
                caseid_stx.caseid_vehicle02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle03':
                caseid_stx.caseid_vehicle03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle04':
                caseid_stx.caseid_vehicle04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle05':
                caseid_stx.caseid_vehicle05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle06':
                caseid_stx.caseid_vehicle06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle07':
                caseid_stx.caseid_vehicle07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle08':
                caseid_stx.caseid_vehicle08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle09':
                caseid_stx.caseid_vehicle09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle10':
                caseid_stx.caseid_vehicle10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle11':
                caseid_stx.caseid_vehicle11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle12':
                caseid_stx.caseid_vehicle12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle13':
                caseid_stx.caseid_vehicle13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle14':
                caseid_stx.caseid_vehicle14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle15':
                caseid_stx.caseid_vehicle15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle16':
                caseid_stx.caseid_vehicle16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle17':
                caseid_stx.caseid_vehicle17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle18':
                caseid_stx.caseid_vehicle18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle19':
                caseid_stx.caseid_vehicle19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle20':
                caseid_stx.caseid_vehicle20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle21':
                caseid_stx.caseid_vehicle21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle22':
                caseid_stx.caseid_vehicle22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle23':
                caseid_stx.caseid_vehicle23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle24':
                caseid_stx.caseid_vehicle24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle25':
                caseid_stx.caseid_vehicle25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle26':
                caseid_stx.caseid_vehicle26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle27':
                caseid_stx.caseid_vehicle27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle28':
                caseid_stx.caseid_vehicle28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle29':
                caseid_stx.caseid_vehicle29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle30':
                caseid_stx.caseid_vehicle30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle31':
                caseid_stx.caseid_vehicle31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle32':
                caseid_stx.caseid_vehicle32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle33':
                caseid_stx.caseid_vehicle33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle34':
                caseid_stx.caseid_vehicle34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle35':
                caseid_stx.caseid_vehicle35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle36':
                caseid_stx.caseid_vehicle36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle37':
                caseid_stx.caseid_vehicle37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle38':
                caseid_stx.caseid_vehicle38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle39':
                caseid_stx.caseid_vehicle39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle40':
                caseid_stx.caseid_vehicle40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle41':
                caseid_stx.caseid_vehicle41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle42':
                caseid_stx.caseid_vehicle42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle43':
                caseid_stx.caseid_vehicle43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle44':
                caseid_stx.caseid_vehicle44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle45':
                caseid_stx.caseid_vehicle45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle46':
                caseid_stx.caseid_vehicle46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle47':
                caseid_stx.caseid_vehicle47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle48':
                caseid_stx.caseid_vehicle48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle49':
                caseid_stx.caseid_vehicle49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle50':
                caseid_stx.caseid_vehicle50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle51':
                caseid_stx.caseid_vehicle51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle52':
                caseid_stx.caseid_vehicle52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle53':
                caseid_stx.caseid_vehicle53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle54':
                caseid_stx.caseid_vehicle54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle55':
                caseid_stx.caseid_vehicle55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle56':
                caseid_stx.caseid_vehicle56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle57':
                caseid_stx.caseid_vehicle57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle58':
                caseid_stx.caseid_vehicle58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle59':
                caseid_stx.caseid_vehicle59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Vehicle60':
                caseid_stx.caseid_vehicle60(self)
        except:
            module_other_stx.swich_tab_0()


        try:
            if case == 'Wallet01':
                caseid_stx.caseid_wallet01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet02':
                caseid_stx.caseid_wallet02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet03':
                caseid_stx.caseid_wallet03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet04':
                caseid_stx.caseid_wallet04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet05':
                caseid_stx.caseid_wallet05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet06':
                caseid_stx.caseid_wallet06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet07':
                caseid_stx.caseid_wallet07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet08':
                caseid_stx.caseid_wallet08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet09':
                caseid_stx.caseid_wallet09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet10':
                caseid_stx.caseid_wallet10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet11':
                caseid_stx.caseid_wallet11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet12':
                caseid_stx.caseid_wallet12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet13':
                caseid_stx.caseid_wallet13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet14':
                caseid_stx.caseid_wallet14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet15':
                caseid_stx.caseid_wallet15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet16':
                caseid_stx.caseid_wallet16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet17':
                caseid_stx.caseid_wallet17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet18':
                caseid_stx.caseid_wallet18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet19':
                caseid_stx.caseid_wallet19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet20':
                caseid_stx.caseid_wallet20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet21':
                caseid_stx.caseid_wallet21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet22':
                caseid_stx.caseid_wallet22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet23':
                caseid_stx.caseid_wallet23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet24':
                caseid_stx.caseid_wallet24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet25':
                caseid_stx.caseid_wallet25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet26':
                caseid_stx.caseid_wallet26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet27':
                caseid_stx.caseid_wallet27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet28':
                caseid_stx.caseid_wallet28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet29':
                caseid_stx.caseid_wallet29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet30':
                caseid_stx.caseid_wallet30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet31':
                caseid_stx.caseid_wallet31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet32':
                caseid_stx.caseid_wallet32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet33':
                caseid_stx.caseid_wallet33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet34':
                caseid_stx.caseid_wallet34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet35':
                caseid_stx.caseid_wallet35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet36':
                caseid_stx.caseid_wallet36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet37':
                caseid_stx.caseid_wallet37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet38':
                caseid_stx.caseid_wallet38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet39':
                caseid_stx.caseid_wallet39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Wallet40':
                caseid_stx.caseid_wallet40(self)
        except:
            module_other_stx.swich_tab_0()


        try:
            if case == 'Promotion01':
                caseid_stx.caseid_promotion01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion02':
                caseid_stx.caseid_promotion02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion03':
                caseid_stx.caseid_promotion03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion04':
                caseid_stx.caseid_promotion04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion05':
                caseid_stx.caseid_promotion05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion05_1':
                caseid_stx.caseid_promotion05_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion05_1':
                caseid_stx.caseid_promotion05_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion06':
                caseid_stx.caseid_promotion06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion07':
                caseid_stx.caseid_promotion07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion08':
                caseid_stx.caseid_promotion08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion09':
                caseid_stx.caseid_promotion09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion10':
                caseid_stx.caseid_promotion10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion11':
                caseid_stx.caseid_promotion11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion12':
                caseid_stx.caseid_promotion12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion13':
                caseid_stx.caseid_promotion13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion14':
                caseid_stx.caseid_promotion14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion15':
                caseid_stx.caseid_promotion15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion16':
                caseid_stx.caseid_promotion16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion17':
                caseid_stx.caseid_promotion17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion18':
                caseid_stx.caseid_promotion18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion19':
                caseid_stx.caseid_promotion19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion20':
                caseid_stx.caseid_promotion20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion21':
                caseid_stx.caseid_promotion21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion22':
                caseid_stx.caseid_promotion22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion23':
                caseid_stx.caseid_promotion23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion24':
                caseid_stx.caseid_promotion24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion25':
                caseid_stx.caseid_promotion25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion26':
                caseid_stx.caseid_promotion26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion27':
                caseid_stx.caseid_promotion27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion28':
                caseid_stx.caseid_promotion28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion29':
                caseid_stx.caseid_promotion29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion30':
                caseid_stx.caseid_promotion30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion31':
                caseid_stx.caseid_promotion31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion32':
                caseid_stx.caseid_promotion32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion33':
                caseid_stx.caseid_promotion33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion34':
                caseid_stx.caseid_promotion34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion35':
                caseid_stx.caseid_promotion35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion36':
                caseid_stx.caseid_promotion36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion37':
                caseid_stx.caseid_promotion37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion38':
                caseid_stx.caseid_promotion38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion39':
                caseid_stx.caseid_promotion39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion40':
                caseid_stx.caseid_promotion40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion41':
                caseid_stx.caseid_promotion41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion42':
                caseid_stx.caseid_promotion42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion43':
                caseid_stx.caseid_promotion43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion44':
                caseid_stx.caseid_promotion44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion45':
                caseid_stx.caseid_promotion45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion46':
                caseid_stx.caseid_promotion46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion47':
                caseid_stx.caseid_promotion47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion48':
                caseid_stx.caseid_promotion48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion49':
                caseid_stx.caseid_promotion49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion50':
                caseid_stx.caseid_promotion50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion51':
                caseid_stx.caseid_promotion51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion52':
                caseid_stx.caseid_promotion52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion53':
                caseid_stx.caseid_promotion53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion54':
                caseid_stx.caseid_promotion54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion55':
                caseid_stx.caseid_promotion55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion56':
                caseid_stx.caseid_promotion56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion57':
                caseid_stx.caseid_promotion57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion58':
                caseid_stx.caseid_promotion58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion59':
                caseid_stx.caseid_promotion59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion60':
                caseid_stx.caseid_promotion60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion61':
                caseid_stx.caseid_promotion61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion62':
                caseid_stx.caseid_promotion62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion63':
                caseid_stx.caseid_promotion63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion64':
                caseid_stx.caseid_promotion64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion65':
                caseid_stx.caseid_promotion65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Promotion66':
                caseid_stx.caseid_promotion66(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Customer01':
                caseid_stx.caseid_customer01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer02':
                caseid_stx.caseid_customer02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer03':
                caseid_stx.caseid_customer03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer04':
                caseid_stx.caseid_customer04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer05':
                caseid_stx.caseid_customer05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer06':
                caseid_stx.caseid_customer06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer07':
                caseid_stx.caseid_customer07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer08':
                caseid_stx.caseid_customer08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer09':
                caseid_stx.caseid_customer09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer10':
                caseid_stx.caseid_customer10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer11':
                caseid_stx.caseid_customer11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer12':
                caseid_stx.caseid_customer12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer13':
                caseid_stx.caseid_customer13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer14':
                caseid_stx.caseid_customer14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer15':
                caseid_stx.caseid_customer15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer16':
                caseid_stx.caseid_customer16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer17':
                caseid_stx.caseid_customer17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer18':
                caseid_stx.caseid_customer18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer19':
                caseid_stx.caseid_customer19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer20':
                caseid_stx.caseid_customer20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer21':
                caseid_stx.caseid_customer21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer22':
                caseid_stx.caseid_customer22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer23':
                caseid_stx.caseid_customer23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer24':
                caseid_stx.caseid_customer24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer25':
                caseid_stx.caseid_customer25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer26':
                caseid_stx.caseid_customer26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer26_1':
                caseid_stx.caseid_customer26_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer26_2':
                caseid_stx.caseid_customer26_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer27':
                caseid_stx.caseid_customer27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer28':
                caseid_stx.caseid_customer28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer29':
                caseid_stx.caseid_customer29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer30':
                caseid_stx.caseid_customer30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer31':
                caseid_stx.caseid_customer31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer32':
                caseid_stx.caseid_customer32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer33':
                caseid_stx.caseid_customer33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer34':
                caseid_stx.caseid_customer34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer35':
                caseid_stx.caseid_customer35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer36':
                caseid_stx.caseid_customer36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer37':
                caseid_stx.caseid_customer37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer38':
                caseid_stx.caseid_customer38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer39':
                caseid_stx.caseid_customer39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer40':
                caseid_stx.caseid_customer40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer41':
                caseid_stx.caseid_customer41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer42':
                caseid_stx.caseid_customer42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer43':
                caseid_stx.caseid_customer43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer44':
                caseid_stx.caseid_customer44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer45':
                caseid_stx.caseid_customer45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer46':
                caseid_stx.caseid_customer46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer47':
                caseid_stx.caseid_customer47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer48':
                caseid_stx.caseid_customer48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer49':
                caseid_stx.caseid_customer49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer50':
                caseid_stx.caseid_customer50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer51':
                caseid_stx.caseid_customer51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer52':
                caseid_stx.caseid_customer52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer53':
                caseid_stx.caseid_customer53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer54':
                caseid_stx.caseid_customer54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer55':
                caseid_stx.caseid_customer55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer56':
                caseid_stx.caseid_customer56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57':
                caseid_stx.caseid_customer57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_1':
                caseid_stx.caseid_customer57_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_2':
                caseid_stx.caseid_customer57_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_3':
                caseid_stx.caseid_customer57_3(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_4':
                caseid_stx.caseid_customer57_4(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer57_5':
                caseid_stx.caseid_customer57_5(self)
        except:
            module_other_stx.swich_tab_0()


        try:
            if case == 'Customer58':
                caseid_stx.caseid_customer58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer59':
                caseid_stx.caseid_customer59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer60':
                caseid_stx.caseid_customer60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer61':
                caseid_stx.caseid_customer61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer62':
                caseid_stx.caseid_customer62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer63':
                caseid_stx.caseid_customer63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer64':
                caseid_stx.caseid_customer64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer65':
                caseid_stx.caseid_customer65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer66':
                caseid_stx.caseid_customer66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer67':
                caseid_stx.caseid_customer67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer68':
                caseid_stx.caseid_customer68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer69':
                caseid_stx.caseid_customer69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer70':
                caseid_stx.caseid_customer70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer71':
                caseid_stx.caseid_customer71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer72':
                caseid_stx.caseid_customer72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer73':
                caseid_stx.caseid_customer73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer74':
                caseid_stx.caseid_customer74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer75':
                caseid_stx.caseid_customer75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer76':
                caseid_stx.caseid_customer76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer77':
                caseid_stx.caseid_customer77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer77_1':
                caseid_stx.caseid_customer77_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer78':
                caseid_stx.caseid_customer78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer79':
                caseid_stx.caseid_customer79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer80':
                caseid_stx.caseid_customer80(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer81':
                caseid_stx.caseid_customer81(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer82':
                caseid_stx.caseid_customer82(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer83':
                caseid_stx.caseid_customer83(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer84':
                caseid_stx.caseid_customer84(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer85':
                caseid_stx.caseid_customer85(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer86':
                caseid_stx.caseid_customer86(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer87':
                caseid_stx.caseid_customer87(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Customer88':
                caseid_stx.caseid_customer88(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer89':
                caseid_stx.caseid_customer89(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer90':
                caseid_stx.caseid_customer90(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer91':
                caseid_stx.caseid_customer91(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer92':
                caseid_stx.caseid_customer92(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Customer93':
                caseid_stx.caseid_customer93(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Report01':
                caseid_stx.caseid_report01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report02':
                caseid_stx.caseid_report02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report03':
                caseid_stx.caseid_report03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report04':
                caseid_stx.caseid_report04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report05':
                caseid_stx.caseid_report05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report06':
                caseid_stx.caseid_report06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report07':
                caseid_stx.caseid_report07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report07_1':
                caseid_stx.caseid_report07_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report07_2':
                caseid_stx.caseid_report07_2(self)
        except:
            module_other_stx.swich_tab_0()


        try:
            if case == 'Report08':
                caseid_stx.caseid_report08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report09':
                caseid_stx.caseid_report09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report10':
                caseid_stx.caseid_report10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report11':
                caseid_stx.caseid_report11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report12':
                caseid_stx.caseid_report12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report13':
                caseid_stx.caseid_report13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report14':
                caseid_stx.caseid_report14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report15':
                caseid_stx.caseid_report15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report16':
                caseid_stx.caseid_report16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report17':
                caseid_stx.caseid_report17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report18':
                caseid_stx.caseid_report18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report19':
                caseid_stx.caseid_report19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report20':
                caseid_stx.caseid_report20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report21':
                caseid_stx.caseid_report21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report22':
                caseid_stx.caseid_report22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report23':
                caseid_stx.caseid_report23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report24':
                caseid_stx.caseid_report24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report25':
                caseid_stx.caseid_report25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report26':
                caseid_stx.caseid_report26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report27':
                caseid_stx.caseid_report27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report28':
                caseid_stx.caseid_report28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report29':
                caseid_stx.caseid_report29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report30':
                caseid_stx.caseid_report30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report31':
                caseid_stx.caseid_report31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report32':
                caseid_stx.caseid_report32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report33':
                caseid_stx.caseid_report33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report34':
                caseid_stx.caseid_report34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report35':
                caseid_stx.caseid_report35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report36':
                caseid_stx.caseid_report36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report37':
                caseid_stx.caseid_report37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report38':
                caseid_stx.caseid_report38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report39':
                caseid_stx.caseid_report39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report40':
                caseid_stx.caseid_report40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report41':
                caseid_stx.caseid_report41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report42':
                caseid_stx.caseid_report42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report43':
                caseid_stx.caseid_report43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report44':
                caseid_stx.caseid_report44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report45':
                caseid_stx.caseid_report45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report46':
                caseid_stx.caseid_report46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report47':
                caseid_stx.caseid_report47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report48':
                caseid_stx.caseid_report48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report49':
                caseid_stx.caseid_report49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report50':
                caseid_stx.caseid_report50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report51':
                caseid_stx.caseid_report51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report52':
                caseid_stx.caseid_report52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report53':
                caseid_stx.caseid_report53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report54':
                caseid_stx.caseid_report54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report55':
                caseid_stx.caseid_report55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report56':
                caseid_stx.caseid_report56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report57':
                caseid_stx.caseid_report57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report58':
                caseid_stx.caseid_report58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report59':
                caseid_stx.caseid_report59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report60':
                caseid_stx.caseid_report60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report61':
                caseid_stx.caseid_report61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report62':
                caseid_stx.caseid_report62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report63':
                caseid_stx.caseid_report63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report64':
                caseid_stx.caseid_report64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report65':
                caseid_stx.caseid_report65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report66':
                caseid_stx.caseid_report66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report67':
                caseid_stx.caseid_report67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report68':
                caseid_stx.caseid_report68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report69':
                caseid_stx.caseid_report69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report70':
                caseid_stx.caseid_report70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report71':
                caseid_stx.caseid_report71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report72':
                caseid_stx.caseid_report72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report73':
                caseid_stx.caseid_report73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report74':
                caseid_stx.caseid_report74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report75':
                caseid_stx.caseid_report75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report76':
                caseid_stx.caseid_report76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report77':
                caseid_stx.caseid_report77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report77_1':
                caseid_stx.caseid_report77_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report77_2':
                caseid_stx.caseid_report77_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report78':
                caseid_stx.caseid_report78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report79':
                caseid_stx.caseid_report79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Report80':
                caseid_stx.caseid_report80(self)
        except:
            module_other_stx.swich_tab_0()

        try:
            if case == 'Admin01':
                caseid_stx.caseid_admin01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin02':
                caseid_stx.caseid_admin02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin03':
                caseid_stx.caseid_admin03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin04':
                caseid_stx.caseid_admin04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin05':
                caseid_stx.caseid_admin05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin06':
                caseid_stx.caseid_admin06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin07':
                caseid_stx.caseid_admin07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08':
                caseid_stx.caseid_admin08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_1':
                caseid_stx.caseid_admin08_1(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_2':
                caseid_stx.caseid_admin08_2(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_3':
                caseid_stx.caseid_admin08_3(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_4':
                caseid_stx.caseid_admin08_4(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_5':
                caseid_stx.caseid_admin08_5(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_6':
                caseid_stx.caseid_admin08_6(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_7':
                caseid_stx.caseid_admin08_7(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_8':
                caseid_stx.caseid_admin08_8(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin08_9':
                caseid_stx.caseid_admin08_9(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin09':
                caseid_stx.caseid_admin09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin10':
                caseid_stx.caseid_admin10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin11':
                caseid_stx.caseid_admin11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin12':
                caseid_stx.caseid_admin12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin13':
                caseid_stx.caseid_admin13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin14':
                caseid_stx.caseid_admin14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin15':
                caseid_stx.caseid_admin15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin16':
                caseid_stx.caseid_admin16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin17':
                caseid_stx.caseid_admin17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin18':
                caseid_stx.caseid_admin18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin19':
                caseid_stx.caseid_admin19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin20':
                caseid_stx.caseid_admin20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin21':
                caseid_stx.caseid_admin21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin22':
                caseid_stx.caseid_admin22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin23':
                caseid_stx.caseid_admin23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin24':
                caseid_stx.caseid_admin24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin25':
                caseid_stx.caseid_admin25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin26':
                caseid_stx.caseid_admin26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin27':
                caseid_stx.caseid_admin27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin28':
                caseid_stx.caseid_admin28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin29':
                caseid_stx.caseid_admin29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin30':
                caseid_stx.caseid_admin30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin31':
                caseid_stx.caseid_admin31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin32':
                caseid_stx.caseid_admin32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin33':
                caseid_stx.caseid_admin33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin34':
                caseid_stx.caseid_admin34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin35':
                caseid_stx.caseid_admin35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin36':
                caseid_stx.caseid_admin36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin37':
                caseid_stx.caseid_admin37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin38':
                caseid_stx.caseid_admin38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin39':
                caseid_stx.caseid_admin39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin40':
                caseid_stx.caseid_admin40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin41':
                caseid_stx.caseid_admin41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin42':
                caseid_stx.caseid_admin42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin43':
                caseid_stx.caseid_admin43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin44':
                caseid_stx.caseid_admin44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin45':
                caseid_stx.caseid_admin45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin46':
                caseid_stx.caseid_admin46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin47':
                caseid_stx.caseid_admin47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin48':
                caseid_stx.caseid_admin48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin49':
                caseid_stx.caseid_admin49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin50':
                caseid_stx.caseid_admin50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin51':
                caseid_stx.caseid_admin51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin52':
                caseid_stx.caseid_admin52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin53':
                caseid_stx.caseid_admin53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin54':
                caseid_stx.caseid_admin54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin55':
                caseid_stx.caseid_admin55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin56':
                caseid_stx.caseid_admin56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin57':
                caseid_stx.caseid_admin57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin58':
                caseid_stx.caseid_admin58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin59':
                caseid_stx.caseid_admin59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin60':
                caseid_stx.caseid_admin60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin61':
                caseid_stx.caseid_admin61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin62':
                caseid_stx.caseid_admin62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin63':
                caseid_stx.caseid_admin63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin64':
                caseid_stx.caseid_admin64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin65':
                caseid_stx.caseid_admin65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin66':
                caseid_stx.caseid_admin66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin67':
                caseid_stx.caseid_admin67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin68':
                caseid_stx.caseid_admin68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin69':
                caseid_stx.caseid_admin69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin70':
                caseid_stx.caseid_admin70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin71':
                caseid_stx.caseid_admin71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin72':
                caseid_stx.caseid_admin72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin73':
                caseid_stx.caseid_admin73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin74':
                caseid_stx.caseid_admin74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin75':
                caseid_stx.caseid_admin75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin76':
                caseid_stx.caseid_admin76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin77':
                caseid_stx.caseid_admin77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin78':
                caseid_stx.caseid_admin78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin79':
                caseid_stx.caseid_admin79(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin80':
                caseid_stx.caseid_admin80(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin81':
                caseid_stx.caseid_admin81(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin82':
                caseid_stx.caseid_admin82(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin83':
                caseid_stx.caseid_admin83(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin84':
                caseid_stx.caseid_admin84(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin85':
                caseid_stx.caseid_admin85(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin86':
                caseid_stx.caseid_admin86(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin87':
                caseid_stx.caseid_admin87(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin88':
                caseid_stx.caseid_admin88(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin89':
                caseid_stx.caseid_admin89(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin90':
                caseid_stx.caseid_admin90(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin91':
                caseid_stx.caseid_admin91(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin92':
                caseid_stx.caseid_admin92(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin93':
                caseid_stx.caseid_admin93(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin94':
                caseid_stx.caseid_admin94(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin95':
                caseid_stx.caseid_admin95(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin96':
                caseid_stx.caseid_admin96(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin97':
                caseid_stx.caseid_admin97(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin98':
                caseid_stx.caseid_admin98(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin99':
                caseid_stx.caseid_admin99(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin100':
                caseid_stx.caseid_admin100(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin101':
                caseid_stx.caseid_admin101(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin102':
                caseid_stx.caseid_admin102(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin103':
                caseid_stx.caseid_admin103(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin104':
                caseid_stx.caseid_admin104(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin105':
                caseid_stx.caseid_admin105(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin106':
                caseid_stx.caseid_admin106(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin107':
                caseid_stx.caseid_admin107(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin108':
                caseid_stx.caseid_admin108(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin109':
                caseid_stx.caseid_admin109(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin110':
                caseid_stx.caseid_admin110(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin111':
                caseid_stx.caseid_admin111(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin112':
                caseid_stx.caseid_admin112(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin113':
                caseid_stx.caseid_admin113(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin114':
                caseid_stx.caseid_admin114(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin115':
                caseid_stx.caseid_admin115(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin116':
                caseid_stx.caseid_admin116(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin117':
                caseid_stx.caseid_admin117(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin118':
                caseid_stx.caseid_admin118(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin119':
                caseid_stx.caseid_admin119(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin120':
                caseid_stx.caseid_admin120(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin121':
                caseid_stx.caseid_admin121(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin122':
                caseid_stx.caseid_admin122(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin123':
                caseid_stx.caseid_admin123(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin124':
                caseid_stx.caseid_admin124(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin125':
                caseid_stx.caseid_admin125(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin126':
                caseid_stx.caseid_admin126(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin127':
                caseid_stx.caseid_admin127(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin128':
                caseid_stx.caseid_admin128(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Admin129':
                caseid_stx.caseid_admin129(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip01':
                caseid_stx.caseid_PartnerTrip01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip02':
                caseid_stx.caseid_PartnerTrip02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip03':
                caseid_stx.caseid_PartnerTrip03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip04':
                caseid_stx.caseid_PartnerTrip04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip05':
                caseid_stx.caseid_PartnerTrip05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'PartnerTrip06':
                caseid_stx.caseid_PartnerTrip06(self)
        except:
            module_other_stx.swich_tab_0()





        try:
            if case == 'Accounting01':
                caseid_stx.caseid_accounting01(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting02':
                caseid_stx.caseid_accounting02(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting03':
                caseid_stx.caseid_accounting03(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting04':
                caseid_stx.caseid_accounting04(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting05':
                caseid_stx.caseid_accounting05(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting06':
                caseid_stx.caseid_accounting06(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting07':
                caseid_stx.caseid_accounting07(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting08':
                caseid_stx.caseid_accounting08(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting09':
                caseid_stx.caseid_accounting09(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting10':
                caseid_stx.caseid_accounting10(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting11':
                caseid_stx.caseid_accounting11(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting12':
                caseid_stx.caseid_accounting12(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting13':
                caseid_stx.caseid_accounting13(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting14':
                caseid_stx.caseid_accounting14(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting15':
                caseid_stx.caseid_accounting15(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting16':
                caseid_stx.caseid_accounting16(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting17':
                caseid_stx.caseid_accounting17(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting18':
                caseid_stx.caseid_accounting18(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting19':
                caseid_stx.caseid_accounting19(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting20':
                caseid_stx.caseid_accounting20(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting21':
                caseid_stx.caseid_accounting21(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting22':
                caseid_stx.caseid_accounting22(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting23':
                caseid_stx.caseid_accounting23(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting24':
                caseid_stx.caseid_accounting24(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting25':
                caseid_stx.caseid_accounting25(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting26':
                caseid_stx.caseid_accounting26(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting27':
                caseid_stx.caseid_accounting27(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting28':
                caseid_stx.caseid_accounting28(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting29':
                caseid_stx.caseid_accounting29(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting30':
                caseid_stx.caseid_accounting30(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting31':
                caseid_stx.caseid_accounting31(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting32':
                caseid_stx.caseid_accounting32(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting33':
                caseid_stx.caseid_accounting33(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting34':
                caseid_stx.caseid_accounting34(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting35':
                caseid_stx.caseid_accounting35(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting36':
                caseid_stx.caseid_accounting36(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting37':
                caseid_stx.caseid_accounting37(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting38':
                caseid_stx.caseid_accounting38(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting39':
                caseid_stx.caseid_accounting39(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting40':
                caseid_stx.caseid_accounting40(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting41':
                caseid_stx.caseid_accounting41(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting42':
                caseid_stx.caseid_accounting42(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting43':
                caseid_stx.caseid_accounting43(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting44':
                caseid_stx.caseid_accounting44(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting45':
                caseid_stx.caseid_accounting45(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting46':
                caseid_stx.caseid_accounting46(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting47':
                caseid_stx.caseid_accounting47(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting48':
                caseid_stx.caseid_accounting48(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting49':
                caseid_stx.caseid_accounting49(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting50':
                caseid_stx.caseid_accounting50(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting51':
                caseid_stx.caseid_accounting51(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting52':
                caseid_stx.caseid_accounting52(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting53':
                caseid_stx.caseid_accounting53(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting54':
                caseid_stx.caseid_accounting54(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting55':
                caseid_stx.caseid_accounting55(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting56':
                caseid_stx.caseid_accounting56(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting57':
                caseid_stx.caseid_accounting57(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting58':
                caseid_stx.caseid_accounting58(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting59':
                caseid_stx.caseid_accounting59(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting60':
                caseid_stx.caseid_accounting60(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting61':
                caseid_stx.caseid_accounting61(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting62':
                caseid_stx.caseid_accounting62(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting63':
                caseid_stx.caseid_accounting63(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting64':
                caseid_stx.caseid_accounting64(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting65':
                caseid_stx.caseid_accounting65(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting66':
                caseid_stx.caseid_accounting66(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting67':
                caseid_stx.caseid_accounting67(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting68':
                caseid_stx.caseid_accounting68(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting69':
                caseid_stx.caseid_accounting69(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting70':
                caseid_stx.caseid_accounting70(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting71':
                caseid_stx.caseid_accounting71(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting72':
                caseid_stx.caseid_accounting72(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting73':
                caseid_stx.caseid_accounting73(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting74':
                caseid_stx.caseid_accounting74(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting75':
                caseid_stx.caseid_accounting75(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting76':
                caseid_stx.caseid_accounting76(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting77':
                caseid_stx.caseid_accounting77(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting78':
                caseid_stx.caseid_accounting78(self)
        except:
            module_other_stx.swich_tab_0()
        try:
            if case == 'Accounting79':
                caseid_stx.caseid_accounting79(self)
        except:
            module_other_stx.swich_tab_0()


def login(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 10
    while (rownum < 30):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Login01':
                        caseid_stx.caseid_login01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_stx.caseid_login02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_stx.caseid_login03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_stx.caseid_login04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_stx.caseid_login05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_stx.caseid_login06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_stx.caseid_login07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_stx.caseid_login08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_stx.caseid_login09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_stx.caseid_login10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_stx.caseid_login11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_stx.caseid_login12(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Login01':
                        caseid_stx.caseid_login01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_stx.caseid_login02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_stx.caseid_login03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_stx.caseid_login04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_stx.caseid_login05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_stx.caseid_login06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_stx.caseid_login07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_stx.caseid_login08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_stx.caseid_login09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_stx.caseid_login10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_stx.caseid_login11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_stx.caseid_login12(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Login01':
                        caseid_stx.caseid_login01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_stx.caseid_login02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_stx.caseid_login03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_stx.caseid_login04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_stx.caseid_login05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_stx.caseid_login06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_stx.caseid_login07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_stx.caseid_login08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_stx.caseid_login09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_stx.caseid_login10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_stx.caseid_login11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_stx.caseid_login12(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Login01':
                        caseid_stx.caseid_login01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_stx.caseid_login02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_stx.caseid_login03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_stx.caseid_login04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_stx.caseid_login05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_stx.caseid_login06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_stx.caseid_login07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_stx.caseid_login08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_stx.caseid_login09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_stx.caseid_login10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_stx.caseid_login11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_stx.caseid_login12(self)
                except:
                    module_other_stx.swich_tab_0()



def minitor(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 27
    while (rownum < 130):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Minitor01':
                        caseid_stx.caseid_minitor01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor02':
                        caseid_stx.caseid_minitor02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor03':
                        caseid_stx.caseid_minitor03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor04':
                        caseid_stx.caseid_minitor04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor05':
                        caseid_stx.caseid_minitor05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor06':
                        caseid_stx.caseid_minitor06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor07':
                        caseid_stx.caseid_minitor07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor08':
                        caseid_stx.caseid_minitor08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor09':
                        caseid_stx.caseid_minitor09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor10':
                        caseid_stx.caseid_minitor10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor11':
                        caseid_stx.caseid_minitor11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor12':
                        caseid_stx.caseid_minitor12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor13':
                        caseid_stx.caseid_minitor13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor14':
                        caseid_stx.caseid_minitor14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor15':
                        caseid_stx.caseid_minitor15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor16':
                        caseid_stx.caseid_minitor16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor17':
                        caseid_stx.caseid_minitor17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor18':
                        caseid_stx.caseid_minitor18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor19':
                        caseid_stx.caseid_minitor19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor20':
                        caseid_stx.caseid_minitor20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor21':
                        caseid_stx.caseid_minitor21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor22':
                        caseid_stx.caseid_minitor22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor23':
                        caseid_stx.caseid_minitor23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor24':
                        caseid_stx.caseid_minitor24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor25':
                        caseid_stx.caseid_minitor25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor26':
                        caseid_stx.caseid_minitor26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor27':
                        caseid_stx.caseid_minitor27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor28':
                        caseid_stx.caseid_minitor28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor29':
                        caseid_stx.caseid_minitor29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor30':
                        caseid_stx.caseid_minitor30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor31':
                        caseid_stx.caseid_minitor31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor32':
                        caseid_stx.caseid_minitor32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor33':
                        caseid_stx.caseid_minitor33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor34':
                        caseid_stx.caseid_minitor34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor35':
                        caseid_stx.caseid_minitor35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor36':
                        caseid_stx.caseid_minitor36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor37':
                        caseid_stx.caseid_minitor37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor38':
                        caseid_stx.caseid_minitor38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor39':
                        caseid_stx.caseid_minitor39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor40':
                        caseid_stx.caseid_minitor40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor41':
                        caseid_stx.caseid_minitor41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor42':
                        caseid_stx.caseid_minitor42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor43':
                        caseid_stx.caseid_minitor43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor44':
                        caseid_stx.caseid_minitor44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor45':
                        caseid_stx.caseid_minitor45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor46':
                        caseid_stx.caseid_minitor46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor47':
                        caseid_stx.caseid_minitor47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor48':
                        caseid_stx.caseid_minitor48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor49':
                        caseid_stx.caseid_minitor49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor50':
                        caseid_stx.caseid_minitor50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor51':
                        caseid_stx.caseid_minitor51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor52':
                        caseid_stx.caseid_minitor52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor53':
                        caseid_stx.caseid_minitor53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor54':
                        caseid_stx.caseid_minitor54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor55':
                        caseid_stx.caseid_minitor55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor56':
                        caseid_stx.caseid_minitor56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor57':
                        caseid_stx.caseid_minitor57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor58':
                        caseid_stx.caseid_minitor58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor59':
                        caseid_stx.caseid_minitor59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor60':
                        caseid_stx.caseid_minitor60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor61':
                        caseid_stx.caseid_minitor61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor62':
                        caseid_stx.caseid_minitor62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor63':
                        caseid_stx.caseid_minitor63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor64':
                        caseid_stx.caseid_minitor64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor65':
                        caseid_stx.caseid_minitor65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor66':
                        caseid_stx.caseid_minitor66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor67':
                        caseid_stx.caseid_minitor67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor68':
                        caseid_stx.caseid_minitor68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor69':
                        caseid_stx.caseid_minitor69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor70':
                        caseid_stx.caseid_minitor70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor71':
                        caseid_stx.caseid_minitor71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor72':
                        caseid_stx.caseid_minitor72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor73':
                        caseid_stx.caseid_minitor73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor74':
                        caseid_stx.caseid_minitor74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor75':
                        caseid_stx.caseid_minitor75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor76':
                        caseid_stx.caseid_minitor76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor77':
                        caseid_stx.caseid_minitor77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor78':
                        caseid_stx.caseid_minitor78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor79':
                        caseid_stx.caseid_minitor79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor80':
                        caseid_stx.caseid_minitor80(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Minitor01':
                        caseid_stx.caseid_minitor01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor02':
                        caseid_stx.caseid_minitor02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor03':
                        caseid_stx.caseid_minitor03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor04':
                        caseid_stx.caseid_minitor04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor05':
                        caseid_stx.caseid_minitor05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor06':
                        caseid_stx.caseid_minitor06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor07':
                        caseid_stx.caseid_minitor07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor08':
                        caseid_stx.caseid_minitor08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor09':
                        caseid_stx.caseid_minitor09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor10':
                        caseid_stx.caseid_minitor10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor11':
                        caseid_stx.caseid_minitor11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor12':
                        caseid_stx.caseid_minitor12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor13':
                        caseid_stx.caseid_minitor13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor14':
                        caseid_stx.caseid_minitor14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor15':
                        caseid_stx.caseid_minitor15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor16':
                        caseid_stx.caseid_minitor16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor17':
                        caseid_stx.caseid_minitor17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor18':
                        caseid_stx.caseid_minitor18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor19':
                        caseid_stx.caseid_minitor19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor20':
                        caseid_stx.caseid_minitor20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor21':
                        caseid_stx.caseid_minitor21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor22':
                        caseid_stx.caseid_minitor22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor23':
                        caseid_stx.caseid_minitor23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor24':
                        caseid_stx.caseid_minitor24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor25':
                        caseid_stx.caseid_minitor25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor26':
                        caseid_stx.caseid_minitor26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor27':
                        caseid_stx.caseid_minitor27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor28':
                        caseid_stx.caseid_minitor28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor29':
                        caseid_stx.caseid_minitor29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor30':
                        caseid_stx.caseid_minitor30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor31':
                        caseid_stx.caseid_minitor31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor32':
                        caseid_stx.caseid_minitor32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor33':
                        caseid_stx.caseid_minitor33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor34':
                        caseid_stx.caseid_minitor34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor35':
                        caseid_stx.caseid_minitor35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor36':
                        caseid_stx.caseid_minitor36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor37':
                        caseid_stx.caseid_minitor37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor38':
                        caseid_stx.caseid_minitor38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor39':
                        caseid_stx.caseid_minitor39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor40':
                        caseid_stx.caseid_minitor40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor41':
                        caseid_stx.caseid_minitor41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor42':
                        caseid_stx.caseid_minitor42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor43':
                        caseid_stx.caseid_minitor43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor44':
                        caseid_stx.caseid_minitor44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor45':
                        caseid_stx.caseid_minitor45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor46':
                        caseid_stx.caseid_minitor46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor47':
                        caseid_stx.caseid_minitor47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor48':
                        caseid_stx.caseid_minitor48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor49':
                        caseid_stx.caseid_minitor49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor50':
                        caseid_stx.caseid_minitor50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor51':
                        caseid_stx.caseid_minitor51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor52':
                        caseid_stx.caseid_minitor52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor53':
                        caseid_stx.caseid_minitor53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor54':
                        caseid_stx.caseid_minitor54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor55':
                        caseid_stx.caseid_minitor55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor56':
                        caseid_stx.caseid_minitor56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor57':
                        caseid_stx.caseid_minitor57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor58':
                        caseid_stx.caseid_minitor58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor59':
                        caseid_stx.caseid_minitor59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor60':
                        caseid_stx.caseid_minitor60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor61':
                        caseid_stx.caseid_minitor61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor62':
                        caseid_stx.caseid_minitor62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor63':
                        caseid_stx.caseid_minitor63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor64':
                        caseid_stx.caseid_minitor64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor65':
                        caseid_stx.caseid_minitor65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor66':
                        caseid_stx.caseid_minitor66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor67':
                        caseid_stx.caseid_minitor67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor68':
                        caseid_stx.caseid_minitor68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor69':
                        caseid_stx.caseid_minitor69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor70':
                        caseid_stx.caseid_minitor70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor71':
                        caseid_stx.caseid_minitor71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor72':
                        caseid_stx.caseid_minitor72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor73':
                        caseid_stx.caseid_minitor73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor74':
                        caseid_stx.caseid_minitor74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor75':
                        caseid_stx.caseid_minitor75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor76':
                        caseid_stx.caseid_minitor76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor77':
                        caseid_stx.caseid_minitor77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor78':
                        caseid_stx.caseid_minitor78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor79':
                        caseid_stx.caseid_minitor79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor80':
                        caseid_stx.caseid_minitor80(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Minitor01':
                        caseid_stx.caseid_minitor01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor02':
                        caseid_stx.caseid_minitor02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor03':
                        caseid_stx.caseid_minitor03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor04':
                        caseid_stx.caseid_minitor04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor05':
                        caseid_stx.caseid_minitor05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor06':
                        caseid_stx.caseid_minitor06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor07':
                        caseid_stx.caseid_minitor07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor08':
                        caseid_stx.caseid_minitor08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor09':
                        caseid_stx.caseid_minitor09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor10':
                        caseid_stx.caseid_minitor10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor11':
                        caseid_stx.caseid_minitor11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor12':
                        caseid_stx.caseid_minitor12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor13':
                        caseid_stx.caseid_minitor13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor14':
                        caseid_stx.caseid_minitor14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor15':
                        caseid_stx.caseid_minitor15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor16':
                        caseid_stx.caseid_minitor16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor17':
                        caseid_stx.caseid_minitor17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor18':
                        caseid_stx.caseid_minitor18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor19':
                        caseid_stx.caseid_minitor19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor20':
                        caseid_stx.caseid_minitor20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor21':
                        caseid_stx.caseid_minitor21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor22':
                        caseid_stx.caseid_minitor22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor23':
                        caseid_stx.caseid_minitor23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor24':
                        caseid_stx.caseid_minitor24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor25':
                        caseid_stx.caseid_minitor25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor26':
                        caseid_stx.caseid_minitor26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor27':
                        caseid_stx.caseid_minitor27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor28':
                        caseid_stx.caseid_minitor28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor29':
                        caseid_stx.caseid_minitor29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor30':
                        caseid_stx.caseid_minitor30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor31':
                        caseid_stx.caseid_minitor31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor32':
                        caseid_stx.caseid_minitor32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor33':
                        caseid_stx.caseid_minitor33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor34':
                        caseid_stx.caseid_minitor34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor35':
                        caseid_stx.caseid_minitor35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor36':
                        caseid_stx.caseid_minitor36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor37':
                        caseid_stx.caseid_minitor37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor38':
                        caseid_stx.caseid_minitor38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor39':
                        caseid_stx.caseid_minitor39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor40':
                        caseid_stx.caseid_minitor40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor41':
                        caseid_stx.caseid_minitor41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor42':
                        caseid_stx.caseid_minitor42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor43':
                        caseid_stx.caseid_minitor43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor44':
                        caseid_stx.caseid_minitor44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor45':
                        caseid_stx.caseid_minitor45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor46':
                        caseid_stx.caseid_minitor46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor47':
                        caseid_stx.caseid_minitor47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor48':
                        caseid_stx.caseid_minitor48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor49':
                        caseid_stx.caseid_minitor49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor50':
                        caseid_stx.caseid_minitor50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor51':
                        caseid_stx.caseid_minitor51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor52':
                        caseid_stx.caseid_minitor52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor53':
                        caseid_stx.caseid_minitor53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor54':
                        caseid_stx.caseid_minitor54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor55':
                        caseid_stx.caseid_minitor55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor56':
                        caseid_stx.caseid_minitor56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor57':
                        caseid_stx.caseid_minitor57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor58':
                        caseid_stx.caseid_minitor58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor59':
                        caseid_stx.caseid_minitor59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor60':
                        caseid_stx.caseid_minitor60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor61':
                        caseid_stx.caseid_minitor61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor62':
                        caseid_stx.caseid_minitor62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor63':
                        caseid_stx.caseid_minitor63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor64':
                        caseid_stx.caseid_minitor64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor65':
                        caseid_stx.caseid_minitor65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor66':
                        caseid_stx.caseid_minitor66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor67':
                        caseid_stx.caseid_minitor67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor68':
                        caseid_stx.caseid_minitor68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor69':
                        caseid_stx.caseid_minitor69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor70':
                        caseid_stx.caseid_minitor70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor71':
                        caseid_stx.caseid_minitor71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor72':
                        caseid_stx.caseid_minitor72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor73':
                        caseid_stx.caseid_minitor73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor74':
                        caseid_stx.caseid_minitor74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor75':
                        caseid_stx.caseid_minitor75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor76':
                        caseid_stx.caseid_minitor76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor77':
                        caseid_stx.caseid_minitor77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor78':
                        caseid_stx.caseid_minitor78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor79':
                        caseid_stx.caseid_minitor79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor80':
                        caseid_stx.caseid_minitor80(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Minitor01':
                        caseid_stx.caseid_minitor01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor02':
                        caseid_stx.caseid_minitor02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor03':
                        caseid_stx.caseid_minitor03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor04':
                        caseid_stx.caseid_minitor04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor05':
                        caseid_stx.caseid_minitor05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor06':
                        caseid_stx.caseid_minitor06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor07':
                        caseid_stx.caseid_minitor07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor08':
                        caseid_stx.caseid_minitor08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor09':
                        caseid_stx.caseid_minitor09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor10':
                        caseid_stx.caseid_minitor10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor11':
                        caseid_stx.caseid_minitor11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor12':
                        caseid_stx.caseid_minitor12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor13':
                        caseid_stx.caseid_minitor13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor14':
                        caseid_stx.caseid_minitor14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor15':
                        caseid_stx.caseid_minitor15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor16':
                        caseid_stx.caseid_minitor16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor17':
                        caseid_stx.caseid_minitor17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor18':
                        caseid_stx.caseid_minitor18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor19':
                        caseid_stx.caseid_minitor19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor20':
                        caseid_stx.caseid_minitor20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor21':
                        caseid_stx.caseid_minitor21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor22':
                        caseid_stx.caseid_minitor22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor23':
                        caseid_stx.caseid_minitor23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor24':
                        caseid_stx.caseid_minitor24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor25':
                        caseid_stx.caseid_minitor25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor26':
                        caseid_stx.caseid_minitor26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor27':
                        caseid_stx.caseid_minitor27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor28':
                        caseid_stx.caseid_minitor28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor29':
                        caseid_stx.caseid_minitor29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor30':
                        caseid_stx.caseid_minitor30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor31':
                        caseid_stx.caseid_minitor31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor32':
                        caseid_stx.caseid_minitor32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor33':
                        caseid_stx.caseid_minitor33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor34':
                        caseid_stx.caseid_minitor34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor35':
                        caseid_stx.caseid_minitor35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor36':
                        caseid_stx.caseid_minitor36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor37':
                        caseid_stx.caseid_minitor37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor38':
                        caseid_stx.caseid_minitor38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor39':
                        caseid_stx.caseid_minitor39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor40':
                        caseid_stx.caseid_minitor40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor41':
                        caseid_stx.caseid_minitor41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor42':
                        caseid_stx.caseid_minitor42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor43':
                        caseid_stx.caseid_minitor43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor44':
                        caseid_stx.caseid_minitor44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor45':
                        caseid_stx.caseid_minitor45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor46':
                        caseid_stx.caseid_minitor46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor47':
                        caseid_stx.caseid_minitor47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor48':
                        caseid_stx.caseid_minitor48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor49':
                        caseid_stx.caseid_minitor49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor50':
                        caseid_stx.caseid_minitor50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor51':
                        caseid_stx.caseid_minitor51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor52':
                        caseid_stx.caseid_minitor52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor53':
                        caseid_stx.caseid_minitor53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor54':
                        caseid_stx.caseid_minitor54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor55':
                        caseid_stx.caseid_minitor55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor56':
                        caseid_stx.caseid_minitor56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor57':
                        caseid_stx.caseid_minitor57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor58':
                        caseid_stx.caseid_minitor58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor59':
                        caseid_stx.caseid_minitor59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor60':
                        caseid_stx.caseid_minitor60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor61':
                        caseid_stx.caseid_minitor61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor62':
                        caseid_stx.caseid_minitor62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor63':
                        caseid_stx.caseid_minitor63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor64':
                        caseid_stx.caseid_minitor64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor65':
                        caseid_stx.caseid_minitor65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor66':
                        caseid_stx.caseid_minitor66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor67':
                        caseid_stx.caseid_minitor67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor68':
                        caseid_stx.caseid_minitor68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor69':
                        caseid_stx.caseid_minitor69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor70':
                        caseid_stx.caseid_minitor70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor71':
                        caseid_stx.caseid_minitor71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor72':
                        caseid_stx.caseid_minitor72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor73':
                        caseid_stx.caseid_minitor73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor74':
                        caseid_stx.caseid_minitor74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor75':
                        caseid_stx.caseid_minitor75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor76':
                        caseid_stx.caseid_minitor76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor77':
                        caseid_stx.caseid_minitor77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor78':
                        caseid_stx.caseid_minitor78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor79':
                        caseid_stx.caseid_minitor79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Minitor80':
                        caseid_stx.caseid_minitor80(self)
                except:
                    module_other_stx.swich_tab_0()



def vehicle(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 125
    while (rownum < 200):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Vehicle01':
                        caseid_stx.caseid_vehicle01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle02':
                        caseid_stx.caseid_vehicle02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle03':
                        caseid_stx.caseid_vehicle03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle04':
                        caseid_stx.caseid_vehicle04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle05':
                        caseid_stx.caseid_vehicle05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle06':
                        caseid_stx.caseid_vehicle06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle07':
                        caseid_stx.caseid_vehicle07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle08':
                        caseid_stx.caseid_vehicle08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle09':
                        caseid_stx.caseid_vehicle09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle10':
                        caseid_stx.caseid_vehicle10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle11':
                        caseid_stx.caseid_vehicle11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle12':
                        caseid_stx.caseid_vehicle12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle13':
                        caseid_stx.caseid_vehicle13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle14':
                        caseid_stx.caseid_vehicle14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle15':
                        caseid_stx.caseid_vehicle15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle16':
                        caseid_stx.caseid_vehicle16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle17':
                        caseid_stx.caseid_vehicle17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle18':
                        caseid_stx.caseid_vehicle18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle19':
                        caseid_stx.caseid_vehicle19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle20':
                        caseid_stx.caseid_vehicle20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle21':
                        caseid_stx.caseid_vehicle21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle22':
                        caseid_stx.caseid_vehicle22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle23':
                        caseid_stx.caseid_vehicle23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle24':
                        caseid_stx.caseid_vehicle24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle25':
                        caseid_stx.caseid_vehicle25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle26':
                        caseid_stx.caseid_vehicle26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle27':
                        caseid_stx.caseid_vehicle27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle28':
                        caseid_stx.caseid_vehicle28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle29':
                        caseid_stx.caseid_vehicle29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle30':
                        caseid_stx.caseid_vehicle30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle31':
                        caseid_stx.caseid_vehicle31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle32':
                        caseid_stx.caseid_vehicle32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle33':
                        caseid_stx.caseid_vehicle33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle34':
                        caseid_stx.caseid_vehicle34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle35':
                        caseid_stx.caseid_vehicle35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle36':
                        caseid_stx.caseid_vehicle36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle37':
                        caseid_stx.caseid_vehicle37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle38':
                        caseid_stx.caseid_vehicle38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle39':
                        caseid_stx.caseid_vehicle39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle40':
                        caseid_stx.caseid_vehicle40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle41':
                        caseid_stx.caseid_vehicle41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle42':
                        caseid_stx.caseid_vehicle42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle43':
                        caseid_stx.caseid_vehicle43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle44':
                        caseid_stx.caseid_vehicle44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle45':
                        caseid_stx.caseid_vehicle45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle46':
                        caseid_stx.caseid_vehicle46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle47':
                        caseid_stx.caseid_vehicle47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle48':
                        caseid_stx.caseid_vehicle48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle49':
                        caseid_stx.caseid_vehicle49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle50':
                        caseid_stx.caseid_vehicle50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle51':
                        caseid_stx.caseid_vehicle51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle52':
                        caseid_stx.caseid_vehicle52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle53':
                        caseid_stx.caseid_vehicle53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle54':
                        caseid_stx.caseid_vehicle54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle55':
                        caseid_stx.caseid_vehicle55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle56':
                        caseid_stx.caseid_vehicle56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle57':
                        caseid_stx.caseid_vehicle57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle58':
                        caseid_stx.caseid_vehicle58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle59':
                        caseid_stx.caseid_vehicle59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle60':
                        caseid_stx.caseid_vehicle60(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Vehicle01':
                        caseid_stx.caseid_vehicle01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle02':
                        caseid_stx.caseid_vehicle02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle03':
                        caseid_stx.caseid_vehicle03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle04':
                        caseid_stx.caseid_vehicle04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle05':
                        caseid_stx.caseid_vehicle05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle06':
                        caseid_stx.caseid_vehicle06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle07':
                        caseid_stx.caseid_vehicle07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle08':
                        caseid_stx.caseid_vehicle08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle09':
                        caseid_stx.caseid_vehicle09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle10':
                        caseid_stx.caseid_vehicle10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle11':
                        caseid_stx.caseid_vehicle11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle12':
                        caseid_stx.caseid_vehicle12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle13':
                        caseid_stx.caseid_vehicle13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle14':
                        caseid_stx.caseid_vehicle14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle15':
                        caseid_stx.caseid_vehicle15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle16':
                        caseid_stx.caseid_vehicle16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle17':
                        caseid_stx.caseid_vehicle17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle18':
                        caseid_stx.caseid_vehicle18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle19':
                        caseid_stx.caseid_vehicle19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle20':
                        caseid_stx.caseid_vehicle20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle21':
                        caseid_stx.caseid_vehicle21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle22':
                        caseid_stx.caseid_vehicle22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle23':
                        caseid_stx.caseid_vehicle23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle24':
                        caseid_stx.caseid_vehicle24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle25':
                        caseid_stx.caseid_vehicle25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle26':
                        caseid_stx.caseid_vehicle26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle27':
                        caseid_stx.caseid_vehicle27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle28':
                        caseid_stx.caseid_vehicle28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle29':
                        caseid_stx.caseid_vehicle29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle30':
                        caseid_stx.caseid_vehicle30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle31':
                        caseid_stx.caseid_vehicle31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle32':
                        caseid_stx.caseid_vehicle32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle33':
                        caseid_stx.caseid_vehicle33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle34':
                        caseid_stx.caseid_vehicle34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle35':
                        caseid_stx.caseid_vehicle35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle36':
                        caseid_stx.caseid_vehicle36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle37':
                        caseid_stx.caseid_vehicle37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle38':
                        caseid_stx.caseid_vehicle38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle39':
                        caseid_stx.caseid_vehicle39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle40':
                        caseid_stx.caseid_vehicle40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle41':
                        caseid_stx.caseid_vehicle41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle42':
                        caseid_stx.caseid_vehicle42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle43':
                        caseid_stx.caseid_vehicle43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle44':
                        caseid_stx.caseid_vehicle44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle45':
                        caseid_stx.caseid_vehicle45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle46':
                        caseid_stx.caseid_vehicle46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle47':
                        caseid_stx.caseid_vehicle47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle48':
                        caseid_stx.caseid_vehicle48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle49':
                        caseid_stx.caseid_vehicle49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle50':
                        caseid_stx.caseid_vehicle50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle51':
                        caseid_stx.caseid_vehicle51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle52':
                        caseid_stx.caseid_vehicle52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle53':
                        caseid_stx.caseid_vehicle53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle54':
                        caseid_stx.caseid_vehicle54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle55':
                        caseid_stx.caseid_vehicle55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle56':
                        caseid_stx.caseid_vehicle56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle57':
                        caseid_stx.caseid_vehicle57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle58':
                        caseid_stx.caseid_vehicle58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle59':
                        caseid_stx.caseid_vehicle59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle60':
                        caseid_stx.caseid_vehicle60(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Vehicle01':
                        caseid_stx.caseid_vehicle01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle02':
                        caseid_stx.caseid_vehicle02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle03':
                        caseid_stx.caseid_vehicle03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle04':
                        caseid_stx.caseid_vehicle04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle05':
                        caseid_stx.caseid_vehicle05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle06':
                        caseid_stx.caseid_vehicle06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle07':
                        caseid_stx.caseid_vehicle07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle08':
                        caseid_stx.caseid_vehicle08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle09':
                        caseid_stx.caseid_vehicle09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle10':
                        caseid_stx.caseid_vehicle10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle11':
                        caseid_stx.caseid_vehicle11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle12':
                        caseid_stx.caseid_vehicle12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle13':
                        caseid_stx.caseid_vehicle13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle14':
                        caseid_stx.caseid_vehicle14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle15':
                        caseid_stx.caseid_vehicle15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle16':
                        caseid_stx.caseid_vehicle16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle17':
                        caseid_stx.caseid_vehicle17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle18':
                        caseid_stx.caseid_vehicle18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle19':
                        caseid_stx.caseid_vehicle19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle20':
                        caseid_stx.caseid_vehicle20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle21':
                        caseid_stx.caseid_vehicle21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle22':
                        caseid_stx.caseid_vehicle22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle23':
                        caseid_stx.caseid_vehicle23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle24':
                        caseid_stx.caseid_vehicle24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle25':
                        caseid_stx.caseid_vehicle25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle26':
                        caseid_stx.caseid_vehicle26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle27':
                        caseid_stx.caseid_vehicle27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle28':
                        caseid_stx.caseid_vehicle28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle29':
                        caseid_stx.caseid_vehicle29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle30':
                        caseid_stx.caseid_vehicle30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle31':
                        caseid_stx.caseid_vehicle31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle32':
                        caseid_stx.caseid_vehicle32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle33':
                        caseid_stx.caseid_vehicle33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle34':
                        caseid_stx.caseid_vehicle34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle35':
                        caseid_stx.caseid_vehicle35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle36':
                        caseid_stx.caseid_vehicle36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle37':
                        caseid_stx.caseid_vehicle37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle38':
                        caseid_stx.caseid_vehicle38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle39':
                        caseid_stx.caseid_vehicle39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle40':
                        caseid_stx.caseid_vehicle40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle41':
                        caseid_stx.caseid_vehicle41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle42':
                        caseid_stx.caseid_vehicle42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle43':
                        caseid_stx.caseid_vehicle43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle44':
                        caseid_stx.caseid_vehicle44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle45':
                        caseid_stx.caseid_vehicle45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle46':
                        caseid_stx.caseid_vehicle46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle47':
                        caseid_stx.caseid_vehicle47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle48':
                        caseid_stx.caseid_vehicle48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle49':
                        caseid_stx.caseid_vehicle49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle50':
                        caseid_stx.caseid_vehicle50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle51':
                        caseid_stx.caseid_vehicle51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle52':
                        caseid_stx.caseid_vehicle52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle53':
                        caseid_stx.caseid_vehicle53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle54':
                        caseid_stx.caseid_vehicle54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle55':
                        caseid_stx.caseid_vehicle55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle56':
                        caseid_stx.caseid_vehicle56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle57':
                        caseid_stx.caseid_vehicle57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle58':
                        caseid_stx.caseid_vehicle58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle59':
                        caseid_stx.caseid_vehicle59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle60':
                        caseid_stx.caseid_vehicle60(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Vehicle01':
                        caseid_stx.caseid_vehicle01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle02':
                        caseid_stx.caseid_vehicle02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle03':
                        caseid_stx.caseid_vehicle03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle04':
                        caseid_stx.caseid_vehicle04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle05':
                        caseid_stx.caseid_vehicle05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle06':
                        caseid_stx.caseid_vehicle06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle07':
                        caseid_stx.caseid_vehicle07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle08':
                        caseid_stx.caseid_vehicle08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle09':
                        caseid_stx.caseid_vehicle09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle10':
                        caseid_stx.caseid_vehicle10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle11':
                        caseid_stx.caseid_vehicle11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle12':
                        caseid_stx.caseid_vehicle12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle13':
                        caseid_stx.caseid_vehicle13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle14':
                        caseid_stx.caseid_vehicle14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle15':
                        caseid_stx.caseid_vehicle15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle16':
                        caseid_stx.caseid_vehicle16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle17':
                        caseid_stx.caseid_vehicle17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle18':
                        caseid_stx.caseid_vehicle18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle19':
                        caseid_stx.caseid_vehicle19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle20':
                        caseid_stx.caseid_vehicle20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle21':
                        caseid_stx.caseid_vehicle21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle22':
                        caseid_stx.caseid_vehicle22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle23':
                        caseid_stx.caseid_vehicle23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle24':
                        caseid_stx.caseid_vehicle24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle25':
                        caseid_stx.caseid_vehicle25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle26':
                        caseid_stx.caseid_vehicle26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle27':
                        caseid_stx.caseid_vehicle27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle28':
                        caseid_stx.caseid_vehicle28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle29':
                        caseid_stx.caseid_vehicle29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle30':
                        caseid_stx.caseid_vehicle30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle31':
                        caseid_stx.caseid_vehicle31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle32':
                        caseid_stx.caseid_vehicle32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle33':
                        caseid_stx.caseid_vehicle33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle34':
                        caseid_stx.caseid_vehicle34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle35':
                        caseid_stx.caseid_vehicle35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle36':
                        caseid_stx.caseid_vehicle36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle37':
                        caseid_stx.caseid_vehicle37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle38':
                        caseid_stx.caseid_vehicle38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle39':
                        caseid_stx.caseid_vehicle39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle40':
                        caseid_stx.caseid_vehicle40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle41':
                        caseid_stx.caseid_vehicle41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle42':
                        caseid_stx.caseid_vehicle42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle43':
                        caseid_stx.caseid_vehicle43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle44':
                        caseid_stx.caseid_vehicle44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle45':
                        caseid_stx.caseid_vehicle45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle46':
                        caseid_stx.caseid_vehicle46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle47':
                        caseid_stx.caseid_vehicle47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle48':
                        caseid_stx.caseid_vehicle48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle49':
                        caseid_stx.caseid_vehicle49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle50':
                        caseid_stx.caseid_vehicle50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle51':
                        caseid_stx.caseid_vehicle51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle52':
                        caseid_stx.caseid_vehicle52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle53':
                        caseid_stx.caseid_vehicle53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle54':
                        caseid_stx.caseid_vehicle54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle55':
                        caseid_stx.caseid_vehicle55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle56':
                        caseid_stx.caseid_vehicle56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle57':
                        caseid_stx.caseid_vehicle57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle58':
                        caseid_stx.caseid_vehicle58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle59':
                        caseid_stx.caseid_vehicle59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Vehicle60':
                        caseid_stx.caseid_vehicle60(self)
                except:
                    module_other_stx.swich_tab_0()



def wallet(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 191
    while (rownum < 250):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Wallet01':
                        caseid_stx.caseid_wallet01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet02':
                        caseid_stx.caseid_wallet02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet03':
                        caseid_stx.caseid_wallet03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet04':
                        caseid_stx.caseid_wallet04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet05':
                        caseid_stx.caseid_wallet05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet06':
                        caseid_stx.caseid_wallet06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet07':
                        caseid_stx.caseid_wallet07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet08':
                        caseid_stx.caseid_wallet08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet09':
                        caseid_stx.caseid_wallet09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet10':
                        caseid_stx.caseid_wallet10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet11':
                        caseid_stx.caseid_wallet11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet12':
                        caseid_stx.caseid_wallet12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet13':
                        caseid_stx.caseid_wallet13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet14':
                        caseid_stx.caseid_wallet14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet15':
                        caseid_stx.caseid_wallet15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet16':
                        caseid_stx.caseid_wallet16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet17':
                        caseid_stx.caseid_wallet17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet18':
                        caseid_stx.caseid_wallet18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet19':
                        caseid_stx.caseid_wallet19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet20':
                        caseid_stx.caseid_wallet20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet21':
                        caseid_stx.caseid_wallet21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet22':
                        caseid_stx.caseid_wallet22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet23':
                        caseid_stx.caseid_wallet23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet24':
                        caseid_stx.caseid_wallet24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet25':
                        caseid_stx.caseid_wallet25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet26':
                        caseid_stx.caseid_wallet26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet27':
                        caseid_stx.caseid_wallet27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet28':
                        caseid_stx.caseid_wallet28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet29':
                        caseid_stx.caseid_wallet29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet30':
                        caseid_stx.caseid_wallet30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet31':
                        caseid_stx.caseid_wallet31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet32':
                        caseid_stx.caseid_wallet32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet33':
                        caseid_stx.caseid_wallet33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet34':
                        caseid_stx.caseid_wallet34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet35':
                        caseid_stx.caseid_wallet35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet36':
                        caseid_stx.caseid_wallet36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet37':
                        caseid_stx.caseid_wallet37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet38':
                        caseid_stx.caseid_wallet38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet39':
                        caseid_stx.caseid_wallet39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet40':
                        caseid_stx.caseid_wallet40(self)
                except:
                    module_other_stx.swich_tab_0()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Wallet01':
                        caseid_stx.caseid_wallet01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet02':
                        caseid_stx.caseid_wallet02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet03':
                        caseid_stx.caseid_wallet03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet04':
                        caseid_stx.caseid_wallet04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet05':
                        caseid_stx.caseid_wallet05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet06':
                        caseid_stx.caseid_wallet06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet07':
                        caseid_stx.caseid_wallet07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet08':
                        caseid_stx.caseid_wallet08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet09':
                        caseid_stx.caseid_wallet09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet10':
                        caseid_stx.caseid_wallet10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet11':
                        caseid_stx.caseid_wallet11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet12':
                        caseid_stx.caseid_wallet12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet13':
                        caseid_stx.caseid_wallet13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet14':
                        caseid_stx.caseid_wallet14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet15':
                        caseid_stx.caseid_wallet15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet16':
                        caseid_stx.caseid_wallet16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet17':
                        caseid_stx.caseid_wallet17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet18':
                        caseid_stx.caseid_wallet18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet19':
                        caseid_stx.caseid_wallet19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet20':
                        caseid_stx.caseid_wallet20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet21':
                        caseid_stx.caseid_wallet21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet22':
                        caseid_stx.caseid_wallet22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet23':
                        caseid_stx.caseid_wallet23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet24':
                        caseid_stx.caseid_wallet24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet25':
                        caseid_stx.caseid_wallet25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet26':
                        caseid_stx.caseid_wallet26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet27':
                        caseid_stx.caseid_wallet27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet28':
                        caseid_stx.caseid_wallet28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet29':
                        caseid_stx.caseid_wallet29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet30':
                        caseid_stx.caseid_wallet30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet31':
                        caseid_stx.caseid_wallet31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet32':
                        caseid_stx.caseid_wallet32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet33':
                        caseid_stx.caseid_wallet33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet34':
                        caseid_stx.caseid_wallet34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet35':
                        caseid_stx.caseid_wallet35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet36':
                        caseid_stx.caseid_wallet36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet37':
                        caseid_stx.caseid_wallet37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet38':
                        caseid_stx.caseid_wallet38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet39':
                        caseid_stx.caseid_wallet39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet40':
                        caseid_stx.caseid_wallet40(self)
                except:
                    module_other_stx.swich_tab_0()




        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Wallet01':
                        caseid_stx.caseid_wallet01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet02':
                        caseid_stx.caseid_wallet02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet03':
                        caseid_stx.caseid_wallet03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet04':
                        caseid_stx.caseid_wallet04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet05':
                        caseid_stx.caseid_wallet05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet06':
                        caseid_stx.caseid_wallet06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet07':
                        caseid_stx.caseid_wallet07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet08':
                        caseid_stx.caseid_wallet08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet09':
                        caseid_stx.caseid_wallet09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet10':
                        caseid_stx.caseid_wallet10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet11':
                        caseid_stx.caseid_wallet11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet12':
                        caseid_stx.caseid_wallet12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet13':
                        caseid_stx.caseid_wallet13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet14':
                        caseid_stx.caseid_wallet14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet15':
                        caseid_stx.caseid_wallet15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet16':
                        caseid_stx.caseid_wallet16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet17':
                        caseid_stx.caseid_wallet17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet18':
                        caseid_stx.caseid_wallet18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet19':
                        caseid_stx.caseid_wallet19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet20':
                        caseid_stx.caseid_wallet20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet21':
                        caseid_stx.caseid_wallet21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet22':
                        caseid_stx.caseid_wallet22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet23':
                        caseid_stx.caseid_wallet23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet24':
                        caseid_stx.caseid_wallet24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet25':
                        caseid_stx.caseid_wallet25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet26':
                        caseid_stx.caseid_wallet26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet27':
                        caseid_stx.caseid_wallet27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet28':
                        caseid_stx.caseid_wallet28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet29':
                        caseid_stx.caseid_wallet29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet30':
                        caseid_stx.caseid_wallet30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet31':
                        caseid_stx.caseid_wallet31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet32':
                        caseid_stx.caseid_wallet32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet33':
                        caseid_stx.caseid_wallet33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet34':
                        caseid_stx.caseid_wallet34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet35':
                        caseid_stx.caseid_wallet35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet36':
                        caseid_stx.caseid_wallet36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet37':
                        caseid_stx.caseid_wallet37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet38':
                        caseid_stx.caseid_wallet38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet39':
                        caseid_stx.caseid_wallet39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet40':
                        caseid_stx.caseid_wallet40(self)
                except:
                    module_other_stx.swich_tab_0()




        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Wallet01':
                        caseid_stx.caseid_wallet01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet02':
                        caseid_stx.caseid_wallet02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet03':
                        caseid_stx.caseid_wallet03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet04':
                        caseid_stx.caseid_wallet04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet05':
                        caseid_stx.caseid_wallet05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet06':
                        caseid_stx.caseid_wallet06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet07':
                        caseid_stx.caseid_wallet07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet08':
                        caseid_stx.caseid_wallet08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet09':
                        caseid_stx.caseid_wallet09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet10':
                        caseid_stx.caseid_wallet10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet11':
                        caseid_stx.caseid_wallet11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet12':
                        caseid_stx.caseid_wallet12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet13':
                        caseid_stx.caseid_wallet13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet14':
                        caseid_stx.caseid_wallet14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet15':
                        caseid_stx.caseid_wallet15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet16':
                        caseid_stx.caseid_wallet16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet17':
                        caseid_stx.caseid_wallet17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet18':
                        caseid_stx.caseid_wallet18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet19':
                        caseid_stx.caseid_wallet19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet20':
                        caseid_stx.caseid_wallet20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet21':
                        caseid_stx.caseid_wallet21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet22':
                        caseid_stx.caseid_wallet22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet23':
                        caseid_stx.caseid_wallet23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet24':
                        caseid_stx.caseid_wallet24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet25':
                        caseid_stx.caseid_wallet25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet26':
                        caseid_stx.caseid_wallet26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet27':
                        caseid_stx.caseid_wallet27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet28':
                        caseid_stx.caseid_wallet28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet29':
                        caseid_stx.caseid_wallet29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet30':
                        caseid_stx.caseid_wallet30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet31':
                        caseid_stx.caseid_wallet31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet32':
                        caseid_stx.caseid_wallet32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet33':
                        caseid_stx.caseid_wallet33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet34':
                        caseid_stx.caseid_wallet34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet35':
                        caseid_stx.caseid_wallet35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet36':
                        caseid_stx.caseid_wallet36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet37':
                        caseid_stx.caseid_wallet37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet38':
                        caseid_stx.caseid_wallet38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet39':
                        caseid_stx.caseid_wallet39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Wallet40':
                        caseid_stx.caseid_wallet40(self)
                except:
                    module_other_stx.swich_tab_0()





def promotion(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 242
    while (rownum < 350):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Promotion01':
                        caseid_stx.caseid_promotion01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion02':
                        caseid_stx.caseid_promotion02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion03':
                        caseid_stx.caseid_promotion03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion04':
                        caseid_stx.caseid_promotion04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05':
                        caseid_stx.caseid_promotion05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05_1':
                        caseid_stx.caseid_promotion05_1(self)
                except:
                    module_other_stx.swich_tab_0()

                try:
                    if case == 'Promotion06':
                        caseid_stx.caseid_promotion06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion07':
                        caseid_stx.caseid_promotion07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion08':
                        caseid_stx.caseid_promotion08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion09':
                        caseid_stx.caseid_promotion09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion10':
                        caseid_stx.caseid_promotion10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion11':
                        caseid_stx.caseid_promotion11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion12':
                        caseid_stx.caseid_promotion12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion13':
                        caseid_stx.caseid_promotion13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion14':
                        caseid_stx.caseid_promotion14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion15':
                        caseid_stx.caseid_promotion15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion16':
                        caseid_stx.caseid_promotion16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion17':
                        caseid_stx.caseid_promotion17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion18':
                        caseid_stx.caseid_promotion18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion19':
                        caseid_stx.caseid_promotion19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion20':
                        caseid_stx.caseid_promotion20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion21':
                        caseid_stx.caseid_promotion21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion22':
                        caseid_stx.caseid_promotion22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion23':
                        caseid_stx.caseid_promotion23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion24':
                        caseid_stx.caseid_promotion24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion25':
                        caseid_stx.caseid_promotion25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion26':
                        caseid_stx.caseid_promotion26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion27':
                        caseid_stx.caseid_promotion27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion28':
                        caseid_stx.caseid_promotion28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion29':
                        caseid_stx.caseid_promotion29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion30':
                        caseid_stx.caseid_promotion30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion31':
                        caseid_stx.caseid_promotion31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion32':
                        caseid_stx.caseid_promotion32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion33':
                        caseid_stx.caseid_promotion33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion34':
                        caseid_stx.caseid_promotion34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion35':
                        caseid_stx.caseid_promotion35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion36':
                        caseid_stx.caseid_promotion36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion37':
                        caseid_stx.caseid_promotion37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion38':
                        caseid_stx.caseid_promotion38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion39':
                        caseid_stx.caseid_promotion39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion40':
                        caseid_stx.caseid_promotion40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion41':
                        caseid_stx.caseid_promotion41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion42':
                        caseid_stx.caseid_promotion42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion43':
                        caseid_stx.caseid_promotion43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion44':
                        caseid_stx.caseid_promotion44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion45':
                        caseid_stx.caseid_promotion45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion46':
                        caseid_stx.caseid_promotion46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion47':
                        caseid_stx.caseid_promotion47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion48':
                        caseid_stx.caseid_promotion48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion49':
                        caseid_stx.caseid_promotion49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion50':
                        caseid_stx.caseid_promotion50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion51':
                        caseid_stx.caseid_promotion51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion52':
                        caseid_stx.caseid_promotion52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion53':
                        caseid_stx.caseid_promotion53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion54':
                        caseid_stx.caseid_promotion54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion55':
                        caseid_stx.caseid_promotion55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion56':
                        caseid_stx.caseid_promotion56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion57':
                        caseid_stx.caseid_promotion57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion58':
                        caseid_stx.caseid_promotion58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion59':
                        caseid_stx.caseid_promotion59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion60':
                        caseid_stx.caseid_promotion60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion61':
                        caseid_stx.caseid_promotion61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion62':
                        caseid_stx.caseid_promotion62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion63':
                        caseid_stx.caseid_promotion63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion64':
                        caseid_stx.caseid_promotion64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion65':
                        caseid_stx.caseid_promotion65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion66':
                        caseid_stx.caseid_promotion66(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Promotion01':
                        caseid_stx.caseid_promotion01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion02':
                        caseid_stx.caseid_promotion02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion03':
                        caseid_stx.caseid_promotion03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion04':
                        caseid_stx.caseid_promotion04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05':
                        caseid_stx.caseid_promotion05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05_1':
                        caseid_stx.caseid_promotion05_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion06':
                        caseid_stx.caseid_promotion06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion07':
                        caseid_stx.caseid_promotion07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion08':
                        caseid_stx.caseid_promotion08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion09':
                        caseid_stx.caseid_promotion09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion10':
                        caseid_stx.caseid_promotion10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion11':
                        caseid_stx.caseid_promotion11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion12':
                        caseid_stx.caseid_promotion12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion13':
                        caseid_stx.caseid_promotion13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion14':
                        caseid_stx.caseid_promotion14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion15':
                        caseid_stx.caseid_promotion15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion16':
                        caseid_stx.caseid_promotion16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion17':
                        caseid_stx.caseid_promotion17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion18':
                        caseid_stx.caseid_promotion18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion19':
                        caseid_stx.caseid_promotion19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion20':
                        caseid_stx.caseid_promotion20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion21':
                        caseid_stx.caseid_promotion21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion22':
                        caseid_stx.caseid_promotion22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion23':
                        caseid_stx.caseid_promotion23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion24':
                        caseid_stx.caseid_promotion24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion25':
                        caseid_stx.caseid_promotion25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion26':
                        caseid_stx.caseid_promotion26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion27':
                        caseid_stx.caseid_promotion27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion28':
                        caseid_stx.caseid_promotion28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion29':
                        caseid_stx.caseid_promotion29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion30':
                        caseid_stx.caseid_promotion30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion31':
                        caseid_stx.caseid_promotion31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion32':
                        caseid_stx.caseid_promotion32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion33':
                        caseid_stx.caseid_promotion33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion34':
                        caseid_stx.caseid_promotion34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion35':
                        caseid_stx.caseid_promotion35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion36':
                        caseid_stx.caseid_promotion36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion37':
                        caseid_stx.caseid_promotion37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion38':
                        caseid_stx.caseid_promotion38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion39':
                        caseid_stx.caseid_promotion39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion40':
                        caseid_stx.caseid_promotion40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion41':
                        caseid_stx.caseid_promotion41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion42':
                        caseid_stx.caseid_promotion42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion43':
                        caseid_stx.caseid_promotion43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion44':
                        caseid_stx.caseid_promotion44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion45':
                        caseid_stx.caseid_promotion45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion46':
                        caseid_stx.caseid_promotion46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion47':
                        caseid_stx.caseid_promotion47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion48':
                        caseid_stx.caseid_promotion48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion49':
                        caseid_stx.caseid_promotion49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion50':
                        caseid_stx.caseid_promotion50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion51':
                        caseid_stx.caseid_promotion51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion52':
                        caseid_stx.caseid_promotion52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion53':
                        caseid_stx.caseid_promotion53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion54':
                        caseid_stx.caseid_promotion54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion55':
                        caseid_stx.caseid_promotion55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion56':
                        caseid_stx.caseid_promotion56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion57':
                        caseid_stx.caseid_promotion57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion58':
                        caseid_stx.caseid_promotion58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion59':
                        caseid_stx.caseid_promotion59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion60':
                        caseid_stx.caseid_promotion60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion61':
                        caseid_stx.caseid_promotion61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion62':
                        caseid_stx.caseid_promotion62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion63':
                        caseid_stx.caseid_promotion63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion64':
                        caseid_stx.caseid_promotion64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion65':
                        caseid_stx.caseid_promotion65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion66':
                        caseid_stx.caseid_promotion66(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Promotion01':
                        caseid_stx.caseid_promotion01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion02':
                        caseid_stx.caseid_promotion02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion03':
                        caseid_stx.caseid_promotion03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion04':
                        caseid_stx.caseid_promotion04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05':
                        caseid_stx.caseid_promotion05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05_1':
                        caseid_stx.caseid_promotion05_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion06':
                        caseid_stx.caseid_promotion06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion07':
                        caseid_stx.caseid_promotion07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion08':
                        caseid_stx.caseid_promotion08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion09':
                        caseid_stx.caseid_promotion09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion10':
                        caseid_stx.caseid_promotion10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion11':
                        caseid_stx.caseid_promotion11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion12':
                        caseid_stx.caseid_promotion12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion13':
                        caseid_stx.caseid_promotion13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion14':
                        caseid_stx.caseid_promotion14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion15':
                        caseid_stx.caseid_promotion15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion16':
                        caseid_stx.caseid_promotion16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion17':
                        caseid_stx.caseid_promotion17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion18':
                        caseid_stx.caseid_promotion18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion19':
                        caseid_stx.caseid_promotion19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion20':
                        caseid_stx.caseid_promotion20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion21':
                        caseid_stx.caseid_promotion21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion22':
                        caseid_stx.caseid_promotion22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion23':
                        caseid_stx.caseid_promotion23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion24':
                        caseid_stx.caseid_promotion24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion25':
                        caseid_stx.caseid_promotion25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion26':
                        caseid_stx.caseid_promotion26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion27':
                        caseid_stx.caseid_promotion27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion28':
                        caseid_stx.caseid_promotion28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion29':
                        caseid_stx.caseid_promotion29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion30':
                        caseid_stx.caseid_promotion30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion31':
                        caseid_stx.caseid_promotion31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion32':
                        caseid_stx.caseid_promotion32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion33':
                        caseid_stx.caseid_promotion33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion34':
                        caseid_stx.caseid_promotion34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion35':
                        caseid_stx.caseid_promotion35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion36':
                        caseid_stx.caseid_promotion36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion37':
                        caseid_stx.caseid_promotion37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion38':
                        caseid_stx.caseid_promotion38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion39':
                        caseid_stx.caseid_promotion39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion40':
                        caseid_stx.caseid_promotion40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion41':
                        caseid_stx.caseid_promotion41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion42':
                        caseid_stx.caseid_promotion42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion43':
                        caseid_stx.caseid_promotion43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion44':
                        caseid_stx.caseid_promotion44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion45':
                        caseid_stx.caseid_promotion45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion46':
                        caseid_stx.caseid_promotion46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion47':
                        caseid_stx.caseid_promotion47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion48':
                        caseid_stx.caseid_promotion48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion49':
                        caseid_stx.caseid_promotion49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion50':
                        caseid_stx.caseid_promotion50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion51':
                        caseid_stx.caseid_promotion51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion52':
                        caseid_stx.caseid_promotion52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion53':
                        caseid_stx.caseid_promotion53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion54':
                        caseid_stx.caseid_promotion54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion55':
                        caseid_stx.caseid_promotion55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion56':
                        caseid_stx.caseid_promotion56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion57':
                        caseid_stx.caseid_promotion57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion58':
                        caseid_stx.caseid_promotion58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion59':
                        caseid_stx.caseid_promotion59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion60':
                        caseid_stx.caseid_promotion60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion61':
                        caseid_stx.caseid_promotion61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion62':
                        caseid_stx.caseid_promotion62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion63':
                        caseid_stx.caseid_promotion63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion64':
                        caseid_stx.caseid_promotion64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion65':
                        caseid_stx.caseid_promotion65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion66':
                        caseid_stx.caseid_promotion66(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Promotion01':
                        caseid_stx.caseid_promotion01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion02':
                        caseid_stx.caseid_promotion02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion03':
                        caseid_stx.caseid_promotion03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion04':
                        caseid_stx.caseid_promotion04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05':
                        caseid_stx.caseid_promotion05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion05_1':
                        caseid_stx.caseid_promotion05_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion06':
                        caseid_stx.caseid_promotion06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion07':
                        caseid_stx.caseid_promotion07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion08':
                        caseid_stx.caseid_promotion08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion09':
                        caseid_stx.caseid_promotion09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion10':
                        caseid_stx.caseid_promotion10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion11':
                        caseid_stx.caseid_promotion11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion12':
                        caseid_stx.caseid_promotion12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion13':
                        caseid_stx.caseid_promotion13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion14':
                        caseid_stx.caseid_promotion14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion15':
                        caseid_stx.caseid_promotion15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion16':
                        caseid_stx.caseid_promotion16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion17':
                        caseid_stx.caseid_promotion17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion18':
                        caseid_stx.caseid_promotion18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion19':
                        caseid_stx.caseid_promotion19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion20':
                        caseid_stx.caseid_promotion20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion21':
                        caseid_stx.caseid_promotion21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion22':
                        caseid_stx.caseid_promotion22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion23':
                        caseid_stx.caseid_promotion23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion24':
                        caseid_stx.caseid_promotion24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion25':
                        caseid_stx.caseid_promotion25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion26':
                        caseid_stx.caseid_promotion26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion27':
                        caseid_stx.caseid_promotion27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion28':
                        caseid_stx.caseid_promotion28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion29':
                        caseid_stx.caseid_promotion29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion30':
                        caseid_stx.caseid_promotion30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion31':
                        caseid_stx.caseid_promotion31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion32':
                        caseid_stx.caseid_promotion32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion33':
                        caseid_stx.caseid_promotion33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion34':
                        caseid_stx.caseid_promotion34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion35':
                        caseid_stx.caseid_promotion35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion36':
                        caseid_stx.caseid_promotion36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion37':
                        caseid_stx.caseid_promotion37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion38':
                        caseid_stx.caseid_promotion38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion39':
                        caseid_stx.caseid_promotion39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion40':
                        caseid_stx.caseid_promotion40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion41':
                        caseid_stx.caseid_promotion41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion42':
                        caseid_stx.caseid_promotion42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion43':
                        caseid_stx.caseid_promotion43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion44':
                        caseid_stx.caseid_promotion44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion45':
                        caseid_stx.caseid_promotion45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion46':
                        caseid_stx.caseid_promotion46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion47':
                        caseid_stx.caseid_promotion47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion48':
                        caseid_stx.caseid_promotion48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion49':
                        caseid_stx.caseid_promotion49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion50':
                        caseid_stx.caseid_promotion50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion51':
                        caseid_stx.caseid_promotion51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion52':
                        caseid_stx.caseid_promotion52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion53':
                        caseid_stx.caseid_promotion53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion54':
                        caseid_stx.caseid_promotion54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion55':
                        caseid_stx.caseid_promotion55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion56':
                        caseid_stx.caseid_promotion56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion57':
                        caseid_stx.caseid_promotion57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion58':
                        caseid_stx.caseid_promotion58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion59':
                        caseid_stx.caseid_promotion59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion60':
                        caseid_stx.caseid_promotion60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion61':
                        caseid_stx.caseid_promotion61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion62':
                        caseid_stx.caseid_promotion62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion63':
                        caseid_stx.caseid_promotion63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion64':
                        caseid_stx.caseid_promotion64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion65':
                        caseid_stx.caseid_promotion65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Promotion66':
                        caseid_stx.caseid_promotion66(self)
                except:
                    module_other_stx.swich_tab_0()



def customer(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 340
    while (rownum < 460):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Customer01':
                        caseid_stx.caseid_customer01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer02':
                        caseid_stx.caseid_customer02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer03':
                        caseid_stx.caseid_customer03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer04':
                        caseid_stx.caseid_customer04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer05':
                        caseid_stx.caseid_customer05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer06':
                        caseid_stx.caseid_customer06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer07':
                        caseid_stx.caseid_customer07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer08':
                        caseid_stx.caseid_customer08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer09':
                        caseid_stx.caseid_customer09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer10':
                        caseid_stx.caseid_customer10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer11':
                        caseid_stx.caseid_customer11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer12':
                        caseid_stx.caseid_customer12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer13':
                        caseid_stx.caseid_customer13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer14':
                        caseid_stx.caseid_customer14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer15':
                        caseid_stx.caseid_customer15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer16':
                        caseid_stx.caseid_customer16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer17':
                        caseid_stx.caseid_customer17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer18':
                        caseid_stx.caseid_customer18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer19':
                        caseid_stx.caseid_customer19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer20':
                        caseid_stx.caseid_customer20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer21':
                        caseid_stx.caseid_customer21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer22':
                        caseid_stx.caseid_customer22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer23':
                        caseid_stx.caseid_customer23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer24':
                        caseid_stx.caseid_customer24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer25':
                        caseid_stx.caseid_customer25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26':
                        caseid_stx.caseid_customer26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_1':
                        caseid_stx.caseid_customer26_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_2':
                        caseid_stx.caseid_customer26_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer27':
                        caseid_stx.caseid_customer27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer28':
                        caseid_stx.caseid_customer28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer29':
                        caseid_stx.caseid_customer29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer30':
                        caseid_stx.caseid_customer30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer31':
                        caseid_stx.caseid_customer31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer32':
                        caseid_stx.caseid_customer32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer33':
                        caseid_stx.caseid_customer33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer34':
                        caseid_stx.caseid_customer34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer35':
                        caseid_stx.caseid_customer35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer36':
                        caseid_stx.caseid_customer36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer37':
                        caseid_stx.caseid_customer37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer38':
                        caseid_stx.caseid_customer38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer39':
                        caseid_stx.caseid_customer39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer40':
                        caseid_stx.caseid_customer40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer41':
                        caseid_stx.caseid_customer41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer42':
                        caseid_stx.caseid_customer42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer43':
                        caseid_stx.caseid_customer43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer44':
                        caseid_stx.caseid_customer44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer45':
                        caseid_stx.caseid_customer45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer46':
                        caseid_stx.caseid_customer46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer47':
                        caseid_stx.caseid_customer47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer48':
                        caseid_stx.caseid_customer48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer49':
                        caseid_stx.caseid_customer49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer50':
                        caseid_stx.caseid_customer50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer51':
                        caseid_stx.caseid_customer51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer52':
                        caseid_stx.caseid_customer52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer53':
                        caseid_stx.caseid_customer53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer54':
                        caseid_stx.caseid_customer54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer55':
                        caseid_stx.caseid_customer55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer56':
                        caseid_stx.caseid_customer56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57':
                        caseid_stx.caseid_customer57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_1':
                        caseid_stx.caseid_customer57_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_2':
                        caseid_stx.caseid_customer57_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_3':
                        caseid_stx.caseid_customer57_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_4':
                        caseid_stx.caseid_customer57_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_5':
                        caseid_stx.caseid_customer57_5(self)
                except:
                    module_other_stx.swich_tab_0()



                try:
                    if case == 'Customer58':
                        caseid_stx.caseid_customer58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer59':
                        caseid_stx.caseid_customer59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer60':
                        caseid_stx.caseid_customer60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer61':
                        caseid_stx.caseid_customer61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer62':
                        caseid_stx.caseid_customer62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer63':
                        caseid_stx.caseid_customer63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer64':
                        caseid_stx.caseid_customer64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer65':
                        caseid_stx.caseid_customer65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer66':
                        caseid_stx.caseid_customer66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer67':
                        caseid_stx.caseid_customer67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer68':
                        caseid_stx.caseid_customer68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer69':
                        caseid_stx.caseid_customer69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer70':
                        caseid_stx.caseid_customer70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer71':
                        caseid_stx.caseid_customer71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer72':
                        caseid_stx.caseid_customer72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer73':
                        caseid_stx.caseid_customer73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer74':
                        caseid_stx.caseid_customer74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer75':
                        caseid_stx.caseid_customer75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer76':
                        caseid_stx.caseid_customer76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77':
                        caseid_stx.caseid_customer77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77_1':
                        caseid_stx.caseid_customer77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer78':
                        caseid_stx.caseid_customer78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer79':
                        caseid_stx.caseid_customer79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer80':
                        caseid_stx.caseid_customer80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer81':
                        caseid_stx.caseid_customer81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer82':
                        caseid_stx.caseid_customer82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer83':
                        caseid_stx.caseid_customer83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer84':
                        caseid_stx.caseid_customer84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer85':
                        caseid_stx.caseid_customer85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer86':
                        caseid_stx.caseid_customer86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer87':
                        caseid_stx.caseid_customer87(self)
                except:
                    module_other_stx.swich_tab_0()

                try:
                    if case == 'Customer88':
                        caseid_stx.caseid_customer88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer89':
                        caseid_stx.caseid_customer89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer90':
                        caseid_stx.caseid_customer90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer91':
                        caseid_stx.caseid_customer91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer92':
                        caseid_stx.caseid_customer92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer93':
                        caseid_stx.caseid_customer93(self)
                except:
                    module_other_stx.swich_tab_0()


        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Customer01':
                        caseid_stx.caseid_customer01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer02':
                        caseid_stx.caseid_customer02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer03':
                        caseid_stx.caseid_customer03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer04':
                        caseid_stx.caseid_customer04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer05':
                        caseid_stx.caseid_customer05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer06':
                        caseid_stx.caseid_customer06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer07':
                        caseid_stx.caseid_customer07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer08':
                        caseid_stx.caseid_customer08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer09':
                        caseid_stx.caseid_customer09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer10':
                        caseid_stx.caseid_customer10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer11':
                        caseid_stx.caseid_customer11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer12':
                        caseid_stx.caseid_customer12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer13':
                        caseid_stx.caseid_customer13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer14':
                        caseid_stx.caseid_customer14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer15':
                        caseid_stx.caseid_customer15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer16':
                        caseid_stx.caseid_customer16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer17':
                        caseid_stx.caseid_customer17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer18':
                        caseid_stx.caseid_customer18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer19':
                        caseid_stx.caseid_customer19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer20':
                        caseid_stx.caseid_customer20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer21':
                        caseid_stx.caseid_customer21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer22':
                        caseid_stx.caseid_customer22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer23':
                        caseid_stx.caseid_customer23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer24':
                        caseid_stx.caseid_customer24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer25':
                        caseid_stx.caseid_customer25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26':
                        caseid_stx.caseid_customer26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_1':
                        caseid_stx.caseid_customer26_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_2':
                        caseid_stx.caseid_customer26_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer27':
                        caseid_stx.caseid_customer27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer28':
                        caseid_stx.caseid_customer28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer29':
                        caseid_stx.caseid_customer29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer30':
                        caseid_stx.caseid_customer30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer31':
                        caseid_stx.caseid_customer31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer32':
                        caseid_stx.caseid_customer32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer33':
                        caseid_stx.caseid_customer33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer34':
                        caseid_stx.caseid_customer34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer35':
                        caseid_stx.caseid_customer35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer36':
                        caseid_stx.caseid_customer36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer37':
                        caseid_stx.caseid_customer37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer38':
                        caseid_stx.caseid_customer38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer39':
                        caseid_stx.caseid_customer39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer40':
                        caseid_stx.caseid_customer40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer41':
                        caseid_stx.caseid_customer41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer42':
                        caseid_stx.caseid_customer42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer43':
                        caseid_stx.caseid_customer43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer44':
                        caseid_stx.caseid_customer44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer45':
                        caseid_stx.caseid_customer45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer46':
                        caseid_stx.caseid_customer46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer47':
                        caseid_stx.caseid_customer47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer48':
                        caseid_stx.caseid_customer48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer49':
                        caseid_stx.caseid_customer49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer50':
                        caseid_stx.caseid_customer50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer51':
                        caseid_stx.caseid_customer51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer52':
                        caseid_stx.caseid_customer52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer53':
                        caseid_stx.caseid_customer53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer54':
                        caseid_stx.caseid_customer54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer55':
                        caseid_stx.caseid_customer55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer56':
                        caseid_stx.caseid_customer56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57':
                        caseid_stx.caseid_customer57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_1':
                        caseid_stx.caseid_customer57_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_2':
                        caseid_stx.caseid_customer57_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_3':
                        caseid_stx.caseid_customer57_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_4':
                        caseid_stx.caseid_customer57_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_5':
                        caseid_stx.caseid_customer57_5(self)
                except:
                    module_other_stx.swich_tab_0()



                try:
                    if case == 'Customer58':
                        caseid_stx.caseid_customer58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer59':
                        caseid_stx.caseid_customer59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer60':
                        caseid_stx.caseid_customer60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer61':
                        caseid_stx.caseid_customer61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer62':
                        caseid_stx.caseid_customer62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer63':
                        caseid_stx.caseid_customer63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer64':
                        caseid_stx.caseid_customer64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer65':
                        caseid_stx.caseid_customer65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer66':
                        caseid_stx.caseid_customer66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer67':
                        caseid_stx.caseid_customer67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer68':
                        caseid_stx.caseid_customer68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer69':
                        caseid_stx.caseid_customer69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer70':
                        caseid_stx.caseid_customer70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer71':
                        caseid_stx.caseid_customer71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer72':
                        caseid_stx.caseid_customer72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer73':
                        caseid_stx.caseid_customer73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer74':
                        caseid_stx.caseid_customer74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer75':
                        caseid_stx.caseid_customer75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer76':
                        caseid_stx.caseid_customer76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77':
                        caseid_stx.caseid_customer77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77_1':
                        caseid_stx.caseid_customer77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer78':
                        caseid_stx.caseid_customer78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer79':
                        caseid_stx.caseid_customer79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer80':
                        caseid_stx.caseid_customer80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer81':
                        caseid_stx.caseid_customer81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer82':
                        caseid_stx.caseid_customer82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer83':
                        caseid_stx.caseid_customer83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer84':
                        caseid_stx.caseid_customer84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer85':
                        caseid_stx.caseid_customer85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer86':
                        caseid_stx.caseid_customer86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer87':
                        caseid_stx.caseid_customer87(self)
                except:
                    module_other_stx.swich_tab_0()

                try:
                    if case == 'Customer88':
                        caseid_stx.caseid_customer88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer89':
                        caseid_stx.caseid_customer89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer90':
                        caseid_stx.caseid_customer90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer91':
                        caseid_stx.caseid_customer91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer92':
                        caseid_stx.caseid_customer92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer93':
                        caseid_stx.caseid_customer93(self)
                except:
                    module_other_stx.swich_tab_0()



        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Customer01':
                        caseid_stx.caseid_customer01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer02':
                        caseid_stx.caseid_customer02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer03':
                        caseid_stx.caseid_customer03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer04':
                        caseid_stx.caseid_customer04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer05':
                        caseid_stx.caseid_customer05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer06':
                        caseid_stx.caseid_customer06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer07':
                        caseid_stx.caseid_customer07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer08':
                        caseid_stx.caseid_customer08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer09':
                        caseid_stx.caseid_customer09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer10':
                        caseid_stx.caseid_customer10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer11':
                        caseid_stx.caseid_customer11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer12':
                        caseid_stx.caseid_customer12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer13':
                        caseid_stx.caseid_customer13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer14':
                        caseid_stx.caseid_customer14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer15':
                        caseid_stx.caseid_customer15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer16':
                        caseid_stx.caseid_customer16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer17':
                        caseid_stx.caseid_customer17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer18':
                        caseid_stx.caseid_customer18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer19':
                        caseid_stx.caseid_customer19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer20':
                        caseid_stx.caseid_customer20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer21':
                        caseid_stx.caseid_customer21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer22':
                        caseid_stx.caseid_customer22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer23':
                        caseid_stx.caseid_customer23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer24':
                        caseid_stx.caseid_customer24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer25':
                        caseid_stx.caseid_customer25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26':
                        caseid_stx.caseid_customer26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_1':
                        caseid_stx.caseid_customer26_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_2':
                        caseid_stx.caseid_customer26_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer27':
                        caseid_stx.caseid_customer27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer28':
                        caseid_stx.caseid_customer28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer29':
                        caseid_stx.caseid_customer29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer30':
                        caseid_stx.caseid_customer30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer31':
                        caseid_stx.caseid_customer31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer32':
                        caseid_stx.caseid_customer32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer33':
                        caseid_stx.caseid_customer33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer34':
                        caseid_stx.caseid_customer34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer35':
                        caseid_stx.caseid_customer35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer36':
                        caseid_stx.caseid_customer36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer37':
                        caseid_stx.caseid_customer37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer38':
                        caseid_stx.caseid_customer38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer39':
                        caseid_stx.caseid_customer39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer40':
                        caseid_stx.caseid_customer40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer41':
                        caseid_stx.caseid_customer41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer42':
                        caseid_stx.caseid_customer42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer43':
                        caseid_stx.caseid_customer43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer44':
                        caseid_stx.caseid_customer44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer45':
                        caseid_stx.caseid_customer45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer46':
                        caseid_stx.caseid_customer46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer47':
                        caseid_stx.caseid_customer47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer48':
                        caseid_stx.caseid_customer48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer49':
                        caseid_stx.caseid_customer49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer50':
                        caseid_stx.caseid_customer50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer51':
                        caseid_stx.caseid_customer51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer52':
                        caseid_stx.caseid_customer52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer53':
                        caseid_stx.caseid_customer53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer54':
                        caseid_stx.caseid_customer54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer55':
                        caseid_stx.caseid_customer55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer56':
                        caseid_stx.caseid_customer56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57':
                        caseid_stx.caseid_customer57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_1':
                        caseid_stx.caseid_customer57_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_2':
                        caseid_stx.caseid_customer57_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_3':
                        caseid_stx.caseid_customer57_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_4':
                        caseid_stx.caseid_customer57_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_5':
                        caseid_stx.caseid_customer57_5(self)
                except:
                    module_other_stx.swich_tab_0()



                try:
                    if case == 'Customer58':
                        caseid_stx.caseid_customer58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer59':
                        caseid_stx.caseid_customer59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer60':
                        caseid_stx.caseid_customer60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer61':
                        caseid_stx.caseid_customer61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer62':
                        caseid_stx.caseid_customer62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer63':
                        caseid_stx.caseid_customer63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer64':
                        caseid_stx.caseid_customer64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer65':
                        caseid_stx.caseid_customer65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer66':
                        caseid_stx.caseid_customer66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer67':
                        caseid_stx.caseid_customer67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer68':
                        caseid_stx.caseid_customer68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer69':
                        caseid_stx.caseid_customer69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer70':
                        caseid_stx.caseid_customer70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer71':
                        caseid_stx.caseid_customer71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer72':
                        caseid_stx.caseid_customer72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer73':
                        caseid_stx.caseid_customer73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer74':
                        caseid_stx.caseid_customer74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer75':
                        caseid_stx.caseid_customer75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer76':
                        caseid_stx.caseid_customer76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77':
                        caseid_stx.caseid_customer77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77_1':
                        caseid_stx.caseid_customer77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer78':
                        caseid_stx.caseid_customer78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer79':
                        caseid_stx.caseid_customer79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer80':
                        caseid_stx.caseid_customer80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer81':
                        caseid_stx.caseid_customer81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer82':
                        caseid_stx.caseid_customer82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer83':
                        caseid_stx.caseid_customer83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer84':
                        caseid_stx.caseid_customer84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer85':
                        caseid_stx.caseid_customer85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer86':
                        caseid_stx.caseid_customer86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer87':
                        caseid_stx.caseid_customer87(self)
                except:
                    module_other_stx.swich_tab_0()

                try:
                    if case == 'Customer88':
                        caseid_stx.caseid_customer88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer89':
                        caseid_stx.caseid_customer89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer90':
                        caseid_stx.caseid_customer90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer91':
                        caseid_stx.caseid_customer91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer92':
                        caseid_stx.caseid_customer92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer93':
                        caseid_stx.caseid_customer93(self)
                except:
                    module_other_stx.swich_tab_0()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Customer01':
                        caseid_stx.caseid_customer01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer02':
                        caseid_stx.caseid_customer02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer03':
                        caseid_stx.caseid_customer03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer04':
                        caseid_stx.caseid_customer04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer05':
                        caseid_stx.caseid_customer05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer06':
                        caseid_stx.caseid_customer06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer07':
                        caseid_stx.caseid_customer07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer08':
                        caseid_stx.caseid_customer08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer09':
                        caseid_stx.caseid_customer09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer10':
                        caseid_stx.caseid_customer10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer11':
                        caseid_stx.caseid_customer11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer12':
                        caseid_stx.caseid_customer12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer13':
                        caseid_stx.caseid_customer13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer14':
                        caseid_stx.caseid_customer14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer15':
                        caseid_stx.caseid_customer15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer16':
                        caseid_stx.caseid_customer16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer17':
                        caseid_stx.caseid_customer17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer18':
                        caseid_stx.caseid_customer18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer19':
                        caseid_stx.caseid_customer19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer20':
                        caseid_stx.caseid_customer20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer21':
                        caseid_stx.caseid_customer21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer22':
                        caseid_stx.caseid_customer22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer23':
                        caseid_stx.caseid_customer23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer24':
                        caseid_stx.caseid_customer24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer25':
                        caseid_stx.caseid_customer25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26':
                        caseid_stx.caseid_customer26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_1':
                        caseid_stx.caseid_customer26_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer26_2':
                        caseid_stx.caseid_customer26_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer27':
                        caseid_stx.caseid_customer27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer28':
                        caseid_stx.caseid_customer28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer29':
                        caseid_stx.caseid_customer29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer30':
                        caseid_stx.caseid_customer30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer31':
                        caseid_stx.caseid_customer31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer32':
                        caseid_stx.caseid_customer32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer33':
                        caseid_stx.caseid_customer33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer34':
                        caseid_stx.caseid_customer34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer35':
                        caseid_stx.caseid_customer35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer36':
                        caseid_stx.caseid_customer36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer37':
                        caseid_stx.caseid_customer37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer38':
                        caseid_stx.caseid_customer38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer39':
                        caseid_stx.caseid_customer39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer40':
                        caseid_stx.caseid_customer40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer41':
                        caseid_stx.caseid_customer41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer42':
                        caseid_stx.caseid_customer42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer43':
                        caseid_stx.caseid_customer43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer44':
                        caseid_stx.caseid_customer44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer45':
                        caseid_stx.caseid_customer45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer46':
                        caseid_stx.caseid_customer46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer47':
                        caseid_stx.caseid_customer47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer48':
                        caseid_stx.caseid_customer48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer49':
                        caseid_stx.caseid_customer49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer50':
                        caseid_stx.caseid_customer50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer51':
                        caseid_stx.caseid_customer51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer52':
                        caseid_stx.caseid_customer52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer53':
                        caseid_stx.caseid_customer53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer54':
                        caseid_stx.caseid_customer54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer55':
                        caseid_stx.caseid_customer55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer56':
                        caseid_stx.caseid_customer56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57':
                        caseid_stx.caseid_customer57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_1':
                        caseid_stx.caseid_customer57_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_2':
                        caseid_stx.caseid_customer57_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_3':
                        caseid_stx.caseid_customer57_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_4':
                        caseid_stx.caseid_customer57_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer57_5':
                        caseid_stx.caseid_customer57_5(self)
                except:
                    module_other_stx.swich_tab_0()



                try:
                    if case == 'Customer58':
                        caseid_stx.caseid_customer58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer59':
                        caseid_stx.caseid_customer59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer60':
                        caseid_stx.caseid_customer60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer61':
                        caseid_stx.caseid_customer61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer62':
                        caseid_stx.caseid_customer62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer63':
                        caseid_stx.caseid_customer63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer64':
                        caseid_stx.caseid_customer64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer65':
                        caseid_stx.caseid_customer65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer66':
                        caseid_stx.caseid_customer66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer67':
                        caseid_stx.caseid_customer67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer68':
                        caseid_stx.caseid_customer68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer69':
                        caseid_stx.caseid_customer69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer70':
                        caseid_stx.caseid_customer70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer71':
                        caseid_stx.caseid_customer71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer72':
                        caseid_stx.caseid_customer72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer73':
                        caseid_stx.caseid_customer73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer74':
                        caseid_stx.caseid_customer74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer75':
                        caseid_stx.caseid_customer75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer76':
                        caseid_stx.caseid_customer76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77':
                        caseid_stx.caseid_customer77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer77_1':
                        caseid_stx.caseid_customer77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer78':
                        caseid_stx.caseid_customer78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer79':
                        caseid_stx.caseid_customer79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer80':
                        caseid_stx.caseid_customer80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer81':
                        caseid_stx.caseid_customer81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer82':
                        caseid_stx.caseid_customer82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer83':
                        caseid_stx.caseid_customer83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer84':
                        caseid_stx.caseid_customer84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer85':
                        caseid_stx.caseid_customer85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer86':
                        caseid_stx.caseid_customer86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer87':
                        caseid_stx.caseid_customer87(self)
                except:
                    module_other_stx.swich_tab_0()

                try:
                    if case == 'Customer88':
                        caseid_stx.caseid_customer88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer89':
                        caseid_stx.caseid_customer89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer90':
                        caseid_stx.caseid_customer90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer91':
                        caseid_stx.caseid_customer91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer92':
                        caseid_stx.caseid_customer92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Customer93':
                        caseid_stx.caseid_customer93(self)
                except:
                    module_other_stx.swich_tab_0()


def report(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 445
    while (rownum < 650):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Report01':
                        caseid_stx.caseid_report01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_stx.caseid_report02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_stx.caseid_report03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_stx.caseid_report04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_stx.caseid_report05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_stx.caseid_report06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_stx.caseid_report07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_1':
                        caseid_stx.caseid_report07_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_2':
                        caseid_stx.caseid_report07_2(self)
                except:
                    module_other_stx.swich_tab_0()


                try:
                    if case == 'Report08':
                        caseid_stx.caseid_report08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_stx.caseid_report09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_stx.caseid_report10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_stx.caseid_report11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_stx.caseid_report12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_stx.caseid_report13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_stx.caseid_report14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_stx.caseid_report15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_stx.caseid_report16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_stx.caseid_report17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_stx.caseid_report18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_stx.caseid_report19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_stx.caseid_report20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_stx.caseid_report21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_stx.caseid_report22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_stx.caseid_report23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_stx.caseid_report24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_stx.caseid_report25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_stx.caseid_report26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_stx.caseid_report27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_stx.caseid_report28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_stx.caseid_report29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_stx.caseid_report30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_stx.caseid_report31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_stx.caseid_report32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_stx.caseid_report33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_stx.caseid_report34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_stx.caseid_report35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_stx.caseid_report36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_stx.caseid_report37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_stx.caseid_report38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_stx.caseid_report39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_stx.caseid_report40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_stx.caseid_report41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_stx.caseid_report42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_stx.caseid_report43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_stx.caseid_report44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_stx.caseid_report45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_stx.caseid_report46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_stx.caseid_report47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_stx.caseid_report48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_stx.caseid_report49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_stx.caseid_report50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_stx.caseid_report51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_stx.caseid_report52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_stx.caseid_report53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_stx.caseid_report54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_stx.caseid_report55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_stx.caseid_report56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_stx.caseid_report57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_stx.caseid_report58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_stx.caseid_report59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_stx.caseid_report60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_stx.caseid_report61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_stx.caseid_report62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_stx.caseid_report63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_stx.caseid_report64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_stx.caseid_report65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_stx.caseid_report66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_stx.caseid_report67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_stx.caseid_report68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_stx.caseid_report69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_stx.caseid_report70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_stx.caseid_report71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_stx.caseid_report72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_stx.caseid_report73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_stx.caseid_report74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_stx.caseid_report75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_stx.caseid_report76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_stx.caseid_report77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_1':
                        caseid_stx.caseid_report77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_2':
                        caseid_stx.caseid_report77_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_stx.caseid_report78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_stx.caseid_report79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_stx.caseid_report80(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Report01':
                        caseid_stx.caseid_report01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_stx.caseid_report02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_stx.caseid_report03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_stx.caseid_report04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_stx.caseid_report05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_stx.caseid_report06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_stx.caseid_report07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_1':
                        caseid_stx.caseid_report07_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_2':
                        caseid_stx.caseid_report07_2(self)
                except:
                    module_other_stx.swich_tab_0()


                try:
                    if case == 'Report08':
                        caseid_stx.caseid_report08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_stx.caseid_report09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_stx.caseid_report10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_stx.caseid_report11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_stx.caseid_report12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_stx.caseid_report13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_stx.caseid_report14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_stx.caseid_report15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_stx.caseid_report16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_stx.caseid_report17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_stx.caseid_report18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_stx.caseid_report19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_stx.caseid_report20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_stx.caseid_report21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_stx.caseid_report22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_stx.caseid_report23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_stx.caseid_report24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_stx.caseid_report25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_stx.caseid_report26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_stx.caseid_report27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_stx.caseid_report28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_stx.caseid_report29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_stx.caseid_report30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_stx.caseid_report31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_stx.caseid_report32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_stx.caseid_report33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_stx.caseid_report34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_stx.caseid_report35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_stx.caseid_report36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_stx.caseid_report37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_stx.caseid_report38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_stx.caseid_report39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_stx.caseid_report40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_stx.caseid_report41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_stx.caseid_report42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_stx.caseid_report43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_stx.caseid_report44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_stx.caseid_report45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_stx.caseid_report46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_stx.caseid_report47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_stx.caseid_report48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_stx.caseid_report49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_stx.caseid_report50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_stx.caseid_report51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_stx.caseid_report52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_stx.caseid_report53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_stx.caseid_report54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_stx.caseid_report55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_stx.caseid_report56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_stx.caseid_report57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_stx.caseid_report58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_stx.caseid_report59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_stx.caseid_report60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_stx.caseid_report61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_stx.caseid_report62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_stx.caseid_report63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_stx.caseid_report64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_stx.caseid_report65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_stx.caseid_report66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_stx.caseid_report67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_stx.caseid_report68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_stx.caseid_report69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_stx.caseid_report70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_stx.caseid_report71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_stx.caseid_report72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_stx.caseid_report73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_stx.caseid_report74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_stx.caseid_report75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_stx.caseid_report76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_stx.caseid_report77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_1':
                        caseid_stx.caseid_report77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_2':
                        caseid_stx.caseid_report77_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_stx.caseid_report78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_stx.caseid_report79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_stx.caseid_report80(self)
                except:
                    module_other_stx.swich_tab_0()


        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Report01':
                        caseid_stx.caseid_report01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_stx.caseid_report02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_stx.caseid_report03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_stx.caseid_report04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_stx.caseid_report05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_stx.caseid_report06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_stx.caseid_report07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_1':
                        caseid_stx.caseid_report07_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_2':
                        caseid_stx.caseid_report07_2(self)
                except:
                    module_other_stx.swich_tab_0()


                try:
                    if case == 'Report08':
                        caseid_stx.caseid_report08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_stx.caseid_report09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_stx.caseid_report10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_stx.caseid_report11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_stx.caseid_report12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_stx.caseid_report13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_stx.caseid_report14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_stx.caseid_report15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_stx.caseid_report16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_stx.caseid_report17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_stx.caseid_report18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_stx.caseid_report19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_stx.caseid_report20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_stx.caseid_report21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_stx.caseid_report22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_stx.caseid_report23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_stx.caseid_report24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_stx.caseid_report25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_stx.caseid_report26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_stx.caseid_report27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_stx.caseid_report28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_stx.caseid_report29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_stx.caseid_report30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_stx.caseid_report31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_stx.caseid_report32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_stx.caseid_report33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_stx.caseid_report34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_stx.caseid_report35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_stx.caseid_report36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_stx.caseid_report37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_stx.caseid_report38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_stx.caseid_report39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_stx.caseid_report40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_stx.caseid_report41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_stx.caseid_report42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_stx.caseid_report43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_stx.caseid_report44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_stx.caseid_report45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_stx.caseid_report46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_stx.caseid_report47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_stx.caseid_report48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_stx.caseid_report49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_stx.caseid_report50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_stx.caseid_report51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_stx.caseid_report52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_stx.caseid_report53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_stx.caseid_report54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_stx.caseid_report55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_stx.caseid_report56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_stx.caseid_report57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_stx.caseid_report58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_stx.caseid_report59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_stx.caseid_report60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_stx.caseid_report61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_stx.caseid_report62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_stx.caseid_report63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_stx.caseid_report64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_stx.caseid_report65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_stx.caseid_report66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_stx.caseid_report67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_stx.caseid_report68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_stx.caseid_report69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_stx.caseid_report70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_stx.caseid_report71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_stx.caseid_report72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_stx.caseid_report73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_stx.caseid_report74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_stx.caseid_report75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_stx.caseid_report76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_stx.caseid_report77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_1':
                        caseid_stx.caseid_report77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_2':
                        caseid_stx.caseid_report77_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_stx.caseid_report78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_stx.caseid_report79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_stx.caseid_report80(self)
                except:
                    module_other_stx.swich_tab_0()


        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Report01':
                        caseid_stx.caseid_report01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_stx.caseid_report02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_stx.caseid_report03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_stx.caseid_report04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_stx.caseid_report05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_stx.caseid_report06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_stx.caseid_report07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_1':
                        caseid_stx.caseid_report07_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report07_2':
                        caseid_stx.caseid_report07_2(self)
                except:
                    module_other_stx.swich_tab_0()



                try:
                    if case == 'Report08':
                        caseid_stx.caseid_report08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_stx.caseid_report09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_stx.caseid_report10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_stx.caseid_report11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_stx.caseid_report12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_stx.caseid_report13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_stx.caseid_report14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_stx.caseid_report15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_stx.caseid_report16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_stx.caseid_report17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_stx.caseid_report18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_stx.caseid_report19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_stx.caseid_report20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_stx.caseid_report21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_stx.caseid_report22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_stx.caseid_report23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_stx.caseid_report24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_stx.caseid_report25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_stx.caseid_report26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_stx.caseid_report27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_stx.caseid_report28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_stx.caseid_report29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_stx.caseid_report30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_stx.caseid_report31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_stx.caseid_report32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_stx.caseid_report33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_stx.caseid_report34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_stx.caseid_report35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_stx.caseid_report36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_stx.caseid_report37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_stx.caseid_report38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_stx.caseid_report39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_stx.caseid_report40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_stx.caseid_report41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_stx.caseid_report42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_stx.caseid_report43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_stx.caseid_report44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_stx.caseid_report45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_stx.caseid_report46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_stx.caseid_report47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_stx.caseid_report48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_stx.caseid_report49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_stx.caseid_report50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_stx.caseid_report51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_stx.caseid_report52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_stx.caseid_report53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_stx.caseid_report54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_stx.caseid_report55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_stx.caseid_report56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_stx.caseid_report57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_stx.caseid_report58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_stx.caseid_report59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_stx.caseid_report60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_stx.caseid_report61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_stx.caseid_report62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_stx.caseid_report63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_stx.caseid_report64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_stx.caseid_report65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_stx.caseid_report66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_stx.caseid_report67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_stx.caseid_report68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_stx.caseid_report69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_stx.caseid_report70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_stx.caseid_report71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_stx.caseid_report72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_stx.caseid_report73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_stx.caseid_report74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_stx.caseid_report75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_stx.caseid_report76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_stx.caseid_report77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_1':
                        caseid_stx.caseid_report77_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report77_2':
                        caseid_stx.caseid_report77_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_stx.caseid_report78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_stx.caseid_report79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_stx.caseid_report80(self)
                except:
                    module_other_stx.swich_tab_0()


def admin(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 586
    while (rownum < 830):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Admin01':
                        caseid_stx.caseid_admin01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_stx.caseid_admin02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_stx.caseid_admin03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_stx.caseid_admin04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_stx.caseid_admin05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_stx.caseid_admin06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_stx.caseid_admin07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_stx.caseid_admin08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_1':
                        caseid_stx.caseid_admin08_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_2':
                        caseid_stx.caseid_admin08_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_3':
                        caseid_stx.caseid_admin08_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_4':
                        caseid_stx.caseid_admin08_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_5':
                        caseid_stx.caseid_admin08_5(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_6':
                        caseid_stx.caseid_admin08_6(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_7':
                        caseid_stx.caseid_admin08_7(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_8':
                        caseid_stx.caseid_admin08_8(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_9':
                        caseid_stx.caseid_admin08_9(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_stx.caseid_admin09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_stx.caseid_admin10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_stx.caseid_admin11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_stx.caseid_admin12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_stx.caseid_admin13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_stx.caseid_admin14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_stx.caseid_admin15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_stx.caseid_admin16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_stx.caseid_admin17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_stx.caseid_admin18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_stx.caseid_admin19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_stx.caseid_admin20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_stx.caseid_admin21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_stx.caseid_admin22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_stx.caseid_admin23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_stx.caseid_admin24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_stx.caseid_admin25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_stx.caseid_admin26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_stx.caseid_admin27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_stx.caseid_admin28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_stx.caseid_admin29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_stx.caseid_admin30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_stx.caseid_admin31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_stx.caseid_admin32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_stx.caseid_admin33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_stx.caseid_admin34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_stx.caseid_admin35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_stx.caseid_admin36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_stx.caseid_admin37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_stx.caseid_admin38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_stx.caseid_admin39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_stx.caseid_admin40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_stx.caseid_admin41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_stx.caseid_admin42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_stx.caseid_admin43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_stx.caseid_admin44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_stx.caseid_admin45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_stx.caseid_admin46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_stx.caseid_admin47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_stx.caseid_admin48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_stx.caseid_admin49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_stx.caseid_admin50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_stx.caseid_admin51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_stx.caseid_admin52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_stx.caseid_admin53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_stx.caseid_admin54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_stx.caseid_admin55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_stx.caseid_admin56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_stx.caseid_admin57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_stx.caseid_admin58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_stx.caseid_admin59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_stx.caseid_admin60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_stx.caseid_admin61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_stx.caseid_admin62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_stx.caseid_admin63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_stx.caseid_admin64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_stx.caseid_admin65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_stx.caseid_admin66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_stx.caseid_admin67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_stx.caseid_admin68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_stx.caseid_admin69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_stx.caseid_admin70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_stx.caseid_admin71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_stx.caseid_admin72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_stx.caseid_admin73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_stx.caseid_admin74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_stx.caseid_admin75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_stx.caseid_admin76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_stx.caseid_admin77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_stx.caseid_admin78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_stx.caseid_admin79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_stx.caseid_admin80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_stx.caseid_admin81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_stx.caseid_admin82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_stx.caseid_admin83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_stx.caseid_admin84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_stx.caseid_admin85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_stx.caseid_admin86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_stx.caseid_admin87(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_stx.caseid_admin88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_stx.caseid_admin89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_stx.caseid_admin90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_stx.caseid_admin91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_stx.caseid_admin92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_stx.caseid_admin93(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_stx.caseid_admin94(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_stx.caseid_admin95(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_stx.caseid_admin96(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_stx.caseid_admin97(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_stx.caseid_admin98(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_stx.caseid_admin99(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_stx.caseid_admin100(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_stx.caseid_admin101(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_stx.caseid_admin102(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_stx.caseid_admin103(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_stx.caseid_admin104(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_stx.caseid_admin105(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_stx.caseid_admin106(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_stx.caseid_admin107(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_stx.caseid_admin108(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_stx.caseid_admin109(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_stx.caseid_admin110(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_stx.caseid_admin111(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_stx.caseid_admin112(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_stx.caseid_admin113(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_stx.caseid_admin114(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_stx.caseid_admin115(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_stx.caseid_admin116(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_stx.caseid_admin117(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_stx.caseid_admin118(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_stx.caseid_admin119(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_stx.caseid_admin120(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_stx.caseid_admin121(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_stx.caseid_admin122(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_stx.caseid_admin123(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_stx.caseid_admin124(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_stx.caseid_admin125(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_stx.caseid_admin126(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_stx.caseid_admin127(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_stx.caseid_admin128(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_stx.caseid_admin129(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip01':
                        caseid_stx.caseid_PartnerTrip01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip02':
                        caseid_stx.caseid_PartnerTrip02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip03':
                        caseid_stx.caseid_PartnerTrip03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip04':
                        caseid_stx.caseid_PartnerTrip04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip05':
                        caseid_stx.caseid_PartnerTrip05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip06':
                        caseid_stx.caseid_PartnerTrip06(self)
                except:
                    module_other_stx.swich_tab_0()



        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Admin01':
                        caseid_stx.caseid_admin01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_stx.caseid_admin02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_stx.caseid_admin03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_stx.caseid_admin04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_stx.caseid_admin05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_stx.caseid_admin06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_stx.caseid_admin07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_stx.caseid_admin08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_1':
                        caseid_stx.caseid_admin08_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_2':
                        caseid_stx.caseid_admin08_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_3':
                        caseid_stx.caseid_admin08_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_4':
                        caseid_stx.caseid_admin08_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_5':
                        caseid_stx.caseid_admin08_5(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_6':
                        caseid_stx.caseid_admin08_6(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_7':
                        caseid_stx.caseid_admin08_7(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_8':
                        caseid_stx.caseid_admin08_8(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_9':
                        caseid_stx.caseid_admin08_9(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_stx.caseid_admin09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_stx.caseid_admin10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_stx.caseid_admin11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_stx.caseid_admin12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_stx.caseid_admin13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_stx.caseid_admin14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_stx.caseid_admin15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_stx.caseid_admin16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_stx.caseid_admin17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_stx.caseid_admin18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_stx.caseid_admin19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_stx.caseid_admin20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_stx.caseid_admin21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_stx.caseid_admin22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_stx.caseid_admin23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_stx.caseid_admin24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_stx.caseid_admin25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_stx.caseid_admin26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_stx.caseid_admin27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_stx.caseid_admin28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_stx.caseid_admin29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_stx.caseid_admin30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_stx.caseid_admin31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_stx.caseid_admin32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_stx.caseid_admin33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_stx.caseid_admin34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_stx.caseid_admin35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_stx.caseid_admin36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_stx.caseid_admin37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_stx.caseid_admin38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_stx.caseid_admin39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_stx.caseid_admin40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_stx.caseid_admin41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_stx.caseid_admin42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_stx.caseid_admin43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_stx.caseid_admin44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_stx.caseid_admin45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_stx.caseid_admin46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_stx.caseid_admin47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_stx.caseid_admin48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_stx.caseid_admin49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_stx.caseid_admin50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_stx.caseid_admin51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_stx.caseid_admin52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_stx.caseid_admin53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_stx.caseid_admin54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_stx.caseid_admin55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_stx.caseid_admin56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_stx.caseid_admin57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_stx.caseid_admin58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_stx.caseid_admin59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_stx.caseid_admin60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_stx.caseid_admin61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_stx.caseid_admin62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_stx.caseid_admin63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_stx.caseid_admin64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_stx.caseid_admin65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_stx.caseid_admin66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_stx.caseid_admin67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_stx.caseid_admin68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_stx.caseid_admin69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_stx.caseid_admin70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_stx.caseid_admin71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_stx.caseid_admin72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_stx.caseid_admin73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_stx.caseid_admin74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_stx.caseid_admin75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_stx.caseid_admin76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_stx.caseid_admin77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_stx.caseid_admin78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_stx.caseid_admin79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_stx.caseid_admin80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_stx.caseid_admin81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_stx.caseid_admin82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_stx.caseid_admin83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_stx.caseid_admin84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_stx.caseid_admin85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_stx.caseid_admin86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_stx.caseid_admin87(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_stx.caseid_admin88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_stx.caseid_admin89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_stx.caseid_admin90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_stx.caseid_admin91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_stx.caseid_admin92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_stx.caseid_admin93(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_stx.caseid_admin94(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_stx.caseid_admin95(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_stx.caseid_admin96(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_stx.caseid_admin97(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_stx.caseid_admin98(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_stx.caseid_admin99(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_stx.caseid_admin100(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_stx.caseid_admin101(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_stx.caseid_admin102(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_stx.caseid_admin103(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_stx.caseid_admin104(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_stx.caseid_admin105(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_stx.caseid_admin106(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_stx.caseid_admin107(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_stx.caseid_admin108(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_stx.caseid_admin109(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_stx.caseid_admin110(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_stx.caseid_admin111(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_stx.caseid_admin112(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_stx.caseid_admin113(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_stx.caseid_admin114(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_stx.caseid_admin115(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_stx.caseid_admin116(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_stx.caseid_admin117(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_stx.caseid_admin118(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_stx.caseid_admin119(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_stx.caseid_admin120(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_stx.caseid_admin121(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_stx.caseid_admin122(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_stx.caseid_admin123(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_stx.caseid_admin124(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_stx.caseid_admin125(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_stx.caseid_admin126(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_stx.caseid_admin127(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_stx.caseid_admin128(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_stx.caseid_admin129(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip01':
                        caseid_stx.caseid_PartnerTrip01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip02':
                        caseid_stx.caseid_PartnerTrip02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip03':
                        caseid_stx.caseid_PartnerTrip03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip04':
                        caseid_stx.caseid_PartnerTrip04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip05':
                        caseid_stx.caseid_PartnerTrip05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip06':
                        caseid_stx.caseid_PartnerTrip06(self)
                except:
                    module_other_stx.swich_tab_0()



        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Admin01':
                        caseid_stx.caseid_admin01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_stx.caseid_admin02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_stx.caseid_admin03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_stx.caseid_admin04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_stx.caseid_admin05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_stx.caseid_admin06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_stx.caseid_admin07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_stx.caseid_admin08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_1':
                        caseid_stx.caseid_admin08_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_2':
                        caseid_stx.caseid_admin08_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_3':
                        caseid_stx.caseid_admin08_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_4':
                        caseid_stx.caseid_admin08_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_5':
                        caseid_stx.caseid_admin08_5(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_6':
                        caseid_stx.caseid_admin08_6(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_7':
                        caseid_stx.caseid_admin08_7(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_8':
                        caseid_stx.caseid_admin08_8(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_9':
                        caseid_stx.caseid_admin08_9(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_stx.caseid_admin09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_stx.caseid_admin10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_stx.caseid_admin11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_stx.caseid_admin12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_stx.caseid_admin13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_stx.caseid_admin14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_stx.caseid_admin15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_stx.caseid_admin16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_stx.caseid_admin17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_stx.caseid_admin18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_stx.caseid_admin19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_stx.caseid_admin20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_stx.caseid_admin21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_stx.caseid_admin22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_stx.caseid_admin23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_stx.caseid_admin24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_stx.caseid_admin25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_stx.caseid_admin26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_stx.caseid_admin27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_stx.caseid_admin28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_stx.caseid_admin29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_stx.caseid_admin30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_stx.caseid_admin31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_stx.caseid_admin32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_stx.caseid_admin33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_stx.caseid_admin34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_stx.caseid_admin35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_stx.caseid_admin36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_stx.caseid_admin37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_stx.caseid_admin38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_stx.caseid_admin39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_stx.caseid_admin40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_stx.caseid_admin41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_stx.caseid_admin42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_stx.caseid_admin43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_stx.caseid_admin44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_stx.caseid_admin45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_stx.caseid_admin46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_stx.caseid_admin47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_stx.caseid_admin48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_stx.caseid_admin49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_stx.caseid_admin50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_stx.caseid_admin51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_stx.caseid_admin52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_stx.caseid_admin53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_stx.caseid_admin54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_stx.caseid_admin55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_stx.caseid_admin56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_stx.caseid_admin57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_stx.caseid_admin58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_stx.caseid_admin59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_stx.caseid_admin60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_stx.caseid_admin61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_stx.caseid_admin62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_stx.caseid_admin63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_stx.caseid_admin64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_stx.caseid_admin65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_stx.caseid_admin66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_stx.caseid_admin67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_stx.caseid_admin68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_stx.caseid_admin69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_stx.caseid_admin70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_stx.caseid_admin71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_stx.caseid_admin72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_stx.caseid_admin73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_stx.caseid_admin74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_stx.caseid_admin75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_stx.caseid_admin76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_stx.caseid_admin77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_stx.caseid_admin78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_stx.caseid_admin79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_stx.caseid_admin80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_stx.caseid_admin81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_stx.caseid_admin82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_stx.caseid_admin83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_stx.caseid_admin84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_stx.caseid_admin85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_stx.caseid_admin86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_stx.caseid_admin87(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_stx.caseid_admin88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_stx.caseid_admin89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_stx.caseid_admin90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_stx.caseid_admin91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_stx.caseid_admin92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_stx.caseid_admin93(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_stx.caseid_admin94(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_stx.caseid_admin95(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_stx.caseid_admin96(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_stx.caseid_admin97(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_stx.caseid_admin98(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_stx.caseid_admin99(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_stx.caseid_admin100(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_stx.caseid_admin101(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_stx.caseid_admin102(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_stx.caseid_admin103(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_stx.caseid_admin104(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_stx.caseid_admin105(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_stx.caseid_admin106(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_stx.caseid_admin107(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_stx.caseid_admin108(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_stx.caseid_admin109(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_stx.caseid_admin110(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_stx.caseid_admin111(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_stx.caseid_admin112(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_stx.caseid_admin113(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_stx.caseid_admin114(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_stx.caseid_admin115(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_stx.caseid_admin116(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_stx.caseid_admin117(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_stx.caseid_admin118(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_stx.caseid_admin119(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_stx.caseid_admin120(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_stx.caseid_admin121(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_stx.caseid_admin122(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_stx.caseid_admin123(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_stx.caseid_admin124(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_stx.caseid_admin125(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_stx.caseid_admin126(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_stx.caseid_admin127(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_stx.caseid_admin128(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_stx.caseid_admin129(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip01':
                        caseid_stx.caseid_PartnerTrip01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip02':
                        caseid_stx.caseid_PartnerTrip02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip03':
                        caseid_stx.caseid_PartnerTrip03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip04':
                        caseid_stx.caseid_PartnerTrip04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip05':
                        caseid_stx.caseid_PartnerTrip05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip06':
                        caseid_stx.caseid_PartnerTrip06(self)
                except:
                    module_other_stx.swich_tab_0()



        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Admin01':
                        caseid_stx.caseid_admin01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_stx.caseid_admin02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_stx.caseid_admin03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_stx.caseid_admin04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_stx.caseid_admin05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_stx.caseid_admin06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_stx.caseid_admin07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_stx.caseid_admin08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_1':
                        caseid_stx.caseid_admin08_1(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_2':
                        caseid_stx.caseid_admin08_2(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_3':
                        caseid_stx.caseid_admin08_3(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_4':
                        caseid_stx.caseid_admin08_4(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_5':
                        caseid_stx.caseid_admin08_5(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_6':
                        caseid_stx.caseid_admin08_6(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_7':
                        caseid_stx.caseid_admin08_7(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_8':
                        caseid_stx.caseid_admin08_8(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin08_9':
                        caseid_stx.caseid_admin08_9(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_stx.caseid_admin09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_stx.caseid_admin10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_stx.caseid_admin11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_stx.caseid_admin12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_stx.caseid_admin13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_stx.caseid_admin14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_stx.caseid_admin15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_stx.caseid_admin16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_stx.caseid_admin17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_stx.caseid_admin18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_stx.caseid_admin19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_stx.caseid_admin20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_stx.caseid_admin21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_stx.caseid_admin22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_stx.caseid_admin23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_stx.caseid_admin24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_stx.caseid_admin25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_stx.caseid_admin26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_stx.caseid_admin27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_stx.caseid_admin28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_stx.caseid_admin29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_stx.caseid_admin30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_stx.caseid_admin31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_stx.caseid_admin32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_stx.caseid_admin33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_stx.caseid_admin34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_stx.caseid_admin35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_stx.caseid_admin36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_stx.caseid_admin37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_stx.caseid_admin38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_stx.caseid_admin39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_stx.caseid_admin40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_stx.caseid_admin41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_stx.caseid_admin42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_stx.caseid_admin43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_stx.caseid_admin44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_stx.caseid_admin45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_stx.caseid_admin46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_stx.caseid_admin47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_stx.caseid_admin48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_stx.caseid_admin49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_stx.caseid_admin50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_stx.caseid_admin51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_stx.caseid_admin52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_stx.caseid_admin53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_stx.caseid_admin54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_stx.caseid_admin55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_stx.caseid_admin56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_stx.caseid_admin57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_stx.caseid_admin58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_stx.caseid_admin59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_stx.caseid_admin60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_stx.caseid_admin61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_stx.caseid_admin62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_stx.caseid_admin63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_stx.caseid_admin64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_stx.caseid_admin65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_stx.caseid_admin66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_stx.caseid_admin67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_stx.caseid_admin68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_stx.caseid_admin69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_stx.caseid_admin70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_stx.caseid_admin71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_stx.caseid_admin72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_stx.caseid_admin73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_stx.caseid_admin74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_stx.caseid_admin75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_stx.caseid_admin76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_stx.caseid_admin77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_stx.caseid_admin78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_stx.caseid_admin79(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_stx.caseid_admin80(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_stx.caseid_admin81(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_stx.caseid_admin82(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_stx.caseid_admin83(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_stx.caseid_admin84(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_stx.caseid_admin85(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_stx.caseid_admin86(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_stx.caseid_admin87(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_stx.caseid_admin88(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_stx.caseid_admin89(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_stx.caseid_admin90(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_stx.caseid_admin91(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_stx.caseid_admin92(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_stx.caseid_admin93(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_stx.caseid_admin94(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_stx.caseid_admin95(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_stx.caseid_admin96(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_stx.caseid_admin97(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_stx.caseid_admin98(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_stx.caseid_admin99(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_stx.caseid_admin100(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_stx.caseid_admin101(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_stx.caseid_admin102(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_stx.caseid_admin103(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_stx.caseid_admin104(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_stx.caseid_admin105(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_stx.caseid_admin106(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_stx.caseid_admin107(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_stx.caseid_admin108(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_stx.caseid_admin109(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_stx.caseid_admin110(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_stx.caseid_admin111(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_stx.caseid_admin112(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_stx.caseid_admin113(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_stx.caseid_admin114(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_stx.caseid_admin115(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_stx.caseid_admin116(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_stx.caseid_admin117(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_stx.caseid_admin118(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_stx.caseid_admin119(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_stx.caseid_admin120(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_stx.caseid_admin121(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_stx.caseid_admin122(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_stx.caseid_admin123(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_stx.caseid_admin124(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_stx.caseid_admin125(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_stx.caseid_admin126(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_stx.caseid_admin127(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_stx.caseid_admin128(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_stx.caseid_admin129(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip01':
                        caseid_stx.caseid_PartnerTrip01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip02':
                        caseid_stx.caseid_PartnerTrip02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip03':
                        caseid_stx.caseid_PartnerTrip03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip04':
                        caseid_stx.caseid_PartnerTrip04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip05':
                        caseid_stx.caseid_PartnerTrip05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'PartnerTrip06':
                        caseid_stx.caseid_PartnerTrip06(self)
                except:
                    module_other_stx.swich_tab_0()




def accounting(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_stx.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 780
    while (rownum < 920):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_stx.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Accounting01':
                        caseid_stx.caseid_accounting01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting02':
                        caseid_stx.caseid_accounting02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting03':
                        caseid_stx.caseid_accounting03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting04':
                        caseid_stx.caseid_accounting04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting05':
                        caseid_stx.caseid_accounting05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting06':
                        caseid_stx.caseid_accounting06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting07':
                        caseid_stx.caseid_accounting07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting08':
                        caseid_stx.caseid_accounting08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting09':
                        caseid_stx.caseid_accounting09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting10':
                        caseid_stx.caseid_accounting10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting11':
                        caseid_stx.caseid_accounting11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting12':
                        caseid_stx.caseid_accounting12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting13':
                        caseid_stx.caseid_accounting13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting14':
                        caseid_stx.caseid_accounting14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting15':
                        caseid_stx.caseid_accounting15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting16':
                        caseid_stx.caseid_accounting16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting17':
                        caseid_stx.caseid_accounting17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting18':
                        caseid_stx.caseid_accounting18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting19':
                        caseid_stx.caseid_accounting19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting20':
                        caseid_stx.caseid_accounting20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting21':
                        caseid_stx.caseid_accounting21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting22':
                        caseid_stx.caseid_accounting22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting23':
                        caseid_stx.caseid_accounting23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting24':
                        caseid_stx.caseid_accounting24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting25':
                        caseid_stx.caseid_accounting25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting26':
                        caseid_stx.caseid_accounting26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting27':
                        caseid_stx.caseid_accounting27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting28':
                        caseid_stx.caseid_accounting28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting29':
                        caseid_stx.caseid_accounting29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting30':
                        caseid_stx.caseid_accounting30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting31':
                        caseid_stx.caseid_accounting31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting32':
                        caseid_stx.caseid_accounting32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting33':
                        caseid_stx.caseid_accounting33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting34':
                        caseid_stx.caseid_accounting34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting35':
                        caseid_stx.caseid_accounting35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting36':
                        caseid_stx.caseid_accounting36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting37':
                        caseid_stx.caseid_accounting37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting38':
                        caseid_stx.caseid_accounting38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting39':
                        caseid_stx.caseid_accounting39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting40':
                        caseid_stx.caseid_accounting40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting41':
                        caseid_stx.caseid_accounting41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting42':
                        caseid_stx.caseid_accounting42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting43':
                        caseid_stx.caseid_accounting43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting44':
                        caseid_stx.caseid_accounting44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting45':
                        caseid_stx.caseid_accounting45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting46':
                        caseid_stx.caseid_accounting46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting47':
                        caseid_stx.caseid_accounting47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting48':
                        caseid_stx.caseid_accounting48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting49':
                        caseid_stx.caseid_accounting49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting50':
                        caseid_stx.caseid_accounting50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting51':
                        caseid_stx.caseid_accounting51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting52':
                        caseid_stx.caseid_accounting52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting53':
                        caseid_stx.caseid_accounting53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting54':
                        caseid_stx.caseid_accounting54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting55':
                        caseid_stx.caseid_accounting55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting56':
                        caseid_stx.caseid_accounting56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting57':
                        caseid_stx.caseid_accounting57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting58':
                        caseid_stx.caseid_accounting58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting59':
                        caseid_stx.caseid_accounting59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting60':
                        caseid_stx.caseid_accounting60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting61':
                        caseid_stx.caseid_accounting61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting62':
                        caseid_stx.caseid_accounting62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting63':
                        caseid_stx.caseid_accounting63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting64':
                        caseid_stx.caseid_accounting64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting65':
                        caseid_stx.caseid_accounting65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting66':
                        caseid_stx.caseid_accounting66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting67':
                        caseid_stx.caseid_accounting67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting68':
                        caseid_stx.caseid_accounting68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting69':
                        caseid_stx.caseid_accounting69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting70':
                        caseid_stx.caseid_accounting70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting71':
                        caseid_stx.caseid_accounting71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting72':
                        caseid_stx.caseid_accounting72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting73':
                        caseid_stx.caseid_accounting73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting74':
                        caseid_stx.caseid_accounting74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting75':
                        caseid_stx.caseid_accounting75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting76':
                        caseid_stx.caseid_accounting76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting77':
                        caseid_stx.caseid_accounting77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting78':
                        caseid_stx.caseid_accounting78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting79':
                        caseid_stx.caseid_accounting79(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Accounting01':
                        caseid_stx.caseid_accounting01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting02':
                        caseid_stx.caseid_accounting02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting03':
                        caseid_stx.caseid_accounting03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting04':
                        caseid_stx.caseid_accounting04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting05':
                        caseid_stx.caseid_accounting05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting06':
                        caseid_stx.caseid_accounting06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting07':
                        caseid_stx.caseid_accounting07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting08':
                        caseid_stx.caseid_accounting08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting09':
                        caseid_stx.caseid_accounting09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting10':
                        caseid_stx.caseid_accounting10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting11':
                        caseid_stx.caseid_accounting11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting12':
                        caseid_stx.caseid_accounting12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting13':
                        caseid_stx.caseid_accounting13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting14':
                        caseid_stx.caseid_accounting14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting15':
                        caseid_stx.caseid_accounting15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting16':
                        caseid_stx.caseid_accounting16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting17':
                        caseid_stx.caseid_accounting17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting18':
                        caseid_stx.caseid_accounting18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting19':
                        caseid_stx.caseid_accounting19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting20':
                        caseid_stx.caseid_accounting20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting21':
                        caseid_stx.caseid_accounting21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting22':
                        caseid_stx.caseid_accounting22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting23':
                        caseid_stx.caseid_accounting23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting24':
                        caseid_stx.caseid_accounting24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting25':
                        caseid_stx.caseid_accounting25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting26':
                        caseid_stx.caseid_accounting26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting27':
                        caseid_stx.caseid_accounting27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting28':
                        caseid_stx.caseid_accounting28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting29':
                        caseid_stx.caseid_accounting29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting30':
                        caseid_stx.caseid_accounting30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting31':
                        caseid_stx.caseid_accounting31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting32':
                        caseid_stx.caseid_accounting32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting33':
                        caseid_stx.caseid_accounting33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting34':
                        caseid_stx.caseid_accounting34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting35':
                        caseid_stx.caseid_accounting35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting36':
                        caseid_stx.caseid_accounting36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting37':
                        caseid_stx.caseid_accounting37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting38':
                        caseid_stx.caseid_accounting38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting39':
                        caseid_stx.caseid_accounting39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting40':
                        caseid_stx.caseid_accounting40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting41':
                        caseid_stx.caseid_accounting41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting42':
                        caseid_stx.caseid_accounting42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting43':
                        caseid_stx.caseid_accounting43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting44':
                        caseid_stx.caseid_accounting44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting45':
                        caseid_stx.caseid_accounting45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting46':
                        caseid_stx.caseid_accounting46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting47':
                        caseid_stx.caseid_accounting47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting48':
                        caseid_stx.caseid_accounting48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting49':
                        caseid_stx.caseid_accounting49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting50':
                        caseid_stx.caseid_accounting50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting51':
                        caseid_stx.caseid_accounting51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting52':
                        caseid_stx.caseid_accounting52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting53':
                        caseid_stx.caseid_accounting53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting54':
                        caseid_stx.caseid_accounting54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting55':
                        caseid_stx.caseid_accounting55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting56':
                        caseid_stx.caseid_accounting56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting57':
                        caseid_stx.caseid_accounting57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting58':
                        caseid_stx.caseid_accounting58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting59':
                        caseid_stx.caseid_accounting59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting60':
                        caseid_stx.caseid_accounting60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting61':
                        caseid_stx.caseid_accounting61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting62':
                        caseid_stx.caseid_accounting62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting63':
                        caseid_stx.caseid_accounting63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting64':
                        caseid_stx.caseid_accounting64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting65':
                        caseid_stx.caseid_accounting65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting66':
                        caseid_stx.caseid_accounting66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting67':
                        caseid_stx.caseid_accounting67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting68':
                        caseid_stx.caseid_accounting68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting69':
                        caseid_stx.caseid_accounting69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting70':
                        caseid_stx.caseid_accounting70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting71':
                        caseid_stx.caseid_accounting71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting72':
                        caseid_stx.caseid_accounting72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting73':
                        caseid_stx.caseid_accounting73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting74':
                        caseid_stx.caseid_accounting74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting75':
                        caseid_stx.caseid_accounting75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting76':
                        caseid_stx.caseid_accounting76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting77':
                        caseid_stx.caseid_accounting77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting78':
                        caseid_stx.caseid_accounting78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting79':
                        caseid_stx.caseid_accounting79(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Accounting01':
                        caseid_stx.caseid_accounting01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting02':
                        caseid_stx.caseid_accounting02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting03':
                        caseid_stx.caseid_accounting03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting04':
                        caseid_stx.caseid_accounting04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting05':
                        caseid_stx.caseid_accounting05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting06':
                        caseid_stx.caseid_accounting06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting07':
                        caseid_stx.caseid_accounting07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting08':
                        caseid_stx.caseid_accounting08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting09':
                        caseid_stx.caseid_accounting09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting10':
                        caseid_stx.caseid_accounting10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting11':
                        caseid_stx.caseid_accounting11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting12':
                        caseid_stx.caseid_accounting12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting13':
                        caseid_stx.caseid_accounting13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting14':
                        caseid_stx.caseid_accounting14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting15':
                        caseid_stx.caseid_accounting15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting16':
                        caseid_stx.caseid_accounting16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting17':
                        caseid_stx.caseid_accounting17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting18':
                        caseid_stx.caseid_accounting18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting19':
                        caseid_stx.caseid_accounting19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting20':
                        caseid_stx.caseid_accounting20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting21':
                        caseid_stx.caseid_accounting21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting22':
                        caseid_stx.caseid_accounting22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting23':
                        caseid_stx.caseid_accounting23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting24':
                        caseid_stx.caseid_accounting24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting25':
                        caseid_stx.caseid_accounting25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting26':
                        caseid_stx.caseid_accounting26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting27':
                        caseid_stx.caseid_accounting27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting28':
                        caseid_stx.caseid_accounting28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting29':
                        caseid_stx.caseid_accounting29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting30':
                        caseid_stx.caseid_accounting30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting31':
                        caseid_stx.caseid_accounting31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting32':
                        caseid_stx.caseid_accounting32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting33':
                        caseid_stx.caseid_accounting33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting34':
                        caseid_stx.caseid_accounting34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting35':
                        caseid_stx.caseid_accounting35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting36':
                        caseid_stx.caseid_accounting36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting37':
                        caseid_stx.caseid_accounting37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting38':
                        caseid_stx.caseid_accounting38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting39':
                        caseid_stx.caseid_accounting39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting40':
                        caseid_stx.caseid_accounting40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting41':
                        caseid_stx.caseid_accounting41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting42':
                        caseid_stx.caseid_accounting42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting43':
                        caseid_stx.caseid_accounting43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting44':
                        caseid_stx.caseid_accounting44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting45':
                        caseid_stx.caseid_accounting45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting46':
                        caseid_stx.caseid_accounting46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting47':
                        caseid_stx.caseid_accounting47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting48':
                        caseid_stx.caseid_accounting48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting49':
                        caseid_stx.caseid_accounting49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting50':
                        caseid_stx.caseid_accounting50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting51':
                        caseid_stx.caseid_accounting51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting52':
                        caseid_stx.caseid_accounting52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting53':
                        caseid_stx.caseid_accounting53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting54':
                        caseid_stx.caseid_accounting54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting55':
                        caseid_stx.caseid_accounting55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting56':
                        caseid_stx.caseid_accounting56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting57':
                        caseid_stx.caseid_accounting57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting58':
                        caseid_stx.caseid_accounting58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting59':
                        caseid_stx.caseid_accounting59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting60':
                        caseid_stx.caseid_accounting60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting61':
                        caseid_stx.caseid_accounting61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting62':
                        caseid_stx.caseid_accounting62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting63':
                        caseid_stx.caseid_accounting63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting64':
                        caseid_stx.caseid_accounting64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting65':
                        caseid_stx.caseid_accounting65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting66':
                        caseid_stx.caseid_accounting66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting67':
                        caseid_stx.caseid_accounting67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting68':
                        caseid_stx.caseid_accounting68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting69':
                        caseid_stx.caseid_accounting69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting70':
                        caseid_stx.caseid_accounting70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting71':
                        caseid_stx.caseid_accounting71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting72':
                        caseid_stx.caseid_accounting72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting73':
                        caseid_stx.caseid_accounting73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting74':
                        caseid_stx.caseid_accounting74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting75':
                        caseid_stx.caseid_accounting75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting76':
                        caseid_stx.caseid_accounting76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting77':
                        caseid_stx.caseid_accounting77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting78':
                        caseid_stx.caseid_accounting78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting79':
                        caseid_stx.caseid_accounting79(self)
                except:
                    module_other_stx.swich_tab_0()

        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Accounting01':
                        caseid_stx.caseid_accounting01(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting02':
                        caseid_stx.caseid_accounting02(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting03':
                        caseid_stx.caseid_accounting03(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting04':
                        caseid_stx.caseid_accounting04(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting05':
                        caseid_stx.caseid_accounting05(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting06':
                        caseid_stx.caseid_accounting06(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting07':
                        caseid_stx.caseid_accounting07(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting08':
                        caseid_stx.caseid_accounting08(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting09':
                        caseid_stx.caseid_accounting09(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting10':
                        caseid_stx.caseid_accounting10(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting11':
                        caseid_stx.caseid_accounting11(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting12':
                        caseid_stx.caseid_accounting12(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting13':
                        caseid_stx.caseid_accounting13(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting14':
                        caseid_stx.caseid_accounting14(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting15':
                        caseid_stx.caseid_accounting15(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting16':
                        caseid_stx.caseid_accounting16(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting17':
                        caseid_stx.caseid_accounting17(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting18':
                        caseid_stx.caseid_accounting18(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting19':
                        caseid_stx.caseid_accounting19(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting20':
                        caseid_stx.caseid_accounting20(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting21':
                        caseid_stx.caseid_accounting21(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting22':
                        caseid_stx.caseid_accounting22(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting23':
                        caseid_stx.caseid_accounting23(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting24':
                        caseid_stx.caseid_accounting24(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting25':
                        caseid_stx.caseid_accounting25(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting26':
                        caseid_stx.caseid_accounting26(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting27':
                        caseid_stx.caseid_accounting27(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting28':
                        caseid_stx.caseid_accounting28(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting29':
                        caseid_stx.caseid_accounting29(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting30':
                        caseid_stx.caseid_accounting30(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting31':
                        caseid_stx.caseid_accounting31(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting32':
                        caseid_stx.caseid_accounting32(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting33':
                        caseid_stx.caseid_accounting33(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting34':
                        caseid_stx.caseid_accounting34(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting35':
                        caseid_stx.caseid_accounting35(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting36':
                        caseid_stx.caseid_accounting36(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting37':
                        caseid_stx.caseid_accounting37(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting38':
                        caseid_stx.caseid_accounting38(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting39':
                        caseid_stx.caseid_accounting39(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting40':
                        caseid_stx.caseid_accounting40(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting41':
                        caseid_stx.caseid_accounting41(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting42':
                        caseid_stx.caseid_accounting42(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting43':
                        caseid_stx.caseid_accounting43(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting44':
                        caseid_stx.caseid_accounting44(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting45':
                        caseid_stx.caseid_accounting45(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting46':
                        caseid_stx.caseid_accounting46(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting47':
                        caseid_stx.caseid_accounting47(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting48':
                        caseid_stx.caseid_accounting48(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting49':
                        caseid_stx.caseid_accounting49(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting50':
                        caseid_stx.caseid_accounting50(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting51':
                        caseid_stx.caseid_accounting51(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting52':
                        caseid_stx.caseid_accounting52(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting53':
                        caseid_stx.caseid_accounting53(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting54':
                        caseid_stx.caseid_accounting54(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting55':
                        caseid_stx.caseid_accounting55(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting56':
                        caseid_stx.caseid_accounting56(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting57':
                        caseid_stx.caseid_accounting57(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting58':
                        caseid_stx.caseid_accounting58(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting59':
                        caseid_stx.caseid_accounting59(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting60':
                        caseid_stx.caseid_accounting60(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting61':
                        caseid_stx.caseid_accounting61(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting62':
                        caseid_stx.caseid_accounting62(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting63':
                        caseid_stx.caseid_accounting63(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting64':
                        caseid_stx.caseid_accounting64(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting65':
                        caseid_stx.caseid_accounting65(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting66':
                        caseid_stx.caseid_accounting66(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting67':
                        caseid_stx.caseid_accounting67(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting68':
                        caseid_stx.caseid_accounting68(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting69':
                        caseid_stx.caseid_accounting69(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting70':
                        caseid_stx.caseid_accounting70(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting71':
                        caseid_stx.caseid_accounting71(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting72':
                        caseid_stx.caseid_accounting72(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting73':
                        caseid_stx.caseid_accounting73(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting74':
                        caseid_stx.caseid_accounting74(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting75':
                        caseid_stx.caseid_accounting75(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting76':
                        caseid_stx.caseid_accounting76(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting77':
                        caseid_stx.caseid_accounting77(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting78':
                        caseid_stx.caseid_accounting78(self)
                except:
                    module_other_stx.swich_tab_0()
                try:
                    if case == 'Accounting79':
                        caseid_stx.caseid_accounting79(self)
                except:
                    module_other_stx.swich_tab_0()








