import logging
import var_stx
import time
import json
from selenium.webdriver.common.by import By
import module_other_stx
import login_stx
from seleniumwire.utils import decode as sw_decode
from selenium.webdriver.common.keys import Keys
import os
import shutil
import re
import openpyxl
from xls2xlsx import XLS2XLSX
import minitor_stx
from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
wait = WebDriverWait(var_stx.driver, 10)















def get_info_web():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table table-hover table-bordered']/thead[1]/tr/th[" + n + "]"
        path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[2]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)



def get_info_web1():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table table-hover table-bordered  cf']/thead/tr/th[" + n + "]"
        path_data = "//*[@class='table table-hover table-bordered  cf']/tbody/tr[1]/td[" + n + "]"
        path_data_a = "//*[@class='table table-hover table-bordered  cf']/tbody/tr[1]/td[" + n + "]/a"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            print("ten cot web:" .format(name_colum))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)

            name_data_a = var_stx.driver.find_element(By.XPATH, path_data_a).text
            print("data cot web a:" .format(name_data_a))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data_a)
        except:
            try:
                name_data = var_stx.driver.find_element(By.XPATH, path_data).text
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
                print("data cot web:" .format(name_data))
            except:
                pass
        n = int(n)



def get_info_web2():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table table-hover table-bordered  cf']/thead/tr/th[" + n + "]"
        path_data = "//*[@class='table table-hover table-bordered  cf']/tbody/tr[1]/td[" + n + "]"
        path_data_a2 = "//*[@class='table table-hover table-bordered  cf']/tbody/tr[1]/td[" + n + "]/a[2]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            print("ten cot web:" .format(name_colum))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)

            name_data_a2 = var_stx.driver.find_element(By.XPATH, path_data_a2).text
            print("data cot web a:" .format(name_data_a2))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data_a2)
        except:
            try:
                name_data = var_stx.driver.find_element(By.XPATH, path_data).text
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
                print("data cot web:" .format(name_data))
            except:
                pass
        n = int(n)



def get_info_web3():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table table-hover table-bordered']/thead[1]/tr/th[" + n + "]"
        path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[1]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)



def get_info_web4():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='ag-header-row ag-header-row-column']/div[" + n + "]"
        path_data = "//*[@class='ag-center-cols-container']/div[1]/div[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)




def get_info_web4_skip(name_skip):
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='ag-header-row ag-header-row-column']/div[" + n + "]"
        path_data = "//*[@class='ag-center-cols-container']/div[1]/div[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))
            if name_colum == "//*[text()='"+name_skip+"']":
                scroll_bar = var_stx.driver.find_element(By.XPATH, "//*[@class='ag-body-horizontal-scroll-viewport']")
                ActionChains(var_stx.driver).click_and_hold(scroll_bar).move_by_offset(800, 0).release().perform()
                n = 1

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)



def get_info_web5():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='text-center table table-hover table-bordered  cf']/thead/tr/th[" + n + "]"
        path_data = "//*[@class='text-center table table-hover table-bordered  cf']/tbody/tr[1]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)




def get_info_web6():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='text-center table table-bordered table-sticky']/thead/tr/th[" + n + "]"
        path_data = "//*[@class='text-center table table-bordered table-sticky']/tbody/tr[1]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)



def get_info_web7():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table table-hover table-bordered']/thead[1]/tr/th[" + n + "]"
        path_data = "//*[@class='table table-hover table-bordered']/tbody/tr[1]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)




def get_info_web8():
    var_stx.driver.implicitly_wait(0.05)
    row = 121
    n = 0
    m = 1
    while (n < 50):
        n += 1
        m += 1
        row += 1
        n = str(n)
        m = str(m)
        path_column = "//*[@id='bookingRevenueAppIndex']/div/div[2]/div[2]/div[1]/div[2]/div/div/div[" + m + "]"#2
        path_data = "//*[@id='bookingRevenueAppIndex']/div/div[2]/div[2]/div[4]/div[1]/div[2]/div/div[1]/div[" + n + "]"#1

        path_column1 = "//*[@id='bookingRevenueAppIndex']/div/div[2]/div[2]/div[1]/div[1]/div[1]/div[1]"
        path_data1 = "//*[@id='bookingRevenueAppIndex']/div/div[2]/div[2]/div[4]/div[1]/div[1]/div[1]/div[1]"

        path_column2 = "//*[@id='bookingRevenueAppIndex']/div/div[2]/div[2]/div[1]/div[1]/div[1]/div[2]"
        path_data2 = "//*[@id='bookingRevenueAppIndex']/div/div[2]/div[2]/div[4]/div[1]/div[1]/div[1]/div[2]"
        print(n)
        name_colum1 = var_stx.driver.find_element(By.XPATH, path_column1).text
        name_data1 = var_stx.driver.find_element(By.XPATH, path_data1).text
        name_colum2 = var_stx.driver.find_element(By.XPATH, path_column2).text
        name_data2 = var_stx.driver.find_element(By.XPATH, path_data2).text
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 120, 1, name_colum1)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 120, 2, name_data1)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 121, 1, name_colum2)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", 121, 2, name_data2)

        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            name_data = var_stx.driver.find_element(By.XPATH, path_data).text
            print("ten cot web: " .format(name_colum))
            print("data cot web:" .format(name_data))

            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)
        m = int(m)




def get_info_web9():
    var_stx.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@class='table-inbox-wrap']/table/tbody/tr[1]/th[" + n + "]"
        path_data = "//*[@class='table-inbox-wrap']/table/tbody/tr[2]/td[" + n + "]"
        path_data_a = "//*[@class='table-inbox-wrap']/table/tbody/tr[2]/td[" + n + "]/a"
        print(n)
        try:
            name_colum = var_stx.driver.find_element(By.XPATH, path_column).text
            print("ten cot web:" .format(name_colum))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 1, name_colum)

            name_data_a = var_stx.driver.find_element(By.XPATH, path_data_a).text
            print("data cot web a:" .format(name_data_a))
            var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data_a)
        except:
            try:
                name_data = var_stx.driver.find_element(By.XPATH, path_data).text
                var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, 2, name_data)
                print("data cot web:" .format(name_data))
            except:
                pass
        n = int(n)




def get_info_web_percent(path_column, path_data, row, column):
    var_stx.driver.implicitly_wait(0.05)
    try:
        name_column = var_stx.driver.find_element(By.XPATH, path_column).text
        name_data = var_stx.driver.find_element(By.XPATH, path_data).text
        text_cleaned = re.sub(r'\s*\(.*?\)', '', name_column)
        print(text_cleaned)
        print("data cot web:" .format(name_data))
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row, column, name_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row + 2, column + 2, name_data)
    except:
        pass






def get_info_excel_data(row, sheet):
    row2 = row + 1

    try:
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))

    except:
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))


    # #Đọc check file excel
    bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    print("r0")
    try:
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)
    except:
        var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)

    print("r1")
    row_tamthoi = 119
    for i in bangchucai:
        row_tamthoi += 1
        if str(sheet[str(i + str(row))].value) == "None":
            break

        cloumn = str(i + str(row))
        print("vị trí tên cột excel: ".format(cloumn))

        cloumn2 = str(i + str(row2))
        print("vị trí data cột excel: ".format(cloumn2))

        name_column = str(sheet[cloumn].value)
        print("Tên cột excel: ".format(name_column))

        data_column = str(sheet[cloumn2].value)
        print(data_column)
        print("Data cột excel: ".format(data_column))

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 3, name_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 4, data_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 5, cloumn)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 6, cloumn2)


def get_info_excel_data1(row, sheet):
    row2 = row + 1

    try:
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xls"))

        x2x = XLS2XLSX(var_stx.excelpath + "/baocao_stx.xls")
        x2x.to_xlsx(var_stx.excelpath + "/baocao_stx.xlsx")
    except:
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xls"))

        x2x = XLS2XLSX(var_stx.excelpath + "/baocao_stx.xls")
        x2x.to_xlsx(var_stx.excelpath + "/baocao_stx.xlsx")


    # #Đọc check file excel
    bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    print("r0")
    try:
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)
    except:
        var_stx.driver.find_element(By.XPATH, var_stx.export_excel).click()
        time.sleep(15)
        filename = max([var_stx.excelpath + "\\" + f for f in os.listdir(var_stx.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_stx.excelpath, r"baocao_stx.xlsx"))
        wordbook = openpyxl.load_workbook(var_stx.excelpath + "/baocao_stx.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)

    print("r1")
    row_tamthoi = 119
    for i in bangchucai:
        row_tamthoi += 1
        if str(sheet[str(i + str(row))].value) == "None":
            break

        cloumn = str(i + str(row))
        print("vị trí tên cột excel: ".format(cloumn))

        cloumn2 = str(i + str(row2))
        print("vị trí data cột excel: ".format(cloumn2))

        name_column = str(sheet[cloumn].value)
        print("Tên cột excel: ".format(name_column))

        data_column = str(sheet[cloumn2].value)
        print(data_column)
        print("Data cột excel: ".format(data_column))

        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 3, name_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 4, data_column)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 5, cloumn)
        var_stx.writeData(var_stx.path_luutamthoi, "Sheet1", row_tamthoi, 6, cloumn2)


def check_info_web_excel_data(code, eventname, result, path_module):
    row = 119
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)


    while (row < 150):
        row += 1
        # name_column_web = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 1))
        data_column_web = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 2))
        # name_column_excel = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 3))
        data_column_excel = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 4))
        # location_name_coloumn = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 5))
        location_data_coloumn = str(var_stx.readData(var_stx.path_luutamthoi, 'Sheet1', row, 6))
        print(f" row1: {row}")

        if data_column_excel == "None":
            break


        logging.info("Dữ liệu web:   " + data_column_web)
        logging.info("Dữ liệu excel: " + data_column_excel)
        if data_column_web == data_column_excel:
            logging.info("True")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.error("False")
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
                                      .format(data_column_web, data_column_excel, location_data_coloumn))
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
            logging.error("Dòng: {}".format(location_data_coloumn))


        # if name_column_web in ['Km GPS', 'NL tiêu thụ', 'Số tiền', 'Tiền nợ', 'Cước phí(cước thật của cuốc)', 'Cước xe (trừ của khách)', 'Số dư còn lại']:
        #     print("name vao 2" + name_column_web)
        #     try:
        #         data_column_web = ''.join(re.findall(r'\d+', data_column_web))[:3]
        #         data_column_excel = ''.join(re.findall(r'\d+', data_column_excel))[:3]
        #     except Exception as e:
        #         logging.error(f"Lỗi khi xử lý dữ liệu: {e}")
        #     if data_column_web == data_column_excel:
        #         logging.info("True")
        #         module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Pass")
        #     else:
        #         logging.error("False")
        #         module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
        #                                   .format(data_column_web, data_column_excel, location_data_coloumn))
        #         module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 7, "Fail")
        #         logging.error("Dòng: {}".format(location_data_coloumn))














class report_8_1:


    def report_8_1_0(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_0).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.0 Báo cáo cuốc khách tổng",
                                                  var_stx.title_page1, "8.1.0 Báo cáo cuốc khách tổng", "_BaoCaoCuocKhachTong.png")


    def report_8_1_0_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_0)
        except:
            report_8_1.report_8_1_0(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("03/06/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("03/06/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.Source).click()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.Source_vanglai).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(8)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.list_data2_1)
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("07/08/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("07/08/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
            time.sleep(1.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(8)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.0 Báo cáo cuốc khách tổng",
                                              var_stx.list_data2_1, "", "_BaoCaoCuocKhachTong_TimKiem.png")


    def report_8_1_0_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_0)
        except:
            report_8_1.report_8_1_0_search(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_1)
        except:
            report_8_1.report_8_1_0_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.search_12_button).click()
        time.sleep(7)
        get_info_web7()
        try:
            minitor_stx.get_info_excel1(5, "Sheet 1")
        except:
            var_stx.driver.refresh()
            time.sleep(7)
            var_stx.driver.find_element(By.XPATH, var_stx.Source_vnpay).click()
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
            var_stx.driver.find_element(By.XPATH, var_stx.search_12_button).click()
            time.sleep(7)
            get_info_web7()
            minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.0 Báo cáo cuốc khách tổng")


    def report_8_1_0_excel_full(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_0)
        except:
            report_8_1.report_8_1_0_search(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_list_data2_1)
        except:
            report_8_1.report_8_1_0_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.search_13_button).click()
        time.sleep(10)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.0 Báo cáo cuốc khách tổng",
                                                        "_BaoCaoCuocKhachTong_ExcelFull.xlsx", "_BaoCaoCuocKhachTong_XuatExcelDayDuThongTin.png")






    def report_8_1_1(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_1).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.1 Tổng cuốc khách",
                                                  var_stx.title_page1, "8.1.1 Tổng cuốc khách", "_TongCuocKhach.png")


    def report_8_1_1_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_1)
        except:
            report_8_1.report_8_1_1(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("17/12/2024 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("19/12/2024 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("02/03/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("03/03/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.1 Tổng cuốc khách",
                                              var_stx.table_1_2, "", "_TongCuocKhach_TimKiem.png")


    def report_8_1_1_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_1)
        except:
            report_8_1.report_8_1_1_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.1 Tổng cuốc khách",
                                                        "_TongCuocKhach.xls", "_TongCuocKhach_XuatExcel.png")


        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 120, 1)    #STT
        # get_info_web_percent(var_stx.column1_2, var_stx.data1_2, 121, 1)    #Ngày
        # get_info_web_percent(var_stx.column1_3, var_stx.data1_3, 122, 1)    #Tổng cuốc
        # get_info_web_percent(var_stx.column2_1, var_stx.data1_4_a1, 123, 1) #Nguồn mobile
        # get_info_web_percent(var_stx.column2_2, var_stx.data1_1, 124, 1)    #Nguồn điều hành
        # get_info_web_percent(var_stx.column2_4, var_stx.data1_1, 125, 1)    #Nguồn vãng lai
        # get_info_web_percent(var_stx.column1_7, var_stx.data1_1, 126, 1)    #Nguồn khách vẫy
        # get_info_web_percent(var_stx.column1_8, var_stx.data1_1, 127, 1)    #Nguồn đối tác
        # get_info_web_percent(var_stx.column1_9, var_stx.data1_1, 128, 1)    #Tổng không đề cử
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 129, 1)    #Tổng khách hủy
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 130, 1)    #Tổng lái xe hủy
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 131, 1)    #Tổng thành công
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 132, 1)    #Không có lái xe nhận
        #
        #
        # get_info_excel_data1(5, "Sheet 1")
        # check_info_web_excel_data(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.1 Tổng cuốc khách")





    def report_8_1_4(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_4).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.4 Tổng cuốc khách theo ngày",
                                                  var_stx.title_page1, "8.1.4 Tổng cuốc khách theo ngày", "_TongCuocKhachTheoNgay.png")


    def report_8_1_4_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_4)
        except:
            report_8_1.report_8_1_4(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("28/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_4_search)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("20/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.4 Tổng cuốc khách theo ngày",
                                              var_stx.check_report_8_1_4_search, "", "_TongCuocKhachTheoNgay_TimKiem.png")


    def report_8_1_4_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_4)
        except:
            report_8_1.report_8_1_4_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.4 Tổng cuốc khách theo ngày",
                                                        "_TongCuocKhachTheoNgay.xls", "_TongCuocKhachTheoNgay_XuatExcel.png")


        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 120, 1)    #STT
        # get_info_web_percent(var_stx.column1_2, var_stx.data1_2, 121, 1)    #Ngày
        # get_info_web_percent(var_stx.column1_3, var_stx.data1_3, 122, 1)    #Tổng cuốc
        # get_info_web_percent(var_stx.column2_1, var_stx.data1_4_a1, 123, 1) #Nguồn mobile
        # get_info_web_percent(var_stx.column2_2, var_stx.data1_1, 124, 1)    #Nguồn điều hành
        # get_info_web_percent(var_stx.column2_4, var_stx.data1_1, 125, 1)    #Nguồn vãng lai
        # get_info_web_percent(var_stx.column1_7, var_stx.data1_1, 126, 1)    #Nguồn khách vẫy
        # get_info_web_percent(var_stx.column1_8, var_stx.data1_1, 127, 1)    #Nguồn đối tác
        # get_info_web_percent(var_stx.column1_9, var_stx.data1_1, 128, 1)    #Tổng không đề cử
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 129, 1)    #Tổng khách hủy
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 130, 1)    #Tổng lái xe hủy
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 131, 1)    #Tổng thành công
        # get_info_web_percent(var_stx.column1_1, var_stx.data1_1, 132, 1)    #Không có lái xe nhận
        #
        #
        # get_info_excel_data1(5, "Sheet 1")
        # check_info_web_excel_data(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.1 Tổng cuốc khách")





    def report_8_1_5(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_5).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.5 Tổng cuốc khách theo nguồn",
                                                  var_stx.title_page1, "8.1.5 Tổng cuốc khách theo nguồn", "_TongCuocKhachTheoNguon.png")


    def report_8_1_5_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_5)
        except:
            report_8_1.report_8_1_5(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("28/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("20/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.5 Tổng cuốc khách theo nguồn",
                                              var_stx.table_1_2, "", "_TongCuocKhachTheoNguon_TimKiem.png")


    def report_8_1_5_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_5)
        except:
            report_8_1.report_8_1_5_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)

        get_info_web1()
        minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.5 Tổng cuốc khách theo nguồn")





    def report_8_1_6(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_6).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.6 Tổng cuốc khách theo cấu hình",
                                                  var_stx.title_page1, "8.1.6 Tổng cuốc khách theo cấu hình", "_TongCuocKhachTheoCauHinh.png")


    def report_8_1_6_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_6)
        except:
            report_8_1.report_8_1_6(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("28/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("20/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.6 Tổng cuốc khách theo cấu hình",
                                              var_stx.table_1_2, "", "_TongCuocKhachTheoCauHinh_TimKiem.png")


    def report_8_1_6_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_6)
        except:
            report_8_1.report_8_1_6_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(6)
        get_info_web1()
        minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.6 Tổng cuốc khách theo cấu hình")







    def report_8_1_7(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_7).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.7 Thống kê cuốc khách yêu thích",
                                                  var_stx.title_page1, "8.1.7 Thống kê cuốc khách yêu thích", "_ThongKeCuocKhachYeuThich.png")


    def report_8_1_7_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_7)
        except:
            report_8_1.report_8_1_7(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.7 Thống kê cuốc khách yêu thích",
                                              var_stx.table_1_2, "", "_ThongKeCuocKhachYeuThich_TimKiem.png")


    def report_8_1_7_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_7)
        except:
            report_8_1.report_8_1_7_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(8)
        get_info_web2()
        minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.7 Thống kê cuốc khách yêu thích")







    def report_8_1_8(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_8).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.check_report_8_1_8)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.8 THỐNG KÊ CHI TIẾT CUỐC KHÁCH",
                                                  var_stx.check_report_8_1_8, "THỐNG KÊ CHI TIẾT CUỐC KHÁCH", "_ThongKeChiTietCuocKhach.png")


    def report_8_1_8_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_8)
        except:
            report_8_1.report_8_1_8(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("27/05/2025 04:30")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.sdt).click()
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("27/05/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.sdt).click()
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.8 THỐNG KÊ CHI TIẾT CUỐC KHÁCH",
                                              var_stx.table_1_2, "", "_ThongKeChiTietCuocKhach_TimKiem.png")


    def report_8_1_8_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_8)
        except:
            report_8_1.report_8_1_8_search(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.data1_2)
        except:
            report_8_1.report_8_1_8_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web2()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.8 THỐNG KÊ CHI TIẾT CUỐC KHÁCH")







    def report_8_1_11(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])
        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_11).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.check_report_8_1_11)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.11 Thống kê cuốc khách theo lái xe",
                                                  var_stx.check_report_8_1_11, "8.1.11 Thống kê cuốc khách theo lái xe", "_ThongKeCuocKhachTheoLaiXe.png")


    def report_8_1_11_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_11)
        except:
            report_8_1.report_8_1_11(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("27/05/2025 04:59")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("28/05/2025 23:59")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.11 Thống kê cuốc khách theo lái xe",
                                              var_stx.table_1_2, "", "_ThongKeCuocKhachTheoLaiXe_TimKiem.png")


    def report_8_1_11_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_11)
        except:
            report_8_1.report_8_1_11_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web2()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.11 Thống kê cuốc khách theo lái xe")






    def report_8_1_12(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_12).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.12 Thống kê cuốc khách theo xe",
                                                  var_stx.title_page1, "8.1.12 Thống kê cuốc khách theo xe", "_ThongKeCuocKhachTheoXe.png")


    def report_8_1_12_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_12)
        except:
            report_8_1.report_8_1_12(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/08/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("29/08/2025 23:59")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_5)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("10/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("20/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.12 Thống kê cuốc khách theo xe",
                                              var_stx.table_1_11, "", "_ThongKeCuocKhachTheoXe_TimKiem.png")


    def report_8_1_12_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_12)
        except:
            report_8_1.report_8_1_12_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web2()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.12 Thống kê cuốc khách theo xe")







    def report_8_1_19(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_1).click()
            time.sleep(2)

        var_stx.driver.find_element(By.XPATH, var_stx.report_8_1_19).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.19 Báo cáo tổng cuốc khách Fixed",
                                                  var_stx.title_page1, "8.1.19 Báo cáo tổng cuốc khách Fixed", "_BaoCaoTongQuocKhachFixed.png")


    def report_8_1_19_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_19)
        except:
            report_8_1.report_8_1_19(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.month_before).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.19 Báo cáo tổng cuốc khách Fixed",
                                              var_stx.table_1_2, "", "_BaoCaoTongQuocKhachFixed_TimKiem.png")


    def report_8_1_19_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_19)
        except:
            report_8_1.report_8_1_19_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        get_info_web3()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.19 Báo cáo tổng cuốc khách Fixed")








class report_8_3:


    def report_8_3_1(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_3_1).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.1 Thống kê đề cử cuốc theo lái xe",
                                                  var_stx.title_page1, "8.3.1 Thống kê đề cử cuốc theo lái xe", "_ThongKeDeCuCuocTheoLaiXe.png")


    def report_8_3_1_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_1)
        except:
            report_8_3.report_8_3_1(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("22/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("24/05/2025 00:00")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.1 Thống kê đề cử cuốc theo lái xe",
                                              var_stx.table_1_2, "", "_ThongKeDeCuCuocTheoLaiXe_TimKiem.png")


    def report_8_3_1_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_1)
        except:
            report_8_3.report_8_3_1_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.1 Thống kê đề cử cuốc theo lái xe",
                                                        "_ThongKeDeCuCuocTheoLaiXe.xls", "_ThongKeDeCuCuocTheoLaiXe_XuatExcel.png")


        # get_info_web3()
        # minitor_stx.get_info_excel(5, "Sheet 1")
        # minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.1 Báo cáo cuốc khách - 8.1.19 Báo cáo tổng cuốc khách Fixed")






    def report_8_3_2(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_3_2).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.2 Thống kê chi tiết đề cử theo lái xe",
                                                  var_stx.title_page1, "8.3.2 Thống kê chi tiết đề cử theo lái xe", "_ThongKeChiTietDeCuTheoLaiXe.png")


    def report_8_3_2_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_2)
        except:
            report_8_3.report_8_3_2(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("28/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("28/05/2025 00:09")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("28/05/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("30/05/2025 23:00")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.2 Thống kê chi tiết đề cử theo lái xe",
                                              var_stx.table_1_2, "", "_ThongKeChiTietDeCuTheoLaiXe_TimKiem.png")


    def report_8_3_2_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_2)
        except:
            report_8_3.report_8_3_2_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web1()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.2 Thống kê chi tiết đề cử theo lái xe")






    def report_8_3_3(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_3_3).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.3 Thống kê cuốc điều chỉ định",
                                                  var_stx.title_page1, "8.3.3 Thống kê cuốc điều chỉ định", "_ThongKeCuocDieuChiDinh.png")


    def report_8_3_3_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_3)
        except:
            report_8_3.report_8_3_3(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("11/08/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("13/08/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/08/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("29/08/2025 23:59")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.3 Thống kê cuốc điều chỉ định",
                                              var_stx.table_1_2, "", "_ThongKeChiTietDeCuTheoLaiXe_TimKiem.png")


    def report_8_3_3_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_3)
        except:
            report_8_3.report_8_3_3_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web1()
        minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.3 Thống kê cuốc điều chỉ định")





    def report_8_3_4(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_3_4).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.4 Thống kê chi tiết cuốc điều chỉ định",
                                                  var_stx.title_page1, "8.3.4 Thống kê chi tiết cuốc điều chỉ định", "_ThongKeChiTietCuocDieuChiDinh.png")


    def report_8_3_4_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_4)
        except:
            report_8_3.report_8_3_4(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("11/08/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("13/08/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/08/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("29/08/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.4 Thống kê chi tiết cuốc điều chỉ định",
                                              var_stx.table_1_2, "", "_ThongKeChiTietCuocDieuChiDinh_TimKiem.png")


    def report_8_3_4_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_4)
        except:
            report_8_3.report_8_3_4_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)
        get_info_web1()
        minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.4 Thống kê chi tiết cuốc điều chỉ định")





    def report_8_3_5(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_3).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_3_5).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.5 Thống kê đề cử lái xe theo tháng",
                                                  var_stx.title_page1, "8.3.5 Thống kê đề cử lái xe theo tháng", "_ThongKeDeCuLaiXeTheoThang.png")


    def report_8_3_5_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_5)
        except:
            report_8_3.report_8_3_5(self, "", "", "")

        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search1).click()
        time.sleep(7)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.5 Thống kê đề cử lái xe theo tháng",
                                              var_stx.listdata2_2, "", "_ThongKeDeCuLaiXeTheoThang_TimKiem.png")


    def report_8_3_5_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_3_5)
        except:
            report_8_3.report_8_3_5_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel4).click()
        time.sleep(7)
        # get_info_web4()
        # minitor_stx.get_info_excel1(5, "Sheet 1")
        # minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.5 Thống kê đề cử lái xe theo tháng")
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.3 Báo cáo đề cử lái xe - 8.3.5 Thống kê đề cử lái xe theo tháng",
                                                        "_ThongKeDeCuLaiXeTheoThang.xlsx", "_ThongKeDeCuLaiXeTheoThang_XuatExcel.png")








class report_8_4:


    def report_8_4_1(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_1).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.1 Báo cáo doanh thu theo lái xe",
                                                  var_stx.title_page1, "8.4.1 Báo cáo doanh thu theo lái xe", "_BaoCaoDoanhThuTheoLaiXe.png")


    def report_8_4_1_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_1)
        except:
            report_8_4.report_8_4_1(self, "", "", "")


        if var_stx.linktest[0:23] == "https://g7test.staxi.vn":
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.reportrange_7day_a).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                time.sleep(7)
                var_stx.driver.find_element(By.XPATH, var_stx.ag_1_2).click()
            except:
                var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.reportrange_30day).click()
                time.sleep(2)
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                time.sleep(7)

            # try:
            #     var_stx.driver.find_element(By.XPATH, var_stx.DriverCode).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            #     time.sleep(0.5)
            #     var_stx.driver.find_element(By.XPATH, var_stx.DriverCode).send_keys("0987")
            #     time.sleep(2)
            #     var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            #     time.sleep(7)
            # except:
            #     try:
            #         var_stx.driver.find_element(By.XPATH, var_stx.ag_1_2)
            #     except:
            #         var_stx.driver.find_element(By.XPATH, var_stx.DriverCode).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            #         time.sleep(0.5)
            #         var_stx.driver.find_element(By.XPATH, var_stx.DriverCode).send_keys("09")
            #         time.sleep(2)
            #         var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            #         time.sleep(7)
            #
            # try:
            #     var_stx.driver.find_element(By.XPATH, var_stx.ag_1_2)
            # except:
            #     try:
            #         var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
            #         time.sleep(2)
            #         var_stx.driver.find_element(By.XPATH, var_stx.yesterday3).click()
            #         time.sleep(2)
            #         var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            #         time.sleep(7)
            #     except:
            #         try:
            #             var_stx.driver.find_element(By.XPATH, var_stx.ag_1_2)
            #         except:
            #             var_stx.driver.refresh()
            #             time.sleep(7)
            #             var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
            #             time.sleep(2)
            #             var_stx.driver.find_element(By.XPATH, var_stx.reportrange_month_before).click()
            #             time.sleep(2)
            #             var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            #             time.sleep(7)
        else:
            var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.reportrange_30day).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.1 Báo cáo doanh thu theo lái xe",
                                              var_stx.ag_1_2, "", "_BaoCaoDoanhThuTheoLaiXe_TimKiem.png")

    def report_8_4_1_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_1)
        except:
            report_8_4.report_8_4_1_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.1 Báo cáo doanh thu theo lái xe",
                                                        "_BaoCaoDoanhThuTheoLaiXe.xlsx", "_BaoCaoDoanhThuTheoLaiXe_XuatExcel.png")

        # get_info_web4()
        # minitor_stx.get_info_excel(5, "Sheet 1")
        # minitor_stx.check_info_web_excel1(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.1 Báo cáo doanh thu theo lái xe")






    def report_8_4_2(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_2).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.2 Chi tiết doanh thu từ app lái xe & định vị",
                                                  var_stx.title_page1, "8.4.2 Chi tiết doanh thu từ app lái xe & định vị", "_ChiTietDoanhThuTuApp.png")


    def report_8_4_2_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_2)
        except:
            report_8_4.report_8_4_2(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(2)
        delete = var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start)
        delete.send_keys(Keys.CONTROL, "a")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("09/09/2025 00:00")
        time.sleep(0.5)

        delete = var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end)
        delete.send_keys(Keys.CONTROL, "a")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("09/09/2025 00:05")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
        time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.ag1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
            time.sleep(2)
            delete = var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start)
            delete.send_keys(Keys.CONTROL, "a")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("10/09/2025 00:00")
            time.sleep(0.5)

            delete = var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end)
            delete.send_keys(Keys.CONTROL, "a")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("10/09/2025 23:05")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(7)


        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.2 Chi tiết doanh thu từ app lái xe & định vị",
                                              var_stx.ag1_2, "", "_ChiTietDoanhThuTuApp_TimKiem.png")


    def report_8_4_2_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_2)
        except:
            report_8_4.report_8_4_2_search(self, "", "", "")

        WebDriverWait(var_stx.driver, 15).until(EC.invisibility_of_element_located((By.CLASS_NAME, "loading-cms")))
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.export_excel5).click()
        time.sleep(10)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.2 Chi tiết doanh thu từ app lái xe & định vị",
                                                        "_ChiTietDoanhThuTuApp.xlsx", "_ChiTietDoanhThuTuApp_XuatExcel.png")
        try:
            not_role = var_stx.driver.find_element(By.XPATH, var_stx.not_role).text
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, not_role)
            var_stx.driver.back()
            time.sleep(4)
        except:
            pass


    def report_8_4_2_excel_full(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_2)
        except:
            report_8_4.report_8_4_2_search(self, "", "", "")

        WebDriverWait(var_stx.driver, 15).until(EC.invisibility_of_element_located((By.CLASS_NAME, "loading-cms")))
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.export_excel_full).click()
        time.sleep(10)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.2 Chi tiết doanh thu từ app lái xe & định vị",
                                                        "_ChiTietDoanhThuTuApp_Full.xlsx", "_ChiTietDoanhThuTuApp_XuatExcelFull.png")

        # try:
        #     var_stx.driver.implicitly_wait(1)
        #     not_role = var_stx.driver.find_element(By.XPATH, var_stx.not_role).text
        #     module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, not_role)
        #     var_stx.driver.back()
        #     time.sleep(4)
        # except:
        #     pass
        # get_info_web4()
        # minitor_stx.get_info_excel1(5, "Sheet 1")
        # minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.2 Chi tiết doanh thu từ app lái xe & định vị")








    def report_8_4_3(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_3).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.3 Thanh toán cuốc khách cho lái xe",
                                                  var_stx.title_page1, "8.4.3 Thanh toán cuốc khách cho lái xe", "_ThanhToanCuocKhachChoLaiXe.png")


    def report_8_4_3_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_3)
        except:
            report_8_4.report_8_4_3(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("25/08/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("26/08/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/07/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("28/08/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.3 Thanh toán cuốc khách cho lái xe",
                                              var_stx.table_1_2, "", "_ThanhToanCuocKhachChoLaiXe_TimKiem.png")


    def report_8_4_3_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_3)
        except:
            report_8_4.report_8_4_3_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web1()
        minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.3 Thanh toán cuốc khách cho lái xe")






    def report_8_4_4(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_4).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.4 Chi tiết thanh toán cuốc khách cho lái xe",
                                                  var_stx.title_page1, "8.4.4 Chi tiết thanh toán cuốc khách cho lái xe", "_ChiTietThanhToanCuocKhachChoLaiXe.png")


    def report_8_4_4_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_4)
        except:
            report_8_4.report_8_4_4(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("28/4/2025 9:22")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("30/5/2025 9:22")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.AbsoluteDriverCode).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.AbsoluteDriverCode).send_keys("bathao")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("1/3/2025 9:22")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("2/6/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.AbsoluteDriverCode).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.AbsoluteDriverCode).send_keys("bathao")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)
        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.4 Chi tiết thanh toán cuốc khách cho lái xe",
                                              var_stx.table_1_2, "", "_ChiTietThanhToanCuocKhachChoLaiXe_TimKiem.png")


    def report_8_4_4_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_4)
        except:
            report_8_4.report_8_4_4_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(8)

        get_info_web1()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.4 Chi tiết thanh toán cuốc khách cho lái xe")






    def report_8_4_5(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_5).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.5 Báo cáo doanh thu theo khách hàng",
                                                  var_stx.title_page1, "8.4.5 Báo cáo doanh thu theo khách hàng", "_BaoCaoDoanhThuTheoKhachHang.png")


    def report_8_4_5_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_5)
        except:
            report_8_4.report_8_4_5(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("20/07/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("01/08/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(5)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("02/04/2025 23:59")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/08/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(5)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.5 Báo cáo doanh thu theo khách hàng",
                                              var_stx.table_1_2, "", "_BaoCaoDoanhThuTheoKhachHang_TimKiem.png")


    def report_8_4_5_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_5)
        except:
            report_8_4.report_8_4_5_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web1()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.5 Báo cáo doanh thu theo khách hàng")






    def report_8_4_6(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_6).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.6 Doanh thu theo nhóm khách hàng",
                                                  var_stx.title_page1, "8.4.6 Doanh thu theo nhóm khách hàng", "_DoanhThuTheoNhomKhachHang.png")


    def report_8_4_6_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_6)
        except:
            report_8_4.report_8_4_6(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("01/05/2025 00:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(6)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("02/04/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("02/06/2025 23:59")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.6 Doanh thu theo nhóm khách hàng",
                                              var_stx.table_1_2, "", "_DoanhThuTheoNhomKhachHang_TimKiem.png")


    def report_8_4_6_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_6)
        except:
            report_8_4.report_8_4_6_search(self, "", "", "")


        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(5)
        get_info_web1()
        minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.6 Doanh thu theo nhóm khách hàng")







    def report_8_4_7(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)

        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_7).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.7 Chi tiết doanh thu từ app lái xe",
                                                  var_stx.title_page1, "8.4.7 Chi tiết doanh thu từ app lái xe", "_ChiTietDoanhThuTuAppLaiXe.png")


    def report_8_4_7_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_7)
        except:
            report_8_4.report_8_4_7(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_start).send_keys("05/08/2025")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.daterangepicker_end).send_keys("05/08/2025")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.apply).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("27/5/2025 14:55")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("28/5/2025 14:55")
            time.sleep(1)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(6)

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_4)
            module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.7 Chi tiết doanh thu từ app lái xe",
                                                  var_stx.table_1_4, "", "_ChiTietDoanhThuTuAppLaiXe_TimKiem.png")
        except:
            module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.7 Chi tiết doanh thu từ app lái xe",
                                                  var_stx.ag1_4, "", "_ChiTietDoanhThuTuAppLaiXe_TimKiem.png")




    def report_8_4_7_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_7)
        except:
            report_8_4.report_8_4_7_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(8)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_4)
            get_info_web1()
        except:
            get_info_web8()

        try:
            minitor_stx.get_info_excel1(5, "Sheet 1")
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
            time.sleep(7)
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.table_1_4)
                get_info_web1()
            except:
                get_info_web8()
            minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.7 Chi tiết doanh thu từ app lái xe")







    def report_8_4_8(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_8).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.8 Chi tiết doanh thu từ định vị",
                                                  var_stx.title_page1, "8.4.8 Chi tiết doanh thu từ định vị", "_ChiTietDoanhThuTuDinhVi.png")


    def report_8_4_8_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_8)
        except:
            report_8_4.report_8_4_8(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("09/09/2025 10:00")
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("09/09/2025 10:09")
        time.sleep(1)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(10)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.table_1_2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("20/07/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys(Keys.CONTROL, "a", Keys.DELETE)
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("21/07/2025 00:00")
            time.sleep(0.5)
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.8 Chi tiết doanh thu từ định vị",
                                              var_stx.table_1_2, "", "_ChiTietDoanhThuTuDinhVi_TimKiem.png")


    def report_8_4_8_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_8)
        except:
            report_8_4.report_8_4_8_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(10)
        module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.8 Chi tiết doanh thu từ định vị",
                                                        "_ChiTietDoanhThuTuDinhVi.xls", "_ChiTietDoanhThuTuDinhVi_XuatExcel.png")

        try:
            no_role = var_stx.driver.find_element(By.XPATH, var_stx.not_role).text
            module_other_stx.writeData(var_stx.checklistpath, "Checklist", code, 6, no_role)
            time.sleep(1)
            var_stx.driver.back()
            time.sleep(4)
        except:
            get_info_web1()
            minitor_stx.get_info_excel1(5, "Sheet 1")
            minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.8 Chi tiết doanh thu từ định vị")







    def report_8_4_17(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_4).click()
            time.sleep(2)
        var_stx.driver.find_element(By.XPATH, var_stx.report_8_4_17).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            var_stx.driver.refresh()
            time.sleep(5)
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.17 Tổng hợp doanh thu lái xe theo ngày",
                                                  var_stx.title_page1, "8.4.17 Tổng hợp doanh thu lái xe theo ngày", "_TongHopDoanhThuLaiXeTheoNgay.png")


    def report_8_4_17_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_17)
        except:
            report_8_4.report_8_4_17(self, "", "", "")

        try:
            var_stx.driver.find_element(By.XPATH, var_stx.search).click()
            time.sleep(4)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_17_search)
        except:
            try:
                var_stx.driver.find_element(By.XPATH, var_stx.FromDate).send_keys(Keys.CONTROL, "a", Keys.DELETE)
                time.sleep(0.5)
                var_stx.driver.find_element(By.XPATH, var_stx.FromDate).send_keys("13/10/2025")
                time.sleep(0.5)
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                time.sleep(4)
                var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_1_17_search)
            except:
                # var_stx.driver.find_element(By.XPATH, var_stx.FromDate).send_keys(Keys.CONTROL, "a", Keys.DELETE)
                # time.sleep(0.5)
                var_stx.driver.find_element(By.XPATH, var_stx.FromDate).send_keys("30/06/2025")
                time.sleep(0.5)
                var_stx.driver.find_element(By.XPATH, var_stx.search).click()
                time.sleep(4)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.17 Tổng hợp doanh thu lái xe theo ngày",
                                              var_stx.check_report_8_1_17_search, "", "_TongHopDoanhThuLaiXeTheoNgay_TimKiem.png")


    def report_8_4_17_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_4_17)
        except:
            report_8_4.report_8_4_17_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(7)

        get_info_web4()
        try:
            minitor_stx.get_info_excel1(5, "Sheet 1")
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
            time.sleep(7)
            minitor_stx.get_info_excel(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.4 Báo cáo doanh thu - 8.4.17 Tổng hợp doanh thu lái xe theo ngày")







class report_8_8:

    def report_8_8_2(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_8).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_8).click()
            time.sleep(2)

        var_stx.driver.find_element(By.XPATH, var_stx.report_8_8_2).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.8 Báo cáo khách hàng - 8.8.2 Chi tiết tin nhắn theo ngày",
                                                  var_stx.title_page1, "8.8.2 Chi tiết tin nhắn theo ngày", "_ChiTietTinNhanTheoNgay.png")

    def report_8_8_2_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_8_2)
        except:
            report_8_8.report_8_8_2(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.reportrange).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.week_before).click()
        time.sleep(1.5)
        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.8 Báo cáo khách hàng - 8.8.2 Chi tiết tin nhắn theo ngày",
                                              var_stx.col_id_date2, "", "_ChiTietTinNhanTheoNgay_TimKiem.png")

    def report_8_8_2_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_8_2)
        except:
            report_8_8.report_8_8_2_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(10)

        get_info_web9()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.8 Báo cáo khách hàng - 8.8.2 Chi tiết tin nhắn theo ngày")
        #
        #
        # module_other_stx.write_result_dowload_file(code, eventname, result, "BÁO CÁO - 8.8 Báo cáo khách hàng - 8.8.2 Chi tiết tin nhắn theo ngày",
        #                                                 "_ChiTietTinNhanTheoNgay.xlsx", "_ChiTietTinNhanTheoNgay_XuatExcel.png")






    def report_8_8_3(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        login_stx.login.login_stx(self, var_stx.data['login']['tk_admin_test'], var_stx.data['login']['mk_admin_test'])

        var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
        time.sleep(2)
        try:
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_8).click()
            time.sleep(2)
        except:
            var_stx.driver.find_element(By.XPATH, var_stx.report1).click()
            time.sleep(2)
            var_stx.driver.find_element(By.XPATH, var_stx.report_8_8).click()
            time.sleep(2)

        var_stx.driver.find_element(By.XPATH, var_stx.report_8_8_3).click()
        time.sleep(2)
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var_stx.title_page1)))
            time.sleep(2)
        except:
            pass
        module_other_stx.write_result_text_try_if(code, eventname, result, "BÁO CÁO - 8.8 Báo cáo khách hàng - 8.8.3 Báo cáo tin nhắn theo ngày",
                                                  var_stx.title_page1, "8.8.3 Báo cáo tin nhắn theo ngày", "_BaoCaoTinNhanTheoNgay.png")

    def report_8_8_3_search(self, code, eventname, result):
        var_stx.driver.implicitly_wait(5)
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_8_3)
        except:
            report_8_8.report_8_8_3(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.from_day).clear()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.from_day).send_keys("05/10/2025")
        time.sleep(0.5)

        var_stx.driver.find_element(By.XPATH, var_stx.to_day).clear()
        time.sleep(0.5)
        var_stx.driver.find_element(By.XPATH, var_stx.to_day).send_keys("14/10/2025")
        time.sleep(0.5)

        var_stx.driver.find_element(By.XPATH, var_stx.search).click()
        time.sleep(7)

        module_other_stx.write_result_text_try_if_other(code, eventname, result, "BÁO CÁO - 8.8 Báo cáo khách hàng - 8.8.3 Báo cáo tin nhắn theo ngày",
                                              var_stx.data_report_2_3, "", "_BaoCaoTinNhanTheoNgay_TimKiem.png")

    def report_8_8_3_excel(self, code, eventname, result):
        var_stx.driver.implicitly_wait(2)
        module_other_stx.delete_excel()
        minitor_stx.clearData_luutamthoi_checkexcel(var_stx.path_luutamthoi, "Sheet1", "", "", "", "", "", "")
        try:
            var_stx.driver.implicitly_wait(2)
            var_stx.driver.find_element(By.XPATH, var_stx.check_report_8_8_3)
        except:
            report_8_8.report_8_8_3_search(self, "", "", "")

        var_stx.driver.find_element(By.XPATH, var_stx.export_excel2).click()
        time.sleep(10)

        get_info_web9()
        minitor_stx.get_info_excel1(5, "Sheet 1")
        minitor_stx.check_info_web_excel(code, eventname, result, "BÁO CÁO - 8.8 Báo cáo khách hàng - 8.8.3 Báo cáo tin nhắn theo ngày")





















